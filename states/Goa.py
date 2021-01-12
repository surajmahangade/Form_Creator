from states import logging,monthdict,Statefolder,read_min_wages_file
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging
from collections import Counter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill, numbers


def Goa(data,contractor_name,contractor_address,filelocation,month,year,report,master):
    Goafilespath = os.path.join(Statefolder,'Goa')
    logging.info('Goa files path is :'+str(Goafilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]
    #wages reg
    input_filelocation=filelocation.split("Registers")[0]
    min_wages_goa=read_min_wages_file("GOA","SEMI-SKILLED",input_filelocation)
    
    def Form_I():

        formIfilepath = os.path.join(Goafilespath,'Form I register of Fine.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formI = data.copy(deep=True)
        data_formI=data_formI.drop_duplicates(subset="Employee Name", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","name&date_of_offence","cause_against_fine","FIXED MONTHLY GROSS",
                                        "Date of payment_fine_released","Date of payment_fine_imposed","remarks"]

        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        data_formI[["name&date_of_offence","cause_against_fine","remarks"]]="---"
        
        data_formI['Fine']=data_formI['Fine'].astype(float)
        data_formI["Date of payment_fine_released"]=data_formI['Date of payment']
        data_formI["Date of payment_fine_imposed"]=data_formI['Date of payment']
        data_formI.loc[data_formI['Fine']==0,["FIXED MONTHLY GROSS","Date of payment_fine_released","Date of payment_fine_imposed","remarks"]]="---"
    
        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 9):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIsheet.row_dimensions[r_idx].height = 20
                else:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIsheet.row_dimensions[r_idx].height = 20

        formIsheet['A4']=formIsheet['A4'].value+" : "+data_formI['Company Name'].unique()[0]
        formIfinalfile = os.path.join(filelocation,'Form I register of Fine.xlsx')
        formIfile.save(filename=formIfinalfile)

    def Form_II():
        formIIfilepath = os.path.join(Goafilespath,'Form II register of damage or loss.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy(deep=True)
        data_formII=data_formII.drop_duplicates(subset="Employee Name", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","attendancefile",
                                        "Damage_loss_with_date","whether_work_showed_cause",
                                        "Date of payment & amount of deduction","num_instalments","Date on which total amount realised","remarks"]

        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII["attendancefile"]="---"
        data_formII[["whether_work_showed_cause","num_instalments"]]="-----"
        data_formII["remarks"]=""
        ######################################
        data_formII['Date of payment']=data_formII['Date of payment'].apply(lambda x: x.strftime('%d-%m-%Y'))
        data_formII['Damage or Loss']=data_formII['Damage or Loss'].astype(float)
        data_formII['Damage or Loss']=data_formII['Damage or Loss'].fillna(0)
        data_formII["Damage_loss_with_date"]=data_formII['Date of payment']+" & "+data_formII['Damage or Loss'].astype(str)
        data_formII.loc[data_formII["Damage or Loss"]==0,"Damage_loss_with_date"]="---"

        data_formII["Date of payment & amount of deduction"]=data_formII["Damage_loss_with_date"]
        data_formII["num_instalments"]="1"
        data_formII["Date on which total amount realised"]=data_formII['Date of payment']
        data_formII.loc[data_formII["Damage or Loss"]==0,["Date on which total amount realised","num_instalments"]]="---"
        ###################################
        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        
        rows = dataframe_to_rows(formII_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 9):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formII):
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formII):
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIIsheet.row_dimensions[r_idx].height = 20
                else:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIIsheet.row_dimensions[r_idx].height = 20

        formIIsheet['A4']=formIIsheet['A4'].value+"  :  "+data_formII['Company Name'].unique()[0]
        formIIfinalfile = os.path.join(filelocation,'Form II register of damage or loss.xlsx')
        formIIfile.save(filename=formIIfinalfile)

    def Form_VIII():
        formVIIIfilepath = os.path.join(Goafilespath,'Form VIII register of Over time.xlsx')
        formVIIIfile = load_workbook(filename=formVIIIfilepath)
        logging.info('Form VIII file has sheet: '+str(formVIIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formVIII = data.copy(deep=True)
        data_formVIII=data_formVIII.drop_duplicates(subset="Employee Name", keep="last")

        data_formVIII['Designation_Dept']=data_formVIII["Designation"]+"_"+data_formVIII["Department"]
        columns=['S.no',"Employee Name","Father's Name","Gender","Designation_Dept","attendancefile",
                                        "extent_of_overtime","total_overtime",
                                        'Normal hrs ','FIXED MONTHLY GROSS',
                                        "overtime rate",'total_earning-overtime',"Overtime",'Total Earning',"date_overtime_paid"]

        data_formVIII['S.no'] = list(range(1,len(data_formVIII)+1))
        data_formVIII[["attendancefile","overtime_rate","ot"]]="---"
        data_formVIII[["extent_of_overtime"]]="---"
        ###
        data_formVIII[['Total\r\nOT Hrs',"Overtime"]]=data_formVIII[['Total\r\nOT Hrs',"Overtime"]].astype(float)
        data_formVIII[['Total\r\nOT Hrs',"Overtime"]]=data_formVIII[['Total\r\nOT Hrs',"Overtime"]].fillna(0)
        
        data_formVIII["total_overtime"]=data_formVIII['Total\r\nOT Hrs']
        
        data_formVIII.loc[data_formVIII['Total\r\nOT Hrs']==0,"total_overtime"]="---"

        data_formVIII['total_earning-overtime']=data_formVIII['Total Earning']-data_formVIII["Overtime"]

        data_formVIII.loc[data_formVIII['Total\r\nOT Hrs']==0,["overtime rate",'total_earning-overtime',"Overtime",'Total Earning']]="---"

        #might need to change
        data_formVIII["date_overtime_paid"]=data_formVIII['Date of payment']
        
        data_formVIII.loc[data_formVIII["Overtime"]==0,"date_overtime_paid"]="---"
        
        formVIII_data=data_formVIII[columns]
        formVIIIsheet = formVIIIfile['Sheet1']
        formVIIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form VIII is ready')

        
        rows = dataframe_to_rows(formVIII_data, index=False, header=False)

        logging.info('rows taken out from data')
        
        row_copy=dataframe_to_rows(formVIII_data, index=False, header=False)
        for i in range(len(list(row_copy))-2):
            i+=12
            formVIIIsheet.merge_cells('C'+str(i)+':D'+str(i))
            formVIIIsheet.merge_cells('F'+str(i)+':H'+str(i))
            formVIIIsheet.merge_cells('I'+str(i)+':K'+str(i))
            formVIIIsheet.merge_cells('L'+str(i)+':N'+str(i))
            formVIIIsheet.merge_cells('O'+str(i)+':R'+str(i))
            formVIIIsheet.merge_cells('S'+str(i)+':T'+str(i))
            formVIIIsheet.merge_cells('U'+str(i)+':V'+str(i))
            formVIIIsheet.merge_cells('W'+str(i)+':X'+str(i))
            formVIIIsheet.merge_cells('Y'+str(i)+':Z'+str(i))
            formVIIIsheet.merge_cells('AA'+str(i)+':AB'+str(i))
            formVIIIsheet.merge_cells('AC'+str(i)+':AD'+str(i))
            formVIIIsheet.merge_cells('AE'+str(i)+':AG'+str(i))
        
        c_idx=0
        row_num=0
        for r_idx, row in enumerate(rows, 10):
            row_iterator=zip(row)
            row_num+=1
            while True:
                c_idx+=1
                if type(formVIIIsheet.cell(row=r_idx, column=c_idx)).__name__ == 'MergedCell':
                    border_sides_thick = Side(style='thick')       
                    border_sides_thin = Side(style='thin')
                    if len(row)==c_idx and row_num==len(data_formVIII):
                        formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                        formVIIIsheet.row_dimensions[r_idx].height = 20
                    elif len(row)==c_idx:
                        formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                        formVIIIsheet.row_dimensions[r_idx].height = 20
                    elif row_num==len(data_formVIII):
                        formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                        formVIIIsheet.row_dimensions[r_idx].height = 20
                    else:
                        formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                        formVIIIsheet.row_dimensions[r_idx].height = 20
                    continue
                try:
                    value=next(row_iterator)[0]
                    
                except:
                    c_idx=0
                    break
                formVIIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formVIIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formVIIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                formVIIIfinalfile = os.path.join(filelocation,'Form VIII register of Over time.xlsx')
                formVIIIfile.save(filename=formVIIIfinalfile)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formVIII):
                    formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formVIIIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formVIIIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formVIII):
                    formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formVIIIsheet.row_dimensions[r_idx].height = 20
                else:
                    formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formVIIIsheet.row_dimensions[r_idx].height = 20
        
        formVIIIsheet['Q4']="Month ending "+month+" "+str(year)
        formVIIIfinalfile = os.path.join(filelocation,'Form VIII register of Over time.xlsx')
        formVIIIfile.save(filename=formVIIIfinalfile)
        
    
    def From_XII():
        formXIIfilepath = os.path.join(Goafilespath,'Form XII Register of leave.xlsx')
        formXIIfile = load_workbook(filename=formXIIfilepath)
        logging.info('Form XII file has sheet: '+str(formXIIfile.sheetnames))
        #print(formXIIfile.sheetnames)
        logging.info('create columns which are now available')

        data_formXII = data.copy(deep=True)
        leave_file_data=data_formXII[["Employee Code","Employee Name","Leave Type","Opening","Monthly Increment","Leave Accrued","Used","Encash","Closing"]]
        
        data_formXII=data_formXII.drop_duplicates(subset="Employee Name", keep="last")

        data_formXII["Employee Name & Code"]=data_formXII["Employee Name"].astype(str)+"||"+data_formXII["Employee Code"].astype(str)

        columns=["Employee Name & Code","Date Joined","Father's Name","Registration_no"]
        data_formXII_columns=list(data_formXII.columns)
        start=data_formXII_columns.index('Arrears salary')
        end=data_formXII_columns.index('Total\r\nDP')
        columns.extend(data_formXII_columns[start+1:end])


        formXII_data=data_formXII[columns]
        formXIIsheet = formXIIfile['Sheet1']

        formXIIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        #for column in  range(ord('A'), ord('G') + 1):
        #    formXIIsheet.unmerge_cells(chr(column)+"11:"+chr(column)+"15")
        formXIIsheet.unmerge_cells("A18:A19")
        formXIIsheet.unmerge_cells("B17:C17")
        formXIIsheet.unmerge_cells("D17:E17")
        formXIIsheet.unmerge_cells("B18:C18")
        formXIIsheet.unmerge_cells("D18:E18")
        formXIIsheet.unmerge_cells("F18:F19")
        formXIIsheet.unmerge_cells("G17:H17")
        formXIIsheet.unmerge_cells("G18:H18")
        formXIIsheet.unmerge_cells("I17:J17")
        formXIIsheet.unmerge_cells("I18:J18")
        
        formXIIsheet.unmerge_cells("A24:A25")
        formXIIsheet.unmerge_cells("B23:C23")
        formXIIsheet.unmerge_cells("D23:E23")
        formXIIsheet.unmerge_cells("B24:C24")
        formXIIsheet.unmerge_cells("D24:E24")
        formXIIsheet.unmerge_cells("F24:F25")
        formXIIsheet.unmerge_cells("G23:H23")
        formXIIsheet.unmerge_cells("G24:H24")
        formXIIsheet.unmerge_cells("I23:J23")
        formXIIsheet.unmerge_cells("I24:J24")

        formXIIsheet.unmerge_cells("A30:A31")
        formXIIsheet.unmerge_cells("B29:C29")
        formXIIsheet.unmerge_cells("B30:C30")
        formXIIsheet.unmerge_cells("D29:E29")
        formXIIsheet.unmerge_cells("D30:E30")
        formXIIsheet.unmerge_cells("F29:G29")
        formXIIsheet.unmerge_cells("F30:G30")

        formXIIsheet.unmerge_cells("E16:F16")
        formXIIsheet.unmerge_cells("E22:F22")
        formXIIsheet.unmerge_cells("C28:D28")
        

        logging.info('data for form I is ready')

        
        #rows_copy = list(dataframe_to_rows(formXII_data, index=False, header=False))
        def cell_write(sheet,r_idx,c_idx,value):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            
        def PL_write(row_index,target,start,end,is_abs_num,name):
            cell_write(target,row_index,1,start)
            cell_write(target,row_index,2,start)
            cell_write(target,row_index,3,end)
            cell_write(target,row_index,4,is_abs_num)
            cell_write(target,row_index,5,start)
            cell_write(target,row_index,6,end)
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==name,:]
            closing=emp_details["Closing"].loc[emp_details["Leave Type"]=="PL"]
            cell_write(target,row_index,7,closing.to_string(index=False))
            cell_write(target,row_index,8,"---")
            cell_write(target,row_index,9,"---")
            cell_write(target,row_index,10,"")
            cell_write(target,row_index,11,"")


        def SL_write(row_index,target,start,end,is_abs_num,name):
            cell_write(target,row_index,1,start)
            cell_write(target,row_index,2,start)
            cell_write(target,row_index,3,end)
            cell_write(target,row_index,4,start)
            cell_write(target,row_index,5,end)
            #balamce due
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==name,:]
            closing=emp_details["Closing"].loc[emp_details["Leave Type"]=="SL"]
            if not closing.empty:
                cell_write(target,row_index,6,closing.to_string(index=False))
            else:
                cell_write(target,row_index,6,"")
            #balance
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==name,:]
            closing=emp_details["Closing"].loc[emp_details["Leave Type"]=="SL"]
            if not closing.empty:
                cell_write(target,row_index,7,closing.to_string(index=False))
            else:
                cell_write(target,row_index,7,"")
            cell_write(target,row_index,8,"")
            cell_write(target,row_index,9,"")
            cell_write(target,row_index,10,"")
            

        def CL_write(row_index,target,start,end,is_abs_num,name):
            cell_write(target,row_index,1,start)
            cell_write(target,row_index,2,start)
            cell_write(target,row_index,3,end)
            cell_write(target,row_index,4,start)
            cell_write(target,row_index,5,end)
            #balamce due
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==name,:]
            closing=emp_details["Closing"].loc[emp_details["Leave Type"]=="CL"]
            if not closing.empty:
                cell_write(target,row_index,6,closing.to_string(index=False))
            else:
                cell_write(target,row_index,6,"")
            #balance
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==name,:]
            closing=emp_details["Closing"].loc[emp_details["Leave Type"]=="CL"]
            if not closing.empty:
                cell_write(target,row_index,7,closing.to_string(index=False))
            else:
                cell_write(target,row_index,7,"")
            cell_write(target,row_index,8,"")
            cell_write(target,row_index,9,"")
            cell_write(target,row_index,10,"")
        
        def ML_write(row_index,target,start,end,is_abs_num,name):
            cell_write(target,row_index,1,start)
            cell_write(target,row_index,2,start)
            cell_write(target,row_index,3,end)
            cell_write(target,row_index,4,start)
            cell_write(target,row_index,5,end)
            cell_write(target,row_index,6,"")
            cell_write(target,row_index,7,"")

        form_write={'PL':PL_write,'SL':SL_write,'CL':CL_write,'ML':ML_write}
        
        def start_end_date_attendance(rows,absent_label,row_offset,initial_offset):  
           # print("infunction",row_offset)
            is_abs_num=0
            row_index=0
            added={}
            for sheet_idx, row in enumerate(rows, 10):
                row_index=0
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        if code =="nan":
                            code=name
                        try:
                            target=formXIIfile[code]
                            added[target.title]=0
                        except:
                            target = formXIIfile.copy_worksheet(formXIIsheet)
                            target.title=code
                            #initial offset
                            row_offset[code]=initial_offset
                            added[target.title]=0
                        target['A4']="Name and address of the Establishment:- "+" "+str(data_formXII['Company Name'].unique()[0])+", "+str(data_formXII['Company Address'].unique()[0])
                        if data_formXII['PE_or_contract'].unique()[0]=="Contractor":
                            target["A5"]="Name of Employer and address:-  "+str(data_formXII['UnitName'].unique()[0])
                        else:
                            target["A5"]="Name of Employer and address:- ---"
                        target['A7']="Name of Employee : "+str(name)
                        added[target.title]=0
                    elif c_idx==2:
                        target['A9']="Date of appointment:- "+str(value)
                    elif c_idx==3:
                        target['A8']="Father's Name : "+str(value)
                    elif c_idx==4:
                        target['A6']="Registration No. :- "+str(value)

                    elif is_abs_num==0 and value==absent_label:
                        is_abs_num=1
                        start=columns[c_idx-1]
                        end=columns[c_idx-1]
                    elif value==absent_label:
                        is_abs_num+=1
                        end=columns[c_idx-1]
                    elif is_abs_num:
                        #target.cell(row=row_index+13, column=1+column_offset, value=is_abs_num)
                     #   print("write",row_index,row_offset,row_index+row_offset[target.title])
                        start_date=start.split("\n")[1].replace("/","-")+"-"+str(year)
                        end_date=end.split("\n")[1].replace("/","-")+"-"+str(year)
                        form_write[absent_label](row_index+row_offset[target.title],target,start_date,end_date,is_abs_num,name)
                        target.insert_rows(row_index+row_offset[target.title]+1)
                        is_abs_num=0
                        row_index+=1
                        added[target.title]+=1

            return added
        offset={}
        initial_offset=14
            
        
        offset=Counter(offset)+Counter(start_end_date_attendance(dataframe_to_rows(formXII_data, index=False, header=False),"PL",offset,initial_offset))
        
        for sheet in formXIIfile.sheetnames:
            offset[sheet]+=20
            initial_offset+=20
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-2)+":A"+str(offset[sheet]-1))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-3)+":C"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-3)+":E"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-2)+":C"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-2)+":E"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("F"+str(offset[sheet]-2)+":F"+str(offset[sheet]-1))
            formXIIfile[sheet].merge_cells("G"+str(offset[sheet]-3)+":H"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("G"+str(offset[sheet]-2)+":H"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("I"+str(offset[sheet]-3)+":J"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("I"+str(offset[sheet]-2)+":J"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-4)+":J"+str(offset[sheet]-4))
            cell_write(sheet=formXIIfile[sheet],r_idx=offset[sheet]-4,c_idx=1,value="Sick Leave")
            
        offset+=Counter(start_end_date_attendance(dataframe_to_rows(formXII_data, index=False, header=False),"SL",offset,initial_offset))
        
        for sheet in formXIIfile.sheetnames:
            offset[sheet]+=6
            initial_offset+=6
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-2)+":A"+str(offset[sheet]-1))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-3)+":C"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-3)+":E"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-2)+":C"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-2)+":E"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("F"+str(offset[sheet]-2)+":F"+str(offset[sheet]-1))
            formXIIfile[sheet].merge_cells("G"+str(offset[sheet]-3)+":H"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("G"+str(offset[sheet]-2)+":H"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("I"+str(offset[sheet]-3)+":J"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("I"+str(offset[sheet]-2)+":J"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-4)+":J"+str(offset[sheet]-4))
            cell_write(sheet=formXIIfile[sheet],r_idx=offset[sheet]-4,c_idx=1,value="Casual Leave")
        
        
        offset+=Counter(start_end_date_attendance(dataframe_to_rows(formXII_data, index=False, header=False),"CL",offset,initial_offset))
        
        for sheet in formXIIfile.sheetnames:
            offset[sheet]+=6
            initial_offset+=6
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-2)+":A"+str(offset[sheet]-1))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-3)+":C"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-2)+":C"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-3)+":E"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-2)+":E"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("F"+str(offset[sheet]-3)+":G"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("F"+str(offset[sheet]-2)+":G"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-4)+":G"+str(offset[sheet]-4))
            cell_write(sheet=formXIIfile[sheet],r_idx=offset[sheet]-4,c_idx=1,value="Maternity Leave")
        offset+=Counter(start_end_date_attendance(dataframe_to_rows(formXII_data, index=False, header=False),"ML",offset,initial_offset))
        formXIIfile.remove(formXIIfile["Sheet1"])
        formXIIfile.remove(formXIIfile["Sheet2"])
        formXIIfile.remove(formXIIfile["Sheet3"])
        formXIIfinalfile = os.path.join(filelocation,'Form XII Register of leave.xlsx')
        formXIIfile.save(filename=formXIIfinalfile)
        
       
    def Form_XXI():
        formXXIfilepath = os.path.join(Goafilespath,'Form XXI Register of Employment.xlsx')
        formXXIfile = load_workbook(filename=formXXIfilepath)
        logging.info('Form XXI file has sheet: '+str(formXXIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formXXI = data.copy(deep=True)
        data_formXXI=data_formXXI.drop_duplicates(subset="Employee Name", keep="last")

        
        columns=['S.no',"Employee Name","Father's Name","Gender","Designation","Date_of_appoinment"]
        
        interval_for_reset_to=data_formXXI.rest_interval.str.split("-",expand=True)[1].unique()[0]
        interval_for_reset_from=data_formXXI.rest_interval.str.split("-",expand=True)[0].unique()[0]
        start_time=data_formXXI["start_time"].unique()[0]
        end_time=data_formXXI["end_time"].unique()[0]

        data_formXXI_columns=list(data_formXXI.columns)
        start=data_formXXI_columns.index('Emp Code')
        end=data_formXXI_columns.index('Total\r\nDP')
        columns.extend(data_formXXI_columns[start+1:end])
        
        less=31-len(data_formXXI_columns[start+1:end])
        for i in range(less):
            columns.extend(["less"+str(i+1)])
            data_formXXI["less"+str(i+1)]=""

        columns.extend(["normal_hours",'Overtime_hrs',"remarks"])
        data_formXXI["Date_of_appoinment"]=data_formXXI['Date Joined']
        data_formXXI["normal_hours"]=len(data_formXXI_columns[start+1:end])-data_formXXI['Total\r\nDP'].astype(float)
        data_formXXI['Overtime_hrs']=data_formXXI['Total\r\nOT Hrs']
        data_formXXI["remarks"]=""
        data_formXXI['S.no'] = list(range(1,len(data_formXXI)+1))

        formXXI_data=data_formXXI[columns]
        formXXIsheet = formXXIfile['Sheet1']
        formXXIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form XXI is ready')

        
        rows = dataframe_to_rows(formXXI_data, index=False, header=False)

        logging.info('rows taken out from data')
        formXXIsheet.unmerge_cells('A23:E23')
        row_num=0
        for r_idx, row in enumerate(rows, 11):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formXXIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formXXI):
                    formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formXXIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formXXIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formXXI):
                    formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formXXIsheet.row_dimensions[r_idx].height = 20
                else:
                    formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formXXIsheet.row_dimensions[r_idx].height = 20
        
        formXXIsheet['AE4']=formXXIsheet['AE4'].value+"   "+str(data_formXXI['Registration_no'].unique()[0])
        formXXIsheet['AG5']=start_time
        formXXIsheet['AK5']=end_time
        formXXIsheet['AG6']="8 hrs"
        #formXXIsheet['AK6']=""
        formXXIsheet['AG7']=interval_for_reset_from
        formXXIsheet['AK7']=interval_for_reset_to
        formXXIsheet['A4']=formXXIsheet['A4'].value+" "+str(data_formXXI['Company Name'].unique()[0])+", "+str(data_formXXI['Company Address'].unique()[0])
        #formXXIsheet['A5']=formXXIsheet['A5'].value+" "+str(data_formXXI['Unit'].unique()[0])+", "+str(data_formXXI['Location'].unique()[0])
        if data_formXXI['PE_or_contract'].unique()[0]=="Contractor":
            formXXIsheet["A5"]="Name of Employer and address:-  "+str(data_formXXI['UnitName'].unique()[0])+", "+str(data_formXXI['Address'].unique()[0])
        else:
            formXXIsheet["A5"]="Name of Employer and address:-  "+"---"
        formXXIfinalfile = os.path.join(filelocation,'Form XXI register of Over time.xlsx')
        formXXIfile.save(filename=formXXIfinalfile)



    def Form_XXIII():
        formXXIIIfilepath = os.path.join(Goafilespath,'Form XXIII Register of wages.xlsx')
        formXXIIIfile = load_workbook(filename=formXXIIIfilepath)
        logging.info('Form XXIII file has sheet: '+str(formXXIIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formXXIII = data.copy(deep=True)
        data_formXXIII=data_formXXIII.drop_duplicates(subset="Employee Name", keep="last")
        
        columns=['S.no',"Employee Name","Father's Name","Designation",'Basic','DA',
                                'Earned Basic','Dearness_Allowance','all_Other_Allowance','Overtime',
                                 'Total Earning','Salary Advance','PF', 'Other_auth_Deduction',
                                 'Total Deductions','Net Paid',"sign",'Date of payment']
        
        data_formXXIII[["sign","remarks"]]=""
        data_formXXIII['Dearness_Allowance']=data_formXXIII['DA']
        
        data_formXXIII["Basic"]=min_wages_goa
        all_other_allowance_columns=['Other Allowance','OtherAllowance1','OtherAllowance2', 'OtherAllowance3', 'OtherAllowance4', 'OtherAllowance5']
        
        data_formXXIII[all_other_allowance_columns]=data_formXXIII[all_other_allowance_columns].astype(float)
        data_formXXIII['all_Other_Allowance']= data_formXXIII.loc[:,all_other_allowance_columns].sum(axis=1)

        Other_auth_Deduction_columns=['Insurance','CSR','ESIC','P.Tax','LWF EE','Loan Deduction','Loan Interest','Other Deduction','TDS',
                                            'OtherDeduction1', 'OtherDeduction2',
                                                    'OtherDeduction3', 'OtherDeduction4', 'OtherDeduction5']
        data_formXXIII[Other_auth_Deduction_columns]=data_formXXIII[Other_auth_Deduction_columns].astype(float)
        data_formXXIII['Other_auth_Deduction']= data_formXXIII.loc[:,Other_auth_Deduction_columns].sum(axis=1)


        data_formXXIII['S.no'] = list(range(1,len(data_formXXIII)+1))

        formXXIII_data=data_formXXIII[columns]
        formXXIIIsheet = formXXIIIfile['Sheet1']
        formXXIIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form XXIII is ready')

        
        rows = dataframe_to_rows(formXXIII_data, index=False, header=False)
        rows_copy = list(dataframe_to_rows(formXXIII_data, index=False, header=False))
        logging.info('rows taken out from data')
        formXXIIIsheet.unmerge_cells('P15:R15')
        formXXIIIsheet["P15"]=""
        row_num=0
        for r_idx, row in enumerate(rows, 10):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                #formXXIIIsheet.cell(row=r_idx, column=c_idx).value=value
                formXXIIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formXXIII):
                    formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formXXIIIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formXXIIIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formXXIII):
                    formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formXXIIIsheet.row_dimensions[r_idx].height = 20
                else:
                    formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formXXIIIsheet.row_dimensions[r_idx].height = 20
        
        
        formXXIIIsheet['P'+str(len(list(rows_copy))+10+5)].value="Signature of Employer"
        
        formXXIIIsheet.merge_cells('P'+str(len(list(rows_copy))+10+5)+':R'+str(len(list(rows_copy))+10+5))
        
        formXXIIIsheet['P4']=formXXIIIsheet['P4'].value+"   "+str(data_formXXIII['Registration_no'].unique()[0])
        formXXIIIsheet['P5']=formXXIIIsheet['P5'].value+"   "+str(month)+" "+str(year)

        if data["PE_or_contract"].unique()[0].upper()=="PE":
            formXXIIIsheet['A4']=" Name of Establishment:-   "+str(data_formXXIII['Company Name'].unique()[0])+" "+str(data_formXXIII['Company Address'].unique()[0])
        else:
            formXXIIIsheet['A4']=" Name of Establishment:-   "+str(data_formXXIII['UnitName'].unique()[0])+" "+str(data_formXXIII['Address'].unique()[0])
            formXXIIIsheet['A5']="Name of Employer and address:-   "+str(data_formXXIII['Contractor_name'].unique()[0])+","+str(data_formXXIII['Contractor_Address'].unique()[0])
        
        
        formXXIIIfinalfile = os.path.join(filelocation,'Form XXIII Register of wages.xlsx')
        formXXIIIfile.save(filename=formXXIIIfinalfile)
    try:  
        Form_I()
        Form_II()
        Form_VIII()
        From_XII()
        Form_XXI()
        Form_XXIII()
    except KeyError as e:
        logging.info("Key error : Check if {} column exsists".format(e))
        print("Key error {}".format(e))
        report.configure(text="Failed: Check input file format  \n column {} not found".format(e))
        master.update()
        raise KeyError