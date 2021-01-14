from states import logging,monthdict,Statefolder
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill, numbers

def Madhya_Pradesh(data,contractor_name,contractor_address,filelocation,month,year,report,master):
    Madhya_Pradeshfilespath = os.path.join(Statefolder,'Madhya Pradesh')
    logging.info('Madhya_Pradesh files path is :'+str(Madhya_Pradeshfilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]

    def Form_I_reg_fine():
        formIfilepath = os.path.join(Madhya_Pradeshfilespath,'Form I register of fine.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')
        
        data_formI = data.copy(deep=True)
        data_formI=data_formI.drop_duplicates(subset="Employee Code", keep="last")

        columns=["S.no","Employee Name","Father's Name","Gender","Department","nature_date_offence","showed_cause_fine","FIXED MONTHLY GROSS",
                                                    "Date of payment","Date of payment","remarks"]
        
        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        
        data_formI[["nature_date_offence","showed_cause_fine"]]="----"
        data_formI["remarks"]=""
        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')
 
        for r_idx, row in enumerate(rows, 7):
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        formIsheet['A4']=formIsheet['A4'].value+" : "+str(data_formI['Unit'].unique()[0])
        
        formIfinalfile = os.path.join(filelocation,'Form I register of fine.xlsx')
        formIfile.save(filename=formIfinalfile)
    
    def Form_I_reg_leave():
        formIfilepath = os.path.join(Madhya_Pradeshfilespath,'Form I register of leave.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')
        
        data_formI = data.copy(deep=True)
        data_formI=data_formI.drop_duplicates(subset="Employee Code", keep="last")

        columns=["Employee Name & Code","Father's Name","Unit_address",'Date Joined',
                                "Acc_balance","Balance_leave","leave_refuse",#"leave_salary_paid"
                                "Salary Advance","return","Date Left",
                                "Date of payment","leave_balance","Used","Closing"]
        
        data_formI["Employee Name & Code"]=data_formI["Employee Name"].astype(str)+"||"+data_formI["Employee Code"].astype(str)
        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        data_formI["Unit_address"]=data_formI['Unit']+", "+data_formI['Address']
        data_formI[['Acc_balance', 'leave_balance', 'Balance_leave', 'return', 'leave_refuse']]=""
        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')
        from string import ascii_uppercase
        for char in ["A","F","G","L","M","N"]:
            formIsheet.unmerge_cells(char+str(15)+':'+char+str(16))
        
        for char in ["BC","DE","HI","JK"]:
            formIsheet.unmerge_cells(char[0]+str(15)+':'+char[1]+str(16))
        row_number=17
        for r_idx, row in enumerate(rows, 16):
            for c_idx, value in enumerate(row, 1):
                value=str(value)
                if c_idx==1:
                    name=value.split("||")[0]
                    code=value.split("||")[1]
                    try:
                        target=formIfile[code]
                    except:
                        target = formIfile.copy_worksheet(formIsheet)
                        target.title=code
                        target["A4"] =target["A4"].value+"  "+value
                elif c_idx==2:
                    target["A5"]=target["A5"].value+" : "+value
                elif c_idx==3:
                    target["A6"]="Name and address of employer/establishment  : "+value
                    target["A6"]=target["A6"].value+"      account for the year : "+" "+str(year)
                elif c_idx==4:
                    target["A8"]=target["A8"].value+" : "+value
                elif c_idx==5:
                    target["A"+str(row_number)]=value
                else:
                    formIsheet.cell(row=row_number, column=c_idx, value=value)
                    target.cell(row=row_number, column=c_idx).font =Font(name ='Bell MT', size =10)
                    target.cell(row=row_number, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=row_number, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        columns=["Employee Name"]
        data_formI_columns=list(data_formI.columns)
        start=data_formI_columns.index('Arrears salary')
        end=data_formI_columns.index('Total\r\nDP')
        columns.extend(data_formI_columns[start+1:end])
        formI_data=data_formI[columns]

        rows = dataframe_to_rows(formI_data, index=False, header=False)        

        def cell_write(sheet,r_idx,c_idx,value):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell IT', size =10)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            
        def start_end_date_attendance(absent_label,offset):  
            is_abs_num=0
            for sheet_idx, row in enumerate(rows, 15):
                row_index=15
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        #try:
                            target=formIfile[value[:31]]
                        #except :
                            #target = formIfile.copy_worksheet(formIsheet)
                            #target.title=value
                    elif is_abs_num==0 and value==absent_label:
                        is_abs_num=1
                        start=columns[c_idx-1]
                        end=columns[c_idx-1]
                    elif value==absent_label:
                        is_abs_num+=1
                        end=columns[c_idx-1]
                    elif is_abs_num:
                        #leave applied
                        cell_write(target,row_index,2+offset,start)
                        cell_write(target,row_index,3+offset,end)
                        #leave granted
                        cell_write(target,row_index,4+offset,start)
                        cell_write(target,row_index,5+offset,end)
                        for i in ["A","F","G","H","I","J","K","L","M","N"]:
                            target[i+str(row_index)]=target[i+"15"].value
                        target.insert_rows(row_index+1)
                        is_abs_num=0
                        row_index+=1
                        
                    
        absent_label="PL"
        column_offset=0           
        start_end_date_attendance(absent_label,column_offset)
        formIfile.remove(formIfile["Sheet1"])
        formIfile.remove(formIfile["Sheet2"])
        formIfile.remove(formIfile["Sheet3"])



        formIfinalfile = os.path.join(filelocation,'Form I register of leave.xlsx')
        formIfile.save(filename=formIfinalfile)
    


    def Form_I_reg_fine_2():
        formIfilepath = os.path.join(Madhya_Pradeshfilespath,'Form I Rgister of fine.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')
        
        data_formI = data.copy(deep=True)
        data_formI=data_formI.drop_duplicates(subset="Employee Code", keep="last")

        columns=["S.no","Employee Name","Father's Name","Department","occupation","act","Designation","against_fine",
                                                    "Net Paid","wages_period","date_amount","date_fine","remarks"]
        
        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        
        data_formI[["occupation","act","against_fine"]]="----"
        data_formI[["wages_period","date_amount","remarks","date_fine"]]=""
        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')
        
        formIsheet.unmerge_cells("A12:M12")
        formIsheet.unmerge_cells("A13:M13")
        formIsheet.insert_rows(8,len(data_formI))

        for r_idx, row in enumerate(rows, 8):
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        formIsheet.merge_cells("A"+str(12+len(data_formI))+":M"+str(12+len(data_formI)))
        formIsheet.merge_cells("A"+str(13+len(data_formI))+":M"+str(13+len(data_formI)))
        
        formIsheet['A4']="Name of the factory : "+str(data_formI['Unit'].unique()[0]) +" Locality "+str(data_formI['Location'].unique()[0])+" District :"+str(data_formI['Location'].unique()[0])
        formIfinalfile = os.path.join(filelocation,'Form I Rgister of fine.xlsx')
        formIfile.save(filename=formIfinalfile)

    def Form_II():
        formIIfilepath = os.path.join(Madhya_Pradeshfilespath,'Form II register of damage or loss.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy(deep=True)
        data_formII = data_formII.drop_duplicates(subset="Employee Code", keep="last")

        columns=['S.no',"Employee Name","Father's Name","Department","occupation","Damage or Loss","cause_deduction","date_amt_deduction",'Total Deductions',
                                                    "num_instalments","date",'Total Deductions',"remarks"]
                                        
        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII[["occupation","cause_deduction","num_instalments"]]="-----"
        data_formII[["remarks","date_amt_deduction","date"]]=""
        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        
        rows = dataframe_to_rows(formII_data, index=False, header=False)

        logging.info('rows taken out from data')

        for r_idx, row in enumerate(rows, 9):
            for c_idx, value in enumerate(row, 1):
                formIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        formIIsheet['A4']=formIIsheet['A4'].value+" : "+data_formII['Unit'].unique()[0]
        formIIfinalfile = os.path.join(filelocation,'Form II register of damage or loss.xlsx')
        formIIfile.save(filename=formIIfinalfile)

    def Form_IV_Overtime():
        formIVfilepath = os.path.join(Madhya_Pradeshfilespath,'Form IV Overtime register.xlsx')
        formIVfile = load_workbook(filename=formIVfilepath)
        logging.info('Form IV file has sheet: '+str(formIVfile.sheetnames))
        logging.info('create columns which are now available')

        data_formIV = data.copy(deep=True)
        data_formIV = data_formIV.drop_duplicates(subset="Employee Code", keep="last")

        columns=['S.no',"Employee Name","Father's Name","Gender","Designation_Dept","Date_overtime_worked",
                                        "Extent of over-time","Total over-time","Normal hrs ",
                                        "FIXED MONTHLY GROSS","overtime rate","Overtime","ot","FIXED MONTHLY GROSS","Date of payment"]
        
        data_formIV['S.no'] = list(range(1,len(data_formIV)+1))
        data_formIV['Designation_Dept']=data_formIV["Designation"]+"_"+data_formIV["Department"]
        data_formIV[["Extent of over-time","Total over-time"]]="-----"
        data_formIV["ot"]=""
        data_formIV["Date_overtime_worked"]=""
        data_formIV["Date of payment& amount of deduction"]=data_formIV["Date of payment"].astype(str)+"\n"+data_formIV["Total Deductions"].astype(str)
        formIV_data=data_formIV[columns]
        formIVsheet = formIVfile['Sheet1']
        formIVsheet.sheet_properties.pageSetUpPr.fitToPage = True
        
        logging.info('data for form IV is ready')

        
        rows = dataframe_to_rows(formIV_data, index=False, header=False)

        logging.info('rows taken out from data')

        for r_idx, row in enumerate(rows, 8):
            for c_idx, value in enumerate(row, 1):
                formIVsheet.cell(row=r_idx, column=c_idx, value=value)
                formIVsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIVsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
    
        formIVsheet['A4']=formIVsheet['A4'].value+" : "+month
        formIVsheet['A6']="Name of the Establishment : "+data_formIV['Unit'].unique()[0]+","+str(data_formIV['Address'].unique()[0])
        formIVfinalfile = os.path.join(filelocation,'Form IV Overtime register.xlsx')
        formIVfile.save(filename=formIVfinalfile)

    def Form_J():
        formJfilepath = os.path.join(Madhya_Pradeshfilespath,'Form J leave book.xlsx')
        formJfile = load_workbook(filename=formJfilepath)
        logging.info('Form J file has sheet: '+str(formJfile.sheetnames))
        logging.info('create columns which are now available')
        
        data_formJ = data.copy(deep=True)
        data_formJ = data_formJ.drop_duplicates(subset="Employee Code", keep="last")

        columns=["Employee Name","Father's Name","Unit_address",'Date Joined',
                                "Acc_balance","Balance_leave","leave_refuse","leave_salary_paid","Date Left",
                                "Date of payment","Opening","Used","Closing"]

        data_formJ['S.no'] = list(range(1,len(data_formJ)+1))
        data_formJ["Unit_address"]=data_formJ['Unit']+", "+data_formJ['Address']
        data_formJ[['Acc_balance',"leave_salary_paid", 'Balance_leave', 'leave_refuse']]=""
        formJ_data=data_formJ[columns]
        formJsheet = formJfile['Sheet1']
        formJsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form J is ready')

        
        rows = dataframe_to_rows(formJ_data, index=False, header=False)

        logging.info('rows taken out from data')
        from string import ascii_uppercase
        
        for char in ["BC","DE","IJ"]:
            formJsheet.unmerge_cells(char[0]+str(17)+':'+char[1]+str(17))
        row_number=17
        for r_idx, row in enumerate(rows, 16):
            for c_idx, value in enumerate(row, 1):
                if c_idx==1:
                    try:
                        target=formJfile[value[:31]]
                    except:
                        target = formJfile.copy_worksheet(formJsheet)
                        target.title=value[:31] 
                        target["A6"] =target["A6"].value+"  "+value
                elif c_idx==2:
                    target["A7"]=target["A7"].value+" : "+value
                elif c_idx==3:
                    target["A8"]="Name and address of employer/establishment  : "+value
                    target["A8"]=target["A8"].value+"      account for the year : "+" "+str(year)
                elif c_idx==4:
                    target["A10"]=target["A10"].value+" : "+value
                elif c_idx==5:
                    target["A"+str(row_number)]=value
                else:
                    formJsheet.cell(row=row_number, column=c_idx, value=value)
                    target.cell(row=row_number, column=c_idx).font =Font(name ='Bell MT', size =10)
                    target.cell(row=row_number, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=row_number, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        columns=["Employee Name"]
        data_formJ_columns=list(data_formJ.columns)
        start=data_formJ_columns.index('Arrears salary')
        end=data_formJ_columns.index('Total\r\nDP')
        columns.extend(data_formJ_columns[start+1:end])
        formJ_data=data_formJ[columns]

        rows = dataframe_to_rows(formJ_data, index=False, header=False)        

        def cell_write(sheet,r_idx,c_idx,value):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell IT', size =10)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            
        def start_end_date_attendance(absent_label,offset):  
            is_abs_num=0
            for sheet_idx, row in enumerate(rows, 15):
                row_index=17
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        #try:
                            target=formJfile[value[:31]]
                        #except :
                            #target = formJfile.copy_worksheet(formJsheet)
                            #target.title=value
                    elif is_abs_num==0 and value==absent_label:
                        is_abs_num=1
                        start=columns[c_idx-1]
                        end=columns[c_idx-1]
                    elif value==absent_label:
                        is_abs_num+=1
                        end=columns[c_idx-1]
                    elif is_abs_num:
                        #leave applied
                        cell_write(target,row_index,2+offset,start)
                        cell_write(target,row_index,3+offset,end)
                        #leave granted
                        cell_write(target,row_index,4+offset,start)
                        cell_write(target,row_index,5+offset,end)
                        for i in ["A","F","G","H","I","J","K","L","M","N"]:
                            target[i+str(row_index)]=target[i+"15"].value
                        target.insert_rows(row_index+1)
                        is_abs_num=0
                        row_index+=1
                        
                    
        absent_label="PL"
        column_offset=0           
        start_end_date_attendance(absent_label,column_offset)
        formJfile.remove(formJfile["Sheet1"])
        formJfile.remove(formJfile["Sheet2"])
        formJfile.remove(formJfile["Sheet3"])



        formJfinalfile = os.path.join(filelocation,'Form J leave book.xlsx')
        formJfile.save(filename=formJfinalfile)



    def Form_N():
        
        formNfilepath = os.path.join(Madhya_Pradeshfilespath,'Form N register of wages.xlsx')
        formNfile = load_workbook(filename=formNfilepath)
        logging.info('Form N file has sheet: '+str(formNfile.sheetnames))
        logging.info('create columns which are now available')

        data_formN = data.copy(deep=True)
        data_formN = data_formN.drop_duplicates(subset="Employee Code", keep="last")

        columns=["Employee Name",'Unit_address',"Father's Name","Age","emp_address",
                    "Designation","rate_wages","Date Joined","Date Left",
                    
                    "Date Joined",'interval_for_reset_from','interval_for_reset_to',
                    "Date Left","ot_from_hrs","ot_tohrs","ot","Earned Basic",
                    "Dearness_Allowance","HRA","Telephone Reimb","Bonus","Fuel Reimb",
                    "Corp Attire Reimb","CCA","ot",'Salary Advance',"Date of payment",
                    "amt_recovered","balance","Fine","Total Deductions","Net Paid",
                    "sign","remarks"]
                                        
        data_formN['interval_for_reset_to']=data_formN.rest_interval.str.split("-",expand=True)[1]
        data_formN['interval_for_reset_from']=data_formN.rest_interval.str.split("-",expand=True)[0]

        data_formN['S.no'] = list(range(1,len(data_formN)+1))
        data_formN[["sign",'ot', 'balance', 'amt_recovered', 'rate_wages', 'Dearness_Allowance', 'ot_tohrs', 'ot_from_hrs', 'remarks']]=""
        data_formN["Unit_address"]=data_formN['Unit']+", "+data_formN['Address']
        data_formN["emp_address"]=data_formN["Permanent Address 1"].apply(str)+data_formN["Permanent Address 2"].apply(str)+data_formN["Permanent Address 3"].apply(str)+data_formN["Permanent Address 4"].apply(str)
        formN_data=data_formN[columns]
        formNsheet = formNfile['Sheet1']
        formNsheet.sheet_properties.pageSetUpPr.fitToPage = True
        
        logging.info('data for form N is ready')

        
        rows = dataframe_to_rows(formN_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_number=20
        for r_idx, row in enumerate(rows,11):
            for c_idx, value in enumerate(row, 1):
                if c_idx==1:
                    try:
                        target=formNfile[value]
                    except:
                        target = formNfile.copy_worksheet(formNsheet)
                        target.title=value
                        target["A6"]="Name of Employee "+value
                        target['A12']="Wage period  Month "+month+" Year "+str(year)
                        target['A5']="For the month of "+month+" Year "+str(year)
                elif c_idx==2:
                    target["A4"]="Name and/or the address of the Establishment "+str(value)
                elif c_idx==3:
                    target["A7"]="Father’s/Husband’s Name "+str(value)
                elif c_idx==4:
                    target["A8"]="  Age "+str(value)
                elif c_idx==5:
                    target["A9"]=" Address of the Employee "+str(value)
                elif c_idx==6:
                    target["A10"]="Nature of Employment :"+str(value)
                elif c_idx==7:
                    target["A11"]=" Rate of wages "+str(value)
                elif c_idx==8:
                    target["A13"]="Date of appointment "+str(value)
                elif c_idx==9:
                    target["A14"]=" Date of discharge "+str(value)
                else:
                    target.cell(row=row_number, column=c_idx-8, value=value)
                    target.cell(row=row_number, column=c_idx-8).font =Font(name ='Bell MT', size =10)
                    target.cell(row=row_number, column=c_idx-8).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=row_number, column=c_idx-8).border = Border(outline= True, right=border_sides, bottom=border_sides)

        formNfile.remove(formNfile["Sheet1"])
        formNfile.remove(formNfile["Sheet2"])
        formNfile.remove(formNfile["Sheet3"])
        formNfinalfile = os.path.join(filelocation,'Form N register of wages.xlsx')
        formNfile.save(filename=formNfinalfile)

       
    def Form_V():
        formVfilepath = os.path.join(Madhya_Pradeshfilespath,'Form v muster roll.xlsx')
        formVfile = load_workbook(filename=formVfilepath)
        logging.info('Form V file has sheet: '+str(formVfile.sheetnames))
        logging.info('create columns which are now available')

        data_formV = data.copy(deep=True)
        data_formV = data_formV.drop_duplicates(subset="Employee Code", keep="last")

        
        columns=['S.no',"Emp Code","Employee Name","Father's Name","Gender","Designation"]
        
        # data_formV_columns=list(data_formV.columns)
        # start=data_formV_columns.index('Arrears salary')
        # end=data_formV_columns.index('Total\r\nDP')
        columnstotake =[]
        days = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31']
        for day in days:
            for col in data_formV.columns:
                if col[5:7]==day:
                    columnstotake.append(col)
        if len(columnstotake)==28:
            columnstotake.append('29')
            columnstotake.append('30')
            columnstotake.append('31')
            data_formV['29'] = ''
            data_formV['30'] = ''
            data_formV['31'] = ''
            
        elif len(columnstotake)==29:
            columnstotake.append('30')
            columnstotake.append('31')
            data_formV['30'] = ''
            data_formV['31'] = ''

        elif len(columnstotake)==30:
            columnstotake.append('31')
            data_formV['31'] = ''
        
        columns.extend(columnstotake)

        columns.extend(["Total\r\nDP"])
        data_formV['S.no'] = list(range(1,len(data_formV)+1))
        formV_data=data_formV[columns]
        formVsheet = formVfile['Muster']
        formVsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form V is ready')

        
        rows = dataframe_to_rows(formV_data, index=False, header=False)
        formVsheet.unmerge_cells("A10:U10")
        logging.info('rows taken out from data')
        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 1):
                formVsheet.cell(row=r_idx, column=c_idx, value=value)
                formVsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formVsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        formVsheet['A5']=formVsheet['A5'].value+"  "+month+" "+str(year)
        formVsheet['A4']=formVsheet['A4'].value+"  "+str(data_formV['Address'].unique()[0])
        formVsheet['A3']=formVsheet['A3'].value+"  "+str(data_formV['Unit'].unique()[0])

        formVfinalfile = os.path.join(filelocation,'Form v muster roll.xlsx')
        formVfile.save(filename=formVfinalfile)

    try:
        Form_I_reg_fine()
        Form_I_reg_leave()
        Form_I_reg_fine_2()
        Form_II()
        Form_IV_Overtime()
        Form_J()
        Form_N()
        Form_V()
    except KeyError as e:
        logging.info("Key error : Check if {} column exsists".format(e))
        print("Key error {}".format(e))
        report.configure(text="Failed: Check input file format  \n column {} not found".format(e))
        master.update()
        raise KeyError
    except FileNotFoundError as e:
        logging.info("File not found : Check if {} exsists".format(e))
        report.configure(text="Failed: File {} not found".format(e))
        master.update()
        raise FileNotFoundError

    