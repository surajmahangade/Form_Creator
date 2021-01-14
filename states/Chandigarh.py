from states import logging,monthdict,Statefolder
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging


def Chandigarh(data,contractor_name,contractor_address,filelocation,month,year,report,master):
    
    Chandigarhfilespath = os.path.join(Statefolder,'Chandigarh')
    logging.info('Chandigarh files path is :'+str(Chandigarhfilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]
    
    def Form_A():
    
        formAfilepath = os.path.join(Chandigarhfilespath,'Form A.xlsx')
        formAfile = load_workbook(filename=formAfilepath)
        logging.info('Form A file has sheet: '+str(formAfile.sheetnames))
        logging.info('create columns which are now available')

        data_formA = data.copy(deep=True)
        data_formA=data_formA.drop_duplicates(subset="Employee Code", keep="last")
        columns=['S.no',"Employee Name","start_time","end_time",'interval_for_reset_from','interval_for_reset_to']
        
        data_formA['interval_for_reset_to']=data_formA.rest_interval.str.split("-",expand=True)[1]
        data_formA['interval_for_reset_from']=data_formA.rest_interval.str.split("-",expand=True)[0]


        data_formA['S.no'] = list(range(1,len(data_formA)+1))
        formA_data=data_formA[columns]
        formAsheet = formAfile['Sheet1']
        formAsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form A is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formA_data, index=False, header=False)

        logging.info('rows taken out from data')
        formAsheet.delete_rows(15,4)
        formAsheet.insert_rows(15,len(data_formA))
        for r_idx, row in enumerate(rows, 15):
            for c_idx, value in enumerate(row, 1):
                formAsheet.cell(row=r_idx, column=c_idx, value=value)
                formAsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formAsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formAsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        formAsheet['A4']=formAsheet['A4'].value+" : "+data_formA['Unit'].unique()[0]
        formAfinalfile = os.path.join(filelocation,'Form A.xlsx')
        formAfile.save(filename=formAfinalfile)

    def Form_C():# - Form T wages slip

        formPath = os.path.join(Chandigarhfilespath,'Form C.xlsx')
        formfile = load_workbook(filename=formPath)
        logging.info('Form T file has sheet: '+str(formfile.sheetnames))
        logging.info('create columns which are now available')

        form_data = data.copy(deep=True)
        form_data=form_data.drop_duplicates(subset="Employee Code", keep="last")
        
        # select sheet to write in and set page properties
        formsheet = formfile['Sheet1']
        formsheet.sheet_properties.pageSetUpPr.fitToPage = True
        
        def write(sheet,r_idx,c_idx,value):
            
            sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
            sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
            return sheet
            
        # for each employee create a new sheet
        for idx in form_data.index:

            empCode = form_data.loc[idx, "Employee Code"]
            
            # create a new sheet for employee
            new_sheet = formfile.copy_worksheet(formsheet)
            new_sheet.title = empCode

            # popoluate every cell with required employee information
            sheet=cell_write(sheet,value,r_idx,c_idx)

            

            new.cell(row= 4, column=3, value=form_data.loc[idx, "Company Name"])# Name of company
            new.cell(row= 5, column=3, value=form_data.loc[idx, "Employee Name"])  #. name of employee
            new.cell(row= 6, column=3, value=form_data.loc[idx, "Father's Name"])  #. Fathers name
            new.cell(row= 7, column=3, value=form_data.loc[idx, "Designation"]) # employee designation C7
            new.cell(row= 8, column=3, value=form_data.loc[idx, "Date Joined"]) # date joined C8
            new.cell(row= 11, column=2, value=form_data.loc[idx, "Earned Basic"]) # earned basic B11
            new.cell(row= 13, column=2, value=form_data.loc[idx, "HRA"])# House rent allowance B13
            new.cell(row= 14, column=2, value=form_data.loc[idx, "Overtime"]) # overtime wages B14
            new.cell(row= 16, column=2, value=form_data.loc[idx,"Other Allowance"]) # otehr allowance B16
            new.cell(row= 17, column=2, value=form_data.loc[idx, "FIXED MONTHLY GROSS"]) # Gross wages B17
            new.cell(row= 12, column=9, value=form_data.loc[idx, "Insurance"]) # employee state insurance I12
            new.cell(row= 13, column=9, value=form_data.loc[idx, "Other Deduction"]) # other deductions I13
            new.cell(row= 17, column=7, value=form_data.loc[idx, "Net Paid"]) # net paid #G17

            
        # delete sheet1 sheet2, sheet3
        formfile.remove(formfile['Sheet1']) 
        formfile.remove(formfile['Sheet2']) 
        formfile.remove(formfile['Sheet3']) 

        # save file
        formCfinalfile = os.path.join(filelocation,'Form C.xlsx')      
        formfile.save(filename=formCfinalfile)

    try:
        Form_A()
    except KeyError as e:
        logging.info("Key error : Check if {} column exsists".format(e))
        print("Key error {}".format(e))
        report.configure(text="Failed: Check input file format  \n column {} not found".format(e))
        master.update()
        raise KeyError
    except FileNotFoundError as e:
        logging.info("File not found : Check if {} exsists".format(e))
        report.configure(text="Failed: File {} not found".format(e))
        master.update()
        raise FileNotFoundError