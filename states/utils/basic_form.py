import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Alignment, Side,numbers


def get_data(data,columns):
    for column in columns:
        if column not in data.columns:
            data[column]=""
    return data
    
#Create borders 
def create_border(sheet,last_row,last_column,start_row,start_column):
    border_sides_thick = Side(style='thick')       
    border_sides_thin = Side(style='thin')
    for c_idx in range(start_column,last_column):
        sheet.cell(row=last_row, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
    for r_idx in range(start_row,last_row):
        sheet.cell(row=r_idx, column=last_column).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin)    
    sheet.cell(row=last_row, column=last_column).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)

#Check if given value is numeric
def if_num(value):
    import numbers
    return isinstance(value, numbers.Number)

#Write value to required cell
def cell_write(sheet,value,r_idx,c_idx):
    sheet.cell(row=r_idx, column=c_idx, value=value)
    sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =15)
    sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
    if if_num(value):
        sheet.cell(row=r_idx, column=c_idx).number_format= numbers.FORMAT_NUMBER

'''
This function writes data like Contrator name,Unit name which is only written once in the entire file
'''

def write_data_once_per_sheet(data_once_per_sheet,sheet):
    for location in data_once_per_sheet.keys():
        if sheet[location].value==None:
            sheet[location]=data_once_per_sheet[location]
        elif not str(data_once_per_sheet[location]).lower() in ["nan","na"]:
            sheet[location]=sheet[location].value+"  "+str(data_once_per_sheet[location])

'''
This function will create basic forms which will have only one sheet and will keep adding data of each employee one below other
data_once_per_sheet is dict such that key is position and value is actual value to be populated,
only pass in data which is only used once, like company name,
contrator address etc
'''
def combine_columns_of_dataframe(dataframe,columns,delimiter=","):
    dataframe.fillna(value="",inplace=True)
    dataframe["combined"]=""
    for column in columns:
        if str(dataframe[column].dtype)[0:8] == 'datetime':
            dataframe[column]=dataframe[column].apply(lambda x:x.strftime('%d-%m-%y'))
        elif str(dataframe[column].dtype)[0:3]!='str':
            dataframe[column]=dataframe[column].astype(str)
        dataframe["combined"]+=dataframe[column]+delimiter
    return dataframe["combined"]

def sum_columns_of_dataframe(dataframe,columns):
    dataframe.fillna(value=0,inplace=True)
    dataframe["sum"]=0
    for column in columns:
        if not dataframe[column].str.isnumeric().all():
            dataframe[column]=dataframe[column].astype(float)
        dataframe["sum"]+=dataframe[column]
    return dataframe["sum"]


def create_basic_form(filename,to_read,to_write,sheet_name,all_employee_data,
                start_row,start_column,report,master,data_once_per_sheet={}):
    #get path from which blank xl file to read
    file_read=os.path.join(to_read,filename)
    #Check if that file exsists
    if not os.path.exists(file_read):
        raise FileNotFoundError(file_read)
    #Load xl file
    work_book=openpyxl.load_workbook(file_read)
    #Check if the specifies sheet is present or not in that file
    if not sheet_name in work_book.sheetnames:
        raise Exception("Sheet {} not found in file {}".format(sheet_name,file_read))
    #Check if data to be populated once is a dictionary or not
    if not isinstance(data_once_per_sheet,dict):
        raise Exception("data_once_per_sheet should be a dictionary such that key is position and value is column name")
    #Get the sheet
    sheet = work_book[sheet_name]
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    #Convert dataframe to rows , to populate the required information
    rows = dataframe_to_rows(all_employee_data, index=False, header=False)
    #Iterate over to populate the values
    r_idx,c_idx=0,0
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, value in enumerate(row, start_column):
            cell_write(sheet,value,r_idx,c_idx)
    #create borders 
    create_border(sheet,last_row=r_idx,last_column=c_idx,start_row=start_row,start_column=start_column)
    #Write data like company name,unit name etc 
    write_data_once_per_sheet(data_once_per_sheet,sheet)
    #Get path to save file
    file_write = os.path.join(to_write,filename)
    #Save the file
    work_book.save(filename=file_write)
    #Return how many lines were written in the file
    return r_idx


def create_per_employee_form(filename,to_read,to_write,sheet_name,start_row,start_column,
                                report,master,employee_codes,data_once_per_sheet={},all_employee_data=None):
    file_read=os.path.join(to_read,filename)
    if not os.path.exists(file_read):
        raise FileNotFoundError(file_read)
    work_book=openpyxl.load_workbook(file_read)
    if not sheet_name in work_book.sheetnames:
        raise Exception("Sheet {} not found".format(sheet_name))
    
    if not isinstance(data_once_per_sheet,dict):
        raise Exception("data_once_per_sheet should be a dictionary such that key is position and value is column name")

    sheet = work_book[sheet_name]
    
    rows_added=1
    r_idx=start_row
    c_idx=0
    if not all_employee_data==None:
        rows = dataframe_to_rows(all_employee_data, index=False, header=False)
        for row,emp_code in zip(rows,employee_codes):
            sheet=work_book.copy_worksheet(sheet_name)
            sheet.title=emp_code
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            write_data_once_per_sheet(data_once_per_sheet,sheet)    
            for c_idx, value in enumerate(row, start_column):
                cell_write(sheet,value,r_idx,c_idx)
            create_border(sheet,last_row=r_idx,last_column=c_idx,start_row=start_row,start_column=start_column)
    else:
        for emp_code in employee_codes:
            sheet=work_book.copy_worksheet(sheet_name)
            sheet.title=emp_code
            write_data_once_per_sheet(data_once_per_sheet,sheet)

    work_book.remove(sheet_name)
    file_write = os.path.join(to_write,filename)
    work_book.save(filename=file_write)
    
    return rows_added

def get_from_to_attendance():
    pass
