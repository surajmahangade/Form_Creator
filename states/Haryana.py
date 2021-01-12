from states import logging,monthdict,Statefolder
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging
from collections import Counter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill, numbers

def Haryana(data,contractor_name,contractor_address,filelocation,month,year,report,master):
    Haryanafilespath = os.path.join(Statefolder,'Haryana')
    logging.info('Haryana files path is :'+str(Haryanafilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]

    def Form_C():
        formCfilepath = os.path.join(Haryanafilespath,'Form C Register of Employees.xlsx')
        formCfile = load_workbook(filename=formCfilepath)

        logging.info('Form C file has sheet: '+str(formCfile.sheetnames))
        logging.info('create columns which are now available')

        data_formC = data.copy()
        
        columns=["Employee Name","Unit","Father's Name","Age","Designation",
                                            "Date Joined","start_month","start_time","end_time","total_hrs_work",
                                           'interval_for_reset_from','interval_for_reset_to',"rest_total_hrs","total_working_hrs",
                                           "overtime_from","overtime_to","overtime_total","renumeration_due"
                                                ]
        
        data_formC['interval_for_reset_to']=data_formC.rest_interval.str.split("-",expand=True)[1]
        data_formC['interval_for_reset_from']=data_formC.rest_interval.str.split("-",expand=True)[0]
        data_formC[['overtime_total', 'rest_total_hrs', 'start_month', 'overtime_to', 'renumeration_due', 'total_working_hrs', 'total_hrs_work', 'overtime_from']]=""
        formC_data=data_formC[columns]
        formCsheet = formCfile['Sheet1']
        formCsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form A is ready')

        
        rows = dataframe_to_rows(formC_data, index=False, header=False)

        logging.info('rows taken out from data')
        added=0
        #formCsheet.insert_rows(9,len(data_formC))
        row_number=12
        for r_idx, row in enumerate(rows, 9):
            for c_idx, value in enumerate(row, 1):
                if c_idx==1:
                    try:
                        target=formCfile[value[:31]]
                    except:
                        target = formCfile.copy_worksheet(formCsheet)
                        target.title=value[:31]
                        employee_name=value 
                        target["A6"] =target["A6"].value+"  "+value
                        target["A5"] =target["A5"].value+" "+month
                elif c_idx==2:
                    target["A4"]="Name of establishment "+str(value)
                    
                elif c_idx==3:
                    target["A5"]="Year and month : "+str(year)+" "+month+" Name of employee :- "+str(employee_name)
                    target["A5"]=target["A5"].value+"   Father’s/Husband’s name : "+str(value)
                elif c_idx==4:
                    target["A5"]=target["A5"].value+" Age : "+str(value)
                elif c_idx==5:
                    target["A6"]=target["A6"].value+str(value)
                elif c_idx==6:
                    target["A7"]="Whether employed on daily, monthly, contract or piece-rate wages, with rate _______________ Date of appointment: "+value
                elif c_idx==7:
                    #don't wrtite since it is already has numerical values
                    continue
                else:
                    target.cell(row=row_number, column=c_idx-6, value=value)
                    target.cell(row=row_number, column=c_idx-6).font =Font(name ='Bell MT', size =10)
                    target.cell(row=row_number, column=c_idx-6).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=row_number, column=c_idx-6).border = Border(outline= True, right=border_sides, bottom=border_sides)

                    #formCfinalfile = os.path.join(filelocation,'Form C Register of Employees.xlsx')
                    #formCfile.save(filename=formCfinalfile)
        
        columns=["Employee Name"]
        
        #columns=[]
        data_formC_columns=list(data_formC.columns)
        start=data_formC_columns.index('Arrears salary')
        end=data_formC_columns.index('Total\r\nDP')
        columns.extend(data_formC_columns[start+1:end])
        #columns.extend(["Used","remarks"])
        formC_data=data_formC[columns]

        rows = dataframe_to_rows(formC_data, index=False, header=False)        

        def cell_write(sheet,r_idx,c_idx,value):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell IT', size =10)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            
        def start_end_date_attendance(absent_label,offset):  
            is_abs_num=0
            for sheet_idx, row in enumerate(rows, 15):
                row_index=12
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        #try:
                            target=formCfile[value[:31]]
                        #except :
                            #target = formCfile.copy_worksheet(formCsheet)
                            #target.title=value
                    elif is_abs_num==0 and value==absent_label:
                        is_abs_num=1
                        start=columns[c_idx-1]
                        end=columns[c_idx-1]
                    elif value==absent_label:
                        is_abs_num+=1
                        end=columns[c_idx-1]
                    elif is_abs_num:
                        #leave applied
                        cell_write(target,row_index,13,is_abs_num)
                        cell_write(target,row_index,14,start)
                        from string import ascii_uppercase
                        for i in ascii_uppercase[:12]:
                            target[i+str(row_index)]=target[i+"12"].value
                        #target.insert_rows(row_index+1)
                        is_abs_num=0
                        row_index+=1
                                
        absent_label="PL"
        column_offset=0           
        start_end_date_attendance(absent_label,column_offset)
        formCfile.remove(formCfile["Sheet1"])

        formCfinalfile = os.path.join(filelocation,'Form C Register of Employees.xlsx')
        formCfile.save(filename=formCfinalfile)


    def Form_D():
        formDfilepath = os.path.join(Haryanafilespath,'Form D register of wages of employees.xlsx')
        formDfile = load_workbook(filename=formDfilepath)

        logging.info('Form D file has sheet: '+str(formDfile.sheetnames))
        logging.info('create columns which are now available')

        data_formD = data.copy()
        
        columns=["Employee Name","Father's Name","FIXED MONTHLY GROSS",'Arrears salary',"FIXED MONTHLY GROSS",
                                                        "ordinary","Overtime","Days Paid",
                                                        "wages_due_rs","wages_due_p","Total Deductions_rs","Total Deductions_p",
                                                        "advance_made_date","advance_rs","advance_p","payment_rs","payment_p"      
                                                ]
        
        data_formD['interval_for_reset_to']=data_formD.rest_interval.str.split("-",expand=True)[1]
        data_formD['interval_for_reset_from']=data_formD.rest_interval.str.split("-",expand=True)[0]
        data_formD[['payment_rs', 'payment_p', 'Total Deductions_rs', 'advance_made_date', 'wages_due_p', 'advance_p', 'Total Deductions_p', 'ordinary', 'advance_rs', 'wages_due_rs']]="test"
        formD_data=data_formD[columns]
        formDsheet = formDfile['Sheet1']
        formDsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form A is ready')

        
        rows = dataframe_to_rows(formD_data, index=False, header=False)

        logging.info('rows taken out from data')
        added=0
        #formDsheet.insert_rows(9,len(data_formD))
        row_number=15
        for r_idx, row in enumerate(rows, 9):
            for c_idx, value in enumerate(row, 1):
                if c_idx==1:
                    try:
                        target=formDfile[value[:31]]
                    except:
                        target = formDfile.copy_worksheet(formDsheet)
                        target.title=value[:31]
                        employee_name=value 
                        target["A6"] =target["A6"].value+"  "+value
                        target["A5"] =target["A5"].value+" "+month
                elif c_idx==2:
                    target["A5"]="Name of employee  "+employee_name+"  father’s name or husband’s name :-"+value
                elif c_idx==3:
                    target["A6"]="Year     "+str(year)+" month    "+month+"   Wages Fixed :- "+str(value)
                elif c_idx==4:
                    target["A7"]=target["A7"].value+"         "+str(value)
                elif c_idx==5:
                    target["A8"]=target["A8"].value+"         "+str(value)
                elif c_idx==6:
                    target["A9"]=target["A9"].value+"        "+str(value)
                elif c_idx==7:
                    target["A10"]=target["A10"].value+"           "+str(value)
                elif c_idx==8:
                    target["A11"]=target["A11"].value+"          "+str(value)
                else:
                    target.cell(row=row_number, column=c_idx-8, value=value)
                    target.cell(row=row_number, column=c_idx-8).font =Font(name ='Bell MT', size =10)
                    target.cell(row=row_number, column=c_idx-8).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=row_number, column=c_idx-8).border = Border(outline= True, right=border_sides, bottom=border_sides)
        formDfile.remove(formDfile["Sheet1"])
        formDfinalfile = os.path.join(filelocation,'Form D register of wages of employees.xlsx')
        formDfile.save(filename=formDfinalfile)

    def Form_E():
        formEfilepath = os.path.join(Haryanafilespath,'Form E register of deduction.xlsx')
        formEfile = load_workbook(filename=formEfilepath)
        logging.info('Form E file has sheet: '+str(formEfile.sheetnames))
        logging.info('create columns which are now available')

        data_formE = data.copy()
        columns=['S.no',"Employee Name","Parentage","month","FIXED MONTHLY GROSS","Total Deductions","fault_deductions","Date of payment ",
                                                            "cause_against deduction","amt_of_deduction","date_utilization","balance_employer"]
                
        
        data_formE['S.no'] = list(range(1,len(data_formE)+1))
        data_formE[["cause_against deduction","amt_of_deduction","date_utilization","balance_employer","fault_deductions","Parentage"]]="-----"
        data_formE["month"]=month
        formE_data=data_formE[columns]
        formEsheet = formEfile['Sheet1']
        formEsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form E is ready')

        
        rows = dataframe_to_rows(formE_data, index=False, header=False)

        logging.info('rows taken out from data')

        for r_idx, row in enumerate(rows, 9):
            for c_idx, value in enumerate(row, 1):
                formEsheet.cell(row=r_idx, column=c_idx, value=value)
                formEsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formEsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formEsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        formEsheet['A5']="Name of the establishment"+data_formE['Unit'][0]+"  "+ month+ "  Year " +str(year)+"   Acts and omission approved by the authorities"
        formEfinalfile = os.path.join(filelocation,'Form E register of deduction.xlsx')
        formEfile.save(filename=formEfinalfile)


    try:
        Form_C()    
        Form_D()
        Form_E()
    except KeyError as e:
        logging.info("Key error : Check if {} column exsists".format(e))
        print("Key error {}".format(e))
        report.configure(text="Failed: Check input file format  \n column {} not found".format(e))
        master.update()
        raise KeyError    

