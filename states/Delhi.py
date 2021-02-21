from states import logging,monthdict,Statefolder
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill, numbers


def Delhi(data,contractor_name,contractor_address,filelocation,month,year,report,master):
    Delhifilespath = os.path.join(Statefolder,'Delhi')
    logging.info('Goa files path is :'+str(Delhifilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]

    #print(sorted(list(data.columns)))
    def Form_G():
        formGfilepath = os.path.join(Delhifilespath,'Form G.xlsx')
        formGfile = load_workbook(filename=formGfilepath)
        logging.info('Form G file has sheet: '+str(formGfile.sheetnames))
        logging.info('create columns which are now available')

        data_formG = data.copy(deep=True)
        # data_formG.fillna(value=0, inplace=True)

        leave_file_data=data_formG[["Employee Code","Employee Name","Leave Type","Opening","Monthly Increment","Leave Accrued","Used","Encash","Closing"]]
        data_formG=data_formG.drop_duplicates(subset="Employee Code", keep="last")
        

        #Part 1 form
        data_formG["Employee Name & Code"]=data_formG["Employee Name"].astype(str)+"||"+data_formG["Employee Code"].astype(str)
        
        columns=["Employee Name & Code",'Nature of work',"Date","start_time","end_time","interval_for_reset_from","interval_for_reset_to","Total_hrs_worked",
                                            'Total\r\nOT Hrs','Overtime',"CL_Sl","leave_due","leave_availed","Balance","sign","remarks"]

        data_formG["leave_due"]=""
        data_formG["leave_availed"]=""
        data_formG["Balance"]=""
        data_formG["remarks"]="---"
        data_formG["sign"]=""
        for employee_name_leave_file in data_formG["Employee Name"]:
            #opening+monthly increment
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==employee_name_leave_file,:]
            opening_pl=emp_details["Opening"].loc[emp_details["Leave Type"]=="PL"].replace("",0).astype(float)
            mon_inr_pl=emp_details["Monthly Increment"].loc[emp_details["Leave Type"]=="PL"].replace("",0).astype(float)
            leave_due=mon_inr_pl.add(opening_pl,fill_value=0).sum()
            data_formG.loc[data_formG["Employee Name"]==employee_name_leave_file,"leave_due"]=leave_due
            ##############################################################################################################################
            #used
            used_pl=emp_details["Used"].loc[emp_details["Leave Type"]=="PL"].replace("",0).astype(float)
            data_formG.loc[data_formG["Employee Name"]==employee_name_leave_file,"leave_availed"]=used_pl
            #closing
            balance_pl=emp_details["Closing"].loc[emp_details["Leave Type"]=="PL"].replace("",0).astype(float)
            data_formG.loc[data_formG["Employee Name"]==employee_name_leave_file,"Balance"]=balance_pl
            ###############################################################################################################################

            
        
        data_formG["Date"]="01"+"-"+str(month)+"-"+str(year)
        #print(data_formG["Date"])
        
        data_formG['Total_hrs_worked']="8 Hours"
        data_formG["CL_Sl"]=data_formG['Total\r\nCL'].replace("",0).astype(float)+data_formG['Total\r\nSL'].replace("",0).astype(float)

        data_formG["Fine_damage_loss"]=data_formG["Fine"].astype(str)+"\n"+data_formG["Damage or Loss"].astype(str)
        data_formG['interval_for_reset_to']=data_formG.rest_interval.str.split("-",expand=True)[1]
        data_formG['interval_for_reset_from']=data_formG.rest_interval.str.split("-",expand=True)[0]

        data_formG_columns=list(data_formG.columns)
        start=data_formG_columns.index('Emp Code')
        end=data_formG_columns.index('Total\r\nDP')
        start_date=data_formG_columns[start+1]
        end_date=data_formG_columns[end-1]
        start_date=start_date.split("\n")[1].replace("/","-")+"-"+str(year)
        end_date=end_date.split("\n")[1].replace("/","-")+"-"+str(year)

        formG_data=data_formG[columns]
        formGsheet = formGfile['Sheet1']
        formGsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form G is ready')



        
        rows = dataframe_to_rows(formG_data, index=False, header=False)

        logging.info('rows taken out from data')
        added=0
        #print("--------------------------------")
        for r_idx, row in enumerate(rows, 14):
            for c_idx, value in enumerate(row, 1):
                if c_idx==1:
                    try:
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        
                        target=formGfile[code]
                    except:
                        target = formGfile.copy_worksheet(formGsheet)
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        target.title=code
                        target["A8"]="Name of Employee "+name
                        target['A7']="Name of Establishment : "+data_formG['Company Name'].unique()[0]
                        target['A4']="Year:- "+str(year)+"  Month:- "+month
                        target['A5']="Wage Period:- "+start_date+" to  "+end_date
                        target["A10"]="Date of Employment   {}".format(data_formG.loc[data_formG["Employee Name"]==name,'Date Joined'].to_string(index=False))
                        
                elif c_idx==2:
                    target["A9"]="Nature of Work:- "+str(value)
                else:
                   # print(value)
                    target.cell(row=15+added, column=c_idx-2, value=value)
                    target.cell(row=15+added, column=c_idx-2).font =Font(name ='Verdana', size =8)
                    target.cell(row=15+added, column=c_idx-2).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=15+added, column=c_idx-2).border = Border(outline= True, right=border_sides, bottom=border_sides)
                    ###
                    border_sides_thick = Side(style='thick')       
                    border_sides_thin = Side(style='thin')

        #print("--------------------------")

        #Part 2 form
        data_formG["Employee Name & Code"]=data_formG["Employee Name"].astype(str)+"||"+data_formG["Employee Code"].astype(str)
        
        columns=["Employee Name & Code","Earned Basic","Overtime","All_Allowance_sum","Total Earning",
                                            "Fine_damage_loss","all_Other_Deduction_sum","date_of_payment",'Salary Advance',
                                            "Total_ded","Net Paid",'Date of payment',"sign"
                                            ]
        
        data_formG['Salary Advance']=data_formG['Salary Advance'].astype(str)
        data_formG=data_formG.replace({'Salary Advance':{"":"0","0.":"0","0.0":"0","nan":"0"}})
        
        
        get_date_of_payment=data_formG['Salary Advance']!="0"
        data_formG["date_of_payment"]=""
        data_formG["date_of_payment"]=data_formG.loc[get_date_of_payment,'Date of payment']
        all_deductions_columns_name=['HRA','Conveyance','Medical Allowance','Telephone Reimb','Tel and Int Reimb',
                                            'Bonus','Other Allowance', 'Fuel Reimb','Prof Dev Reimb','Corp Attire Reimb',
                                            'Meal Allowance','Special Allowance','Personal Allowance','CCA','Other Reimb',
                                            'Arrears','Other Earning',"Retention Pay",'Variable Pay','Leave Encashment',
                                            'Stipend','Consultancy Fees','Covid Deduction','OtherAllowance1', 
                                            'OtherAllowance2', 'OtherAllowance3', 'OtherAllowance4', 'OtherAllowance5'
                                            ]
        if "Covid Deduction" not in data_formG.columns:
            data_formG["Covid Deduction"]=0
        if "Retention Pay" not in data_formG.columns:
            data_formG["Retention Pay"]=0
            
        data_formG[all_deductions_columns_name]=data_formG[all_deductions_columns_name].replace("",0).astype(float)
        data_formG['All_Allowance_sum']= data_formG.loc[:,all_deductions_columns_name].sum(axis=1)

        data_formG["Fine_damage_loss"]=data_formG['Fine'].replace("",0).astype(float)+data_formG['Damage or Loss'].replace("",0).astype(float)

        other_deductions_columns_name=['Other Deduction','OtherDeduction1', 'OtherDeduction2',
                                                        'OtherDeduction3', 'OtherDeduction4', 'OtherDeduction5']

        data_formG[other_deductions_columns_name]=data_formG[other_deductions_columns_name].replace("",0).astype(float)
        data_formG["all_Other_Deduction_sum"]= data_formG.loc[:,other_deductions_columns_name].sum(axis=1)

        data_formG["Total_ded"]=data_formG["all_Other_Deduction_sum"]-data_formG['Salary Advance'].replace("",0).astype(float)
        
        data_formG["sign"]=""

        formG_data=data_formG[columns]
        formGsheet = formGfile['Sheet1']
        formGfile.remove(formGfile["Sheet1"])
        formGsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form G is ready')



        
        rows = dataframe_to_rows(formG_data, index=False, header=False)

        added=0
        for r_idx, row in enumerate(rows, 28):
            for c_idx, value in enumerate(row, 1):
                if c_idx==1:
                    try:
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        target=formGfile[code]
                    except:
                        target = formGfile.copy_worksheet(formGsheet)
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        
                        target.title=code
                        target["A8"]=target["A8"].value+" "+name
                        target['A7']=target['A7'].value+" : "+data_formG['Company Name'].unique()[0]
                        target['A4']="Year "+str(year)+"Month "+month
                        target['A5']="Wage Period:- "+start_date+"-"+end_date
                        target["A10"]="" #"Date of Employment   {}".format(data_formG.loc[data_formG["Employee Name"]==name,'Date Joined'])
                else:
                    target.cell(row=28+added, column=c_idx-1, value=value)
                    target.cell(row=28+added, column=c_idx-1).font =Font(name ='Verdana', size =8)
                    target.cell(row=28+added, column=c_idx-1).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=28+added, column=c_idx-1).border = Border(outline= True, right=border_sides, bottom=border_sides)
                    border_sides_thick = Side(style='thick')       
                    border_sides_thin = Side(style='thin')


        formGfinalfile = os.path.join(filelocation,'Form G.xlsx')
        formGfile.save(filename=formGfinalfile)



    def Form_H():
        formHfilepath = os.path.join(Delhifilespath,'Form H.xlsx')
        formHfile = load_workbook(filename=formHfilepath)
        logging.info('Form H file has sheet: '+str(formHfile.sheetnames))
        logging.info('create columns which are now available')

        def Part_I():
            data_formH = data.copy(deep=True)
            data_formH=data_formH.drop_duplicates(subset="Employee Code", keep="last")
            columns=['S.no',"Employee Name",'Nature of work']

            # data_formH_columns=list(data_formH.columns)
            
            # start=data_formH_columns.index('Emp Code')
            # end=data_formH_columns.index('Total\r\nDP')
            columnstotake =[]
            days = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31']
            for day in days:
                for col in data_formH.columns:
                    if col[5:7]==day:
                        columnstotake.append(col)
            if len(columnstotake)==28:

                columnstotake.append('29')
                columnstotake.append('30')
                columnstotake.append('31')
                data_formH['29'] = ''
                data_formH['30'] = ''
                data_formH['31'] = ''
                
            elif len(columnstotake)==29:
                columnstotake.append('30')
                columnstotake.append('31')
                data_formH['30'] = ''
                data_formH['31'] = ''

            elif len(columnstotake)==30:
                columnstotake.append('31')
                data_formH['31'] = ''
            
            columns.extend(columnstotake)
                
            
            columns.extend(["remarks"])


            data_formH['S.no'] = list(range(1,len(data_formH)+1))
            data_formH[["remarks"]]=""
            formH_data=data_formH[columns]
            formHsheet = formHfile['Sheet1']
            formHsheet.sheet_properties.pageSetUpPr.fitToPage = True
            logging.info('data for form H is ready')

            
            rows = dataframe_to_rows(formH_data, index=False, header=False)
            rows_copy = list(dataframe_to_rows(formH_data, index=False, header=False))
            

            logging.info('rows taken out from data')
            formHsheet.unmerge_cells("A15:N15")
            formHsheet.unmerge_cells("A18:A19")
            formHsheet.unmerge_cells("B18:B19")
            
            formHsheet.unmerge_cells("C18:G18")
            formHsheet.unmerge_cells("H18:K18")
            formHsheet.unmerge_cells("L18:L19")
            formHsheet.unmerge_cells("M18:M19")
            formHsheet.unmerge_cells("N18:N19")
            
            from string import ascii_uppercase
            for char in ascii_uppercase[:14]:
                formHsheet.unmerge_cells(char+str(20)+':'+char+str(22))
            formHsheet.insert_rows(10,len(rows_copy))
            #formHsheet.delete_rows(18,2)
            formHsheet.merge_cells("C"+str(len(rows_copy)+18)+":G"+str(len(rows_copy)+18))
            formHsheet.merge_cells("H"+str(len(rows_copy)+18)+":K"+str(len(rows_copy)+18))
            formHsheet.merge_cells("A"+str(len(rows_copy)+18)+":A"+str(len(rows_copy)+19))
            formHsheet.merge_cells("B"+str(len(rows_copy)+18)+":B"+str(len(rows_copy)+19))
            formHsheet.merge_cells("L"+str(len(rows_copy)+18)+":L"+str(len(rows_copy)+19))
            formHsheet.merge_cells("M"+str(len(rows_copy)+18)+":M"+str(len(rows_copy)+19))
            formHsheet.merge_cells("N"+str(len(rows_copy)+18)+":N"+str(len(rows_copy)+19))
            formHsheet.merge_cells("A"+str(len(rows_copy)+15)+":N"+str(len(rows_copy)+15))

            row_num=0
            for r_idx, row in enumerate(rows, 10):
                row_num+=1
                for c_idx, value in enumerate(row, 1):
                    formHsheet.cell(row=r_idx, column=c_idx, value=value)
                    formHsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                    formHsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                    border_sides_thick = Side(style='thick')       
                    border_sides_thin = Side(style='thin')
                    if len(row)==c_idx and row_num==len(data_formH):
                        formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                        formHsheet.row_dimensions[r_idx].height = 20
                    elif len(row)==c_idx:
                        formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                        formHsheet.row_dimensions[r_idx].height = 20
                    elif row_num==len(data_formH):
                        formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                        formHsheet.row_dimensions[r_idx].height = 20
                    else:
                        formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                        formHsheet.row_dimensions[r_idx].height = 20
                    
        def Part_II():
            data_formH = data.copy(deep=True)
            data_formH=data_formH.drop_duplicates(subset="Employee Code", keep="last")
            columns=["Employee Name","Designation",'Earned Basic','DA','All_Other_Allowance',
                                'Total Earning','Overtime','Salary Advance',"sal_fine_damage",
                                'All_other_deductions','Total Deductions','Net Paid','sign','Date of payment']

            all_other_allowance_columns=['Other Allowance','OtherAllowance1', 'OtherAllowance2', 'OtherAllowance3', 'OtherAllowance4', 'OtherAllowance5']
            
            data_formH[all_other_allowance_columns]=data_formH[all_other_allowance_columns].replace("",0).astype(float)
            data_formH['All_Other_Allowance']= data_formH.loc[:,all_other_allowance_columns].sum(axis=1)
            data_formH["sal_fine_damage"]=data_formH["Fine"].replace("",0).apply(float)+data_formH["Damage or Loss"].replace("",0).apply(float)


            other_deductions_columns=['Insurance','CSR','PF','ESIC','P.Tax','LWF EE','Loan Deduction','Loan Interest','Other Deduction','TDS']
            data_formH[other_deductions_columns]=data_formH[other_deductions_columns].replace("",0).astype(float)
            data_formH['All_other_deductions']= data_formH.loc[:,other_deductions_columns].sum(axis=1)
            
            data_formH[["remarks",'Amount_Due','sign','Dearness_Allowance']]=""
            formH_data=data_formH[columns]
            formHsheet = formHfile['Sheet1']
            formHsheet.sheet_properties.pageSetUpPr.fitToPage = True
            logging.info('data for form H is ready')

            
            rows = dataframe_to_rows(formH_data, index=False, header=False)
            rows_copy = list(dataframe_to_rows(formH_data, index=False, header=False))
            

            logging.info('rows taken out from data')
            formHsheet.insert_rows(len(rows_copy)+20,len(rows_copy))
            row_num=0
            for r_idx, row in enumerate(rows, len(rows_copy)+20):
                row_num+=1
                for c_idx, value in enumerate(row, 1):
                    formHsheet.cell(row=r_idx, column=c_idx, value=value)
                    formHsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                    formHsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                    border_sides_thick = Side(style='thick')       
                    border_sides_thin = Side(style='thin')
                    if len(row)==c_idx and row_num==len(data_formH):
                        formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                        formHsheet.row_dimensions[r_idx].height = 20
                    elif len(row)==c_idx:
                        formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                        formHsheet.row_dimensions[r_idx].height = 20
                    elif row_num==len(data_formH):
                        formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                        formHsheet.row_dimensions[r_idx].height = 20
                    else:
                        formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                        formHsheet.row_dimensions[r_idx].height = 20

            formHsheet['A5']="Name of Establishment   "+str(data_formH['Company Name'].unique()[0])
            formHsheet['H5']=str(data_formH['start_time'].unique()[0])
            
            formHsheet['A6']="Registration No   "+str(data_formH['Registration_no'].unique()[0])
            formHsheet['H6']=str(data_formH['end_time'].unique()[0])
            formHsheet['Q7']=str(data_formH.rest_interval.str.split("-",expand=True)[0].unique()[0])
            formHsheet['U7']=str(data_formH.rest_interval.str.split("-",expand=True)[1].unique()[0])
            

            formHsheet['A'+str(len(rows_copy)+16)]="Name of Establishment   "+str(data_formH['Company Name'].unique()[0])
            formHsheet['A'+str(len(rows_copy)+17)]="Registration No   "+str(data_formH['Registration_no'].unique()[0])
            formHsheet['E'+str(len(rows_copy)+17)]="Wage Period :  "+str(month)+"  "+str(year)
        Part_I()
        Part_II()
        
        formHfinalfile = os.path.join(filelocation,'Form H.xlsx')
        formHfile.save(filename=formHfinalfile)
        

    def Form_I_reg():
        formIfilepath = os.path.join(Delhifilespath,'Form I register of Fine.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formI = data.copy(deep=True)
        data_formI=data_formI.drop_duplicates(subset="Employee Code", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","nature_of_offence","cause_against_fine","FIXED MONTHLY GROSS",
                                        "Date of payment&Fine",'Date_fine',"remarks"]

        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        data_formI[["nature_of_offence","cause_against_fine",'Date_fine']]="-----"
        data_formI["remarks"]=""
        
        data_formI['Fine']=data_formI['Fine'].replace("",0).astype(float)
        data_formI['Fine']=data_formI['Fine'].fillna(0)
        data_formI["Date of payment&Fine"]=data_formI['Date of payment']
        data_formI.loc[data_formI['Fine']==0,"Date of payment&Fine"]="---"
        


        #data_formI['Date of payment']+"\n"+data_formI["Fine"]
        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 7):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIsheet.row_dimensions[r_idx].height = 20
                else:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIsheet.row_dimensions[r_idx].height = 20

        formIsheet['A4']=formIsheet['A4'].value+" : "+data_formI['Company Name'].unique()[0]
        formIfinalfile = os.path.join(filelocation,'Form I register of Fine.xlsx')
        formIfile.save(filename=formIfinalfile)

    def Form_I():
        formIfilepath = os.path.join(Delhifilespath,'Form I.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formI = data.copy(deep=True)
        leave_file_data=data_formI[["Employee Code","Employee Name","Leave Type","Opening","Monthly Increment","Leave Accrued","Used","Encash","Closing"]]
        
        data_formI=data_formI.drop_duplicates(subset="Employee Code", keep="last")
        columns=["Employee Name & Code","Date Joined"]
        data_formI["Employee Name & Code"]=data_formI["Employee Name"].astype(str)+"||"+data_formI["Employee Code"].astype(str)
        data_formI_columns=list(data_formI.columns)
        start=data_formI_columns.index('Emp Code')
        end=data_formI_columns.index('Total\r\nDP')
        columns.extend(data_formI_columns[start+1:end])


        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']

        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        # for column in  range(ord('A'), ord('G') + 1):
        #     formIsheet.unmerge_cells(chr(column)+"11:"+chr(column)+"15")
        # formIsheet.unmerge_cells("H11:I15")
        # formIsheet.unmerge_cells("J11:J15")
        # formIsheet.unmerge_cells("K11:K15")
        logging.info('data for form I is ready')

        
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')

        def cell_write(sheet,r_idx,c_idx,value):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            
        def start_end_date_attendance(absent_label,column_offset,row_offset,initial_offset):  
            is_abs_num=0
            row_index=0
            all_start_dates={}
            all_end_dates={}
            added={}
            for sheet_idx, row in enumerate(dataframe_to_rows(formI_data, index=False, header=False), 12):
                row_index=0
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        if code =="nan":
                            code=name
                        try:
                            target=formIfile[code]
                            added[target.title]=0
                        except:
                            target = formIfile.copy_worksheet(formIsheet)
                            target.title=code
                            #initial offset
                            row_offset[code]=initial_offset
                            added[target.title]=0
                        target['A6']="Name of Employee : "+name
                        target['A4']="Name of Establishment : "+data_formI['Company Name'].unique()[0]
                        target['A7']="Period "+str(month)+" "+str(year)
                    elif c_idx==2:
                        target['A5']="Date of Employment : "+value
                    elif is_abs_num==0 and value==absent_label:
                        is_abs_num=1
                        start=columns[c_idx-1]
                        end=columns[c_idx-1]
                    elif value==absent_label:
                        is_abs_num+=1
                        end=columns[c_idx-1]
                    elif is_abs_num:
                        #target.cell(row=row_index+13, column=1+column_offset, value=is_abs_num)
                        start=start.split("\n")[1].replace("/","-")+"-"+str(year)
                        end=end.split("\n")[1].replace("/","-")+"-"+str(year)
                        
                        cell_write(target,row_index+row_offset[target.title],3+column_offset,start)
                        cell_write(target,row_index+row_offset[target.title],4+column_offset,end)
                        emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==name,:]
                        if absent_label=="PL":
                            total=emp_details["Used"].loc[emp_details["Leave Type"]=="PL"]
                            Closing_pl=emp_details["Closing"].loc[emp_details["Leave Type"]=="PL"].replace("",0).astype(float)
                            Closing_cl=emp_details["Closing"].loc[emp_details["Leave Type"]=="CL"].replace("",0).astype(float)
                            Closing_sl=emp_details["Closing"].loc[emp_details["Leave Type"]=="SL"].replace("",0).astype(float)
                            balance=Closing_cl.add(Closing_pl.add(Closing_sl,fill_value=0), fill_value=0).sum()
                            cell_write(target,row_index+row_offset[target.title],6,"----")
                            cell_write(target,row_index+row_offset[target.title],7,"----")
                            cell_write(target,row_index+row_offset[target.title],10,total.to_string(index=False))
                            cell_write(target,row_index+row_offset[target.title],11,balance)
                        else:
                            temp=data_formI.loc[data_formI["Employee Name"]==name,'Total\r\nCL'].replace("",0).astype(float)
                            amt_leave_requested=data_formI.loc[data_formI["Employee Name"]==name,'Total\r\nSL'].replace("",0).astype(float)+temp
                            #print(amt_leave_requested)
                            Used_cl=emp_details["Used"].loc[emp_details["Leave Type"]=="CL"].replace("",0).astype(float)
                            Used_sl=emp_details["Used"].loc[emp_details["Leave Type"]=="SL"].replace("",0).astype(float)
                            availed=Used_cl.add(Used_sl,fill_value=0).sum()
                            cell_write(target,row_index+row_offset[target.title],1,amt_leave_requested.to_string(index=False))
                            cell_write(target,row_index+row_offset[target.title],2,"----")
                            cell_write(target,row_index+row_offset[target.title],5,availed)
                            
                        # cell_write(target,row_index+11,5+offset,is_abs_num)
                        is_abs_num=0
                        row_index+=1
                        added[target.title]+=1
                        # border_sides_thick = Side(style='thick')       
                        # border_sides_thin = Side(style='thin')
                        # if len(row)==c_idx and len(row)==len(data_formI):
                        #     formIsheet.cell(row=row_index+row_offset[target.title], column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                        #     formIsheet.row_dimensions[row_index+row_offset[target.title]].height = 20
                        # elif len(row)==c_idx:
                        #     formIsheet.cell(row=row_index+row_offset[target.title], column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                        #     formIsheet.row_dimensions[row_index+row_offset[target.title]].height = 20
                        # elif len(row)==len(data_formI):
                        #     formIsheet.cell(row=row_index+row_offset[target.title], column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                        #     formIsheet.row_dimensions[row_index+row_offset[target.title]].height = 20
                        # else:
                        #     formIsheet.cell(row=row_index+row_offset[target.title], column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                        #     formIsheet.row_dimensions[row_index+row_offset[target.title]].height = 20
            return added
                    
        absent_label="PL"
        column_offset=5
        initial_offset=12
        row_offset={}          
        row_offset=start_end_date_attendance(absent_label,column_offset,row_offset,initial_offset)
        #reset row_offset since it was for PL
        
        row_offset = {x: initial_offset for x in row_offset}
        
        absent_label="CL"
        column_offset=0
        from collections import Counter
        #increment = {x: initial_offset for x in row_offset}
        row_offset=Counter(start_end_date_attendance(absent_label,column_offset,row_offset,initial_offset))+Counter(row_offset)
        absent_label="SL"
        column_offset=0
        row_offset=start_end_date_attendance(absent_label,column_offset,row_offset,initial_offset)
        

        formIfile.remove(formIfile["Sheet1"])
        formIfinalfile = os.path.join(filelocation,'Form I.xlsx')
        formIfile.save(filename=formIfinalfile)


    def Form_II():
        formIIfilepath = os.path.join(Delhifilespath,'Form II.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy(deep=True)
        data_formII=data_formII.drop_duplicates(subset="Employee Code", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Gender","Department",
                                        "Damage_loss_with_date","whether_work_showed_cause",
                                        "Date of payment & amount of deduction","num_instalments",'Date on which total amount realised',"remarks"]
        
        data_formII["Damage or Loss"]=data_formII["Damage or Loss"].replace("",0).astype(float)
        data_formII["Damage or Loss"]=data_formII["Damage or Loss"].fillna(0)
        
        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        if str(data_formII['Date of payment'].dtype)[0:8] == 'datetime':
            data_formII['Date of payment']=data_formII['Date of payment'].apply(lambda x: x.strftime('%d-%m-%Y'))
        else:
            data_formII['Date of payment']=data_formII['Date of payment'].astype(str)

        data_formII["Damage_loss_with_date"]=data_formII['Date of payment']+" & "+data_formII["Damage or Loss"].astype(str)
        data_formII.loc[data_formII["Damage or Loss"]==0,"Damage_loss_with_date"]="---"

        data_formII["Date of payment & amount of deduction"]=data_formII["Damage_loss_with_date"]
        data_formII["num_instalments"]="1"
        data_formII.loc[data_formII["Damage or Loss"]==0,"num_instalments"]="---"

        data_formII["Date on which total amount realised"]=data_formII['Date of payment']
        data_formII.loc[data_formII["Damage or Loss"]==0,"Date on which total amount realised"]="---"

        data_formII[["remarks","whether_work_showed_cause"]]="-----"
        

        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        
        rows = dataframe_to_rows(formII_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 7):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                
                if len(row)==c_idx and row_num==len(data_formII):
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formII):
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIIsheet.row_dimensions[r_idx].height = 20
                else:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIIsheet.row_dimensions[r_idx].height = 20

        formIIsheet['A4']=formIIsheet['A4'].value+" : "+data_formII['Company Name'].unique()[0]
        formIIfinalfile = os.path.join(filelocation,'Form II.xlsx')
        formIIfile.save(filename=formIIfinalfile)

    def Form_IV():
        formIVfilepath = os.path.join(Delhifilespath,'Form IV.xlsx')
        formIVfile = load_workbook(filename=formIVfilepath)
        logging.info('Form IV file has sheet: '+str(formIVfile.sheetnames))
        logging.info('create columns which are now available')

        data_formIV = data.copy(deep=True)
        data_formIV=data_formIV.drop_duplicates(subset="Employee Code", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Gender","Designation_Dept","Date_overtime_worked",
                                        "Extent of over-time","Total over-time","Normal hrs ",
                                        "FIXED MONTHLY GROSS","overtime rate",
                                        "normal_earning","Overtime",'Total Earning',"date_overtime_paid"]
                                        
        data_formIV[['Total\r\nOT Hrs',"Overtime",'Total Earning']]=data_formIV[['Total\r\nOT Hrs',"Overtime",'Total Earning']].replace("",0).astype(float)
        data_formIV["Total over-time"]=data_formIV['Total\r\nOT Hrs']
        data_formIV["normal_earning"]=data_formIV['Total Earning']-data_formIV["Overtime"]
        data_formIV.loc[data_formIV['Total\r\nOT Hrs']==0,["Total over-time","Normal hrs ",
                                        "FIXED MONTHLY GROSS","overtime rate",
                                        "normal_earning","Overtime",'Total Earning']]="---"

        data_formIV["date_overtime_paid"]=data_formIV['Date of payment']
        data_formIV.loc[data_formIV["Overtime"]==0,"date_overtime_paid"]="---"
        data_formIV.loc[data_formIV['Total\r\nOT Hrs']==0,"date_overtime_paid"]="---"

        data_formIV['S.no'] = list(range(1,len(data_formIV)+1))
        data_formIV['Designation_Dept']=data_formIV["Designation"]+"_"+data_formIV["Department"]
        data_formIV["Extent of over-time"]="-----"

        data_formIV["Date_overtime_worked"]="-----"
        # data_formIV["Date of payment & amount of deduction"]=data_formIV['Date of payment']+"\n"+data_formIV["Total Deductions"]
        formIV_data=data_formIV[columns]
        formIVsheet = formIVfile['Sheet1']
        formIVsheet.sheet_properties.pageSetUpPr.fitToPage = True
        for column in  range(ord('A'), ord('O') + 1):
            formIVsheet.unmerge_cells(chr(column)+"7:"+chr(column)+"14")

        logging.info('data for form IV is ready')

        
        rows = dataframe_to_rows(formIV_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 7):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formIVsheet.cell(row=r_idx, column=c_idx, value=value)
                formIVsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIVsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formIV):
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIVsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIVsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formIV):
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIVsheet.row_dimensions[r_idx].height = 20
                else:
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIVsheet.row_dimensions[r_idx].height = 20

        formIVsheet['A4']=formIVsheet['A4'].value+"  "+data_formIV['Company Name'].unique()[0]+"  "+data_formIV['Company Address'].unique()[0]+"                                Month Ending: "+month+" "+str(year)
        # formIVsheet.cell(row=4, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
        #formIVsheet['A4']="Month Ending: "+month+" "+str(year)
        formIVfinalfile = os.path.join(filelocation,'Form IV.xlsx')
        formIVfile.save(filename=formIVfinalfile)
    try:   
        Form_H()
        master.update()
        Form_I_reg()
        master.update()
        Form_I()
        master.update()
        master.update()
        Form_II()
        master.update()
        Form_IV()
        master.update()
        Form_G()
        master.update()
    except KeyError as e:
        logging.info("Key error : Check if {} column exsists".format(e))
        print("Key error {}".format(e))
        report.configure(text="Failed: Check input file format  \n column {} not found".format(e))
        master.update()
        raise KeyError
    except FileNotFoundError as e:
        logging.info("File not found : Check if {} exsists".format(e))
        report.configure(text="Failed: File {} not found".format(e))
        master.update()
        raise FileNotFoundError