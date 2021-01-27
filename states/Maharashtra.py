from states import logging,monthdict,Statefolder,read_min_wages_file
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging
from collections import Counter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill, numbers
from dateutil import parser
from states import Register_folder  
from states.utils import forms_template
employee_name_column="Employee Name"
fathers_name_column="Father's Name"
gender_column="Gender"
employee_code_column="Employee Code"
contractor_name_column='Contractor_name'
department_column='Department'
fix_monthly_gross_column="FIXED MONTHLY GROSS"
date_of_payment_column='Date of payment'
company_name_column="Company Name"
address_column="Address"
PE_or_contract_column='PE_or_contract'

def Maharashtra(data,contractor_name,contractor_address,filelocation,month,year,report,master):
    logging.info('Maharashtra forms')

    Maharashtrafilespath = os.path.join(Statefolder,'Maharashtra')
    logging.info('Maharashtra files path is :'+str(Maharashtrafilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]
    #Min wages xl
    input_filelocation=filelocation.split(Register_folder)[0]
    min_wages_maharashtra=read_min_wages_file("MAHARASHTRA","SEMI-SKILLED",input_filelocation)
    
    templates=forms_template.Templates(to_read=Maharashtrafilespath,to_write=filelocation,month=month,year=year,report=report,master=master)
    
    def Read_Holiday_file():
        inputfolder = filelocation.split(Register_folder)[0]
        file_list = os.listdir(inputfolder)
        logging.info('input folder is '+str(inputfolder))
        for f in file_list:
            if 'HOLIDAY' in f.upper() and 'LIST' in f.upper():
                holidayfilename = f
                logging.info('holidayfilename is :'+f)
        if 'holidayfilename' in locals():
            holidayfile = os.path.join(inputfolder,holidayfilename)
            holiday=pd.read_excel(holidayfile).dropna()
            holiday.columns=[ "SN.", "Date"," Day"," Occasion"]
            convert=lambda variable: datetime.datetime.strptime(variable,'%d%m%Y')
            holiday=holiday[1:]
            holiday["Date"]=holiday["Date"].apply(str)
            holiday["Date"]=holiday["Date"].apply(parser.parse)
            holiday=holiday.set_index(pd.PeriodIndex(data=holiday.Date,freq='D'))
        else:
            holiday_columns = [ "SN.", "Date"," Day"," Occasion"]
            holiday = pd.DataFrame(columns = holiday_columns)
            holiday=holiday.set_index(pd.PeriodIndex(data=holiday.Date,freq='D'))
        return holiday.sort_index()
        
    def Form_I():
        logging.info('create columns which are now available')
        data_formI = data.copy(deep=True)
        data_formI=data_formI.drop_duplicates(subset=employee_code_column, keep="last")

        columns=['S.no',employee_name_column,fathers_name_column,gender_column,department_column,"name&date_of_offence","cause_against_fine",
                                        fix_monthly_gross_column,date_of_payment_column,"Date of Fine","remarks"]

        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        data_formI[["name&date_of_offence","cause_against_fine",fix_monthly_gross_column,date_of_payment_column,"Date of Fine","remarks"]]="NIL"
        formI_data=data_formI[columns]
        data_once_per_sheet={'A5':data_formI[company_name_column].unique()[0],'A6':str(month)+" "+str(year)}
        templates.create_basic_form(filename='Form I register of fine.xlsx',
                                    sheet_name='Sheet1',all_employee_data=formI_data,start_row=8,start_column=1,
                                    data_once_per_sheet=data_once_per_sheet)
    
    def Form_II_Muster_Roll():
        formIIfilepath = os.path.join(Maharashtrafilespath,'Form II muster roll.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy(deep=True)
        data_formII=data_formII.drop_duplicates(subset=employee_code_column, keep="last")

        columns=['S.no',employee_code_column,employee_name_column,"start_time","end_time",
                                        "interval_for_reset_from","interval_for_reset_to"]
        
        columns.extend(templates.get_attendance_columns(data_formII))

        columns.extend(["Total\r\nDP"])
        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII['interval_for_reset_to']="2:00 PM"
        data_formII['interval_for_reset_from']="1:00 PM"
        data_formII["start_time"]="9:30 AM"
        data_formII["end_time"]="6:30 PM"
        
        formII_data=templates.get_data(data_formII,columns)
        data_once_per_sheet={'A2':month+str(year)}
        if not data[PE_or_contract_column].unique()[0].upper()=="PE":
            data_once_per_sheet['A3']=templates.combine_columns_of_dataframe(data_formII,[contractor_name_column,'Contractor_Address']).unique()[0]
            data_once_per_sheet['A4']=templates.combine_columns_of_dataframe(data_formII,['Unit',address_column]).unique()[0]
        
        templates.create_basic_form(filename='Form II muster roll.xlsx',sheet_name='Sheet1',all_employee_data=formII_data,
                                    start_row=9,start_column=1,data_once_per_sheet=data_once_per_sheet)

    def Form_II_reg_damage_loss():
        data_formII = data.copy(deep=True)
        data_formII=data_formII.drop_duplicates(subset=employee_code_column, keep="last")

        columns=['S.no',employee_name_column,fathers_name_column,gender_column,department_column,"Damage or Loss","whether_work_showed_cause",
                                        "Date of payment & amount of deduction","num_instalments",date_of_payment_column,"remarks"]
        
        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII[["Damage or Loss","whether_work_showed_cause","Date of payment & amount of deduction",
                                                    "num_instalments",date_of_payment_column,"remarks"]]="NIL"
        formII_data=templates.get_data(data_formII,columns)
        data_once_per_sheet={'A5':str(data_formII[company_name_column].unique()[0])+","+str(data_formII[address_column].unique()[0]),
                            'A6':str(month)+" "+str(year)}
        templates.create_basic_form(filename='Form II register of damage or losses.xlsx',sheet_name='Sheet1',all_employee_data=formII_data,
                                    start_row=9,start_column=1,data_once_per_sheet=data_once_per_sheet)

    def Form_II_wages_reg():
        
        data_formII = data.copy(deep=True)
        leave_file_data=templates.get_data(data_formII,[employee_code_column,employee_name_column,"Leave Type","Opening","Monthly Increment","Leave Accrued","Used","Encash","Closing"])
        data_formII=data_formII.drop_duplicates(subset=employee_code_column, keep="last")
        data_formII.fillna(value=0, inplace=True)
        #print(sorted(data_formII.columns))
        columns=['S.no',employee_code_column,employee_name_column,'Age',gender_column,"Designation","Date Joined","Days Paid",
                                    "min_wages",fix_monthly_gross_column,"Total_Production_Piece_Rate",'Total\r\nOT Hrs',
                                    fix_monthly_gross_column,"Earned Basic","HRA/Earned_basic","HRA","Tel and Int Reimb",
                                    "Bonus","Fuel Reimb","Corp Attire Reimb","CCA","Overtime","Total Earning",
                                    "PF","P.Tax","Insurance","sal_fine_damage","Total Deductions","Net Paid",
                                    "Prev_balance","Earned_during_month","Availed","colsing_bal",date_of_payment_column,
                                    "Bank A/c Number",'Cheque No - NEFT date',"Net Paid","sign"]
        # print(leave_file_data)
        data_formII[["Prev_balance","Earned_during_month","Availed","colsing_bal"]]=""
        
        def date_format_change(val):
            return val.strftime('%d-%m-%y')
        
        for employee_name_leave_file in data_formII[employee_name_column]:
            #opening
            emp_details=leave_file_data.loc[leave_file_data[employee_name_column]==employee_name_leave_file,:]
            opening_pl=emp_details["Opening"].loc[emp_details["Leave Type"]=="PL"].astype(float)
            opening_cl=emp_details["Opening"].loc[emp_details["Leave Type"]=="CL"].astype(float)
            opening_sl=emp_details["Opening"].loc[emp_details["Leave Type"]=="SL"].astype(float)
            prev_bal=opening_pl.add(opening_cl.add(opening_sl,fill_value=0), fill_value=0).sum()
            
            data_formII.loc[data_formII[employee_name_column]==employee_name_leave_file,"Prev_balance"]=prev_bal
            
            #monthly_inr
            mon_inr_pl=emp_details["Monthly Increment"].loc[emp_details["Leave Type"]=="PL"].astype(float)
            mon_inr_cl=emp_details["Monthly Increment"].loc[emp_details["Leave Type"]=="CL"].astype(float)
            mon_inr_sl=emp_details["Monthly Increment"].loc[emp_details["Leave Type"]=="SL"].astype(float)
            earned=mon_inr_cl.add(mon_inr_pl.add(mon_inr_sl,fill_value=0), fill_value=0).sum()
            data_formII.loc[data_formII[employee_name_column]==employee_name_leave_file,"Earned_during_month"]=earned
            #availed during month
            Used_pl=emp_details["Used"].loc[emp_details["Leave Type"]=="PL"].astype(float)
            Used_cl=emp_details["Used"].loc[emp_details["Leave Type"]=="CL"].astype(float)
            Used_sl=emp_details["Used"].loc[emp_details["Leave Type"]=="SL"].astype(float)
            availed=Used_cl.add(Used_pl.add(Used_sl,fill_value=0), fill_value=0).sum()
            
            data_formII.loc[data_formII[employee_name_column]==employee_name_leave_file,"Availed"]=availed
            #closing
            Closing_pl=emp_details["Closing"].loc[emp_details["Leave Type"]=="PL"].astype(float)
            Closing_cl=emp_details["Closing"].loc[emp_details["Leave Type"]=="CL"].astype(float)
            Closing_sl=emp_details["Closing"].loc[emp_details["Leave Type"]=="SL"].astype(float)
            closing=Closing_cl.add(Closing_pl.add(Closing_sl,fill_value=0), fill_value=0).sum()
            
            data_formII.loc[data_formII[employee_name_column]==employee_name_leave_file,"colsing_bal"]=closing

        if str(data_formII[date_of_payment_column].dtype)[0:8] == 'datetime':
            data_formII[date_of_payment_column]=data_formII[date_of_payment_column].apply(date_format_change)

        data_formII["HRA/Earned_basic"]=((data_formII["HRA"].apply(float)/data_formII["Earned Basic"].apply(float))*100.0)
        data_formII["sal_fine_damage"]=templates.sum_columns_of_dataframe(data_formII,["Fine","Damage or Loss"])
        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII[["Total_Production_Piece_Rate"]]="----"
        data_formII["min_wages"]=min_wages_maharashtra
        data_formII[["sign"]]=""
        
        formII_data=templates.get_data(data_formII,columns)
        data_once_per_sheet={'A2':str(month)+" "+str(year),
                            'A4':str(data_formII[company_name_column].unique()[0])}
        templates.create_basic_form(filename='Form II wages register.xlsx',sheet_name='Sheet1',all_employee_data=formII_data,
                                    start_row=7,start_column=1,data_once_per_sheet=data_once_per_sheet)

    def Form_VI_Overtime():
        
        data_formIV = data.copy(deep=True)
        data_formIV=data_formIV.drop_duplicates(subset=employee_code_column, keep="last")

        columns=['S.no',employee_name_column,fathers_name_column,gender_column,"Designation_Dept","Date_overtime_worked",
                                        "Extent of over-time",'Total\r\nOT Hrs','Normal hrs ',
                                        fix_monthly_gross_column,"overtime rate","Total Earning-Overtime","Overtime",'Total Earning',date_of_payment_column]
        
        data_formIV['S.no'] = list(range(1,len(data_formIV)+1))
        data_formIV["Overtime"]=data_formIV["Overtime"].fillna(value=0)
        data_formIV["Overtime"]=data_formIV["Overtime"].astype(float)

        data_formIV.loc[data_formIV["Overtime"]==0,date_of_payment_column]="---"
        data_formIV[date_of_payment_column]=data_formIV[date_of_payment_column].replace(0,"---")
        data_formIV['Designation_Dept']=data_formIV["Designation"]+"_"+data_formIV[department_column]
        data_formIV["Total Earning-Overtime"]=data_formIV['Total Earning'].astype(float)-data_formIV["Overtime"].astype(float)
        data_formIV[["Date_overtime_worked","Extent of over-time"]]="NIL"
        formIV_data=templates.get_data(data_formIV,columns)

        data_once_per_sheet={'A5':str(month)+" "+str(year),"A7":str(data_formIV[contractor_name_column].unique()[0])}
        templates.create_basic_form(filename='Form IV Overtime register.xlsx',sheet_name='Sheet1',all_employee_data=formIV_data,
                                    start_row=10,start_column=1,data_once_per_sheet=data_once_per_sheet)

        # for i in range(1,16):
        #     formIVsheet.cell(row=7, column=i).border = Border(outline= True,bottom=border_sides_thick)
        
    def Form_VI_reg_advance():
        
        data_formIV = data.copy(deep=True)
        data_formIV=data_formIV.drop_duplicates(subset=employee_code_column, keep="last")

        data_formIV.fillna(value=0, inplace=True)
        columns=['S.no',employee_name_column,fathers_name_column,department_column,"Salary Advance","purpose_advance",
                                        "num_installments_advance","Postponement_granted",
                                        "Date repaid","remarks"]
                                        
                                        
        data_formIV['S.no'] = list(range(1,len(data_formIV)+1))
        
        data_formIV[["purpose_advance","num_installments_advance","Postponement_granted","Date repaid","remarks"]]="NIL"
        formIV_data=templates.get_data(data_formIV,columns)
        
        data_once_per_sheet={'A7':str(month)+" "+str(year),"A6":str(data_formIV[company_name_column].unique()[0])}
        templates.create_basic_form(filename='Form IV register of advance.xlsx',sheet_name='Sheet1',all_employee_data=formIV_data,
                                    start_row=13,start_column=1,data_once_per_sheet=data_once_per_sheet)


    def From_O():
        formOfilepath = os.path.join(Maharashtrafilespath,'Form O leave book.xlsx')
        formOfile = load_workbook(filename=formOfilepath)
        logging.info('Form O file has sheet: '+str(formOfile.sheetnames))
        #print(formOfile.sheetnames)
        logging.info('create columns which are now available')

        data_formO = data.copy(deep=True)
        leave_file_data=data_formO[[employee_code_column,employee_name_column,"Leave Type","Opening","Monthly Increment","Leave Accrued","Used","Encash","Closing"]]
        data_formO=data_formO.drop_duplicates(subset=employee_code_column, keep="last")

        data_formO.fillna(value=0, inplace=True)
        columns=["Employee Name & Code","Date Joined",department_column,"Registration_no"]
        data_formO["Employee Name & Code"]=data_formO[employee_name_column].astype(str)+"||"+data_formO[employee_code_column].astype(str)

        data_formO[["num_days","Earned_during_month","Availed","colsing_bal",'Cheque No - NEFT date']]=""
        for employee_name_leave_file in data_formO[employee_name_column]:
            #opening
            emp_details=leave_file_data.loc[leave_file_data[employee_name_column]==employee_name_leave_file,:]
            opening_pl=emp_details["Opening"].loc[emp_details["Leave Type"]=="PL"]
            if opening_pl.empty:
                opening_pl="0"
            else:
                opening_pl=opening_pl.to_string(index=False)
            if opening_pl.lower()=="nan":
                opening_pl="0"
            
            data_formO.loc[data_formO[employee_name_column]==employee_name_leave_file,"num_days"]=opening_pl
            
           
     
        data_formO_columns=list(data_formO.columns)
        start_month = datetime.date(year,month_num,1)
        end_month = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
        
        
        columns.extend(templates.get_attendance_columns(data_formO))


        formO_data=data_formO[columns]
        formOsheet = formOfile['Sheet1']

        formOsheet.sheet_properties.pageSetUpPr.fitToPage = True

        #for column in  range(ord('A'), ord('G') + 1):
        #    formOsheet.unmerge_cells(chr(column)+"11:"+chr(column)+"15")
        formOsheet.unmerge_cells("A22:H22")
        formOsheet.unmerge_cells("A23:B23")
        formOsheet.unmerge_cells("C23:C24")
        formOsheet.unmerge_cells("D23:D24")
        formOsheet.unmerge_cells("E23:E24")
        formOsheet.unmerge_cells("F23:G24")
        formOsheet.unmerge_cells("H23:H24")
        formOsheet.unmerge_cells("F25:G25")
        formOsheet.unmerge_cells("F26:G26")
        formOsheet.unmerge_cells("F27:G27")
        
        formOsheet.unmerge_cells("A28:F28")
        formOsheet.unmerge_cells("A29:B30")
        formOsheet.unmerge_cells("C29:C31")
        formOsheet.unmerge_cells("D29:D31")
        formOsheet.unmerge_cells("E29:E31")
        formOsheet.unmerge_cells("F29:F31")
        
        
        logging.info('data for form I is ready')

        
        #rows_copy = list(dataframe_to_rows(formO_data, index=False, header=False))
        def cell_write(sheet,r_idx,c_idx,value):
            if not (str(value)=="nan" or str(value)=="NaN"):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =15)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            
        def PL_write(row_index,target,start,end,is_abs_num):

            cell_write(target,row_index,3,start+"--"+end)
            target.row_dimensions[row_index].height = 50
            cell_write(target,row_index , 1,start_month)
            cell_write(target,row_index , 4,"----")
            cell_write(target,row_index , 5,"----")
            # cell_write(target,row_index , 6,"----")
            # cell_write(target,row_index , 7,"----")
            # cell_write(target,row_index , 8,"----")
            # #print(data_formO.loc[data_formO[columns[0]]],emp_name)
            def get_emp_name(var):
                return var.split("||")[0]
            temp=str(data_formO.loc[data_formO[columns[0]].apply(get_emp_name)==emp_name,"Date Left"].tolist()[0])
            if not (temp=="nan" or temp=="0"):
                cell_write(target,row_index , 9,temp)
            else:
                cell_write(target,row_index , 9,"---")
            cell_write(target,row_index ,10,data_formO.loc[data_formO[columns[0]].apply(get_emp_name)==emp_name,"Leave Encashment"].to_string(index=False))
            #cell_write(target,row_index,4,is_abs_num)
            #cell_write(target,row_index,5,start)
            #cell_write(target,row_index,6,end)

        # def FL_write(row_index,target,start,end,is_abs_num):
        #     cell_write(target,row_index,1,start)
        #     cell_write(target,row_index,2,end)
        #     cell_write(target,row_index, 6, "-----")
        #     formOfile[sheet].merge_cells("F"+str(row_index)+":G"+str(row_index))
        #     #print("F"+str(row_index)+":G"+str(row_index))
        #     #cell_write(target,row_index,4,is_abs_num)
        #     #cell_write(target,row_index,5,start)
        #     #cell_write(target,row_index,6,end)
        
        # def CL_write(row_index,target,start,end,is_abs_num):
        #     cell_write(target,row_index,2,start)
        #     cell_write(target,row_index,3,end)
        #     #cell_write(target,row_index,5,start)
        #     #cell_write(target,row_index,6,end)

        form_write={'PL':PL_write}#,'FL':FL_write,'CL':CL_write}
        
        def start_end_date_attendance(rows,absent_label,row_offset,initial_offset):  
           # print("infunction",row_offset)
            is_abs_num=0
            row_index=0
            added={}
            for sheet_idx, row in enumerate(rows, 10):
                row_index=0
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        try:
                            target=formOfile[code]
                        except:
                            target = formOfile.copy_worksheet(formOsheet)
                            target.title=code
                            #initial offset
                            row_offset[code]=initial_offset
                        
                        target['A4']="Name and address of the Establishment:- "" "+str(data_formO[company_name_column].unique()[0])#+","+str(data_formO[address_column].unique()[0])
                        #target['A5']="Name of Employer:- "" "+str(data_formO['Unit'].unique()[0])
                        target["H4"]="Name of the employee:- "+str(name)+"\n"+" Receipt of leave book - "
                        target['A7']="Name of worker : "+str(name)
                        global emp_name
                        emp_name=str(name)
                        added[target.title]=0
                        form_write[absent_label](row_index+row_offset[target.title],target,"","","")
                        cell_write(target,row_index+row_offset[target.title] , 1, str("01"+"-"+str(month_num)+"-"+str(year)))
                        num=data_formO.loc[data_formO[employee_name_column]==emp_name,"num_days"]
                        if num.empty:
                            cell_write(target,row_index+row_offset[target.title] , 2,0)
                        else:
                            cell_write(target,row_index+row_offset[target.title] , 2,num.to_string(index=False))
                        #cell_write(target,row_index+row_offset[target.title] , 2,data_formO.loc[data_formO[employee_name_column]==emp_name,"Opening"].to_string(index=False))
                        
                    elif c_idx==2:
                        target['H8']="Date of entry into service :- "+str(value)
                    elif c_idx==3:
                        target['A8']="Description of the Department (If Applicable) :-  "+str(value)
                    elif c_idx==4:
                        target['A6']="Registration No. :- "+str(value)
                    elif is_abs_num==0 and value==absent_label:
                        is_abs_num=1
                        start=columns[c_idx-1]
                        end=columns[c_idx-1]
                    elif value==absent_label:
                        is_abs_num+=1
                        end=columns[c_idx-1]
                    elif is_abs_num:
                        start=start.split("\n")[1].replace("/","-")+"-"+str(year)
                        end=end.split("\n")[1].replace("/","-")+"-"+str(year)
                        
                        form_write[absent_label](row_index+row_offset[target.title],target,start,end,is_abs_num)
                        
                        #Uncomment these lines if there are too many lines in the first part of the form(This will cause border problems)
                        #target.insert_rows(row_index+row_offset[target.title]+1)
                        #added[target.title]+=1
                        
                        is_abs_num=0
                        num=data_formO.loc[data_formO[employee_name_column]==emp_name,"num_days"]
                        if num.empty:
                            cell_write(target,row_index+row_offset[target.title] , 2,0)
                        else:
                            cell_write(target,row_index+row_offset[target.title] , 2,num.to_string(index=False))
                        cell_write(target,row_index+row_offset[target.title] , 1, str("01"+"-"+str(month_num)+"-"+str(year)))
                        row_index+=1
                    
            # print(added)
            return added
        offset={}
        initial_offset=13
        #for sheet in formOfile.sheetnames:
        #    offset[sheet]=initial_offset
        offset=Counter(offset)+Counter(start_end_date_attendance(dataframe_to_rows(formO_data, index=False, header=False),"PL",offset,initial_offset))
        
        for sheet in formOfile.sheetnames:
            offset[sheet]+=25
            initial_offset+=25
            formOfile[sheet].merge_cells("A"+str(offset[sheet]-3)+":H"+str(offset[sheet]-3))
            formOfile[sheet].merge_cells("A"+str(offset[sheet]-2)+":B"+str(offset[sheet]-2))
            formOfile[sheet].merge_cells("C"+str(offset[sheet]-2)+":C"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("D"+str(offset[sheet]-2)+":D"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("E"+str(offset[sheet]-2)+":E"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("F"+str(offset[sheet]-2)+":G"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("H"+str(offset[sheet]-2)+":H"+str(offset[sheet]-1))
        columns=["Employee Name & Code"]
        data_formO["Employee Name & Code"]=data_formO[employee_name_column].astype(str)+"||"+data_formO[employee_code_column].astype(str)
        formO_data=data_formO[columns]
        
        rows = dataframe_to_rows(formO_data, index=False, header=False)
        logging.info('rows taken out from data')
        holidays=Read_Holiday_file()
        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 1):
                    name=value.split("||")[0]
                    code=value.split("||")[1]
                    if code =="nan":
                        code=name
                    target=formOfile[code]
                    last_day=calendar.monthrange(int(year),month_num)[1]
                    start_date = str(year)+"-"+str(month_num)+"-01"
                    end_date = str(year)+"-"+str(month_num)+"-"+str(last_day)

                    after_start_date = holidays.index >= start_date
                    before_end_date = holidays.index <= end_date
                    between_two_dates = after_start_date & before_end_date

                    filtered_dates = holidays.loc[after_start_date,"Date"]
                    start_date=datetime.datetime.strptime(start_date,'%Y-%m-%d')
                    end_date=datetime.datetime.strptime(end_date,'%Y-%m-%d')
                    
                    #for index,date in enumerate(filtered_dates):
                    index=0
                    target.cell(row=offset[code]+index, column=1, value=start_date.date().strftime('%d-%m-%y'))
                    target.cell(row=offset[code]+index, column=1).font =Font(name ='Verdana', size =8)
                    target.cell(row=offset[code]+index, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=offset[code]+index, column=1).border = Border(outline= True, right=border_sides, bottom=border_sides)
                    
                    target.cell(row=offset[code]+index, column=2, value=end_date.date().strftime('%d-%m-%y'))
                    target.cell(row=offset[code]+index, column=2).font =Font(name ='Verdana', size =8)
                    target.cell(row=offset[code]+index, column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=offset[code]+index, column=2).border = Border(outline= True, right=border_sides, bottom=border_sides)

                    target.cell(row=offset[code]+index, column=3, value=len(holidays.loc[holidays.index >=str(start_date)]))
                    target.cell(row=offset[code]+index, column=3).font =Font(name ='Verdana', size =8)
                    target.cell(row=offset[code]+index, column=3).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=offset[code]+index, column=3).border = Border(outline= True, right=border_sides, bottom=border_sides)    

                    target.cell(row=offset[code]+index, column=4, value=between_two_dates.sum())
                    target.cell(row=offset[code]+index, column=4).font =Font(name ='Verdana', size =8)
                    target.cell(row=offset[code]+index, column=4).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=offset[code]+index, column=4).border = Border(outline= True, right=border_sides, bottom=border_sides)    

                    target.cell(row=offset[code]+index, column=5, value=len(holidays.loc[holidays.index >=str(start_date)])-between_two_dates.sum())
                    target.cell(row=offset[code]+index, column=5).font =Font(name ='Verdana', size =8)
                    target.cell(row=offset[code]+index, column=5).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=offset[code]+index, column=5).border = Border(outline= True, right=border_sides, bottom=border_sides)    
                
                    #offset[code]+=1
                    
        #offset+=Counter(start_end_date_attendance(dataframe_to_rows(formO_data, index=False, header=False),"FL",offset,initial_offset))
        
        for sheet in formOfile.sheetnames:
            offset[sheet]+=7
            initial_offset+=7
            formOfile[sheet].merge_cells("A"+str(offset[sheet]-4)+":F"+str(offset[sheet]-4))
            formOfile[sheet].merge_cells("A"+str(offset[sheet]-3)+":B"+str(offset[sheet]-2))
            formOfile[sheet].merge_cells("C"+str(offset[sheet]-3)+":C"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("D"+str(offset[sheet]-3)+":D"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("E"+str(offset[sheet]-3)+":E"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("F"+str(offset[sheet]-3)+":F"+str(offset[sheet]-1))

        columns=["Employee Name & Code","total_leave","availed","balance","remarks"]
        data_formO[["total_leave","availed","balance"]]=""
        for employee_name_leave_file in data_formO[employee_name_column]:
            #opening
            emp_details=leave_file_data.loc[leave_file_data[employee_name_column]==employee_name_leave_file,:]
            opening_cl=emp_details["Opening"].loc[emp_details["Leave Type"]=="CL"]
            if opening_cl.empty:
                data_formO.loc[data_formO[employee_name_column]==employee_name_leave_file,"total_leave"]="0"
            else:
                opening_cl=opening_cl.to_string(index=False)
                data_formO.loc[data_formO[employee_name_column]==employee_name_leave_file,"total_leave"]=opening_cl if not opening_cl=="" else "0"

            availed=emp_details["Used"].loc[emp_details["Leave Type"]=="CL"]
            
            if availed.empty:
                data_formO.loc[data_formO[employee_name_column]==employee_name_leave_file,"availed"]="0"
            else:
                availed=availed.to_string(index=False)
                data_formO.loc[data_formO[employee_name_column]==employee_name_leave_file,"availed"]=availed if not availed=="" else "0"
                # print("------------------------------------------------------------------------------------------------------")
                # print("availed")
                # print(availed)

            balance=emp_details["Closing"].loc[emp_details["Leave Type"]=="CL"]
           
            if balance.empty:
                data_formO.loc[data_formO[employee_name_column]==employee_name_leave_file,"balance"]="0"
            else:
                balance=balance.to_string(index=False)
                data_formO.loc[data_formO[employee_name_column]==employee_name_leave_file,"balance"]=balance if not balance=="" else "0"
                # print("balance")
                # print(balance)

            
            
        data_formO[["remarks"]]=""
        data_formO["Employee Name & Code"]=data_formO[employee_name_column].astype(str)+"||"+data_formO[employee_code_column].astype(str)
        formO_data=data_formO[columns]
        
        
        rows = dataframe_to_rows(formO_data, index=False, header=False)
        logging.info('rows taken out from data')
        # offset[code]+=1
        border_sides_thin = Side(style='thin')
        border_sides_thick = Side(style='thick')
        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 1):
                if c_idx==1:
                    name=value.split("||")[0]
                    code=value.split("||")[1]
                    target=formOfile[code]
                    start_date = "01"+"-"+str(month_num)+"-"+str(year)
                    end_date = str(last_day)+"-"+str(month_num)+"-"+str(year)
                    target.cell(row=offset[code], column=1, value=start_date)
                    target.cell(row=offset[code], column=1).font =Font(name ='Verdana', size =8)
                    target.cell(row=offset[code], column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=offset[code], column=1).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    
                    target.cell(row=offset[code], column=2, value=end_date)
                    target.cell(row=offset[code], column=2).font =Font(name ='Verdana', size =8)
                    target.cell(row=offset[code], column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=offset[code], column=2).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    target.row_dimensions[offset[code]].height = 20
                else:
                   # print("--------------------------------------------------------")
                   # print(offset[code],c_idx+1)
                    target.cell(row=offset[code], column=c_idx+1, value=str(value))
                    target.cell(row=offset[code], column=c_idx+1).font =Font(name ='Verdana', size =8)
                    target.cell(row=offset[code], column=c_idx+1).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=offset[code], column=c_idx+1).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    target.row_dimensions[offset[code]].height = 20
            #offset[code]+=1

        #offset+=Counter(start_end_date_attendance(dataframe_to_rows(formO_data, index=False, header=False),"CL",offset,initial_offset))
        formOfile.remove(formOfile["Sheet1"])
        formOfile.remove(formOfile["Sheet2"])
        formOfile.remove(formOfile["Sheet3"])
        formOfinalfile = os.path.join(filelocation,'Form O leave book.xlsx')
        formOfile.save(filename=formOfinalfile)
    
    try:
        Form_I()
        Form_II_Muster_Roll()
        Form_II_reg_damage_loss()
        Form_II_wages_reg()
        Form_VI_Overtime()
        Form_VI_reg_advance()
        From_O()
    except KeyError as e:
        logging.info("Key error : Check if {} column exsists".format(e))
        print("Key error {}".format(e))
        report.configure(text="Failed: Check input file format  \n column {} not found".format(e))
        master.update()
        raise KeyError
    except FileNotFoundError as e:
        logging.info("File not found : Check if {} exsists".format(e))
        report.configure(text="Failed: File {} not found".format(e))
        master.update()
        raise FileNotFoundError