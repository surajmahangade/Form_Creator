import os
from pathlib import Path
import logging
import datetime
import pandas as pd
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill, numbers
#backend code starts here

from pathlib import Path
import sys
path = Path(sys.executable)
systemdrive = path.drive + path.root
if systemdrive=="/":
    systemdrive="/home/"+os.getlogin()

# systemdrive = os.getenv('WINDIR')[0:3]
# print(systemdrive)

dbfolder = os.path.join(systemdrive,'Forms','DB')
#dbfolder = "D:\Company Projects\Form creator\DB"
State_forms = os.path.join(systemdrive,'Forms','State forms')
#State_forms = "D:\Company Projects\Form creator\State forms"
Statefolder = Path(State_forms)
logfolder = os.path.join(systemdrive,'Forms','logs')

if not os.path.isdir(logfolder):
    os.makedirs(logfolder)
    

#logfolder = "D:\Company Projects\Form creator\logs"

Register_folder='Registers'
monthdict= {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}


log_filename = datetime.datetime.now().strftime(os.path.join(logfolder,'logfile_%d_%m_%Y_%H_%M_%S.log'))
logging.basicConfig(filename=log_filename, level=logging.INFO)

def read_min_wages_file(state_name,type_skilled,input_filelocation):
    formsfilelist = os.listdir(input_filelocation)
    for f in formsfilelist:
        if f[0:13].upper()=='MINIMUM WAGES':
            min_wagesfilename = f
            logging.info('min_wagesfilename is :'+f)
    min_wages=pd.read_excel(os.path.join(input_filelocation,min_wagesfilename))
    min_wages=min_wages.dropna(how="all",axis=1)
    min_wages=min_wages.dropna(how="all",axis=0)
    min_wages.columns=["SR NO" ,"STATE" ,"HIGHLY SKILLED" , "SKILLED" ,"SEMI-SKILLED" ,"UNSKILLED"]
    min_wages=min_wages.drop(["SR NO"],axis=1)
    min_wages=min_wages.set_index("STATE")
    min_wages=min_wages.loc[state_name,type_skilled]
    return min_wages

border_sides_thick = Side(style='thick')       
border_sides_thin = Side(style='thin')

def create_border(sheet,last_row,last_column,start_row):
    
    for c_idx in range(1,last_column):
        sheet.cell(row=last_row, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
    
    for r_idx in range(start_row,last_row):
        sheet.cell(row=r_idx, column=last_column).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin)
    
    sheet.cell(row=last_row, column=last_column).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
    
    return sheet

def cell_write(sheet,value,r_idx,c_idx):
    sheet.cell(row=r_idx, column=c_idx, value=value)
    sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
    sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
    border_sides = Side(style='thin')
    sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
    return sheet