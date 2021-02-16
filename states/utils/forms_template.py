from states import Testing
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Alignment, Side, numbers
employee_code_column = "Employee Code"


class Helper_functions():
    def __init__(self):
        self.columns_not_found = []
        self.merged_cells_bounds = []
    # Check if given value is numeric

    def if_num(self, value):
        import numbers
        return isinstance(value, numbers.Number)

    def get_data(self, data, columns):
        for column in columns:
            if column not in data.columns:
                data[column] = ""
                if column not in self.columns_not_found:
                    self.columns_not_found.append(column)
                if Testing:
                    raise Exception(
                        "{} not found in excel file,please add this header in one of the files".format(column))
        return data[columns]

    # Write value to required cell
    def cell_write(self, sheet, value, r_idx, c_idx):
        sheet.cell(row=r_idx, column=c_idx, value=value)
        sheet.cell(row=r_idx, column=c_idx).font = Font(
            name='Bell MT', size=15)
        sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        border_sides = Side(style='thin')
        sheet.cell(row=r_idx, column=c_idx).border = Border(outline=True, left=border_sides, top=border_sides,
                                                            right=border_sides, bottom=border_sides)
        if self.if_num(value):
            sheet.cell(
                row=r_idx, column=c_idx).number_format = numbers.FORMAT_NUMBER

    # Create borders
    def create_border(self, sheet, last_row, last_column, start_row, start_column):
        border_sides_thick = Side(style='thick')
        border_sides_thin = Side(style='thin')
        for c_idx in range(start_column, last_column):
            sheet.cell(row=last_row, column=c_idx).border = Border(
                outline=True, right=border_sides_thin, bottom=border_sides_thick)
        for r_idx in range(start_row, last_row):
            sheet.cell(row=r_idx, column=last_column).border = Border(
                outline=True, right=border_sides_thick, bottom=border_sides_thin)
        sheet.cell(row=last_row, column=last_column).border = Border(
            outline=True, right=border_sides_thick, bottom=border_sides_thick)

    def write_to_column(self, sheet, values, c_idx, start_row):
        for r_idx, value in enumerate(values, start_row):
            self.cell_write(sheet, value, r_idx, c_idx)

    def write_to_row(self, sheet, values, r_idx, start_column):
        for c_idx, value in enumerate(values, start_column):
            self.cell_write(sheet, value, r_idx, c_idx)

    '''
    This function writes data like Contrator name,Unit name which is only written once in the entire file
    '''

    def write_data_once_per_sheet(self, data_once_per_sheet, sheet):
        for location in data_once_per_sheet.keys():
            if sheet[location].value == None:
                sheet[location] = data_once_per_sheet[location]
            elif not str(data_once_per_sheet[location]).lower() in ["nan", "na"]:
                sheet[location] = sheet[location].value + \
                    "  "+str(data_once_per_sheet[location])

    '''
    Use if data of columns is to be combined, usefull since it will take care of datatypes of the columns.
    '''

    def combine_columns_of_dataframe(self, dataframe, columns, delimiter=","):
        columns_data = self.get_data(dataframe, columns).copy()
        columns_data.fillna(value="", inplace=True)
        columns_data["combined"] = ""
        for index, column in enumerate(columns, 1):
            if str(columns_data[column].dtype)[0:8] == 'datetime':
                columns_data[column] = columns_data[column].apply(
                    lambda x: x.strftime('%d-%m-%y'))
            elif str(columns_data[column].dtype)[0:3] != 'str':
                columns_data[column] = columns_data[column].astype(str)
            if not index == len(columns):
                columns_data["combined"] += columns_data[column]+delimiter
            else:
                columns_data["combined"] += columns_data[column]
        return columns_data["combined"]

    '''
    Use if data of columns is to be summed, usefull since it will take care of datatypes of the columns.
    '''

    def sum_columns_of_dataframe(self, dataframe, columns):
        columns_data = self.get_data(dataframe, columns).copy()
        columns_data.fillna(value=0, inplace=True)
        columns_data["sum"] = 0
        for column in columns:
            columns_data[column] = columns_data[column].astype(float)
            columns_data["sum"] += columns_data[column]
        return columns_data["sum"]

    '''
    Used to get attendance columns()
    '''

    def get_attendance_columns(self, data):
        columnstotake = []
        days = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18',
                '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31']
        for day in days:
            for col in data.columns:
                if col[5:7] == day:
                    columnstotake.append(col)
        if len(columnstotake) == 28:
            columnstotake.append('29')
            columnstotake.append('30')
            columnstotake.append('31')
            data['29'] = ''
            data['30'] = ''
            data['31'] = ''

        elif len(columnstotake) == 29:
            columnstotake.append('30')
            columnstotake.append('31')
            data['30'] = ''
            data['31'] = ''

        elif len(columnstotake) == 30:
            columnstotake.append('31')
            data['31'] = ''
        elif len(columnstotake) == 31:
            pass
        else:
            raise Exception(
                "Didnot find all attendance columns, please check format")
        return columnstotake

    '''
    This function is used to get data if for every sheet(new sheet per employee) 
    data to be published once per sheet depends on the employee.
    So this will return a dictionary where key is the employee code and value 
    is the dictionary containing location and the value to be published.
    It takes input the dataframe and a mapping dictionary which should have key as 
    location on the sheet and value as the column name from the datatframe
    '''

    def get_data_once_persheet_peremployee(self, data, mapping):
        data_once_per_sheet = {}
        for index, row in data.iterrows():
            emp_code = row[employee_code_column]
            temp = {}
            for location, column in mapping.items():
                temp[location] = row[column]
            data_once_per_sheet[emp_code] = temp
        return data_once_per_sheet

    # Not used functions
    def unmerge_cells(self, sheet, start_row):
        for item in sheet.merged_cell_ranges:
            if item.bounds[1] >= start_row:
                self.merged_cells_bounds.append(item.bounds)
                sheet.unmerge_cells(str(item))

    def merge_cells(self, sheet, num_rows_added):
        for cell in self.merged_cells_bounds:
            sheet.merge_cells(start_row=cell[1]+num_rows_added, start_column=cell[0],
                              end_row=cell[3]+num_rows_added, end_column=cell[2])
        self.merged_cells_bounds = []


class Templates(Helper_functions):
    def __init__(self, to_read, to_write, month, year, report, master):
        self.to_read = to_read
        self.to_write = to_write
        self.report = report
        self.master = master
        self.month = month
        self.year = year
        super().__init__()

    '''
    This function will create basic forms which will have only one sheet and 
    will keep adding data of each employee one below other
    data_once_per_sheet is dict such that key is position and value is actual value to be populated,
    only pass in data which is only used once, like company name,
    contrator address etc
    '''

    def create_basic_form(self, filename, sheet_name, all_employee_data,
                          start_row, start_column, data_once_per_sheet):
        # get path from which blank xl file to read
        file_read = os.path.join(self.to_read, filename)
        # Check if that file exsists
        if not os.path.exists(file_read):
            raise FileNotFoundError(file_read)
        # Load xl file
        work_book = openpyxl.load_workbook(file_read)
        # Check if the specifies sheet is present or not in that file
        if not sheet_name in work_book.sheetnames:
            raise Exception("Sheet {} not found in file {}".format(
                sheet_name, file_read))
        # Get the sheet
        sheet = work_book[sheet_name]
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        # Convert dataframe to rows , to populate the required information
        rows = dataframe_to_rows(all_employee_data, index=False, header=False)
        # Iterate over to populate the values
        r_idx, c_idx = 0, 0
        for r_idx, row in enumerate(rows, start_row):
            for c_idx, value in enumerate(row, start_column):
                self.cell_write(sheet, value, r_idx, c_idx)
            sheet.insert_rows(r_idx+1)
        # create borders
        self.create_border(sheet, last_row=r_idx, last_column=c_idx,
                           start_row=start_row, start_column=start_column)
        # Write data like company name,unit name etc
        self.write_data_once_per_sheet(data_once_per_sheet, sheet)
        # Get path to save file
        file_write = os.path.join(self.to_write, filename)
        # Save the file
        work_book.save(filename=file_write)
        # Return how many lines were written in the file
        return r_idx

    '''
    This function can be used to create forms which require to create new sheet for each employee,
    Here all employee data can be none and one can only send data_once_persheet which will 
    create new sheet per employee and populate the data_once_persheetdata
    If per_employee_diff is True then data_once_persheet should be dictionary(This 
    one can get using the get_data_once_persheet_peremployee fucntion) 
    such that key is employee code and for each key there will be a dictionary which will 
    be used to populate the sheet, hence is usefull to populate employee specific data for each new sheet of the employee
    If all_amployee data is given then one can populate information same as basic form for each sheet per employee
    '''

    def create_per_employee_basic_form(self, filename, sheet_name, start_row, start_column,
                                       employee_codes, data_once_per_sheet, per_employee_diff_data, all_employee_data=None):
        file_read = os.path.join(self.to_read, filename)
        if not os.path.exists(file_read):
            raise FileNotFoundError(file_read)
        work_book = openpyxl.load_workbook(file_read)
        if not sheet_name in work_book.sheetnames:
            raise Exception("Sheet {} not found".format(sheet_name))

        if not sheet_name in work_book.sheetnames:
            raise Exception("Sheet {} not found in file {}".format(
                sheet_name, file_read))
        original_sheet = work_book[sheet_name]

        rows_added = 1
        r_idx = start_row
        c_idx = 0
        if not all_employee_data == None:
            rows = dataframe_to_rows(
                all_employee_data, index=False, header=False)
            for row, emp_code in zip(rows, employee_codes):
                sheet = work_book.copy_worksheet(original_sheet)
                sheet.title = emp_code
                sheet.sheet_properties.pageSetUpPr.fitToPage = True
                if per_employee_diff_data:
                    self.write_data_once_per_sheet(
                        data_once_per_sheet[emp_code], sheet)
                else:
                    self.write_data_once_per_sheet(data_once_per_sheet, sheet)
                for c_idx, value in enumerate(row, start_column):
                    self.cell_write(sheet, value, r_idx, c_idx)
                self.create_border(sheet, last_row=r_idx, last_column=c_idx,
                                   start_row=start_row, start_column=start_column)
        else:
            for emp_code in employee_codes:
                sheet = work_book.copy_worksheet(original_sheet)
                sheet.title = emp_code
                if per_employee_diff_data:
                    self.write_data_once_per_sheet(
                        data_once_per_sheet[emp_code], sheet)
                else:
                    self.write_data_once_per_sheet(data_once_per_sheet, sheet)

        #work_book.remove(original_sheet)
        file_write = os.path.join(self.to_write, filename)
        work_book.save(filename=file_write)

        return rows_added

    '''
    This function is used to get the from and to dates from the attendacne file for which the person was absent
    if sno_column is not empty strig then it should be a column which is the sr. number to be 
    populated in the form,will be used ful if one wants sr number in proper order.
    If it is empty string then sr number column (if to be published) will be very messy
    '''

    def get_from_to_dates_attendance(self, data, absent_label, sno_column=""):
        data = data.drop_duplicates(subset=employee_code_column, keep="last")
        columns = [employee_code_column]
        columns.extend(self.get_attendance_columns(data))
        attendance = self.get_data(data, columns)
        rows = dataframe_to_rows(attendance, index=False, header=False)

        data[["from", "to", "numdays"]] = ""
        temp_df = None
        is_abs_num = 0
        num = {}
        for row in rows:
            for idx, value in enumerate(row):
                if idx == 0:
                    emp_code = value
                    if emp_code not in num.keys():
                        num[emp_code] = 0
                    temp_df = data.loc[data[employee_code_column]
                                       == emp_code].iloc[0].copy(deep=True)
                elif is_abs_num == 0 and value == absent_label:
                    is_abs_num = 1
                    start = columns[idx]
                    end = columns[idx]
                elif value == absent_label:
                    is_abs_num += 1
                    end = columns[idx]
                elif is_abs_num:
                    start = start.split("\n")[1].replace(
                        "/", "-")+"-"+str(self.year)
                    end = end.split("\n")[1].replace(
                        "/", "-")+"-"+str(self.year)
                    temp_df["from"] = start
                    temp_df["to"] = end
                    temp_df["numdays"] = is_abs_num
                    if not (sno_column == ""):
                        num[emp_code] += 1
                        temp_df[sno_column] = num[emp_code]
                    data = data.append([temp_df], ignore_index=True)
                    is_abs_num = 0

            if is_abs_num:
                start = start.split("\n")[1].replace(
                    "/", "-")+"-"+str(self.year)
                end = end.split("\n")[1].replace("/", "-")+"-"+str(self.year)
                temp_df["from"] = start
                temp_df["to"] = end
                temp_df["numdays"] = is_abs_num
                if not (sno_column == ""):
                    num[emp_code] += 1
                    temp_df[sno_column] = num[emp_code]
                data = data.append([temp_df], ignore_index=True)
                is_abs_num = 0
            temp_df = None

        data = data.loc[data["numdays"] != "", :]
        return data

    '''
    This is used topopulate a attendance file such that new sheet is created per employee.
    This function is useful as it will take care of the data which eventually comes from get_from_to_dates_attendance which will create duplicate
    rows for each employee for each from and to date found
    So this fucntion will populate the from and to date properly in each sheet
    '''

    def create_attendance_form_per_employee(self, filename, sheet_name, start_row, start_column,
                                            data_with_attendance, columns, data_once_per_sheet, per_employee_diff_data):

        all_employee_data = self.get_data(data_with_attendance, columns)
        employee_codes = data_with_attendance[employee_code_column]

        file_read = os.path.join(self.to_read, filename)
        if not os.path.exists(file_read):
            raise FileNotFoundError(file_read)
        work_book = openpyxl.load_workbook(file_read)
        if not sheet_name in work_book.sheetnames:
            raise Exception("Sheet {} not found".format(sheet_name))

        original_sheet = work_book[sheet_name]

        rows_added = {}
        r_idx = start_row
        c_idx = 0

        rows = dataframe_to_rows(all_employee_data, index=False, header=False)

        for row, emp_code in zip(rows, employee_codes):
            if not emp_code in work_book.sheetnames:
                sheet = work_book.copy_worksheet(original_sheet)
                sheet.title = emp_code
                sheet.sheet_properties.pageSetUpPr.fitToPage = True
                rows_added[emp_code] = 0
                if per_employee_diff_data:
                    self.write_data_once_per_sheet(
                        data_once_per_sheet[emp_code], sheet)
                else:
                    self.write_data_once_per_sheet(data_once_per_sheet, sheet)
            else:
                sheet = work_book[emp_code]
                rows_added[emp_code] += 1

            for c_idx, value in enumerate(row, start_column):
                self.cell_write(sheet, value, r_idx +
                                rows_added[emp_code], c_idx)
        for emp_code in employee_codes:
            self.create_border(sheet=work_book[emp_code], last_row=r_idx+rows_added[emp_code], last_column=c_idx,
                               start_row=start_row, start_column=start_column)

        #work_book.remove(original_sheet)
        file_write = os.path.join(self.to_write, filename)
        work_book.save(filename=file_write)

        return rows_added