from states import logging,monthdict,Statefolder,read_min_wages_file
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill, numbers

def Gujarat(data,contractor_name,contractor_address,filelocation,month,year,report,master):
    Gujaratfilespath = os.path.join(Statefolder,'Gujarat')
    logging.info('Gujarat files path is :'+str(Gujaratfilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]
    input_filelocation=filelocation.split("Registers")[0]
    min_wages_gujarat=read_min_wages_file("GUJARAT","SEMI-SKILLED",input_filelocation)

    def Form_F():
        formFfilepath = os.path.join(Gujaratfilespath,'Form F Register of refusal of leave.xlsx')
        formFfile = load_workbook(filename=formFfilepath)
        logging.info('Form F file has sheet: '+str(formFfile.sheetnames))
        logging.info('create columns which are now available')

        data_formF = data.copy(deep=True)
        data_formF=data_formF.drop_duplicates(subset="Employee Code", keep="last")

        columns=['S.no',"name_employer",'Company Name',"Address","Employee Name","Leave_due","Encash","Date_of_refusal","sign","remarks"]
        
        data_formF["name_employer"]=""
        data_formF["PE_or_contract"]=data_formF["PE_or_contract"].astype(str)
        data_formF.loc[data_formF["PE_or_contract"]=="PE","name_employer"]=data_formF.loc[data_formF["PE_or_contract"]=="PE",'Company Name']
        data_formF.loc[data_formF["PE_or_contract"]=="Contract","name_employer"]=data_formF.loc[data_formF["PE_or_contract"]=="Contract",'UnitName']
        
        data_formF[["Leave_due","Encash","Date_of_refusal","sign","remarks"]]="---"
        data_formF['S.no'] = list(range(1,len(data_formF)+1))
        formF_data=data_formF[columns]
        formFsheet = formFfile['Sheet1']
        formFsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form F is ready')

        
        rows = dataframe_to_rows(formF_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 10):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formFsheet.cell(row=r_idx, column=c_idx, value=value)
                formFsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formFsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formFsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formF):
                    formFsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formFsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formFsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formFsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formF):
                    formFsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formFsheet.row_dimensions[r_idx].height = 20
                else:
                    formFsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formFsheet.row_dimensions[r_idx].height = 20

        formFsheet['A4']=formFsheet['A4'].value+"   "+str(data_formF['Company Name'].unique()[0])+","+str(data_formF['Company Address'].unique()[0])
        if data_formF["PE_or_contract"].unique()[0].upper()=="CL":
            formFsheet['A5']=formFsheet['A5'].value+"   "+data_formF['UnitName'].unique()[0]+" "+data_formF["Address"].unique()[0]
        
        formFsheet['A6']=formFsheet['A6'].value+"   "+str(data_formF['Branch'].unique()[0])
        formFsheet['A7']="WAGE PERIOD    "+str(month)+" "+str(year)

        formFfinalfile = os.path.join(filelocation,'Form F Register of refusal of leave.xlsx')
        formFfile.save(filename=formFfinalfile)

    def Form_I():
        formIfilepath = os.path.join(Gujaratfilespath,'Form I Register of employment in a shop.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')
        
        data_formI = data.copy(deep=True)
        data_formI=data_formI.drop_duplicates(subset="Employee Code", keep="last")
        
        columns=["Employee Name","Gender","Age","start_time","end_time","rest_interval","mon","tue","wed","thu","Fri","sat","sun",
                                                "days_overtime","extent_of_overtime","extent_of_overtime_previously"]
    
        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        
        data_formI[["mon","tue","wed","thu","Fri","sat","sun","days_overtime","extent_of_overtime","extent_of_overtime_previously"]]=""
        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')

        formIsheet.unmerge_cells("A8:P8")
        formIsheet.unmerge_cells("A9:P9")
        
        formIsheet.insert_rows(7,len(data_formI))
        row_num=0
        for r_idx, row in enumerate(rows, 7):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIsheet.row_dimensions[r_idx].height = 20
                else:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIsheet.row_dimensions[r_idx].height = 20
        
        formIsheet.merge_cells("A"+str(8+len(data_formI))+":P"+str(8+len(data_formI)))
        formIsheet.merge_cells("A"+str(9+len(data_formI))+":P"+str(9+len(data_formI)))

        formIfinalfile = os.path.join(filelocation,'Form I Register of employment in a shop.xlsx')
        formIfile.save(filename=formIfinalfile)

    def Form_IV():
        
        formIVfilepath = os.path.join(Gujaratfilespath,'Form IV A register  of wages.xlsx')
        formIVfile = load_workbook(filename=formIVfilepath)
        logging.info('Form IV file has sheet: '+str(formIVfile.sheetnames))
        logging.info('create columns which are now available')

        data_formIV = data.copy(deep=True)
        data_formIV=data_formIV.drop_duplicates(subset="Employee Code", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Designation","basic","DA","Earned Basic","DA","Days Paid",
                                        "Overtime","HRA",'Tel and Int Reimb',"Bonus","Fuel Reimb","Prof Dev Reimb","Corp Attire Reimb","CCA",
                                        "deductions-advance",'Total Earning','PF',"H.R.","all_Other_deductions","Insurance","P.Tax","Total Deductions","Net Paid",
                                        "Date of payment","Bank A/c Number","sign"]
        
        remove_point=lambda input_str: input_str.split(".")[0]
        data_formIV["Bank A/c Number"]=data_formIV["Bank A/c Number"].apply(str).apply(remove_point)

        data_formIV['S.no'] = list(range(1,len(data_formIV)+1))
        data_formIV["basic"]=min_wages_gujarat
        #others_columns=[]
        others_columns=['HRA','Conveyance','Medical Allowance','Telephone Reimb','Tel and Int Reimb',
                                            'Bonus','Other Allowance', 'Fuel Reimb','Prof Dev Reimb','Corp Attire Reimb',
                                            'Meal Allowance','Special Allowance','Personal Allowance','CCA','Other Reimb',
                                            'Arrears','Other Earning',"Retention Pay",'Variable Pay','Leave Encashment',
                                            'Stipend','Consultancy Fees','Covid Deduction','OtherAllowance1', 
                                            'OtherAllowance2', 'OtherAllowance3', 'OtherAllowance4', 'OtherAllowance5'
                                            ]
        if "Covid Deduction" not in data_formIV.columns:
            data_formIV["Covid Deduction"]=0
        if "Retention Pay" not in data_formIV.columns:
            data_formIV["Retention Pay"]=0
        # data_formIV["deductions-advance"]=data
        data_formIV[others_columns]=data_formIV[others_columns].replace("",0).astype(float)
        data_formIV[others_columns]=data_formIV[others_columns].fillna(0)
        data_formIV['Salary Advance']=data_formIV['Salary Advance'].replace("",0).astype(float)
        data_formIV['Salary Advance']=data_formIV['Salary Advance'].fillna(0)
        
        data_formIV["deductions-advance"]= data_formIV.loc[:,others_columns].sum(axis=1)-data_formIV['Salary Advance']
        data_formIV["H.R."]=0

        all_Other_deductions_columns=['Other Deduction','OtherDeduction1', 'OtherDeduction2','OtherDeduction3', 'OtherDeduction4', 'OtherDeduction5']
        
        data_formIV[all_Other_deductions_columns]=data_formIV[all_Other_deductions_columns].replace("",0).astype(float)
        data_formIV[all_Other_deductions_columns]=data_formIV[all_Other_deductions_columns].fillna(0)

        data_formIV["all_Other_deductions"]=data_formIV.loc[:,all_Other_deductions_columns].sum(axis=1)

        data_formIV["sign"]=""
        #data_formIV["Date_overtime_worked"]=month
        formIV_data=data_formIV[columns]
        formIVsheet = formIVfile['Sheet1']
        formIVsheet.sheet_properties.pageSetUpPr.fitToPage = True
        #for column in  range(ord('A'), ord('O') + 1):
        #    formIVsheet.unmerge_cells(chr(column)+"7:"+chr(column)+"14")

        logging.info('data for form IV is ready')

        
        rows = dataframe_to_rows(formIV_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows,10):
            row_num+=1
        
            for c_idx, value in enumerate(row, 1):
                formIVsheet.cell(row=r_idx, column=c_idx, value=value)
                formIVsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIVsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formIV):
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIVsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIVsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formIV):
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIVsheet.row_dimensions[r_idx].height = 20
                else:
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIVsheet.row_dimensions[r_idx].height = 20

        formIVsheet['A3']=formIVsheet['A3'].value+" "+str(data_formIV['Company Name'].unique()[0])
        if data["PE_or_contract"].unique()[0].upper()=="CL":
            formXXIIIsheet['A4']=" Name of Establishment:-   "+str(data_formIV['UnitName'].unique()[0])

        formIVsheet['A5']="PLACE "+data_formIV['Branch'].unique()[0]
        formIVsheet['A6']=formIVsheet['A6'].value+" "+str(month)+" "+str(year)
        formIVfinalfile = os.path.join(filelocation,'Form IV A register  of wages.xlsx')
        formIVfile.save(filename=formIVfinalfile)
                    
    def Form_M():
        formMfilepath = os.path.join(Gujaratfilespath,'Form M Register of leave.xlsx')
        formMfile = load_workbook(filename=formMfilepath)
        logging.info('Form M file has sheet: '+str(formMfile.sheetnames))
        logging.info('create columns which are now available')

        data_formM = data.copy(deep=True)
        leave_file_data=data_formM[["Employee Code","Employee Name","Leave Type","Opening","Monthly Increment","Leave Accrued","Used","Encash","Closing"]]
        
        data_formM=data_formM.drop_duplicates(subset="Employee Code", keep="last")
        data_formM["Employee Name & Code"]=data_formM["Employee Name"].astype(str)+"||"+data_formM["Employee Code"].astype(str)

        columns=["Employee Name & Code","Department","Date Joined","month_year","num_days","balance_days","Date Left",'Leave Encashment']
        data_formM["month_year"]=str(month)+" "+str(year)

        for employee_name_leave_file in data_formM["Employee Name"]:
            #opening
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==employee_name_leave_file,:]
            opening_pl=emp_details["Opening"].loc[emp_details["Leave Type"]=="PL"].replace("",0).astype(float)
            opening_cl=emp_details["Opening"].loc[emp_details["Leave Type"]=="CL"].replace("",0).astype(float)
            opening_sl=emp_details["Opening"].loc[emp_details["Leave Type"]=="SL"].replace("",0).astype(float)
            prev_bal=opening_pl.add(opening_cl.add(opening_sl,fill_value=0), fill_value=0).sum()
            
            data_formM.loc[data_formM["Employee Name"]==employee_name_leave_file,'num_days']=prev_bal


            #####
            #monthly_inr
            mon_inr_pl=emp_details["Monthly Increment"].loc[emp_details["Leave Type"]=="PL"].replace("",0).astype(float)
            mon_inr_cl=emp_details["Monthly Increment"].loc[emp_details["Leave Type"]=="CL"].replace("",0).astype(float)
            mon_inr_sl=emp_details["Monthly Increment"].loc[emp_details["Leave Type"]=="SL"].replace("",0).astype(float)
            earned=mon_inr_cl.add(mon_inr_pl.add(mon_inr_sl,fill_value=0), fill_value=0).sum()

            #availed during month
            Used_pl=emp_details["Used"].loc[emp_details["Leave Type"]=="PL"].replace("",0).astype(float)
            Used_cl=emp_details["Used"].loc[emp_details["Leave Type"]=="CL"].replace("",0).astype(float)
            Used_sl=emp_details["Used"].loc[emp_details["Leave Type"]=="SL"].replace("",0).astype(float)
            availed=Used_cl.add(Used_pl.add(Used_sl,fill_value=0), fill_value=0).sum()

            data_formM.loc[data_formM["Employee Name"]==employee_name_leave_file,"balance_days"]=prev_bal+earned-availed

            


        data_formM[['balance_days']]=""
        # data_formM_columns=list(data_formM.columns)
        # start=data_formM_columns.index('Arrears salary')
        # end=data_formM_columns.index('Total\r\nDP')

        columnstotake =[]
        days = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31']
        for day in days:
            for col in data_formM.columns:
                if col[5:7]==day:
                    columnstotake.append(col)
        if len(columnstotake)==28:

            columnstotake.append('29')
            columnstotake.append('30')
            columnstotake.append('31')
            data_formM['29'] = ''
            data_formM['30'] = ''
            data_formM['31'] = ''
            
        elif len(columnstotake)==29:
            columnstotake.append('30')
            columnstotake.append('31')
            data_formM['30'] = ''
            data_formM['31'] = ''

        elif len(columnstotake)==30:
            columnstotake.append('31')
            data_formM['31'] = ''
        
        columns.extend(columnstotake)

        formM_data=data_formM[columns]
        formMsheet = formMfile['Sheet1']

        formMsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form M is ready')

        
        rows = dataframe_to_rows(formM_data, index=False, header=False)

        logging.info('rows taken out from data')

        def cell_write(sheet,r_idx,c_idx,value):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            
        def start_end_date_attendance(absent_label,offset,row_offset,initial_offset):  
            is_abs_num=0
            row_index=0
            added={}
            for sheet_idx, row in enumerate(rows, 10):
                row_index=0
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        if code =="nan":
                            code=name
                        try:
                            target=formMfile[code]
                            added[target.title]=0
                        except:
                            target = formMfile.copy_worksheet(formMsheet)
                            target.title=code
                            #initial offset
                            row_offset[code]=initial_offset
                            added[target.title]=initial_offset
                        target['A5']="Name of Employee : "+name
                        formMsheet['A4']="Name of the employer or the establishment:- "+str(data_formM['Company Name'].unique()[0])+","+str(data_formM['Company Address'].unique()[0])    
                    elif c_idx==2:
                        target['A6']="Description of the department (if applicable):    "+value
                    elif c_idx==3:
                        target['A7']="Date of entry into service:  "+value
                    elif c_idx==4:
                        Leave_Accrued=value 
                        cell_write(target,row_index+row_offset[target.title],1,Leave_Accrued)
                    elif c_idx==5:
                        num_days=value
                        cell_write(target,row_index+row_offset[target.title],2,num_days)
                    elif c_idx==6:
                        balance_days=value
                        cell_write(target,row_index+row_offset[target.title],5,balance_days)
                    elif c_idx==7:
                        Date_Left=value
                        cell_write(target,row_index+row_offset[target.title],9,Date_Left)
                    elif c_idx==8:
                        Date_of_payment=value
                        cell_write(target,row_index+row_offset[target.title],10,Date_of_payment)
                    elif is_abs_num==0 and value==absent_label:
                        is_abs_num=1
                        start=columns[c_idx-1]
                        end=columns[c_idx-1]
                    elif value==absent_label:
                        is_abs_num+=1
                        end=columns[c_idx-1]
                    elif is_abs_num:
                        #target.cell(row=row_index+13, column=1+column_offset, value=is_abs_num)
                        start=start.split("\n")[1].replace("/","-")+"-"+str(year)
                        end=end.split("\n")[1].replace("/","-")+"-"+str(year)
                        cell_write(target,row_index+row_offset[target.title],3+offset,start)
                        cell_write(target,row_index+row_offset[target.title],4+offset,end)
                        cell_write(target,row_index+row_offset[target.title],10,Date_of_payment)
                        cell_write(target,row_index+row_offset[target.title],9,Date_Left)
                        cell_write(target,row_index+row_offset[target.title],5,balance_days)
                        cell_write(target,row_index+row_offset[target.title],2,num_days)
                        cell_write(target,row_index+row_offset[target.title],1,Leave_Accrued)

                        target['F'+str(row_index+row_offset[target.title])]="----"
                        target['G'+str(row_index+row_offset[target.title])]="----"
                        target['H'+str(row_index+row_offset[target.title])]="----"
                        target.insert_rows(row_index+row_offset[target.title]+1)
                        is_abs_num=0
                        row_index+=1
                        added[target.title]+=1
            return added            
                    
        absent_label="PL"
        column_offset=0
        initial_offset=13
        row_offset={}
        from collections import Counter
        row_offset=Counter(start_end_date_attendance(absent_label,column_offset,row_offset,initial_offset))
        absent_label="SL"
        
        row_offset+=Counter(start_end_date_attendance(absent_label,column_offset,row_offset,initial_offset))
        absent_label="CL"
        
        row_offset+=Counter(start_end_date_attendance(absent_label,column_offset,row_offset,initial_offset))
        
        formMfile.remove(formMfile["Sheet1"])
        formMfile.remove(formMfile["Sheet2"])
        formMfile.remove(formMfile["Sheet3"])
        formMfinalfile = os.path.join(filelocation,'Form M Register of leave.xlsx')
        formMfile.save(filename=formMfinalfile)

    def Form_P():
        formPfilepath = os.path.join(Gujaratfilespath,'Form P Muster roll.xlsx')
        formPfile = load_workbook(filename=formPfilepath)
        logging.info('Form P file has sheet: '+str(formPfile.sheetnames))
        logging.info('create columns which are now available')

        data_formP = data.copy(deep=True)
        data_formP=data_formP.drop_duplicates(subset="Employee Code", keep="last")
        
        columns=['S.no',"Employee Name","Designation","Age","Gender","Date Joined","start_time",
                                                                "end_time",'interval_for_reset_from','interval_for_reset_to']
        data_formP['interval_for_reset_to']=data_formP.rest_interval.str.split("-",expand=True)[1]
        data_formP['interval_for_reset_from']=data_formP.rest_interval.str.split("-",expand=True)[0]
        # data_formP_columns=list(data_formP.columns)
        # start=data_formP_columns.index('Emp Code')
        # end=data_formP_columns.index('Total\r\nDP')
        
        columnstotake =[]
        days = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31']
        for day in days:
            for col in data_formP.columns:
                if col[5:7]==day:
                    columnstotake.append(col)
        if len(columnstotake)==28:

            columnstotake.append('29')
            columnstotake.append('30')
            columnstotake.append('31')
            data_formP['29'] = ''
            data_formP['30'] = ''
            data_formP['31'] = ''
            
        elif len(columnstotake)==29:
            columnstotake.append('30')
            columnstotake.append('31')
            data_formP['30'] = ''
            data_formP['31'] = ''

        elif len(columnstotake)==30:
            columnstotake.append('31')
            data_formP['31'] = ''
        
        columns.extend(columnstotake)
        
        columns.append('Total\r\nDP')
        data_formP['S.no'] = list(range(1,len(data_formP)+1))

        formP_data=data_formP[columns]
        formPsheet = formPfile['Sheet1']
        formPsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form P is ready')
        
        for i in range(9,20):
            formPsheet["A"+str(i)]=""

        
        rows = dataframe_to_rows(formP_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 10):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formPsheet.cell(row=r_idx, column=c_idx, value=value)
                formPsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formPsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formPsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formP):
                    formPsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formPsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formPsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formPsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formP):
                    formPsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formPsheet.row_dimensions[r_idx].height = 20
                else:
                    formPsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formPsheet.row_dimensions[r_idx].height = 20

        
        #formPsheet['AE4']=formPsheet['AE4'].value+"   "+str(data_formP['Registration_no'].unique()[0])
        
        formPsheet['A4']="Name of establishment :-  "+str(data_formP['Company Name'].unique()[0])+","+str(data_formP['Company Address'].unique()[0])
        if data["PE_or_contract"].unique()[0].upper()=="CL":
            formPsheet['A5']="Name of the employer:-    "+str(data_formP['UnitName'].unique()[0])


        formPsheet['A6']=formPsheet['A6'].value+" "+str(month)+" "+str(year)
        formPsheet['N4']="Place   "+data_formP['Branch'].unique()[0]
        formPfinalfile = os.path.join(filelocation,'Form P Muster roll.xlsx')
        formPfile.save(filename=formPfinalfile)

    

    def Form_Notice_holiday():
        formNotice_holidayfilepath = os.path.join(Gujaratfilespath,'Notice of holiday.xlsx')
        formNotice_holidayfile = load_workbook(filename=formNotice_holidayfilepath)
        logging.info('Form Notice_holiday file has sheet: '+str(formNotice_holidayfile.sheetnames))
        logging.info('create columns which are now available')

        data_formNotice_holiday = data.copy(deep=True)
        data_formNotice_holiday=data_formNotice_holiday.drop_duplicates(subset="Employee Code", keep="last")
        columns=["Employee Name","day_holiday_allowed"]

        data_formNotice_holiday["day_holiday_allowed"]="Sunday , Saturday"
        data_formNotice_holiday['S.no'] = list(range(1,len(data_formNotice_holiday)+1))

        formNotice_holiday_data=data_formNotice_holiday[columns]
        formNotice_holidaysheet = formNotice_holidayfile['Sheet1']
        formNotice_holidaysheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form Notice_holiday is ready')
        
        
        rows = dataframe_to_rows(formNotice_holiday_data, index=False, header=False)

        logging.info('rows taken out from data')
        if len(data_formNotice_holiday)>7:
            formNotice_holidaysheet.insert_rows(15,len(data_formNotice_holiday)-8)

        for r_idx, row in enumerate(rows, 14):
            for c_idx, value in enumerate(row, 1):
                formNotice_holidaysheet.cell(row=r_idx, column=c_idx, value=value)
                formNotice_holidaysheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formNotice_holidaysheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formNotice_holidaysheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                
        
        formNotice_holidaysheet['A3']=formNotice_holidaysheet['A3'].value+" "+str(data_formNotice_holiday['Unit'].unique()[0])+", "+str(data_formNotice_holiday['Address'].unique()[0])
        formNotice_holidayfinalfile = os.path.join(filelocation,'Notice of holiday.xlsx')
        formNotice_holidayfile.save(filename=formNotice_holidayfinalfile)

    try:
        Form_F()
        master.update()
        Form_IV()
        master.update()
        Form_M()
        master.update()
        Form_P()
        master.update()
        Form_Notice_holiday()
        master.update()
        #No need
        # Form_I()
    except KeyError as e:
        logging.info("Key error : Check if {} column exsists".format(e))
        print("Key error {}".format(e))
        report.configure(text="Failed: Check input file format  \n column {} not found".format(e))
        master.update()
        raise KeyError
    except FileNotFoundError as e:
        logging.info("File not found : Check if {} exsists".format(e))
        report.configure(text="Failed: File {} not found".format(e))
        master.update()
        raise FileNotFoundError