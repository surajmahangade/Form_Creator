from states import logging,monthdict,Statefolder
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill, numbers


def Central_Process(data,contractor_name,contractor_address,filelocation,month,year,report,master):
    Centralfilespath = os.path.join(Statefolder,'Central')
    logging.info('Central files path is :'+str(Centralfilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]
    #Comment this line if in future leave file data is needed in any of functions below
    data=data.drop_duplicates(subset='Employee Code', keep="last")
    
    def Form_C():
        formCfilepath = os.path.join(Centralfilespath,'Form C Format of register of loan.xlsx')
        formCfile = load_workbook(filename=formCfilepath)
        logging.info('Form C file has sheet: '+str(formCfile.sheetnames))
        logging.info('create columns which are now available')

        data_formC = data.copy(deep=True)
        data_formC=data_formC.drop_duplicates(subset="Employee Code", keep="last")
        columns=['Employee Code',"Employee Name","Recovery_Type","Particulars","Date of payment","amount","whether_show_cause_issue","explaination_heard_in_presence_of",
                                    "num_installments","first_month_year","last_month_year","Date_of_complete_recovery","remarks"]
        

        Recovery_Type_columns_name=['Other Deduction','OtherDeduction1', 'OtherDeduction2',
                                                        'OtherDeduction3', 'OtherDeduction4', 'OtherDeduction5','Damage or Loss','Fine','Salary Advance']

        data_formC["Recovery_Type"]=data_formC.loc[:,Recovery_Type_columns_name].sum(axis=1)
        data_formC["amount"]=data_formC["Recovery_Type"]
        data_formC[["Particulars","whether_show_cause_issue","explaination_heard_in_presence_of",
                    "num_installments","first_month_year","last_month_year","Date_of_complete_recovery"]]="---"
        data_formC["remarks"]=""
        formC_data=data_formC[columns]
        formCsheet = formCfile['Sheet1']
        formCsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        
        rows = dataframe_to_rows(formC_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 9):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formCsheet.cell(row=r_idx, column=c_idx, value=value)
                formCsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formCsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formCsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formC):
                    formCsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formCsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formCsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formCsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formC):
                    formCsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formCsheet.row_dimensions[r_idx].height = 20
                else:
                    formCsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formCsheet.row_dimensions[r_idx].height = 20

        formCsheet['A4']=formCsheet['A4'].value+" : "+str(data_formC['UnitName'].unique()[0])
        formCfinalfile = os.path.join(filelocation,'Form C Format of register of loan.xlsx')
        formCfile.save(filename=formCfinalfile)
    
    def Form_I():
        formIfilepath = os.path.join(Centralfilespath,'Form I register of Fine.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formI = data.copy(deep=True)
        data_formI=data_formI.drop_duplicates(subset="Employee Code", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","name&date_of_offence","cause_against_fine","FIXED MONTHLY GROSS",
                                        "Date of payment_fine_released","Date of payment_fine_imposed","remarks"]

        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        data_formI[["name&date_of_offence","cause_against_fine"]]="---"
        
        data_formI['Fine']=data_formI['Fine'].astype(float)
        data_formI["Date of payment_fine_released"]=data_formI['Date of payment']
        data_formI["Date of payment_fine_imposed"]=data_formI['Date of payment']
        # data_formI.loc[data_formI['Fine']==0,["FIXED MONTHLY GROSS","Date of payment_fine_released","Date of payment_fine_imposed"]]="---"
        data_formI.loc[:,["FIXED MONTHLY GROSS","Date of payment_fine_released","Date of payment_fine_imposed"]]="---"
        
        data_formI["remarks"]=""

        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 8):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIsheet.row_dimensions[r_idx].height = 20
                else:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIsheet.row_dimensions[r_idx].height = 20

        formIsheet['A4']=formIsheet['A4'].value+" : "+str(data_formI['UnitName'].unique()[0])
        formIfinalfile = os.path.join(filelocation,'Form I register of Fine.xlsx')
        formIfile.save(filename=formIfinalfile)

    def Form_II_reg_damage_loss():
        formIIfilepath = os.path.join(Centralfilespath,'Form II Register of deductions for damage or loss.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy(deep=True)
        data_formII=data_formII.drop_duplicates(subset="Employee Code", keep="last")

        data_formII.fillna(value=0, inplace=True)
        #print(sorted(data_formII.columns))
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","Damage or Loss","whether_work_showed_cause",
                                        "Date of payment & amount of deduction","num_instalments","Date of payment","remarks"]
        
        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII[["Damage or Loss","whether_work_showed_cause","Date of payment & amount of deduction","num_instalments","Date of payment","remarks"]]="---"
        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        
        rows = dataframe_to_rows(formII_data, index=False, header=False)
        border_sides_thick = Side(style='thick')       
        border_sides_thin = Side(style='thin')
        
        logging.info('rows taken out from data')
        for r_idx, row in enumerate(rows, 9):
            for c_idx, value in enumerate(row, 1):
                formIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                if len(row)==c_idx and int(row[0])==len(data_formII):
                       formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                elif len(row)==c_idx:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                elif int(row[0])==len(data_formII):
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                else:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                #border_sides = Side(style='thin')
                #formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        
        formIIsheet['A4']=formIIsheet['A5'].value+str(data_formII['UnitName'].unique()[0])
        formIIsheet['A5']="PERIOD "+str(month)+" "+str(year)
        formIIfinalfile = os.path.join(filelocation,'Form II Register of deductions for damage or loss.xlsx')
        formIIfile.save(filename=formIIfinalfile)

    def Form_IV():
        formIVfilepath = os.path.join(Centralfilespath,'Form IV Overtime register.xlsx')
        formIVfile = load_workbook(filename=formIVfilepath)
        logging.info('Form IV file has sheet: '+str(formIVfile.sheetnames))
        logging.info('create columns which are now available')

        data_formIV = data.copy(deep=True)
        data_formIV=data_formIV.drop_duplicates(subset="Employee Code", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Gender","Designation_Dept","Date_overtime_worked",
                                        "Extent of over-time","Total over-time","Normal hrs ",
                                        "FIXED MONTHLY GROSS","overtime rate",
                                        "normal_earning","Overtime",'Total Earning',"date_overtime_paid"]
        
        # data_formIV['Total\r\nOT Hrs']=data_formIV[['Total\r\nOT Hrs',"Overtime",'Total Earning']].astype(float)
        # data_formIV["Total over-time"]=data_formIV['Total\r\nOT Hrs']
        # data_formIV["normal_earning"]=data_formIV['FIXED MONTHLY GROSS']-data_formIV["Overtime"]
        # data_formIV.loc[data_formIV['Total\r\nOT Hrs']==0,["Total over-time","Normal hrs ",
        #                                 "FIXED MONTHLY GROSS","overtime rate",
        #                                 "normal_earning","Overtime",'Total Earning']]="---"

        # data_formIV["date_overtime_paid"]=data_formIV['Date of payment']
        # data_formIV.loc[data_formIV["Overtime"]==0,"date_overtime_paid"]="---"
        # data_formIV.loc[data_formIV['Total\r\nOT Hrs']==0,"date_overtime_paid"]="---"
        # data_formIV["Extent of over-time"]="-----"
        # data_formIV["Date_overtime_worked"]="-----"
        
        data_formIV["Date_overtime_worked","Extent of over-time","Total over-time","Normal hrs ",
                    "FIXED MONTHLY GROSS","overtime rate","normal_earning","Overtime",'Total Earning',"date_overtime_paid"]="---"

        data_formIV['S.no'] = list(range(1,len(data_formIV)+1))
        data_formIV['Designation_Dept']=data_formIV["Designation"]+"_"+data_formIV["Department"]
        # data_formIV["Date of payment & amount of deduction"]=data_formIV['Date of payment']+"\n"+data_formIV["Total Deductions"]
        data_formIV[['Extent of over-time', 'Date_overtime_worked', 'date_overtime_paid', 'normal_earning', 'Total over-time']]="---"
        formIV_data=data_formIV[columns]
        formIVsheet = formIVfile['Sheet1']
        formIVsheet.sheet_properties.pageSetUpPr.fitToPage = True
        # for column in  range(ord('A'), ord('O') + 1):
        #     formIVsheet.unmerge_cells(chr(column)+"7:"+chr(column)+"14")

        logging.info('data for form IV is ready')

        
        rows = dataframe_to_rows(formIV_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 8):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formIVsheet.cell(row=r_idx, column=c_idx, value=value)
                formIVsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIVsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formIV):
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIVsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIVsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formIV):
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIVsheet.row_dimensions[r_idx].height = 20
                else:
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIVsheet.row_dimensions[r_idx].height = 20

        # formIVsheet['A4']=formIVsheet['A4'].value+"  "+data_formIV['Company Name'].unique()[0]+"  "+data_formIV['Company Address'].unique()[0]+"                                Month Ending: "+month+" "+str(year)
        # formIVsheet.cell(row=4, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
        formIVsheet['A4']="Month Ending: "+month+" "+str(year)
        formIVfinalfile = os.path.join(filelocation,'Form IV Overtime register.xlsx')
        formIVfile.save(filename=formIVfinalfile)

    def Form_V():
        formVfilepath = os.path.join(Centralfilespath,'Form V Muster Roll.xlsx')
        formVfile = load_workbook(filename=formVfilepath)
        logging.info('Form V file has sheet: '+str(formVfile.sheetnames))
        logging.info('create columns which are now available')

        data_formV = data.copy(deep=True)
        data_formV=data_formV.drop_duplicates(subset="Employee Code", keep="last")
        
        columns=['S.no',"Employee Name","Father's Name","Gender",'Nature of work']
        
        # data_formV_columns=list(data_formV.columns)
        # start=data_formV_columns.index('Emp Code')
        # end=data_formV_columns.index('Total\r\nDP')
        columnstotake =[]
        days = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31']
        for day in days:
            for col in data_formV.columns:
                if col[5:7]==day:
                    columnstotake.append(col)
        if len(columnstotake)==28:
            columnstotake.append('29')
            columnstotake.append('30')
            columnstotake.append('31')
            data_formV['29'] = ''
            data_formV['30'] = ''
            data_formV['31'] = ''
            
        elif len(columnstotake)==29:
            columnstotake.append('30')
            columnstotake.append('31')
            data_formV['30'] = ''
            data_formV['31'] = ''

        elif len(columnstotake)==30:
            columnstotake.append('31')
            data_formV['31'] = ''
        
        columns.extend(columnstotake)
        columns.append('Total\r\nDP')
        data_formV['S.no'] = list(range(1,len(data_formV)+1))

        formV_data=data_formV[columns]
        formVsheet = formVfile['Sheet1']
        formVsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form V is ready')

        
        rows = dataframe_to_rows(formV_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 9):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formVsheet.cell(row=r_idx, column=c_idx, value=value)
                formVsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formVsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formV):
                    formVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formVsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formVsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formV):
                    formVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formVsheet.row_dimensions[r_idx].height = 20
                else:
                    formVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formVsheet.row_dimensions[r_idx].height = 20

        
        #formPsheet['AE4']=formPsheet['AE4'].value+"   "+str(data_formP['Registration_no'].unique()[0])
        
        formVsheet['A4']="Name of establishment :-  "+str(data_formV['UnitName'].unique()[0])
        monthstart = datetime.date(year,month_num,1)
        monthend = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
        
        formVsheet['A6']="  From:  "+str(monthstart)+"       "+"To:  "+str(monthend)
        
        formVsheet['A5']="Place   "+data_formV['Branch'].unique()[0]
        formVfinalfile = os.path.join(filelocation,'Form V Muster Roll.xlsx')
        formVfile.save(filename=formVfinalfile)

    def Form_X():
        formXfilepath = os.path.join(Centralfilespath,'Form X register of wages.xlsx')
        formXfile = load_workbook(filename=formXfilepath)
        logging.info('Form X file has sheet: '+str(formXfile.sheetnames))
        logging.info('create columns which are now available')

        data_formX = data.copy(deep=True)
        
        columns=['S.no',"Employee Name","Father's Name","Designation",'Earned Basic','DA',
                                'Earned Basic','DA',"Days Paid","Overtime","FIXED MONTHLY GROSS","PF","HRA",
                                "all_Other_Deduction_sum",'Total Deductions','Net Paid','Date of payment',
                                'sign']

        other_deductions_columns_name=['Other Deduction','OtherDeduction1', 'OtherDeduction2',
                                                        'OtherDeduction3', 'OtherDeduction4', 'OtherDeduction5']

        data_formX[other_deductions_columns_name]=data_formX[other_deductions_columns_name].astype(float)
        data_formX["all_Other_Deduction_sum"]= data_formX.loc[:,other_deductions_columns_name].sum(axis=1)

        data_formX[["sign"]]=""

        data_formX['S.no'] = list(range(1,len(data_formX)+1))

        formX_data=data_formX[columns]
        formXsheet = formXfile['Sheet1']
        formXsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form X is ready')

        
        rows = dataframe_to_rows(formX_data, index=False, header=False)
        rows_copy = list(dataframe_to_rows(formX_data, index=False, header=False))
        logging.info('rows taken out from data')

        row_num=0
        for r_idx, row in enumerate(rows, 9):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                #formXXIIIsheet.cell(row=r_idx, column=c_idx).value=value
                formXsheet.cell(row=r_idx, column=c_idx, value=value)
                formXsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formX):
                    formXsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formXsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formXsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formXsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formX):
                    formXsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formXsheet.row_dimensions[r_idx].height = 20
                else:
                    formXsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formXsheet.row_dimensions[r_idx].height = 20
        
        
        # formXsheet['P4']=formXsheet['P4'].value+"   "+str(data_formX['Registration_no'].unique()[0])
        # formXsheet['P5']=formXsheet['P5'].value+"   "+month
        formXsheet['A3']=" Name of Establishment:-   "+str(data_formX['UnitName'].unique()[0])
        formXsheet['A4']="Place : "+str(data_formX['Branch'].unique()[0])
        monthstart = datetime.date(year,month_num,1)
        monthend = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
        
        formXsheet['A5']="Wage perod  From:  "+str(monthstart)+"       "+"To:  "+str(monthend)
        
        formXfinalfile = os.path.join(filelocation,'Form X register of wages.xlsx')
        formXfile.save(filename=formXfinalfile)

    def create_ecard():
    
        ecardfilepath = os.path.join(Centralfilespath,'Form XI wages slip.xlsx')
        ecardfile = load_workbook(filename=ecardfilepath)
        logging.info('Employment card file has sheet: '+str(ecardfile.sheetnames))
        sheetecard = ecardfile['Sheet1']

        logging.info('create columns which are now available')

        data_ecard = data.drop_duplicates(subset=['Employee Code']).copy(deep=True)
        data_ecard.fillna(value=0, inplace=True)

        emp_count = len(data_ecard.index)
        
        for i in range(0,emp_count):
            key = (data_ecard).iloc[i]['Employee Code']
            sheet_key = 'Employment card_'+str(key)

            emp_data = (data_ecard).iloc[i]
            emp_data.fillna(value='', inplace=True)

            sheet1 = ecardfile.copy_worksheet(sheetecard)
            sheet1.title = sheet_key

            sheet1['B4'] = emp_data['UnitName']
            sheet1['B5'] = emp_data['Location']
            sheet1['B6'] = emp_data['Employee Name']+" / "+emp_data["Father's Name"]
            sheet1['B7'] = emp_data['Designation']
            sheet1['B8'] = str(month)+" "+str(year)
            sheet1['B9'] = emp_data['FIXED MONTHLY GROSS']
            sheet1['B10'] = emp_data['Earned Basic']
            sheet1['B11'] = emp_data['DA']
            sheet1['B12'] = emp_data['Days Paid']
            sheet1['B13'] = emp_data["Overtime"]
            sheet1['B14'] = emp_data["FIXED MONTHLY GROSS"]
            sheet1['B15'] = emp_data['Total Deductions']
            sheet1['B16'] = emp_data['Net Paid']
            

        ecardfinalfile = os.path.join(filelocation,'Form XI wages slip.xlsx')
        ecardfile.remove(sheetecard)
        ecardfile.save(filename=ecardfinalfile)
    try:
        Form_C()
        master.update()
        Form_I()
        master.update()
        Form_II_reg_damage_loss()
        master.update()
        Form_IV()
        master.update()
        Form_V()
        master.update()
        Form_X()
        master.update()
        create_ecard()
        master.update()
    except KeyError as e:
        logging.info("Key error : Check if {} column exsists".format(e))
        print("Key error {}".format(e))
        report.configure(text="Failed: Check input file format  \n column {} not found".format(e))
        master.update()
        raise KeyError
    except FileNotFoundError as e:
        logging.info("File not found : Check if {} exsists".format(e))
        report.configure(text="Failed: File {} not found".format(e))
        master.update()
        raise FileNotFoundError
        