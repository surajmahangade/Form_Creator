
# This code is used to develop the form P,R,T section of Tamil Nadu state
# Author: Riddhi Pravin Shah
from states import logging,monthdict,Statefolder
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging
from collections import Counter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill, numbers

def Tamilnadu(data,contractor_name,contractor_address,filelocation,month,year):
    
    tnfilespath = os.path.join(Statefolder,'Tamilnadu')
    logging.info('Tamilnadu files path is :'+ str(tnfilespath))
    
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]

    print("inside")
    
    
    def Form_p():

        # Importing file path of Form P
        formPFilePath = os.path.join(tnfilespath,'Form P register of deduction.xlsx')
        formPFile = load_workbook(filename = formPFilePath)
        logging.info('Form P file has sheet: ' + str(formPFile.sheetnames))
        logging.info('create columns which are now available')

        # Preparing Master data for form P with columns required.
        data_formP = data.copy()

        ## Adding placeholder columns where data isnot found
        data_formP['S.no'] = list(range(1,len(data_formP)+1))
        colToAdd = ['Number of Instalments to be recovered',
                                    'Date on which recovery completed','Date of Show Cause Notice'
                                    ,'Date on which deduction completed','Act or omission','Date on which fine recovery completed'
                                    ,'Signature or thumb impression of the person employed']
        data_formP = data_formP.reindex(data_formP.columns.tolist() + colToAdd, axis=1)
        data_formP.loc[:,colToAdd] = '----'

        ## Appending columns in order required for FORM P
        columns = ['S.no','Employee Name','Father\'s Name','Employee Code','Designation','Date of payment ',
                   'Net Paid','Number of Instalments to be recovered','Date on which recovery completed',
                   'Damage or Loss','Date of Show Cause Notice','Total Deductions','Number of Instalments to be recovered',
                   'Date on which deduction completed','Act or omission','Date of Show Cause Notice','Fine',
                   'Date on which fine recovery completed','Signature or thumb impression of the person employed','Remarks']


        formP_data = data_formP[columns]

        formPsheet = formPFile['Sheet1']
        formPsheet.sheet_properties.pageSetUpPr.fitToPage = True
        print("Here")
        
        logging.info('data for form P is ready')

        # Load data into rows of Form P
        rows = dataframe_to_rows(formP_data, index=False, header=False)
        logging.info('rows taken out from data')

        for r_idx, row in enumerate(rows, 8):
            for c_idx, value in enumerate(row, 1):
                formPsheet.cell(row=r_idx, column=c_idx, value = value)
                formPsheet.cell(row=r_idx, column=c_idx).fill = PatternFill(fill_type=None)
                formPsheet.cell(row=r_idx, column=c_idx).font = Font(name ='Bell MT', size = 10)
                formPsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formPsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        # Adding Unitfile Unitname and address
        formPsheet.cell(row=5, column=4, value = 'PLACEHOLDER') # Add variable that needs to be added
        formPsheet.merge_cells('D5:G5')

        # Formatting the table 
        formPsheet.merge_cells('A5:C5')
        formPsheet.merge_cells('D5:T5')
        formPsheet.merge_cells('A4:T4')
        formPsheet.row_dimensions[5].height = 30

        for row in formPsheet.iter_rows():
            for cell in row:      
                cell.alignment =  cell.alignment.copy(wrapText=True)

        for row_id in range(1,7):
            for col_id in range(1,21):
                formPsheet.cell(row=row_id, column=col_id).border = Border(left=Side(style='thick'), 
                                                                           right=Side(style='thick'), 
                                                                           top=Side(style='thick'), 
                                                                           bottom=Side(style='thick'))

        #formPsheet['A4'] = formPsheet['A4'].value + " : " + data_formP['Unit'][0]
        formPfinalfile = os.path.join(outputPath,'Form P register of deduction.xlsx')
        formPFile.save(filename = formPfinalfile)  

    def Form_r():
        
        form_data = data.copy()   # will be passed inside the func
        
        # cls required for form R from main DF
        # whatever has "-" has no corresponding match

        # TODO: update as and when you h=get new cols
        colsRequired = {1:'Employee Name',2:'Gender',3:'Designation',4:'-',
                        5:'-',6:'Days Paid',7:'-',8:'-',9:'overtime rate',10:'-',
                        11:'-',12:'-',13:'-',14:'CHECK CTC Gross',15:'FIXED MONTHLY GROSS',
                        16:'-',17:'-',18:'-',19:'Fine',20:'Net Paid',21:'-',22:'-',}

        # load form
        formPath = os.path.join(tnfilespath,'Form R register of wages.xlsx')
        formfile = load_workbook(filename=formPath)

        # select sheet to write in and set page properties
        formsheet = formfile['Sheet1']
        formsheet.sheet_properties.pageSetUpPr.fitToPage = True
        
        
        ## initialize variables
        rownum = 9 # starting row count
        pushDownIdx = 11
        serialNum = 0

        # populate ros with required employee information
        for idx in form_data.index:

            formsheet.cell(row= rownum, column=1, value=serialNum)
            for col_idx in colsRequired.keys():
                col = colsRequired[col_idx]

                if  col == "-":  # if no value exisits
                    populate = "-----"
                else: # fetch value from dataframe
                    populate = form_data.loc[idx, col]

                # write to excel
                print("rownum",rownum, "column", col_idx+1)
                formsheet.cell(row= rownum, column=col_idx+1, value=populate)
                formsheet.cell(row= rownum, column=col_idx+1).font =Font(name ='Bell MT', size =10)
                formsheet.cell(row= rownum, column=col_idx+1).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formsheet.cell(row=rownum, column=col_idx +1).border = Border(outline= True, right=border_sides, bottom=border_sides)

            # format table
            for row_id in range(1,rownum+1):
                for col_id in range(1,24):
                    formsheet.cell(row=row_id, column=col_id).border = Border(left=Side(style='medium'), 
                                                                               right=Side(style='medium'), 
                                                                                   top=Side(style='medium'), 
                                                                                   bottom=Side(style='medium'))
            # insery rows to push down the note
            formsheet.insert_rows(idx=pushDownIdx, amount=1)
            rownum +=1
            pushDownIdx+=1
            serialNum +=1

        # save file
        formIfinalfile = os.path.join(outputPath,'Form R register of wages.xlsx')      
        formfile.save(filename=formIfinalfile)

    def Form_t():# - Form T wages slip

        form_data = data.copy()   # will be passed inside the func

        # load form
        formPath = os.path.join(tnfilespath,'Form T wages slip.xlsx')
        formfile = load_workbook(filename=formPath)

        # select sheet to write in and set page properties
        formsheet = formfile['Sheet1']
        formsheet.sheet_properties.pageSetUpPr.fitToPage = True

        # for each employee create a new sheet
        for idx in form_data.index:

            empName = form_data.loc[idx, "Employee Name"]
            print("Employee:", empName)

            # create a new sheet for employee
            new = formfile.copy_worksheet(formsheet)
            new.title = empName

            # popoluate every cell with required employee information 
            new.cell(row= 4, column=3, value=form_data.loc[idx, "Company Name"])# Name of company
            new.cell(row= 5, column=3, value=form_data.loc[idx, "Employee Name"])  #. name of employee
            new.cell(row= 6, column=3, value=form_data.loc[idx, "Father's Name"])  #. Fathers name
            new.cell(row= 7, column=3, value=form_data.loc[idx, "Designation"]) # employee designation C7
            new.cell(row= 8, column=3, value=form_data.loc[idx, "Date Joined"]) # date joined C8
            new.cell(row= 11, column=2, value=form_data.loc[idx, "Earned Basic"]) # earned basic B11
            new.cell(row= 13, column=2, value=form_data.loc[idx, "HRA"])# House rent allowance B13
            new.cell(row= 14, column=2, value=form_data.loc[idx, "Overtime"]) # overtime wages B14
            new.cell(row= 16, column=2, value=form_data.loc[idx,"Other Allowance"]) # otehr allowance B16
            new.cell(row= 17, column=2, value=form_data.loc[idx, "FIXED MONTHLY GROSS"]) # Gross wages B17
            new.cell(row= 12, column=9, value=form_data.loc[idx, "Insurance"]) # employee state insurance I12
            new.cell(row= 13, column=9, value=form_data.loc[idx, "Other Deduction"]) # other deductions I13
            new.cell(row= 17, column=7, value=form_data.loc[idx, "Net Paid"]) # net paid #G17

            # format table A1 - K19
            for row_id in range(1,20):
                for col_id in range(1,12):
                    new.cell(row=row_id, column=col_id).border = Border(left=Side(style='medium'), 
                                                                                       right=Side(style='medium'), 
                                                                                           top=Side(style='medium'), 
                                                                                         bottom=Side(style='medium'))

        # delete sheet1 sheet2, sheet3
        formfile.remove(formfile['Sheet1']) 
        formfile.remove(formfile['Sheet2']) 
        formfile.remove(formfile['Sheet3']) 

        # save file
        formIfinalfile = os.path.join(outputPath,'Form T wages slip.xlsx')      
        formfile.save(filename=formIfinalfile)




    ## --------FUNCTION CALL-------------------------
    Form_p()  ## Call this function in the main def
    Form_r()
    Form_t()