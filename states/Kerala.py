from states import logging,monthdict,Statefolder
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging
from states.utils import forms_template

create_border = forms_template.Helper_functions.create_border


def Kerala(data,contractor_name,contractor_address,filelocation,month,year,report,master):
    Keralafilespath = os.path.join(Statefolder,'Kerala')
    logging.info('Kerala files path is :'+str(Keralafilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]

    def Form_A():
        formAfilepath = os.path.join(Keralafilespath,'Form A Register of employment.xlsx')
        formAfile = load_workbook(filename=formAfilepath)
        logging.info('Form A file has sheet: '+str(formAfile.sheetnames))
        logging.info('create columns which are now available')

        data_formA = data.copy(deep=True)
        data_formA=data_formA.drop_duplicates(subset="Employee Code", keep="last")
        
        columns=['S.no',"Employee Name","young_or_not","start_time","end_time","rest_interval","Hrs_worked","days_overtime","Overtime"]
        # data_formA_columns=list(data_formA.columns)
        # start=data_formA_columns.index('Arrears salary')
        # end=data_formA_columns.index('Total\r\nDP')
        columnstotake =[]
        days = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31']
        for day in days:
            for col in data_formA.columns:
                if col[5:7]==day:
                    columnstotake.append(col)
        if len(columnstotake)==28:
            columnstotake.append('29')
            columnstotake.append('30')
            columnstotake.append('31')
            data_formA['29'] = ''
            data_formA['30'] = ''
            data_formA['31'] = ''
            
        elif len(columnstotake)==29:
            columnstotake.append('30')
            columnstotake.append('31')
            data_formA['30'] = ''
            data_formA['31'] = ''

        elif len(columnstotake)==30:
            columnstotake.append('31')
            data_formA['31'] = ''
        
        columns.extend(columnstotake)

        columns.extend(["Total\r\nDP"])
        data_formA['S.no'] = list(range(1,len(data_formA)+1))
        data_formA[['young_or_not',"Hrs_worked","days_overtime"]]=""
        formA_data=data_formA[columns]
        formAsheet = formAfile['Sheet1']
        formAsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form A is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formA_data, index=False, header=False)

        logging.info('rows taken out from data')
        r_idx=0
        c_idx=0
        start_row=7
        for r_idx, row in enumerate(rows, start_row):
            for c_idx, value in enumerate(row, 1):
                formAsheet.cell(row=r_idx, column=c_idx, value=value)
                formAsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formAsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formAsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        create_border(formAsheet, last_row=r_idx, last_column=c_idx, start_row=7, start_column=1)
        formAsheet['A4']="Month : "+month+"  Year:  "+str(year)
        formAfinalfile = os.path.join(filelocation,'Form A Register of employment.xlsx')
        formAfile.save(filename=formAfinalfile)

    def Form_C():
        formCfilepath = os.path.join(Keralafilespath,'Form C notice of working day.xlsx')
        formCfile = load_workbook(filename=formCfilepath)
        logging.info('Form C file has sheet: '+str(formCfile.sheetnames))
        logging.info('create columns which are now available')

        data_formC = data.copy(deep=True)
        data_formC=data_formC.drop_duplicates(subset="Employee Code", keep="last")
        
        
        columns=["Employee Name","young_or_not","Employee_to_commence","intervals_meal_rest","employment_to_cease"]
        
        #data_formC['S.no'] = list(range(1,len(data_formC)+1))
        data_formC["young_or_not"]="----"
        data_formC[["Employee_to_commence","intervals_meal_rest","employment_to_cease"]]=""
        #data_formC[["Date of payment & amount of deduction","remarks"]]=""
        formC_data=data_formC[columns]
        formCsheet = formCfile['Sheet1']
        formCsheet.sheet_properties.pageSetUpPr.fitToPage = True

        formCsheet.unmerge_cells("A9:A12")
        formCsheet.unmerge_cells("B9:B12")
        formCsheet.unmerge_cells("C9:C12")
        formCsheet.unmerge_cells("D9:D12")
        formCsheet.unmerge_cells("E9:E12")

        formCsheet.unmerge_cells("B14:E14")
        formCsheet.unmerge_cells("B16:E16")
        formCsheet.unmerge_cells("B17:E17")
        formCsheet.unmerge_cells("B18:E18")
        formCsheet.unmerge_cells("A19:F19")
        formCsheet.unmerge_cells("A20:F20")
        formCsheet.unmerge_cells("A21:F21")

        logging.info('data for form A is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formC_data, index=False, header=False)

        logging.info('rows taken out from data')
        added=0
        formCsheet.insert_rows(9,len(data_formC))
        start_row=9
        r_idx=0
        c_dix=0
        for r_idx, row in enumerate(rows, start_row):
            for c_idx, value in enumerate(row, 1):
                formCsheet.cell(row=r_idx, column=c_idx, value=value)
                formCsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formCsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formCsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                added+=1
        
        create_border(formCsheet, last_row=r_idx, last_column=c_idx, start_row=9, start_column=1)
        
        formCsheet.merge_cells("B"+str(14+len(data_formC))+":E"+str(14+len(data_formC)))
        formCsheet.merge_cells("B"+str(16+len(data_formC))+":E"+str(16+len(data_formC)))
        formCsheet.merge_cells("B"+str(17+len(data_formC))+":E"+str(17+len(data_formC)))
        formCsheet.merge_cells("B"+str(18+len(data_formC))+":E"+str(18+len(data_formC)))
        formCsheet.merge_cells("A"+str(19+len(data_formC))+":F"+str(19+len(data_formC)))
        formCsheet.merge_cells("A"+str(20+len(data_formC))+":F"+str(20+len(data_formC)))
        formCsheet.merge_cells("A"+str(21+len(data_formC))+":F"+str(21+len(data_formC)))
        
        # formCsheet['A5']="Name and Address of the Establishment "+str(data_formC['Unit'].unique()[0])+","+str(data_formC['Address'].unique()[0])
        formCfinalfile = os.path.join(filelocation,'Form C notice of working day.xlsx')
        formCfile.save(filename=formCfinalfile)
    

    def Form_I():
        formIfilepath = os.path.join(Keralafilespath,'Form I Register of fines.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formI = data.copy(deep=True)
        data_formI=data_formI.drop_duplicates(subset="Employee Code", keep="last")

        columns=['S.no',"Employee Name","Father's Name","Department","act_of_fine","Fine","Designation","Date of payment","cause_against_fine",
                                        "amount","month","Fine","Date of payment","remarks"]
    
        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        data_formI["month"]=month
        data_formI["act_of_fine"]="-----"
        data_formI["cause_against_fine"]="-----"
        data_formI[["Date of Fine","amount","remarks"]]=""
        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')

        from string import ascii_uppercase
        for char in ascii_uppercase[:14]:
            formIsheet.unmerge_cells(char+str(7)+':'+char+str(9))
        
        formIsheet.unmerge_cells("A11:N11")
        formIsheet.unmerge_cells("A12:N12")
        formIsheet.insert_rows(7,len(data_formI))
        start_row=7
        r_dix=0
        c_idx=0
        for r_idx, row in enumerate(rows, start_row):
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        

        create_border(formIsheet, last_row=r_idx, last_column=c_idx, start_row=7, start_column=1)
        

        formIsheet.merge_cells("A"+str(11+len(data_formI))+":N"+str(11+len(data_formI)))
        formIsheet.merge_cells("A"+str(12+len(data_formI))+":N"+str(12+len(data_formI)))


        formIsheet['A4']=formIsheet['A4'].value+" : "+data_formI['Unit'].unique()[0]
        formIfinalfile = os.path.join(filelocation,'Form I Register of fines.xlsx')
        formIfile.save(filename=formIfinalfile)


    def Form_II():
        formIIfilepath = os.path.join(Keralafilespath,'Form II Register of damage or loss.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy(deep=True)
        data_formII=data_formII.drop_duplicates(subset="Employee Code", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Department","Damage or Loss",'Total Deductions',
                                        "Designation","Date of payment","whether_work_showed_cause","num_instalments","date_amt_deduction","remarks"]

        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII[["whether_work_showed_cause","num_instalments","date_amt_deduction"]]="-----"
        data_formII["remarks"]=""
        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formII_data, index=False, header=False)

        logging.info('rows taken out from data')
        start_row=7
        r_dix=0
        c_idx=0
        for r_idx, row in enumerate(rows, start_row):
            for c_idx, value in enumerate(row, 1):
                formIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        
        create_border(formIIsheet, last_row=r_idx, last_column=c_idx, start_row=7, start_column=1)

        formIIsheet['A4']=formIIsheet['A4'].value+" : "+data_formII['Unit'].unique()[0]
        formIIfinalfile = os.path.join(filelocation,'Form II Register of damage or loss.xlsx')
        formIIfile.save(filename=formIIfinalfile)


    def Form_III():

        formIIIfilepath = os.path.join(Keralafilespath,'Form III register of advance.xlsx')
        formIIIfile = load_workbook(filename=formIIIfilepath)
        logging.info('Form III file has sheet: '+str(formIIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formIII = data.copy(deep=True)
        data_formIII = data_formIII.drop_duplicates(subset="Employee Code", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Department","Date of payment",
                                        "purpose_advance","num_instalments","postponements_granted","date_total_repaid","remarks"]
        
        data_formIII['S.no'] = list(range(1,len(data_formIII)+1))
        data_formIII[["purpose_advance","num_instalments","postponements_granted","date_total_repaid"]]="---"
        data_formIII["remarks"]=""
        formIII_data=data_formIII[columns]
        formIIIsheet = formIIIfile['Sheet1']
        formIIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form III is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formIII_data, index=False, header=False)

        logging.info('rows taken out from data')
        r_idx=0
        c_idx=0
        start_row=7
        for r_idx, row in enumerate(rows, start_row):
            for c_idx, value in enumerate(row, 1):
                formIIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        create_border(formIIIsheet, last_row=r_idx, last_column=c_idx, start_row=7, start_column=1)
        formIIIsheet['A4']=formIIIsheet['A4'].value+" : "+data_formIII['Unit'].unique()[0]
        formIIIfinalfile = os.path.join(filelocation,'Form III register of advance.xlsx')
        formIIIfile.save(filename=formIIIfinalfile)


    def Form_XIV():
    
        formXIVfilepath = os.path.join(Keralafilespath,'Form XIV register of employment and wages.xlsx')
        formXIVfile = load_workbook(filename=formXIVfilepath)
        logging.info('Form XIV file has sheet: '+str(formXIVfile.sheetnames))
        logging.info('create columns which are now available')

        data_formXIV = data.copy(deep=True)
        data_formXIV=data_formXIV.drop_duplicates(subset="Employee Code", keep="last")

        columns=['S.no',"Employee Code","Father's Name","Gender","Date of Birth","Designation","Designation_code","Date Joined","Mobile Tel No.",
                                                    "E-Mail","Bank Name","IFSC_code","Bank A/c Number","Days Paid","LOP","num_weekly_granted",
                                                    "num_leave","Earned Basic","DA","HRA","City_compensation","FIXED MONTHLY GROSS",
                                                    "Overtime","Leave Encashment","Festival_wages","Arrears","Bonus",
                                                    "Maternity_benefit","Other Allowance","Salary Advance","Total Earning",
                                                    "PF","ESIC","Salary Advance","LWF EE","P.Tax","TDS","Fine",
                                                    "Damage or Loss","Other Deduction","Total Deductions","Net Paid","Date of payment",
                                                    "remarks"]
        
        remove_point=lambda input_str: input_str.split(".")[0]
        data_formXIV["Bank A/c Number"]=data_formXIV["Bank A/c Number"].apply(str).apply(remove_point)
        data_formXIV['S.no'] = list(range(1,len(data_formXIV)+1))
        #data_formXIV[["purpose_advance","num_instalments","postponements_granted","date_total_repaid"]]="-----"
        data_formXIV[["IFSC_code","LOP","num_weekly_granted","num_leave","DA","City_compensation","Festival_wages","Maternity_benefit",'Designation_code','remarks']]=""
        formXIV_data=data_formXIV[columns]
        formXIVsheet = formXIVfile['Sheet1']
        formXIVsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form XIV is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formXIV_data, index=False, header=False)

        logging.info('rows taken out from data')
        r_idx=0
        c_idx=0
        start_row=18
        for r_idx, row in enumerate(rows, start_row):
            for c_idx, value in enumerate(row, 1):
                formXIVsheet.cell(row=r_idx, column=c_idx, value=value)
                formXIVsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formXIVsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        create_border(formXIVsheet, last_row=r_idx, last_column=c_idx, start_row=18, start_column=1)

        formXIVsheet['A4']=formXIVsheet['A4'].value+" : "+str(data_formXIV['Location'].unique()[0])
        formXIVsheet['A5']=formXIVsheet['A5'].value+" : "+str(data_formXIV['Unit'].unique()[0])+", "+str(data_formXIV['Location'].unique()[0])
        formXIVsheet['A6']=formXIVsheet['A6'].value+" : "+str(data_formXIV['Unit'].unique()[0])+", "+str(data_formXIV['Location'].unique()[0])
        formXIVsheet['A7']=formXIVsheet['A7'].value+" : "+str(data_formXIV['Unit'].unique()[0])+", "+str(data_formXIV['Location'].unique()[0])
        #formXIVsheet['A7']=formXIVsheet['A7'].value+" : "+"Bank Name"
        formXIVsheet['A10']=formXIVsheet['A10'].value+" : "+month+" "+str(year)
        formXIVfinalfile = os.path.join(filelocation,'Form XIV register of employment and wages.xlsx')
        formXIVfile.save(filename=formXIVfinalfile)


    try:
        Form_A()
        master.update()
        Form_C()
        master.update()
        Form_I()
        master.update()
        Form_II()
        master.update()
        Form_III()
        master.update()
        Form_XIV()
        master.update()
    except KeyError as e:
        logging.info("Key error : Check if {} column exsists".format(e))
        print("Key error {}".format(e))
        report.configure(text="Failed: Check input file format  \n column {} not found".format(e))
        master.update()
        raise KeyError
    except FileNotFoundError as e:
        logging.info("File not found : Check if {} exsists".format(e))
        report.configure(text="Failed: File {} not found".format(e))
        master.update()
        raise FileNotFoundError