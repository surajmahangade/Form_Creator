from states import logging,monthdict,Statefolder
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill, numbers

def Contractor_Process(data,contractor_name,contractor_address,filelocation,month,year,report,master):
    Contractorfilespath = os.path.join(Statefolder,'CLRA')
    logging.info('Contractor files path is :'+str(Contractorfilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]
    # data_formIV = data.copy(deep=True)
    #Comment this line if in future leave file data is needed in any of functions below
    data=data.drop_duplicates(subset='Employee Code', keep="last")
    
    def create_form_A():
    
        formAfilepath = os.path.join(Contractorfilespath,'Form A Employee register.xlsx')
        formAfile = load_workbook(filename=formAfilepath)
        logging.info('Form A file has sheet: '+str(formAfile.sheetnames))

        
        logging.info('create columns which are now available')
        data_formA=data.copy(deep=True)
        
        data_formA.fillna(value=0, inplace=True)

        data_formA['S.no'] = list(range(1,len(data_formA)+1))

        formA_columns = ["S.no",'Employee Code','Employee Name',"Gender","Father's Name",'Date of Birth',"Nationality","Education Level",'Date Joined',
                        'Designation','Category Address',"Type of Employment",'Mobile Tel No.','UAN Number',"PAN Number","ESIC Number","LWF EE","Aadhar Number",
                        "Bank A/c Number","Bank Name",'Branch',"Present_Address","Permanent_Address",'Service Book No',"Date Left","Reason for Leaving",'Identification mark',
                        "photo","sign","remarks"]
        
        data_formA[["photo","sign","remarks"]]=""
        remove_point=lambda input_str: input_str.split(".")[0]
        data_formA["Bank A/c Number"]=data_formA["Bank A/c Number"].apply(str).apply(remove_point)
        data_formA["Aadhar Number"]=data_formA["Aadhar Number"].apply(str).apply(remove_point)
        data_formA['Category Address']=""
        data_formA[['Local Address 1', 'Local Address 2','Local Address 3', 'Local Address 4']]=data_formA[['Local Address 1', 'Local Address 2','Local Address 3', 'Local Address 4']].astype(str)
        data_formA[['Permanent Address 1', 'Permanent Address 2','Permanent Address 3', 'Permanent Address 4']]=data_formA[['Permanent Address 1', 'Permanent Address 2','Permanent Address 3', 'Permanent Address 4']].astype(str)
        data_formA["Present_Address"]=data_formA['Local Address 1']+data_formA['Local Address 2']+data_formA['Local Address 3']+ data_formA['Local Address 4']
        data_formA["Permanent_Address"]=data_formA['Permanent Address 1']+data_formA['Permanent Address 2']+data_formA['Permanent Address 3']+ data_formA['Permanent Address 4']
        
        formA_data = data_formA[formA_columns]
        formAsheet = formAfile['Sheet1']
        formAsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form A is ready')
        rows = dataframe_to_rows(formA_data, index=False, header=False)

        logging.info('rows taken out from data')

        for r_idx, row in enumerate(rows, 11):
            for c_idx, value in enumerate(row, 1):
                formAsheet.cell(row=r_idx, column=c_idx, value=value)
                formAsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formAsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formAsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                # if c_idx==15 or c_idx==16 or c_idx==20 or c_idx==21:
                #     formAsheet.cell(row=r_idx, column=c_idx).number_format= numbers.FORMAT_NUMBER

        logging.info('')


        if str(data_formA['Company Name'].dtype)[0:3] != 'obj':
            data_formA['Company Name'] = data_formA['Company Name'].astype(str)

        if str(data_formA['Company Address'].dtype)[0:3] != 'obj':
            data_formA['Company Address'] = data_formA['Company Address'].astype(str)

        if str(data_formA['Contractor_name'].dtype)[0:3] != 'obj':
            data_formA['Contractor_name'] = data_formA['Contractor_name'].astype(str)

        if str(data_formA['Contractor_Address'].dtype)[0:3] != 'obj':
            data_formA['Contractor_Address'] = data_formA['Contractor_Address'].astype(str)

        if str(data_formA['Unit'].dtype)[0:3] != 'obj':
            data_formA['Unit'] = data_formA['Unit'].astype(str)

        if str(data_formA['Branch'].dtype)[0:3] != 'obj':
            data_formA['Branch'] = data_formA['Branch'].astype(str)

        establishment = formAsheet['A5'].value
        
        if data_formA['PE_or_contract'].unique()[0] == 'PE':
            A5_data = establishment+' '+data_formA['Company Name'].unique()[0] +', '+data_formA['Company Address'].unique()[0]  
        else:    
            A5_data = establishment+' '+data_formA['Contractor_name'].unique()[0]+', '+data_formA['Contractor_Address'].unique()[0]
        formAsheet['A5'] = A5_data


        # company = formAsheet['A10'].value
        # A10_data = company+' '+data_formA['Unit'].unique()[0]+', '+data_formA['Branch'].unique()[0]
        # formAsheet['A10'] = A10_data

        
        formAfinalfile = os.path.join(filelocation,'Form A Employee register.xlsx')
        logging.info('Form A file is' +str(formAfinalfile))
        formAfile.save(filename=formAfinalfile)
        

    def create_form_B():
        formBfilepath = os.path.join(Contractorfilespath,'Form B wage register equal remuniration.xlsx')
        formBfile = load_workbook(filename=formBfilepath)
        logging.info('Form B file has sheet: '+str(formBfile.sheetnames))

        
        logging.info('create columns which are now available')
        data_formB=data.copy(deep=True)
        data_formB.fillna(value=0, inplace=True)

        #data_formB['OT hours'] = 0
        #data_formB['Pay OT'] = 0
        data_formB['DA']= data_formB['DA'].replace("",0).astype(float)
        data_formB['Earned Basic']=data_formB['Earned Basic'].replace("",0).astype(float)
        
        data_formB['basic_and_allo'] = data_formB['Earned Basic']+ data_formB['DA']
        #data_formB['Other EAR'] = data_formB['Other Reimb']+data_formB['Arrears']+data_formB['Other Earning']+data_formB['Variable Pay']+data_formB['Stipend'] +data_formB['Consultancy Fees']
        #data_formB['VPF']=0
        data_formB['Society']="---"
        data_formB['Income Tax']="---"
        
        data_formB['Other Deduction']= data_formB['Other Deduction'].replace("",0).astype(float)
        data_formB['Salary Advance']=data_formB['Salary Advance'].replace("",0).astype(float)

        data_formB['Other Deduc']= data_formB['Other Deduction']+ data_formB['Salary Advance']
        data_formB['EMP PF'] = data_formB['PF']
        #data_formB['BankID'] = ''
        #data_formB['Pay Date'] = ''
        data_formB['Remarks'] =''

        formB_columns = ['Employee Code','Employee Name','FIXED MONTHLY GROSS',	'Days Paid','Total\r\nOT Hrs',	'basic_and_allo', 'Overtime',	
                        'HRA',	'Tel and Int Reimb', 'Bonus', 'Fuel Reimb',	'Prof Dev Reimb', 'Corp Attire Reimb',	'CCA',	
                        'all_Other_Allowance','Total Earning','PF','VPF','P.Tax','Society','LWF EE','Insurance',
                        'TDS','advance+deductions','Recoveries',"Total Deductions",'Net Paid',"PF","Bank A/c Number",
                        'Date of payment','Remarks']

        remove_point=lambda input_str: input_str.split(".")[0]
        data_formB["Bank A/c Number"]=data_formB["Bank A/c Number"].apply(str).apply(remove_point)

        all_other_allowance_columns=['Other Allowance','OtherAllowance1','OtherAllowance2', 'OtherAllowance3', 'OtherAllowance4', 'OtherAllowance5']
        
        data_formB[all_other_allowance_columns]=data_formB[all_other_allowance_columns].replace("",0).astype(float)
        data_formB['all_Other_Allowance']= data_formB.loc[:,all_other_allowance_columns].sum(axis=1)

        all_Other_deductions_columns=['Other Deduction','OtherDeduction1', 'OtherDeduction2','OtherDeduction3', 'OtherDeduction4', 'OtherDeduction5']
        
        data_formB[all_Other_deductions_columns]=data_formB[all_Other_deductions_columns].replace("",0).astype(float)
        data_formB[all_Other_deductions_columns]=data_formB[all_Other_deductions_columns].fillna(0)

        data_formB['Salary Advance']=data_formB['Salary Advance'].replace("",0).astype(float)
        data_formB['Salary Advance']=data_formB['Salary Advance'].fillna(0)
        data_formB['advance+deductions']=data_formB.loc[:,all_Other_deductions_columns].sum(axis=1)+data_formB['Salary Advance']
        
        data_formB['Recoveries']=""


        formB_data = data_formB[formB_columns]

        formBsheet = formBfile['Sheet1']

        formBsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form B is ready')

        
        rows = dataframe_to_rows(formB_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 17):
            for c_idx, value in enumerate(row, 1):
                formBsheet.cell(row=r_idx, column=c_idx, value=value)
                formBsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formBsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formBsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                # if c_idx==45:
                #     formBsheet.cell(row=r_idx, column=c_idx).number_format= numbers.FORMAT_NUMBER

        contractline = formBsheet['A8'].value
        A8_data = contractline+' '+contractor_name+', '+contractor_address
        formBsheet['A8'] = A8_data

        if str(data_formB['Nature of work'].dtype)[0:3] != 'obj':
            data_formB['Nature of work'] = data_formB['Nature of work'].astype(str)

        if str(data_formB['Location'].dtype)[0:3] != 'obj':
            data_formB['Location'] = data_formB['Location'].astype(str)

        if str(data_formB['Company Name'].dtype)[0:3] != 'obj':
            data_formB['Company Name'] = data_formB['Company Name'].astype(str)

        if str(data_formB['Company Address'].dtype)[0:3] != 'obj':
            data_formB['Company Address'] = data_formB['Company Address'].astype(str)

        if str(data_formB['Unit'].dtype)[0:3] != 'obj':
            data_formB['Unit'] = data_formB['Unit'].astype(str)

        if str(data_formB['Address'].dtype)[0:3] != 'obj':
            data_formB['Address'] = data_formB['Address'].astype(str)

        locationline = formBsheet['A9'].value
        A9_data = locationline+' '+data_formB['Nature of work'].unique()[0]+', '+data_formB['Location'].unique()[0]
        formBsheet['A9'] = A9_data

        establine = formBsheet['A10'].value
        if data_formB['PE_or_contract'].unique()[0]== 'PE':
            A10_data = establine+' '+data_formB['Company Name'].unique()[0]+', '+data_formB['Company Address'].unique()[0]  
        else:    
            A10_data = establine+' '+data_formB['Unit'].unique()[0]+', '+data_formB['Address'].unique()[0]
        formBsheet['A10'] = A10_data

        peline = formBsheet['A11'].value
        A11_data = peline+' '+data_formB['Unit'].unique()[0]+', '+data_formB['Address'].unique()[0]
        formBsheet['A11'] = A11_data

        monthstart = datetime.date(year,month_num,1)
        monthend = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
        formBsheet['A12'] = 'Wage period From: '+str(monthstart)+' to '+str(monthend)

        formBfinalfile = os.path.join(filelocation,'Form B wage register equal remuniration.xlsx')
        formBfile.save(filename=formBfinalfile)

    def Form_C():
        formCfilepath = os.path.join(Contractorfilespath,'Form C register of loan or recoveries.xlsx')
        formCfile = load_workbook(filename=formCfilepath)
        logging.info('Form C file has sheet: '+str(formCfile.sheetnames))
        logging.info('create columns which are now available')

        data_formC = data.copy(deep=True)
        
        columns=['Employee Code',"Employee Name","Recovery_Type","Particulars","Date of payment and damage loss",'Damage or Loss',"whether_show_cause_issue","explaination_heard_in_presence_of",
                                    "num_installments","first_month_year","last_month_year","Date_of_complete_recovery","remarks"]
        

        Recovery_Type_columns_name=['Other Deduction','OtherDeduction1', 'OtherDeduction2',
                                                        'OtherDeduction3', 'OtherDeduction4', 'OtherDeduction5','Damage or Loss','Fine','Salary Advance']

        data_formC["Recovery_Type"]=data_formC.loc[:,Recovery_Type_columns_name].sum(axis=1)
        data_formC["amount"]=data_formC["Recovery_Type"]
        data_formC[["Particulars","whether_show_cause_issue","explaination_heard_in_presence_of",
                    "num_installments","first_month_year","last_month_year","Date_of_complete_recovery"]]="---"
        
        data_formC["remarks"]=""
        data_formC['Damage or Loss']=data_formC['Damage or Loss'].replace("",0).astype(float)
        data_formC['Damage or Loss']=data_formC['Damage or Loss'].fillna(0)

        data_formC["Date of payment and damage loss"]=data_formC["Date of payment"].astype(str)+"/"+data_formC['Damage or Loss'].astype(str)
        
        formC_data=data_formC[columns]
        formCsheet = formCfile['Sheet1']
        formCsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        
        rows = dataframe_to_rows(formC_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 9):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formCsheet.cell(row=r_idx, column=c_idx, value=value)
                formCsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formCsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formCsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formC):
                    formCsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formCsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formCsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formCsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formC):
                    formCsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formCsheet.row_dimensions[r_idx].height = 20
                else:
                    formCsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formCsheet.row_dimensions[r_idx].height = 20

        formCsheet['A4']=formCsheet['A4'].value+" : "+str(data_formC['UnitName'].unique()[0])
        formCfinalfile = os.path.join(filelocation,'Form C register of loan or recoveries.xlsx')
        formCfile.save(filename=formCfinalfile)

    def Form_D():
        formDfilepath = os.path.join(Contractorfilespath,'Form D Register of attendance.xlsx')
        formDfile = load_workbook(filename=formDfilepath)
        logging.info('Form D file has sheet: '+str(formDfile.sheetnames))
        logging.info('create columns which are now available')

        data_formD = data.copy(deep=True)
        
        columns=['S.no',"Employee Name","Relay_or_set_work",'Branch']
        
        
        # data_formD_columns=list(data_formD.columns)
        # start=data_formD_columns.index('Emp Code')
        # end=data_formD_columns.index('Total\r\nDP')
        columnstotake =[]
        days = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31']
        for day in days:
            for col in data_formD.columns:
                if col[5:7]==day:
                    columnstotake.append(col)
        if len(columnstotake)==28:
            columnstotake.append('29')
            columnstotake.append('30')
            columnstotake.append('31')
            data_formD['29'] = ''
            data_formD['30'] = ''
            data_formD['31'] = ''
            
        elif len(columnstotake)==29:
            columnstotake.append('30')
            columnstotake.append('31')
            data_formD['30'] = ''
            data_formD['31'] = ''

        elif len(columnstotake)==30:
            columnstotake.append('31')
            data_formD['31'] = ''
        
        columns.extend(columnstotake)
        
        columns.extend(['in','out','Total\r\nDP','num_hours','sign'])

        data_formD[["Relay_or_set_work","in","out",'num_hours','sign']]=""

        data_formD['S.no'] = list(range(1,len(data_formD)+1))

        formD_data=data_formD[columns]
        formDsheet = formDfile['Sheet1']
        formDsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form V is ready')

        
        rows = dataframe_to_rows(formD_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 12):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formDsheet.cell(row=r_idx, column=c_idx, value=value)
                formDsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formDsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formDsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formD):
                    formDsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formDsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formDsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formDsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formD):
                    formDsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formDsheet.row_dimensions[r_idx].height = 20
                else:
                    formDsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formDsheet.row_dimensions[r_idx].height = 20

        
        #formPsheet['AE4']=formPsheet['AE4'].value+"   "+str(data_formP['Registration_no'].unique()[0])
        
        formDsheet['A4']="Name of establishment :-  "+str(data_formD['UnitName'].unique()[0])
        
        # formDsheet['A6']="From:  01"+"-"+str(month)+"-"+str(year)+"       "+"From:  01"+"-"+str(month)+"-"+str(year)
        formDsheet['A5']="Name of Owner:- "+str(data_formD['UnitName'].unique()[0])
        
        formDfinalfile = os.path.join(filelocation,'Form D Register of attendance.xlsx')
        formDfile.save(filename=formDfinalfile)

    def Form_E():
        formEfilepath = os.path.join(Contractorfilespath,'Form E Register of Rest,Leave,leave wages.xlsx')
        formEfile = load_workbook(filename=formEfilepath)
        logging.info('Form E file has sheet: '+str(formEfile.sheetnames))
        logging.info('create columns which are now available')

        data_formE = data.copy(deep=True)
        
        columns=['Employee Code',"Employee Name","Days Paid","opening_bal","added","rest_allowed","rest_availed",
                    "closing_bal","Opening","Monthly Increment","Leave Accrued","Closing",
                    "Monthly Increment","Leave Accrued","Closing","openeing_bal","added","leave_availed","closing_bal",
                    "remarks"]

        data_formE[["opening_bal","added","rest_allowed","rest_availed","closing_bal","openeing_bal","added","leave_availed","closing_bal"]]="---"
        data_formE["remarks"]=""
        formE_data=data_formE[columns]
        formEsheet = formEfile['Sheet1']
        formEsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form E is ready')

        
        rows = dataframe_to_rows(formE_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 13):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formEsheet.cell(row=r_idx, column=c_idx, value=value)
                formEsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formEsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formEsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formE):
                    formEsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formEsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formEsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formEsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formE):
                    formEsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formEsheet.row_dimensions[r_idx].height = 20
                else:
                    formEsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formEsheet.row_dimensions[r_idx].height = 20

        formEsheet['A5']=formEsheet['A5'].value+" : "+str(data_formE['UnitName'].unique()[0])
        formEsheet['A6']=formEsheet['A6'].value+" : "+str(data_formE['UnitName'].unique()[0])
        formEsheet['A8']=formEsheet['A8'].value+" : "+str(month)+" "+str(year)
        
        formEfinalfile = os.path.join(filelocation,'Form E Register of Rest,Leave,leave wages.xlsx')
        formEfile.save(filename=formEfinalfile)

    def create_formXIX():
        formXIXfilepath = os.path.join(Contractorfilespath,'Form XIX Wages slip.xlsx')
        formXIXfile = load_workbook(filename=formXIXfilepath)
        logging.info('Form XIX file has sheet: '+str(formXIXfile.sheetnames))
        sheetformXIX = formXIXfile['Sheet1']

        
        logging.info('create columns which are now available')

        data_formXIX=data.copy(deep=True)
        data_formXIX.fillna(value=0, inplace=True)

        emp_count = len(data_formXIX.index)
        
        for i in range(0,emp_count):
            key = (data_formXIX).iloc[i]['Employee Code']
            sheet_key = 'FORM XIX_'+str(key)

            emp_data = (data_formXIX).iloc[i]

            sheet1 = formXIXfile.copy_worksheet(sheetformXIX)
            sheet1.title = sheet_key
            sheet1['B4'] = contractor_name+', '+contractor_address
            sheet1['B5'] = str(emp_data['Nature of work'])+', '+str(emp_data['Location'])
            if emp_data['PE_or_contract'][0]== 'PE':
                sheet1['B6'] = emp_data['Company Name']+', '+emp_data['Company Address']
                sheet1['B6'] = emp_data['Company Name']+', '+emp_data['Company Address']
            else:
                sheet1['B6'] = emp_data['Unit']+', '+emp_data['Address']
                sheet1['B7'] = emp_data['Unit']+', '+emp_data['Address']
            
            sheet1['B8'] = month+'-'+str(year)
            sheet1['B9'] = key
            
            sheet1['B10'] = emp_data['Employee Name']
            sheet1['B11'] = emp_data['Days Paid']
            sheet1['B12'] = ""
            sheet1['B13'] = ""
            sheet1['B14'] = emp_data['Earned Basic']
            sheet1['B15'] = emp_data['HRA']
            sheet1['B16'] = emp_data['Tel and Int Reimb']
            sheet1['B17'] = emp_data['Bonus']
            sheet1['B18'] = emp_data['Fuel Reimb']
            sheet1['B19'] = emp_data['Corp Attire Reimb']
            sheet1['B20'] = emp_data['CCA']
            sheet1['B21'] = emp_data['Total Earning']
            sheet1['B22'] = emp_data['Insurance']
            sheet1['B23'] = emp_data['PF']
            sheet1['B24'] = emp_data['P.Tax']
            sheet1['B25'] = emp_data['Total Deductions']
            sheet1['B26'] = emp_data['Net Paid']

        formXIXfinalfile = os.path.join(filelocation,'Form XIX Wages slip.xlsx')
        formXIXfile.remove(sheetformXIX)
        formXIXfile.save(filename=formXIXfinalfile)

    def create_formXV():
        formXIXfilepath = os.path.join(Contractorfilespath,'Form XV Service certificate.xlsx')
        formXIXfile = load_workbook(filename=formXIXfilepath)
        logging.info('Form XV file has sheet: '+str(formXIXfile.sheetnames))
        sheetformXIX = formXIXfile['Sheet1']

        
        logging.info('create columns which are now available')

        data_formXIX=data.copy(deep=True)
        
        data_formXIX.fillna(value=0, inplace=True)

        emp_count = len(data_formXIX.index)
        
        for i in range(0,emp_count):
            key = (data_formXIX).iloc[i]['Employee Code']
            sheet_key = 'FORM XV_'+str(key)

            emp_data = (data_formXIX).iloc[i]

            sheet1 = formXIXfile.copy_worksheet(sheetformXIX)
            sheet1.title = sheet_key
            sheet1['A5'] = sheet1['A5'].value+" "+ contractor_name+', '+contractor_address
            sheet1['A6'] = sheet1['A6'].value+" " +emp_data['Unit']+', '+emp_data['Address']
            sheet1['A7'] = sheet1['A7'].value+" " +str(emp_data['Nature of work'])+', '+str(emp_data['Location'])
            sheet1['A8'] = sheet1['A8'].value+" " +emp_data['Unit']+', '+emp_data['Address']
            sheet1['A9'] = sheet1['A9'].value+" " +emp_data['Unit']+', '+emp_data['Address']
            sheet1['A10'] = sheet1['A10'].value+" " + str(emp_data["Age"])
            sheet1['A11'] = sheet1['A11'].value+" " + str(emp_data["Identification mark"])
            sheet1['A12'] = sheet1['A12'].value+" " + str(emp_data["Father's Name"])
            sheet1['A18'] = "1"
            monthstart = datetime.date(year,month_num,1)
            monthend = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
            sheet1['B18'] = monthstart
            sheet1['C18'] = monthend
            sheet1['D18'] =emp_data['Designation']
        
        formXIXfinalfile = os.path.join(filelocation,'Form XV Service certificate.xlsx')
        formXIXfile.remove(sheetformXIX)
        formXIXfile.save(filename=formXIXfinalfile)

    def create_form_XX():
        formXXfilepath = os.path.join(Contractorfilespath,'Form XX Register of Deduction for damage or loss.xlsx')
        formXXfile = load_workbook(filename=formXXfilepath)
        logging.info('Form XX file has sheet: '+str(formXXfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXX=data.copy(deep=True)
        
        data_formXX.fillna(value=0, inplace=True)

        data_formXX['S.no'] = list(range(1,len(data_formXX)+1))

        data_formXX['c'] ='---'
        data_formXX['d'] ='---'
        data_formXX['f'] ='---'
        data_formXX['g'] ='---'
        data_formXX['h'] ='---'
        data_formXX['i'] =''

        formXX_columns = ['S.no','Employee Name',"Father's Name",'Designation','Damage or Loss',"Date of payment",'c','d',"all_Other_Deduction_sum",'f','g','h','i']

        other_deductions_columns_name=['Other Deduction','OtherDeduction1', 'OtherDeduction2',
                                                        'OtherDeduction3', 'OtherDeduction4', 'OtherDeduction5']

        data_formXX[other_deductions_columns_name]=data_formXX[other_deductions_columns_name].replace("",0).astype(float)
        data_formXX["all_Other_Deduction_sum"]= data_formXX.loc[:,other_deductions_columns_name].sum(axis=1)

        formXX_data = data_formXX[formXX_columns]

        formXXsheet = formXXfile['Sheet1']

        formXXsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XX is ready')

        
        rows = dataframe_to_rows(formXX_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 13):
            for c_idx, value in enumerate(row, 1):
                formXXsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXsheet['A5'].value
        A5_data = contractline+' '+contractor_name+', '+contractor_address
        formXXsheet['A5'] = A5_data

        if str(data_formXX['Nature of work'].dtype)[0:3] != 'obj':
            data_formXX['Nature of work'] = data_formXX['Nature of work'].astype(str)

        if str(data_formXX['Location'].dtype)[0:3] != 'obj':
            data_formXX['Location'] = data_formXX['Location'].astype(str)

        if str(data_formXX['Company Name'].dtype)[0:3] != 'obj':
            data_formXX['Company Name'] = data_formXX['Company Name'].astype(str)

        if str(data_formXX['Company Address'].dtype)[0:3] != 'obj':
            data_formXX['Company Address'] = data_formXX['Company Address'].astype(str)

        if str(data_formXX['Unit'].dtype)[0:3] != 'obj':
            data_formXX['Unit'] = data_formXX['Unit'].astype(str)

        if str(data_formXX['Address'].dtype)[0:3] != 'obj':
            data_formXX['Address'] = data_formXX['Address'].astype(str)

        
        locationline = formXXsheet['A6'].value
        A6_data = locationline+' '+data_formXX['Nature of work'].unique()[0]+', '+data_formXX['Location'].unique()[0]
        formXXsheet['A6'] = A6_data

        
        establine = formXXsheet['A7'].value
        A7_data = establine+' '+data_formXX['Unit'].unique()[0]+', '+data_formXX['Address'].unique()[0]
        formXXsheet['A7'] = A7_data

        peline = formXXsheet['A8'].value
        A8_data = peline+' '+data_formXX['Unit'].unique()[0]+', '+data_formXX['Address'].unique()[0]
        formXXsheet['A8'] = A8_data

        #border the region
        count1 = len(data_formXX)
        border_1 = Side(style='thick')
        # for i in range(1,17):
        #     formXXsheet.cell(row=2, column=i).border = Border(outline= True, top=border_1)
        #     formXXsheet.cell(row=count1+16, column=i).border = Border(outline= True, bottom=border_1)
        # for i in range(1,count1+17):
        #     formXXsheet.cell(row=i, column=2).border = Border(outline= True, left=border_1)
        #     formXXsheet.cell(row=i, column=16).border = Border(outline= True, right=border_1)

        formXXfinalfile = os.path.join(filelocation,'Form XX Register of Deduction for damage or loss.xlsx')
        formXXfile.save(filename=formXXfinalfile)


    def create_form_XXI():
        formXXIfilepath = os.path.join(Contractorfilespath,'Form XXI register of fine.xlsx')
        formXXIfile = load_workbook(filename=formXXIfilepath)
        logging.info('Form XXI file has sheet: '+str(formXXIfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXXI=data.copy(deep=True)
        
        data_formXXI.fillna(value=0, inplace=True)

        data_formXXI['S.no'] = list(range(1,len(data_formXXI)+1))

        data_formXXI['a'] ='---'
        data_formXXI['b'] ='---'
        data_formXXI['c'] ='---'
        #data_formXXI['e'] ='---'
        data_formXXI['f'] ='---'
        data_formXXI['g'] =''

        formXXI_columns = ['S.no','Employee Name',"Father's Name",'Designation','a','Date of payment','c','Employee Name','Date of payment and FIXED MONTHLY GROSS','Fine','Date of payment','g']

        data_formXXI['Date of payment and FIXED MONTHLY GROSS']=data_formXXI['Date of payment'].astype(str)+" / "+data_formXXI['FIXED MONTHLY GROSS'].astype(str)
        formXXI_data = data_formXXI[formXXI_columns]

        formXXIsheet = formXXIfile['Sheet1']

        formXXIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XXI is ready')

        
        rows = dataframe_to_rows(formXXI_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 12):
            for c_idx, value in enumerate(row, 1):
                formXXIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXIsheet['A5'].value
        A5_data = contractline+' '+contractor_name+', '+contractor_address
        formXXIsheet['A5'] = A5_data

        if str(data_formXXI['Nature of work'].dtype)[0:3] != 'obj':
            data_formXXI['Nature of work'] = data_formXXI['Nature of work'].astype(str)

        if str(data_formXXI['Location'].dtype)[0:3] != 'obj':
            data_formXXI['Location'] = data_formXXI['Location'].astype(str)

        if str(data_formXXI['Company Name'].dtype)[0:3] != 'obj':
            data_formXXI['Company Name'] = data_formXXI['Company Name'].astype(str)

        if str(data_formXXI['Company Address'].dtype)[0:3] != 'obj':
            data_formXXI['Company Address'] = data_formXXI['Company Address'].astype(str)

        if str(data_formXXI['Unit'].dtype)[0:3] != 'obj':
            data_formXXI['Unit'] = data_formXXI['Unit'].astype(str)

        if str(data_formXXI['Address'].dtype)[0:3] != 'obj':
            data_formXXI['Address'] = data_formXXI['Address'].astype(str)

        locationline = formXXIsheet['A6'].value
        A6_data = locationline+' '+data_formXXI['Nature of work'].unique()[0]+', '+data_formXXI['Location'].unique()[0]
        formXXIsheet['A6'] = A6_data

        establine = formXXIsheet['A7'].value
        A7_data = establine+' '+data_formXXI['Unit'].unique()[0]+', '+data_formXXI['Address'].unique()[0]
        formXXIsheet['A7'] = A7_data

        establine = formXXIsheet['A8'].value
        A8_data = establine+' '+data_formXXI['Unit'].unique()[0]+', '+data_formXXI['Address'].unique()[0]
        formXXIsheet['A8'] = A8_data


        #border the region
        count1 = len(data_formXXI)
        border_1 = Side(style='thick')
        # for i in range(1,12):
        #     formXXIsheet.cell(row=2, column=i).border = Border(outline= True, top=border_1)
        #     formXXIsheet.cell(row=count1+15, column=i).border = Border(outline= True, bottom=border_1)
        # for i in range(1,count1+13):
        #     formXXIsheet.cell(row=i, column=2).border = Border(outline= True, left=border_1)
        #     formXXIsheet.cell(row=i, column=14).border = Border(outline= True, right=border_1)

        formXXIfinalfile = os.path.join(filelocation,'Form XXI register of fine.xlsx')
        formXXIfile.save(filename=formXXIfinalfile)



    def create_form_XXII():
        
        formXXIIfilepath = os.path.join(Contractorfilespath,'Form XXII Register of Advances.xlsx')
        formXXIIfile = load_workbook(filename=formXXIIfilepath)
        logging.info('Form XXII file has sheet: '+str(formXXIIfile.sheetnames))

        
        logging.info('create columns which are now available')
        
        data_formXXII=data.copy(deep=True)
        
        data_formXXII.fillna(value=0, inplace=True)

        data_formXXII['S.no'] = list(range(1,len(data_formXXII)+1))

        data_formXXII['c'] ='---'
        data_formXXII['d'] ='---'
        data_formXXII['e'] ='---'
        data_formXXII['f'] ='---'
        data_formXXII['g'] =''

        formXXII_columns = ['S.no','Employee Name',"Father's Name",'Designation','FIXED MONTHLY GROSS','Date of payment','c','d','e','f','g']

        formXXII_data = data_formXXII[formXXII_columns]

        formXXIIsheet = formXXIIfile['Sheet1']

        formXXIIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XXII is ready')

        
        rows = dataframe_to_rows(formXXII_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 12):
            for c_idx, value in enumerate(row,1):
                formXXIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXIIsheet['A5'].value
        A5_data = contractline+' '+contractor_name+', '+contractor_address
        formXXIIsheet['A5'] = A5_data


        if str(data_formXXII['Nature of work'].dtype)[0:3] != 'obj':
            data_formXXII['Nature of work'] = data_formXXII['Nature of work'].astype(str)

        if str(data_formXXII['Location'].dtype)[0:3] != 'obj':
            data_formXXII['Location'] = data_formXXII['Location'].astype(str)

        if str(data_formXXII['Company Name'].dtype)[0:3] != 'obj':
            data_formXXII['Company Name'] = data_formXXII['Company Name'].astype(str)

        if str(data_formXXII['Company Address'].dtype)[0:3] != 'obj':
            data_formXXII['Company Address'] = data_formXXII['Company Address'].astype(str)

        if str(data_formXXII['Unit'].dtype)[0:3] != 'obj':
            data_formXXII['Unit'] = data_formXXII['Unit'].astype(str)

        if str(data_formXXII['Address'].dtype)[0:3] != 'obj':
            data_formXXII['Address'] = data_formXXII['Address'].astype(str)

        locationline = formXXIIsheet['A6'].value
        A6_data = locationline+' '+data_formXXII['Nature of work'].unique()[0]+', '+data_formXXII['Location'].unique()[0]
        formXXIIsheet['A6'] = A6_data

        establine = formXXIIsheet['A7'].value
        A7_data = establine+' '+data_formXXII['Unit'].unique()[0]+', '+data_formXXII['Address'].unique()[0]
        formXXIIsheet['A7'] = A7_data

        establine = formXXIIsheet['A8'].value
        A8_data = establine+' '+data_formXXII['Unit'].unique()[0]+', '+data_formXXII['Address'].unique()[0]
        formXXIIsheet['A8'] = A8_data

        #border the region
        count1 = len(data_formXXII)
        border_1 = Side(style='thick')
        # for i in range(1,12):
        #     formXXIIsheet.cell(row=2, column=i).border = Border(outline= True, top=border_1)
        #     formXXIIsheet.cell(row=count1+16, column=i).border = Border(outline= True, bottom=border_1)
        # for i in range(1,count1+14):
        #     formXXIIsheet.cell(row=i, column=2).border = Border(outline= True, left=border_1)
        #     formXXIIsheet.cell(row=i, column=14).border = Border(outline= True, right=border_1)

        formXXIIfinalfile = os.path.join(filelocation,'Form XXII Register of Advances.xlsx')
        formXXIIfile.save(filename=formXXIIfinalfile)


    def create_form_XXIII():
        formXXIIIfilepath = os.path.join(Contractorfilespath,'Form XXIII register of overtime.xlsx')
        formXXIIIfile = load_workbook(filename=formXXIIIfilepath)
        logging.info('Form XXIII file has sheet: '+str(formXXIIIfile.sheetnames))

        
        logging.info('create columns which are now available')
        data_formXXIII=data.copy(deep=True)
        
        data_formXXIII.fillna(value=0, inplace=True)

        data_formXXIII['S.no'] = list(range(1,len(data_formXXIII)+1))

        data_formXXIII['a'] ='---'
        data_formXXIII['g'] =''

        formXXIII_columns = ['S.no','Employee Name',"Father's Name",'Gender','Designation','Date of payment','a','FIXED MONTHLY GROSS',"overtime rate","Overtime",'Date of payment','g']

        formXXIII_data = data_formXXIII[formXXIII_columns]

        formXXIIIsheet = formXXIIIfile['Sheet1']

        formXXIIIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XXIII is ready')

        
        rows = dataframe_to_rows(formXXIII_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 12):
            for c_idx, value in enumerate(row, 1):
                formXXIIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXIIIsheet['A5'].value
        A5_data = contractline+' '+contractor_name+', '+contractor_address
        formXXIIIsheet['A5'] = A5_data

        if str(data_formXXIII['Nature of work'].dtype)[0:3] != 'obj':
            data_formXXIII['Nature of work'] = data_formXXIII['Nature of work'].astype(str)

        if str(data_formXXIII['Location'].dtype)[0:3] != 'obj':
            data_formXXIII['Location'] = data_formXXIII['Location'].astype(str)

        if str(data_formXXIII['Company Name'].dtype)[0:3] != 'obj':
            data_formXXIII['Company Name'] = data_formXXIII['Company Name'].astype(str)

        if str(data_formXXIII['Company Address'].dtype)[0:3] != 'obj':
            data_formXXIII['Company Address'] = data_formXXIII['Company Address'].astype(str)

        if str(data_formXXIII['Unit'].dtype)[0:3] != 'obj':
            data_formXXIII['Unit'] = data_formXXIII['Unit'].astype(str)

        if str(data_formXXIII['Address'].dtype)[0:3] != 'obj':
            data_formXXIII['Address'] = data_formXXIII['Address'].astype(str)

        locationline = formXXIIIsheet['A6'].value
        A6_data = locationline+' '+data_formXXIII['Nature of work'].unique()[0]+', '+data_formXXIII['Location'].unique()[0]
        formXXIIIsheet['A6'] = A6_data

        establine = formXXIIIsheet['A7'].value
        A7_data = establine+' '+data_formXXIII['Unit'].unique()[0]+', '+data_formXXIII['Address'].unique()[0]
        formXXIIIsheet['A7'] = A7_data

        establine = formXXIIIsheet['A8'].value
        A8_data = establine+' '+data_formXXIII['Unit'].unique()[0]+', '+data_formXXIII['Address'].unique()[0]
        formXXIIIsheet['A8'] = A8_data


        #border the region
        count1 = len(data_formXXIII)
        border_1 = Side(style='thick')
        # for i in range(1,13):
        #     formXXIIIsheet.cell(row=2, column=i).border = Border(outline= True, top=border_1)
        #     formXXIIIsheet.cell(row=count1+13, column=i).border = Border(outline= True, bottom=border_1)
        # for i in range(1,count1+12):
        #     formXXIIIsheet.cell(row=i, column=2).border = Border(outline= True, left=border_1)
        #     formXXIIIsheet.cell(row=i, column=14).border = Border(outline= True, right=border_1)


        formXXIIIfinalfile = os.path.join(filelocation,'Form XXIII register of overtime.xlsx')
        formXXIIIfile.save(filename=formXXIIIfinalfile)


    def create_ecard():

        ecardfilepath = os.path.join(Contractorfilespath,'FormXII Employment Card.xlsx')
        ecardfile = load_workbook(filename=ecardfilepath)
        logging.info('Employment card file has sheet: '+str(ecardfile.sheetnames))
        sheetecard = ecardfile['Sheet1']

        
        logging.info('create columns which are now available')
        data_ecard=data.copy(deep=True)
        data_ecard.fillna(value=0, inplace=True)

        emp_count = len(data_ecard.index)
        
        for i in range(0,emp_count):
            key = (data_ecard).iloc[i]['Employee Code']
            sheet_key = 'Employment card_'+str(key)

            emp_data = (data_ecard).iloc[i]
            
            sheet1 = ecardfile.copy_worksheet(sheetecard)
            sheet1.title = sheet_key
            sheet1['B4'] = contractor_name
            sheet1['B5'] = str(emp_data['Contractor_LIN'])+' / '+str(emp_data['Contractor_PAN'])
            sheet1['B6'] = emp_data['Contractor_email']
            sheet1['B7'] = emp_data['Contractor_mobile']
            sheet1['B7'].number_format= numbers.FORMAT_NUMBER
            sheet1['B8'] = emp_data['Nature of work']
            sheet1['B9'] = contractor_address
            sheet1['B10'] = emp_data['Unit']
            sheet1['B11'] = str(emp_data['Unit_LIN'])+' / '+str(emp_data['Unit_PAN'])
            sheet1['B12'] = emp_data['Unit_email']
            sheet1['B13'] = emp_data['Unit_mobile']
            sheet1['B13'].number_format= numbers.FORMAT_NUMBER
            sheet1['B14'] = emp_data['Employee Name']
            sheet1['B15'] = emp_data['Aadhar Number']
            sheet1['B15'].number_format= numbers.FORMAT_NUMBER
            sheet1['B16'] = emp_data['Mobile Tel No.']
            sheet1['B16'].number_format= numbers.FORMAT_NUMBER
            sheet1['B17'] = key
            sheet1['B18'] = emp_data['Designation']
            sheet1['B19'] = "" #emp_data['FIXED MONTHLY GROSS']
            sheet1['B20'] = emp_data['Date Joined']
            sheet1['B21'] = '-'
            

        ecardfinalfile = os.path.join(filelocation,'FormXII Employment Card.xlsx')
        ecardfile.remove(sheetecard)
        ecardfile.save(filename=ecardfinalfile)
            
    try:
        create_form_A()
        master.update()
        create_form_B()
        master.update()
        Form_C()
        master.update()
        Form_D()
        master.update()
        Form_E()
        master.update()
        create_formXIX()
        master.update()
        create_formXV()
        master.update()
        create_form_XX()
        master.update()
        create_form_XXI()
        master.update()
        create_form_XXII()
        master.update()
        create_form_XXIII()
        master.update()
        create_ecard()
        master.update()
    except KeyError as e:
        logging.info("Key error : Check if {} column exsists".format(e))
        print("Key error {}".format(e))
        report.configure(text="Failed: Check input file format  \n column {} not found".format(e))
        master.update()
        raise KeyError
    except FileNotFoundError as e:
        logging.info("File not found : Check if {} exsists".format(e))
        report.configure(text="Failed: File {} not found".format(e))
        master.update()
        raise FileNotFoundError