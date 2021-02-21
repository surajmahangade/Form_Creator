
# This code is used to develop the form P,R,T section of Tamil Nadu state
# Author: Riddhi Pravin Shah
from states import logging,monthdict,Statefolder,create_border
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging
from collections import Counter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill, numbers


def Tamilnadu(data,contractor_name,contractor_address,filelocation,month,year,report,master):
    Tamilnadufilespath = os.path.join(Statefolder,'Tamilnadu')
    logging.info('Tamilnadu files path is :'+str(Tamilnadufilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]

    def Form_P():
        formPfilepath = os.path.join(Tamilnadufilespath,'Form P register of deduction.xlsx')
        formPfile = load_workbook(filename=formPfilepath)
        logging.info('Form P file has sheet: '+str(formPfile.sheetnames))
        logging.info('create columns which are now available')
        
        data_formP = data.copy(deep=True)
        data_formP=data_formP.drop_duplicates(subset="Employee Code", keep="last")
        
        columns=['S.no',"Employee Name","Father's Name","Employee Code","Designation","Date of payment","Net Paid","num_instalments_recovered","Date_recovery_completed",
                                    "Damage or Loss","Date_show_cause_notice","Total Deductions","num_installments_to_recovered","Date_deduction_completed",
                                    "act_or_ommision","date_of_show_cause_notice","Fine","Date_fine_recovery_completed","sign","remarks"]
        
        data_formP['S.no'] = list(range(1,len(data_formP)+1))
        data_formP[["num_instalments_recovered","Date_recovery_completed","Date_show_cause_notice","num_installments_to_recovered","Date_deduction_completed",
                                    "act_or_ommision","date_of_show_cause_notice","Date_fine_recovery_completed","sign","remarks"]]=""
        
        formP_data=data_formP[columns]
        formPsheet = formPfile['Sheet1']
        formPsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form E is ready')

        
        rows = dataframe_to_rows(formP_data, index=False, header=False)

        logging.info('rows taken out from data')
        start_row=9
        r_idx=0
        c_idx=0
        for r_idx, row in enumerate(rows, start_row):
            for c_idx, value in enumerate(row, 1):
                formPsheet.cell(row=r_idx, column=c_idx, value=value)
                formPsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formPsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formPsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        formPsheet=create_border(formPsheet,r_idx,c_idx,start_row)
        formPsheet['A6']=str(data_formP['Unit'].unique()[0])+","+str(data_formP['Address'].unique()[0])
        formPfinalfile = os.path.join(filelocation,'Form P register of deduction.xlsx')
        formPfile.save(filename=formPfinalfile)

    def Form_Q():
        formQfilepath = os.path.join(Tamilnadufilespath,'Form Q register of employment.xlsx')
        formQfile = load_workbook(filename=formQfilepath)
        logging.info('Form I file has sheet: '+str(formQfile.sheetnames))
        logging.info('create columns which are now available')
        
        data_formQ = data.copy(deep=True)
        data_formQ=data_formQ.drop_duplicates(subset="Employee Code", keep="last")
        
        columns=['S.no',"Employee Name","Date Joined","Date of Birth","Designation",
        
        
                    "Gender","Age","start_time","end_time","rest_interval","mon","tue","wed","thu","Fri","sat","sun",
                                                "days_overtime","extent_of_overtime","extent_of_overtime_previously"]
    
        data_formQ['S.no'] = list(range(1,len(data_formQ)+1))
        
        formQ_data=data_formQ[columns]
        formQsheet = formQfile['Sheet1']
        formQsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form Q is ready')

        
        rows = dataframe_to_rows(formQ_data, index=False, header=False)

        logging.info('rows taken out from data')

        row_num=0
        for r_idx, row in enumerate(rows, 7):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formQsheet.cell(row=r_idx, column=c_idx, value=value)
                formQsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formQsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formQsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formI):
                    formQsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formQsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formQsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formQsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formI):
                    formQsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formQsheet.row_dimensions[r_idx].height = 20
                else:
                    formQsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formQsheet.row_dimensions[r_idx].height = 20
        
        
        formQfinalfile = os.path.join(filelocation,'Form Q register of employment.xlsx')
        formQfile.save(filename=formQfinalfile)


    def Form_R():
        formRfilepath = os.path.join(Tamilnadufilespath,'Form R register of wages.xlsx')
        formRfile = load_workbook(filename=formRfilepath)
        logging.info('Form R file has sheet: '+str(formRfile.sheetnames))
        logging.info('create columns which are now available')

        data_formR = data.copy(deep=True)
        data_formR=data_formR.drop_duplicates(subset="Employee Code", keep="last")
        
        columns=['S.no',"Employee Name","Gender","Designation","Daily_rated","wages_period","Days Paid","units_of_work_done",
                    "Daily_rate_wages","Overtime","Earned Basic",'DA',"all_Other_Allowance","Overtime","leave_wages","FIXED MONTHLY GROSS",
                    'PF',"Insurance","all_Other_deductions","Fine","Net Paid","sign","total_unpaid_amt"]
        
        
        data_formR[["Daily_rate_wages"]]="--"
        data_formR[["Daily_rated","units_of_work_done","leave_wages","sign","total_unpaid_amt"]]=""
        data_formR["wages_period"]=str(month)+" "+str(year)
        data_formR['Dearness_Allowance']=data_formR['DA']
        
        # data_formR["Basic"]=min_wages_goa
        all_other_allowance_columns=['Other Allowance','OtherAllowance1','OtherAllowance2', 'OtherAllowance3', 'OtherAllowance4', 'OtherAllowance5']
        
        data_formR[all_other_allowance_columns]=data_formR[all_other_allowance_columns].replace("",0).astype(float)
        data_formR['all_Other_Allowance']= data_formR.loc[:,all_other_allowance_columns].sum(axis=1)

        all_Other_deductions_columns=['Other Deduction','OtherDeduction1', 'OtherDeduction2','OtherDeduction3', 
                                        'OtherDeduction4', 'OtherDeduction5']
        
        data_formR[all_Other_deductions_columns]=data_formR[all_Other_deductions_columns].replace("",0).astype(float)
        data_formR[all_Other_deductions_columns]=data_formR[all_Other_deductions_columns].fillna(0)

        data_formR["all_Other_deductions"]=data_formR.loc[:,all_Other_deductions_columns].sum(axis=1)

        data_formR['S.no'] = list(range(1,len(data_formR)+1))

        formR_data=data_formR[columns]
        formRsheet = formRfile['Sheet1']
        formRsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form R is ready')

        
        rows = dataframe_to_rows(formR_data, index=False, header=False)
        rows_copy = list(dataframe_to_rows(formR_data, index=False, header=False))
        logging.info('rows taken out from data')
        r_idx=0
        c_idx=0
        start_row=9
        formRsheet.insert_rows(start_row,len(data_formR)-2)
        for r_idx, row in enumerate(rows, start_row):
            for c_idx, value in enumerate(row, 1):
                formRsheet.cell(row=r_idx, column=c_idx, value=value)
                formRsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formRsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formRsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        formRsheet=create_border(formRsheet,r_idx,c_idx,start_row)
        
        formRsheet['A4']=" Name of Establishment:-   "+str(data_formR['UnitName'].unique()[0])
        formRsheet['A5']="Name of Employer and address:-   "+str(data_formR['UnitName'].unique()[0])+","+str(data_formR['Address'].unique()[0])+" / "+str(data_formR['Contractor_name'].unique()[0])
        
        formRfinalfile = os.path.join(filelocation,'Form R Register of wages.xlsx')
        formRfile.save(filename=formRfinalfile)

    def Form_T():# - Form T wages slip

        formPath = os.path.join(Tamilnadufilespath,'Form T wages slip.xlsx')
        formfile = load_workbook(filename=formPath)
        logging.info('Form T file has sheet: '+str(formfile.sheetnames))
        logging.info('create columns which are now available')

        form_data = data.copy(deep=True)
        form_data=form_data.drop_duplicates(subset="Employee Code", keep="last")
        
        # select sheet to write in and set page properties
        formsheet = formfile['Sheet1']
        formsheet.sheet_properties.pageSetUpPr.fitToPage = True

        # for each employee create a new sheet
        for idx in form_data.index:

            empCode = form_data.loc[idx, "Employee Code"]
            
            # create a new sheet for employee
            new = formfile.copy_worksheet(formsheet)
            new.title = empCode

            # popoluate every cell with required employee information 
            new.cell(row= 4, column=3, value=form_data.loc[idx, "Company Name"])# Name of company
            new.cell(row= 5, column=3, value=form_data.loc[idx, "Employee Name"])  #. name of employee
            new.cell(row= 6, column=3, value=form_data.loc[idx, "Father's Name"])  #. Fathers name
            new.cell(row= 7, column=3, value=form_data.loc[idx, "Designation"]) # employee designation C7
            new.cell(row= 8, column=3, value=form_data.loc[idx, "Date Joined"]) # date joined C8
            new.cell(row= 11, column=2, value=form_data.loc[idx, "Earned Basic"]) # earned basic B11
            new.cell(row= 13, column=2, value=form_data.loc[idx, "HRA"])# House rent allowance B13
            new.cell(row= 14, column=2, value=form_data.loc[idx, "Overtime"]) # overtime wages B14
            new.cell(row= 16, column=2, value=form_data.loc[idx,"Other Allowance"]) # otehr allowance B16
            new.cell(row= 17, column=2, value=form_data.loc[idx, "FIXED MONTHLY GROSS"]) # Gross wages B17
            new.cell(row= 12, column=9, value=form_data.loc[idx, "Insurance"]) # employee state insurance I12
            new.cell(row= 13, column=9, value=form_data.loc[idx, "Other Deduction"]) # other deductions I13
            new.cell(row= 17, column=7, value=form_data.loc[idx, "Net Paid"]) # net paid #G17

            # format table A1 - K19
            for row_id in range(1,20):
                for col_id in range(1,12):
                    new.cell(row=row_id, column=col_id).border = Border(left=Side(style='medium'), 
                                                                                       right=Side(style='medium'), 
                                                                                           top=Side(style='medium'), 
                                                                                         bottom=Side(style='medium'))

        # delete sheet1 sheet2, sheet3
        formfile.remove(formfile['Sheet1']) 
        formfile.remove(formfile['Sheet2']) 
        formfile.remove(formfile['Sheet3']) 

        # save file
        formTfinalfile = os.path.join(filelocation,'Form T wages slip.xlsx')      
        formfile.save(filename=formTfinalfile)
    ## --------FUNCTION CALL-------------------------
    try:
        Form_P()  ## Call this function in the main def
        master.update()
        Form_R()
        master.update()
        Form_T()
        master.update()
        # Form_Q()
    except KeyError as e:
        logging.info("Key error : Check if {} column exsists".format(e))
        report.configure(text="Failed: Check input file format  \n column {} not found".format(e))
        master.update()
        raise KeyError
    except FileNotFoundError as e:
        logging.info("File not found : Check if {} exsists".format(e))
        report.configure(text="Failed: File {} not found".format(e))
        master.update()
        raise FileNotFoundError