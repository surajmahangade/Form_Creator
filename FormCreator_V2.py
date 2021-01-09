from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from dateutil import parser
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill, numbers
import calendar
import logging
from collections import Counter

master = Tk()
master.title("Form Creator")
master.minsize(640,400)


#backend code starts here

systemdrive = os.getenv('WINDIR')[0:3]
dbfolder = os.path.join(systemdrive,'Forms\DB')
#dbfolder = "D:\Company Projects\Form creator\DB"
State_forms = os.path.join(systemdrive,'Forms\State forms')
#State_forms = "D:\Company Projects\Form creator\State forms"
Statefolder = Path(State_forms)
logfolder = os.path.join(systemdrive,'Forms\logs')
#logfolder = "D:\Company Projects\Form creator\logs"


monthdict= {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}


log_filename = datetime.datetime.now().strftime(os.path.join(logfolder,'logfile_%d_%m_%Y_%H_%M_%S.log'))
logging.basicConfig(filename=log_filename, level=logging.INFO)

def create_pdf(folderlocation,file_name):
    import win32com.client
    from pywintypes import com_error



    excel_filename = file_name
    pdf_filename = file_name.split('.')[0]+'.pdf'

    


    # Path to original excel file
    WB_PATH=os.path.join(folderlocation,excel_filename)
    # PDF path when saving
    PATH_TO_PDF =os.path.join(folderlocation,pdf_filename)

    logging.info(WB_PATH)
    logging.info(PATH_TO_PDF)


    excel = win32com.client.Dispatch("Excel.Application")

    excel.Visible = False

    try:
        logging.info('Start conversion to PDF')

        # Open
        wb = excel.Workbooks.Open(WB_PATH)

        sheetnumbers= len(pd.ExcelFile(WB_PATH).sheet_names)

        # Specify the sheet you want to save by index. 1 is the first (leftmost) sheet.
        ws_index_list = list(range(1,sheetnumbers+1))
        wb.WorkSheets(ws_index_list).Select()

        # Save
        wb.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF)
    except com_error as e:
        logging.info('failed.')
    else:
        logging.info('Succeeded.')
    finally:
        wb.Close()
        excel.Quit()


def Rajasthan(data,contractor_name,contractor_address,filelocation,month,year):
    logging.info("Rajasthan form creation")
    



def Madhya_Pradesh(data,contractor_name,contractor_address,filelocation,month,year):
    logging.info("Madhya Pradesh form creation")
    

def Kerala(data,contractor_name,contractor_address,filelocation,month,year):
    logging.info("Kerala form creation")



def read_min_wages_file(state_name,type_skilled,input_filelocation):
    formsfilelist = os.listdir(input_filelocation)
    for f in formsfilelist:
        if f[0:13].upper()=='MINIMUM WAGES':
            min_wagesfilename = f
            logging.info('min_wagesfilename is :'+f)
    min_wages=pd.read_excel(os.path.join(input_filelocation,min_wagesfilename))
    min_wages=min_wages.dropna(how="all",axis=1)
    min_wages=min_wages.dropna(how="all",axis=0)
    min_wages.columns=["SR NO" ,"STATE" ,"HIGHLY SKILLED" , "SKILLED" ,"SEMI-SKILLED" ,"UNSKILLED"]
    min_wages=min_wages.drop(["SR NO"],axis=1)
    min_wages=min_wages.set_index("STATE")
    min_wages=min_wages.loc["GUJARAT",type_skilled]
    return min_wages

def Gujarat(data,contractor_name,contractor_address,filelocation,month,year):
    Gujaratfilespath = os.path.join(Statefolder,'Gujarat')
    logging.info('Gujarat files path is :'+str(Gujaratfilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]
    input_filelocation=filelocation.split("Registers")[0]
    min_wages_gujarat=read_min_wages_file("GUJARAT","SEMI-SKILLED",input_filelocation)

    def Form_F():
        formFfilepath = os.path.join(Gujaratfilespath,'Form F Register of refusal of leave.xlsx')
        formFfile = load_workbook(filename=formFfilepath)
        logging.info('Form F file has sheet: '+str(formFfile.sheetnames))
        logging.info('create columns which are now available')

        data_formF = data.copy(deep=True)
        data_formF=data_formF.drop_duplicates(subset="Employee Name", keep="last")

        columns=['S.no',"name_employer",'Company Name',"Address","Employee Name","Leave_due","Encash","Date_of_refusal","sign","remarks"]
        
        data_formF["name_employer"]=""
        data_formF["PE_or_contract"]=data_formF["PE_or_contract"].astype(str)
        data_formF.loc[data_formF["PE_or_contract"]=="PE","name_employer"]=data_formF.loc[data_formF["PE_or_contract"]=="PE",'Company Name']
        data_formF.loc[data_formF["PE_or_contract"]=="Contract","name_employer"]=data_formF.loc[data_formF["PE_or_contract"]=="Contract",'UnitName']
        
        data_formF[["Leave_due","Encash","Date_of_refusal","sign","remarks"]]="---"
        data_formF['S.no'] = list(range(1,len(data_formF)+1))
        formF_data=data_formF[columns]
        formFsheet = formFfile['Sheet1']
        formFsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form F is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formF_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 10):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formFsheet.cell(row=r_idx, column=c_idx, value=value)
                formFsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formFsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formFsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formF):
                    formFsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formFsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formFsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formFsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formF):
                    formFsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formFsheet.row_dimensions[r_idx].height = 20
                else:
                    formFsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formFsheet.row_dimensions[r_idx].height = 20

        formFsheet['A4']=formFsheet['A4'].value+"   "+str(data_formF['Company Name'].unique()[0])+","+str(data_formF['Company Address'].unique()[0])
        if data_formF["PE_or_contract"].unique()[0].upper()=="CL":
            formFsheet['A5']=formFsheet['A5'].value+"   "+data_formF['UnitName'][0]+" "+data_formF["Address"][0]
        
        formFsheet['A6']=formFsheet['A6'].value+"   "+str(data_formF['Branch'].unique()[0])
        formFsheet['A7']="WAGE PERIOD    "+str(month)+" "+str(year)

        formFfinalfile = os.path.join(filelocation,'Form F Register of refusal of leave.xlsx')
        formFfile.save(filename=formFfinalfile)

    def Form_I():
        formIfilepath = os.path.join(Gujaratfilespath,'Form I Register of employment in a shop.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')
        
        data_formI = data.copy(deep=True)
        data_formI=data_formI.drop_duplicates(subset="Employee Name", keep="last")
        
        columns=["Employee Name","Gender","Age","start_time","end_time","rest_interval","mon","tue","wed","thu","Fri","sat","sun",
                                                "days_overtime","extent_of_overtime","extent_of_overtime_previously"]
    
        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        
        data_formI[["mon","tue","wed","thu","Fri","sat","sun","days_overtime","extent_of_overtime","extent_of_overtime_previously"]]=""
        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')

        formIsheet.unmerge_cells("A8:P8")
        formIsheet.unmerge_cells("A9:P9")
        
        formIsheet.insert_rows(7,len(data_formI))
        row_num=0
        for r_idx, row in enumerate(rows, 7):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIsheet.row_dimensions[r_idx].height = 20
                else:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIsheet.row_dimensions[r_idx].height = 20
        
        formIsheet.merge_cells("A"+str(8+len(data_formI))+":P"+str(8+len(data_formI)))
        formIsheet.merge_cells("A"+str(9+len(data_formI))+":P"+str(9+len(data_formI)))

        formIfinalfile = os.path.join(filelocation,'Form I Register of employment in a shop.xlsx')
        formIfile.save(filename=formIfinalfile)

    def Form_IV():
        
        formIVfilepath = os.path.join(Gujaratfilespath,'Form IV A register  of wages.xlsx')
        formIVfile = load_workbook(filename=formIVfilepath)
        logging.info('Form IV file has sheet: '+str(formIVfile.sheetnames))
        logging.info('create columns which are now available')

        data_formIV = data.copy(deep=True)
        data_formIV=data_formIV.drop_duplicates(subset="Employee Name", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Designation","basic","DA","Earned Basic","DA","Days Paid",
                                        "Overtime","HRA",'Tel and Int Reimb',"Bonus","Fuel Reimb","Prof Dev Reimb","Corp Attire Reimb","CCA",
                                        "deductions-advance",'Total Earning','PF',"H.R.","all_Other_deductions","Insurance","P.Tax","Total Deductions","Net Paid",
                                        "Date of payment","Bank A/c Number","sign"]
        
        remove_point=lambda input_str: input_str.split(".")[0]
        data_formIV["Bank A/c Number"]=data_formIV["Bank A/c Number"].apply(str).apply(remove_point)

        data_formIV['S.no'] = list(range(1,len(data_formIV)+1))
        data_formIV["basic"]=min_wages_gujarat
        #others_columns=[]
        others_columns=['HRA','Conveyance','Medical Allowance','Telephone Reimb','Tel and Int Reimb',
                                            'Bonus','Other Allowance', 'Fuel Reimb','Prof Dev Reimb','Corp Attire Reimb',
                                            'Meal Allowance','Special Allowance','Personal Allowance','CCA','Other Reimb',
                                            'Arrears','Other Earning',"Retention Pay",'Variable Pay','Leave Encashment',
                                            'Stipend','Consultancy Fees','Covid Deduction','OtherAllowance1', 
                                            'OtherAllowance2', 'OtherAllowance3', 'OtherAllowance4', 'OtherAllowance5'
                                            ]
        if "Covid Deduction" not in data_formIV.columns:
            data_formIV["Covid Deduction"]=0
        if "Retention Pay" not in data_formIV.columns:
            data_formIV["Retention Pay"]=0
        # data_formIV["deductions-advance"]=data
        data_formIV[others_columns]=data_formIV[others_columns].astype(float)
        data_formIV[others_columns]=data_formIV[others_columns].fillna(0)
        data_formIV['Salary Advance']=data_formIV['Salary Advance'].astype(float)
        data_formIV['Salary Advance']=data_formIV['Salary Advance'].fillna(0)
        
        data_formIV["deductions-advance"]= data_formIV.loc[:,others_columns].sum(axis=1)-data_formIV['Salary Advance']
        data_formIV["H.R."]=0

        all_Other_deductions_columns=['Other Deduction','OtherDeduction1', 'OtherDeduction2','OtherDeduction3', 'OtherDeduction4', 'OtherDeduction5']
        
        data_formIV[all_Other_deductions_columns]=data_formIV[all_Other_deductions_columns].astype(float)
        data_formIV[all_Other_deductions_columns]=data_formIV[all_Other_deductions_columns].fillna(0)

        data_formIV["all_Other_deductions"]=data_formIV.loc[:,all_Other_deductions_columns].sum(axis=1)

        data_formIV["sign"]=""
        #data_formIV["Date_overtime_worked"]=month
        formIV_data=data_formIV[columns]
        formIVsheet = formIVfile['Sheet1']
        formIVsheet.sheet_properties.pageSetUpPr.fitToPage = True
        #for column in  range(ord('A'), ord('O') + 1):
        #    formIVsheet.unmerge_cells(chr(column)+"7:"+chr(column)+"14")

        logging.info('data for form IV is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formIV_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows,11):
            row_num+=1
        
            for c_idx, value in enumerate(row, 1):
                formIVsheet.cell(row=r_idx, column=c_idx, value=value)
                formIVsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIVsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formIV):
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIVsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIVsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formIV):
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIVsheet.row_dimensions[r_idx].height = 20
                else:
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIVsheet.row_dimensions[r_idx].height = 20

        formIVsheet['A3']=formIVsheet['A3'].value+" "+str(data_formIV['Company Name'].unique()[0])
        if data["PE_or_contract"].unique()[0].upper()=="CL":
            formXXIIIsheet['A4']=" Name of Establishment:-   "+str(data_formIV['UnitName'].unique()[0])

        formIVsheet['A5']="PLACE "+data_formIV['Branch'].unique()[0]
        formIVsheet['A6']=formIVsheet['A6'].value+" "+str(month)+" "+str(year)
        formIVfinalfile = os.path.join(filelocation,'Form IV A register  of wages.xlsx')
        formIVfile.save(filename=formIVfinalfile)
                    
    def Form_M():
        formMfilepath = os.path.join(Gujaratfilespath,'Form M Register of leave.xlsx')
        formMfile = load_workbook(filename=formMfilepath)
        logging.info('Form M file has sheet: '+str(formMfile.sheetnames))
        logging.info('create columns which are now available')

        data_formM = data.copy(deep=True)
        leave_file_data=data_formM[["Employee Code","Employee Name","Leave Type","Opening","Monthly Increment","Leave Accrued","Used","Encash","Closing"]]
        
        data_formM=data_formM.drop_duplicates(subset="Employee Name", keep="last")
        data_formM["Employee Name & Code"]=data_formM["Employee Name"].astype(str)+"||"+data_formM["Employee Code"].astype(str)

        columns=["Employee Name & Code","Department","Date Joined","month_year","num_days","balance_days","Date Left",'Leave Encashment']
        data_formM["month_year"]=str(month)+" "+str(year)

        for employee_name_leave_file in data_formM["Employee Name"]:
            #opening
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==employee_name_leave_file,:]
            opening_pl=emp_details["Opening"].loc[emp_details["Leave Type"]=="PL"].astype(float)
            opening_cl=emp_details["Opening"].loc[emp_details["Leave Type"]=="CL"].astype(float)
            opening_sl=emp_details["Opening"].loc[emp_details["Leave Type"]=="SL"].astype(float)
            prev_bal=opening_pl.add(opening_cl.add(opening_sl,fill_value=0), fill_value=0).sum()
            
            data_formM.loc[data_formM["Employee Name"]==employee_name_leave_file,'num_days']=prev_bal


            #####
            #monthly_inr
            mon_inr_pl=emp_details["Monthly Increment"].loc[emp_details["Leave Type"]=="PL"].astype(float)
            mon_inr_cl=emp_details["Monthly Increment"].loc[emp_details["Leave Type"]=="CL"].astype(float)
            mon_inr_sl=emp_details["Monthly Increment"].loc[emp_details["Leave Type"]=="SL"].astype(float)
            earned=mon_inr_cl.add(mon_inr_pl.add(mon_inr_sl,fill_value=0), fill_value=0).sum()

            #availed during month
            Used_pl=emp_details["Used"].loc[emp_details["Leave Type"]=="PL"].astype(float)
            Used_cl=emp_details["Used"].loc[emp_details["Leave Type"]=="CL"].astype(float)
            Used_sl=emp_details["Used"].loc[emp_details["Leave Type"]=="SL"].astype(float)
            availed=Used_cl.add(Used_pl.add(Used_sl,fill_value=0), fill_value=0).sum()

            data_formM.loc[data_formM["Employee Name"]==employee_name_leave_file,"balance_days"]=prev_bal+earned-availed

            


        data_formM[['balance_days']]=""
        data_formM_columns=list(data_formM.columns)
        start=data_formM_columns.index('Arrears salary')
        end=data_formM_columns.index('Total\r\nDP')
        columns.extend(data_formM_columns[start+1:end])


        formM_data=data_formM[columns]
        formMsheet = formMfile['Sheet1']

        formMsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form M is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formM_data, index=False, header=False)

        logging.info('rows taken out from data')

        def cell_write(sheet,r_idx,c_idx,value):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            
        def start_end_date_attendance(absent_label,offset,row_offset,initial_offset):  
            is_abs_num=0
            row_index=0
            added={}
            for sheet_idx, row in enumerate(rows, 10):
                row_index=0
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        if code =="nan":
                            code=name
                        try:
                            target=formMfile[code]
                            added[target.title]=0
                        except:
                            target = formMfile.copy_worksheet(formMsheet)
                            target.title=code
                            #initial offset
                            row_offset[code]=initial_offset
                            added[target.title]=initial_offset
                        target['A5']="Name of Employee : "+name
                        formMsheet['A4']="Name of the employer or the establishment:- "+str(data_formM['Company Name'].unique()[0])+","+str(data_formM['Company Address'].unique()[0])    
                    elif c_idx==2:
                        target['A6']="Description of the department (if applicable):    "+value
                    elif c_idx==3:
                        target['A7']="Date of entry into service:  "+value
                    elif c_idx==4:
                        Leave_Accrued=value 
                        cell_write(target,row_index+row_offset[target.title],1,Leave_Accrued)
                    elif c_idx==5:
                        num_days=value
                        cell_write(target,row_index+row_offset[target.title],2,num_days)
                    elif c_idx==6:
                        balance_days=value
                        cell_write(target,row_index+row_offset[target.title],5,balance_days)
                    elif c_idx==7:
                        Date_Left=value
                        cell_write(target,row_index+row_offset[target.title],9,Date_Left)
                    elif c_idx==8:
                        Date_of_payment=value
                        cell_write(target,row_index+row_offset[target.title],10,Date_of_payment)
                    elif is_abs_num==0 and value==absent_label:
                        is_abs_num=1
                        start=columns[c_idx-1]
                        end=columns[c_idx-1]
                    elif value==absent_label:
                        is_abs_num+=1
                        end=columns[c_idx-1]
                    elif is_abs_num:
                        #target.cell(row=row_index+13, column=1+column_offset, value=is_abs_num)
                        start=start.split("\n")[1].replace("/","-")+"-"+str(year)
                        end=end.split("\n")[1].replace("/","-")+"-"+str(year)
                        cell_write(target,row_index+row_offset[target.title],3+offset,start)
                        cell_write(target,row_index+row_offset[target.title],4+offset,end)
                        cell_write(target,row_index+row_offset[target.title],10,Date_of_payment)
                        cell_write(target,row_index+row_offset[target.title],9,Date_Left)
                        cell_write(target,row_index+row_offset[target.title],5,balance_days)
                        cell_write(target,row_index+row_offset[target.title],2,num_days)
                        cell_write(target,row_index+row_offset[target.title],1,Leave_Accrued)

                        target['F'+str(row_index+row_offset[target.title])]="----"
                        target['G'+str(row_index+row_offset[target.title])]="----"
                        target['H'+str(row_index+row_offset[target.title])]="----"
                        target.insert_rows(row_index+row_offset[target.title]+1)
                        is_abs_num=0
                        row_index+=1
                        added[target.title]+=1
            return added            
                    
        absent_label="PL"
        column_offset=0
        initial_offset=13
        row_offset={}
        from collections import Counter
        row_offset=Counter(start_end_date_attendance(absent_label,column_offset,row_offset,initial_offset))
        absent_label="SL"
        
        row_offset+=Counter(start_end_date_attendance(absent_label,column_offset,row_offset,initial_offset))
        absent_label="CL"
        
        row_offset+=Counter(start_end_date_attendance(absent_label,column_offset,row_offset,initial_offset))
        
        formMfile.remove(formMfile["Sheet1"])
        formMfile.remove(formMfile["Sheet2"])
        formMfile.remove(formMfile["Sheet3"])
        formMfinalfile = os.path.join(filelocation,'Form M Register of leave.xlsx')
        formMfile.save(filename=formMfinalfile)

    def Form_P():
        formPfilepath = os.path.join(Gujaratfilespath,'Form P Muster roll.xlsx')
        formPfile = load_workbook(filename=formPfilepath)
        logging.info('Form P file has sheet: '+str(formPfile.sheetnames))
        logging.info('create columns which are now available')

        data_formP = data.copy(deep=True)
        data_formP=data_formP.drop_duplicates(subset="Employee Name", keep="last")
        
        columns=['S.no',"Employee Name","Designation","Age","Gender","Date Joined","start_time",
                                                                "end_time",'interval_for_reset_from','interval_for_reset_to']
        data_formP['interval_for_reset_to']=data_formP.rest_interval.str.split("-",expand=True)[1]
        data_formP['interval_for_reset_from']=data_formP.rest_interval.str.split("-",expand=True)[0]
        data_formP_columns=list(data_formP.columns)
        start=data_formP_columns.index('Emp Code')
        end=data_formP_columns.index('Total\r\nDP')
        columns.extend(data_formP_columns[start+1:end])
        
        less=31-len(data_formP_columns[start+1:end])
            
        for i in range(less):
            columns.extend(["less"+str(i+1)])
            data_formP["less"+str(i+1)]=""
        columns.append('Total\r\nDP')
        data_formP['S.no'] = list(range(1,len(data_formP)+1))

        formP_data=data_formP[columns]
        formPsheet = formPfile['Sheet1']
        formPsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form P is ready')
        
        for i in range(9,20):
            formPsheet["A"+str(i)]=""

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formP_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 11):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formPsheet.cell(row=r_idx, column=c_idx, value=value)
                formPsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formPsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formPsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formP):
                    formPsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formPsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formPsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formPsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formP):
                    formPsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formPsheet.row_dimensions[r_idx].height = 20
                else:
                    formPsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formPsheet.row_dimensions[r_idx].height = 20

        
        #formPsheet['AE4']=formPsheet['AE4'].value+"   "+str(data_formP['Registration_no'].unique()[0])
        
        formPsheet['A4']="Name of establishment :-  "+str(data_formP['Company Name'].unique()[0])+","+str(data_formP['Company Address'].unique()[0])
        if data["PE_or_contract"].unique()[0].upper()=="CL":
            formPsheet['A5']="Name of the employer:-    "+str(data_formP['UnitName'].unique()[0])


        formPsheet['A6']=formPsheet['A6'].value+" "+str(month)+" "+str(year)
        formPsheet['N4']="Place   "+data_formP['Branch'].unique()[0]
        formPfinalfile = os.path.join(filelocation,'Form P Muster roll.xlsx')
        formPfile.save(filename=formPfinalfile)

    

    def Form_Notice_holiday():
        formNotice_holidayfilepath = os.path.join(Gujaratfilespath,'Notice of holiday.xlsx')
        formNotice_holidayfile = load_workbook(filename=formNotice_holidayfilepath)
        logging.info('Form Notice_holiday file has sheet: '+str(formNotice_holidayfile.sheetnames))
        logging.info('create columns which are now available')

        data_formNotice_holiday = data.copy(deep=True)
        data_formNotice_holiday=data_formNotice_holiday.drop_duplicates(subset="Employee Name", keep="last")
        columns=["Employee Name","day_holiday_allowed"]

        data_formNotice_holiday["day_holiday_allowed"]="Sunday , Saturday"
        data_formNotice_holiday['S.no'] = list(range(1,len(data_formNotice_holiday)+1))

        formNotice_holiday_data=data_formNotice_holiday[columns]
        formNotice_holidaysheet = formNotice_holidayfile['Sheet1']
        formNotice_holidaysheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form Notice_holiday is ready')
        
        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formNotice_holiday_data, index=False, header=False)

        logging.info('rows taken out from data')
        if len(data_formNotice_holiday)>7:
            formNotice_holidaysheet.insert_rows(15,len(data_formNotice_holiday)-8)

        for r_idx, row in enumerate(rows, 14):
            for c_idx, value in enumerate(row, 1):
                formNotice_holidaysheet.cell(row=r_idx, column=c_idx, value=value)
                formNotice_holidaysheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formNotice_holidaysheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formNotice_holidaysheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                
        
        formNotice_holidaysheet['A3']=formNotice_holidaysheet['A3'].value+" "+str(data_formNotice_holiday['Unit'].unique()[0])+", "+str(data_formNotice_holiday['Address'].unique()[0])
        formNotice_holidayfinalfile = os.path.join(filelocation,'Notice of holiday.xlsx')
        formNotice_holidayfile.save(filename=formNotice_holidayfinalfile)


    Form_F()
    Form_IV()
    Form_M()
    Form_P()
    Form_Notice_holiday()
    #No need
    # Form_I()

    

def Haryana(data,contractor_name,contractor_address,filelocation,month,year):
    logging.info("Haryana form creation")



def Goa(data,contractor_name,contractor_address,filelocation,month,year):
    Goafilespath = os.path.join(Statefolder,'Goa')
    logging.info('Goa files path is :'+str(Goafilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]
    #wages reg
    input_filelocation=filelocation.split("Registers")[0]
    min_wages_goa=read_min_wages_file("GOA","SEMI-SKILLED",input_filelocation)
    
    def Form_I():

        formIfilepath = os.path.join(Goafilespath,'Form I register of Fine.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formI = data.copy(deep=True)
        data_formI=data_formI.drop_duplicates(subset="Employee Name", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","name&date_of_offence","cause_against_fine","FIXED MONTHLY GROSS",
                                        "Date of payment_fine_released","Date of payment_fine_imposed","remarks"]

        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        data_formI[["name&date_of_offence","cause_against_fine","remarks"]]="---"
        
        data_formI['Fine']=data_formI['Fine'].astype(float)
        data_formI["Date of payment_fine_released"]=data_formI['Date of payment']
        data_formI["Date of payment_fine_imposed"]=data_formI['Date of payment']
        data_formI.loc[data_formI['Fine']==0,["FIXED MONTHLY GROSS","Date of payment_fine_released","Date of payment_fine_imposed","remarks"]]="---"

        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 10):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIsheet.row_dimensions[r_idx].height = 20
                else:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIsheet.row_dimensions[r_idx].height = 20

        formIsheet['A4']=formIsheet['A4'].value+" : "+data_formI['Company Name'].unique()[0]
        formIfinalfile = os.path.join(filelocation,'Form I register of Fine.xlsx')
        formIfile.save(filename=formIfinalfile)

    def Form_II():
        formIIfilepath = os.path.join(Goafilespath,'Form II register of damage or loss.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy(deep=True)
        data_formII=data_formII.drop_duplicates(subset="Employee Name", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","attendancefile",
                                        "Damage_loss_with_date","whether_work_showed_cause",
                                        "Date of payment & amount of deduction","num_instalments","Date on which total amount realised","remarks"]

        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII["attendancefile"]="---"
        data_formII[["whether_work_showed_cause","num_instalments"]]="-----"
        data_formII["remarks"]=""
        ######################################
        data_formII['Date of payment']=data_formII['Date of payment'].apply(lambda x: x.strftime('%d-%m-%Y'))
        data_formII['Damage or Loss']=data_formII['Damage or Loss'].astype(float)
        data_formII['Damage or Loss']=data_formII['Damage or Loss'].fillna(0)
        data_formII["Damage_loss_with_date"]=data_formII['Date of payment']+" & "+data_formII['Damage or Loss'].astype(str)
        data_formII.loc[data_formII["Damage or Loss"]==0,"Damage_loss_with_date"]="---"

        data_formII["Date of payment & amount of deduction"]=data_formII["Damage_loss_with_date"]
        data_formII["num_instalments"]="1"
        data_formII["Date on which total amount realised"]=data_formII['Date of payment']
        data_formII.loc[data_formII["Damage or Loss"]==0,["Date on which total amount realised","num_instalments"]]="---"
        ###################################
        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formII_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 10):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formII):
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formII):
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIIsheet.row_dimensions[r_idx].height = 20
                else:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIIsheet.row_dimensions[r_idx].height = 20

        formIIsheet['A4']=formIIsheet['A4'].value+"  :  "+data_formII['Company Name'].unique()[0]
        formIIfinalfile = os.path.join(filelocation,'Form II register of damage or loss.xlsx')
        formIIfile.save(filename=formIIfinalfile)

    def Form_VIII():
        formVIIIfilepath = os.path.join(Goafilespath,'Form VIII register of Over time.xlsx')
        formVIIIfile = load_workbook(filename=formVIIIfilepath)
        logging.info('Form VIII file has sheet: '+str(formVIIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formVIII = data.copy(deep=True)
        data_formVIII=data_formVIII.drop_duplicates(subset="Employee Name", keep="last")

        data_formVIII['Designation_Dept']=data_formVIII["Designation"]+"_"+data_formVIII["Department"]
        columns=['S.no',"Employee Name","Father's Name","Gender","Designation_Dept","attendancefile",
                                        "extent_of_overtime","total_overtime",
                                        'Normal hrs ','FIXED MONTHLY GROSS',
                                        "overtime rate",'total_earning-overtime',"Overtime",'Total Earning',"date_overtime_paid"]

        data_formVIII['S.no'] = list(range(1,len(data_formVIII)+1))
        data_formVIII[["attendancefile","overtime_rate","ot"]]="---"
        data_formVIII[["extent_of_overtime"]]="---"
        ###
        data_formVIII[['Total\r\nOT Hrs',"Overtime"]]=data_formVIII[['Total\r\nOT Hrs',"Overtime"]].astype(float)
        data_formVIII[['Total\r\nOT Hrs',"Overtime"]]=data_formVIII[['Total\r\nOT Hrs',"Overtime"]].fillna(0)
        
        data_formVIII["total_overtime"]=data_formVIII['Total\r\nOT Hrs']
        
        data_formVIII.loc[data_formVIII['Total\r\nOT Hrs']==0,"total_overtime"]="---"

        data_formVIII['total_earning-overtime']=data_formVIII['Total Earning']-data_formVIII["Overtime"]

        data_formVIII.loc[data_formVIII['Total\r\nOT Hrs']==0,["overtime rate",'total_earning-overtime',"Overtime",'Total Earning']]="---"

        #might need to change
        data_formVIII["date_overtime_paid"]=data_formVIII['Date of payment']
        
        data_formVIII.loc[data_formVIII["Overtime"]==0,"date_overtime_paid"]="---"
        
        formVIII_data=data_formVIII[columns]
        formVIIIsheet = formVIIIfile['Sheet1']
        formVIIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form VIII is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formVIII_data, index=False, header=False)

        logging.info('rows taken out from data')
        
        row_copy=dataframe_to_rows(formVIII_data, index=False, header=False)
        for i in range(len(list(row_copy))-2):
            i+=12
            formVIIIsheet.merge_cells('C'+str(i)+':D'+str(i))
            formVIIIsheet.merge_cells('F'+str(i)+':H'+str(i))
            formVIIIsheet.merge_cells('I'+str(i)+':K'+str(i))
            formVIIIsheet.merge_cells('L'+str(i)+':N'+str(i))
            formVIIIsheet.merge_cells('O'+str(i)+':R'+str(i))
            formVIIIsheet.merge_cells('S'+str(i)+':T'+str(i))
            formVIIIsheet.merge_cells('U'+str(i)+':V'+str(i))
            formVIIIsheet.merge_cells('W'+str(i)+':X'+str(i))
            formVIIIsheet.merge_cells('Y'+str(i)+':Z'+str(i))
            formVIIIsheet.merge_cells('AA'+str(i)+':AB'+str(i))
            formVIIIsheet.merge_cells('AC'+str(i)+':AD'+str(i))
            formVIIIsheet.merge_cells('AE'+str(i)+':AG'+str(i))
        
        c_idx=0
        row_num=0
        for r_idx, row in enumerate(rows, 10):
            row_iterator=zip(row)
            row_num+=1
            while True:
                c_idx+=1
                if type(formVIIIsheet.cell(row=r_idx, column=c_idx)).__name__ == 'MergedCell':
                    border_sides_thick = Side(style='thick')       
                    border_sides_thin = Side(style='thin')
                    if len(row)==c_idx and row_num==len(data_formVIII):
                        formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                        formVIIIsheet.row_dimensions[r_idx].height = 20
                    elif len(row)==c_idx:
                        formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                        formVIIIsheet.row_dimensions[r_idx].height = 20
                    elif row_num==len(data_formVIII):
                        formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                        formVIIIsheet.row_dimensions[r_idx].height = 20
                    else:
                        formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                        formVIIIsheet.row_dimensions[r_idx].height = 20
                    continue
                try:
                    value=next(row_iterator)[0]
                    
                except:
                    c_idx=0
                    break
                formVIIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formVIIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formVIIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                formVIIIfinalfile = os.path.join(filelocation,'Form VIII register of Over time.xlsx')
                formVIIIfile.save(filename=formVIIIfinalfile)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formVIII):
                    formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formVIIIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formVIIIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formVIII):
                    formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formVIIIsheet.row_dimensions[r_idx].height = 20
                else:
                    formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formVIIIsheet.row_dimensions[r_idx].height = 20
        
        formVIIIsheet['Q4']="Month ending "+month+" "+str(year)
        formVIIIfinalfile = os.path.join(filelocation,'Form VIII register of Over time.xlsx')
        formVIIIfile.save(filename=formVIIIfinalfile)
        
    
    def From_XII():
        formXIIfilepath = os.path.join(Goafilespath,'Form XII Register of leave.xlsx')
        formXIIfile = load_workbook(filename=formXIIfilepath)
        logging.info('Form XII file has sheet: '+str(formXIIfile.sheetnames))
        #print(formXIIfile.sheetnames)
        logging.info('create columns which are now available')

        data_formXII = data.copy(deep=True)
        leave_file_data=data_formXII[["Employee Code","Employee Name","Leave Type","Opening","Monthly Increment","Leave Accrued","Used","Encash","Closing"]]
        
        data_formXII=data_formXII.drop_duplicates(subset="Employee Name", keep="last")

        data_formXII["Employee Name & Code"]=data_formXII["Employee Name"].astype(str)+"||"+data_formXII["Employee Code"].astype(str)

        columns=["Employee Name & Code","Date Joined","Father's Name","Registration_no"]
        data_formXII_columns=list(data_formXII.columns)
        start=data_formXII_columns.index('Arrears salary')
        end=data_formXII_columns.index('Total\r\nDP')
        columns.extend(data_formXII_columns[start+1:end])


        formXII_data=data_formXII[columns]
        formXIIsheet = formXIIfile['Sheet1']

        formXIIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        #for column in  range(ord('A'), ord('G') + 1):
        #    formXIIsheet.unmerge_cells(chr(column)+"11:"+chr(column)+"15")
        formXIIsheet.unmerge_cells("A18:A19")
        formXIIsheet.unmerge_cells("B17:C17")
        formXIIsheet.unmerge_cells("D17:E17")
        formXIIsheet.unmerge_cells("B18:C18")
        formXIIsheet.unmerge_cells("D18:E18")
        formXIIsheet.unmerge_cells("F18:F19")
        formXIIsheet.unmerge_cells("G17:H17")
        formXIIsheet.unmerge_cells("G18:H18")
        formXIIsheet.unmerge_cells("I17:J17")
        formXIIsheet.unmerge_cells("I18:J18")
        
        formXIIsheet.unmerge_cells("A24:A25")
        formXIIsheet.unmerge_cells("B23:C23")
        formXIIsheet.unmerge_cells("D23:E23")
        formXIIsheet.unmerge_cells("B24:C24")
        formXIIsheet.unmerge_cells("D24:E24")
        formXIIsheet.unmerge_cells("F24:F25")
        formXIIsheet.unmerge_cells("G23:H23")
        formXIIsheet.unmerge_cells("G24:H24")
        formXIIsheet.unmerge_cells("I23:J23")
        formXIIsheet.unmerge_cells("I24:J24")

        formXIIsheet.unmerge_cells("A30:A31")
        formXIIsheet.unmerge_cells("B29:C29")
        formXIIsheet.unmerge_cells("B30:C30")
        formXIIsheet.unmerge_cells("D29:E29")
        formXIIsheet.unmerge_cells("D30:E30")
        formXIIsheet.unmerge_cells("F29:G29")
        formXIIsheet.unmerge_cells("F30:G30")

        formXIIsheet.unmerge_cells("E16:F16")
        formXIIsheet.unmerge_cells("E22:F22")
        formXIIsheet.unmerge_cells("C28:D28")
        

        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        #rows_copy = list(dataframe_to_rows(formXII_data, index=False, header=False))
        def cell_write(sheet,r_idx,c_idx,value):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            
        def PL_write(row_index,target,start,end,is_abs_num,name):
            cell_write(target,row_index,1,start)
            cell_write(target,row_index,2,start)
            cell_write(target,row_index,3,end)
            cell_write(target,row_index,4,is_abs_num)
            cell_write(target,row_index,5,start)
            cell_write(target,row_index,6,end)
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==name,:]
            closing=emp_details["Closing"].loc[emp_details["Leave Type"]=="PL"]
            cell_write(target,row_index,7,closing.to_string(index=False))
            cell_write(target,row_index,8,"---")
            cell_write(target,row_index,9,"---")
            cell_write(target,row_index,10,"")
            cell_write(target,row_index,11,"")


        def SL_write(row_index,target,start,end,is_abs_num,name):
            cell_write(target,row_index,1,start)
            cell_write(target,row_index,2,start)
            cell_write(target,row_index,3,end)
            cell_write(target,row_index,4,start)
            cell_write(target,row_index,5,end)
            #balamce due
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==name,:]
            closing=emp_details["Closing"].loc[emp_details["Leave Type"]=="SL"]
            if not closing.empty:
                cell_write(target,row_index,6,closing.to_string(index=False))
            else:
                cell_write(target,row_index,6,"")
            #balance
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==name,:]
            closing=emp_details["Closing"].loc[emp_details["Leave Type"]=="SL"]
            if not closing.empty:
                cell_write(target,row_index,7,closing.to_string(index=False))
            else:
                cell_write(target,row_index,7,"")
            cell_write(target,row_index,8,"")
            cell_write(target,row_index,9,"")
            cell_write(target,row_index,10,"")
            

        def CL_write(row_index,target,start,end,is_abs_num,name):
            cell_write(target,row_index,1,start)
            cell_write(target,row_index,2,start)
            cell_write(target,row_index,3,end)
            cell_write(target,row_index,4,start)
            cell_write(target,row_index,5,end)
            #balamce due
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==name,:]
            closing=emp_details["Closing"].loc[emp_details["Leave Type"]=="CL"]
            if not closing.empty:
                cell_write(target,row_index,6,closing.to_string(index=False))
            else:
                cell_write(target,row_index,6,"")
            #balance
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==name,:]
            closing=emp_details["Closing"].loc[emp_details["Leave Type"]=="CL"]
            if not closing.empty:
                cell_write(target,row_index,7,closing.to_string(index=False))
            else:
                cell_write(target,row_index,7,"")
            cell_write(target,row_index,8,"")
            cell_write(target,row_index,9,"")
            cell_write(target,row_index,10,"")
        
        def ML_write(row_index,target,start,end,is_abs_num,name):
            cell_write(target,row_index,1,start)
            cell_write(target,row_index,2,start)
            cell_write(target,row_index,3,end)
            cell_write(target,row_index,4,start)
            cell_write(target,row_index,5,end)
            cell_write(target,row_index,6,"")
            cell_write(target,row_index,7,"")

        form_write={'PL':PL_write,'SL':SL_write,'CL':CL_write,'ML':ML_write}
        
        def start_end_date_attendance(rows,absent_label,row_offset,initial_offset):  
           # print("infunction",row_offset)
            is_abs_num=0
            row_index=0
            added={}
            for sheet_idx, row in enumerate(rows, 10):
                row_index=0
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        if code =="nan":
                            code=name
                        try:
                            target=formXIIfile[code]
                            added[target.title]=0
                        except:
                            target = formXIIfile.copy_worksheet(formXIIsheet)
                            target.title=code
                            #initial offset
                            row_offset[code]=initial_offset
                            added[target.title]=0
                        target['A4']="Name and address of the Establishment:- "+" "+str(data_formXII['Company Name'].unique()[0])+", "+str(data_formXII['Company Address'].unique()[0])
                        if data_formXII['PE_or_contract'].unique()[0]=="Contractor":
                            target["A5"]="Name of Employer and address:-  "+str(data_formXII['UnitName'].unique()[0])
                        else:
                            target["A5"]="Name of Employer and address:- ---"
                        target['A7']="Name of Employee : "+str(name)
                        added[target.title]=0
                    elif c_idx==2:
                        target['A9']="Date of appointment:- "+str(value)
                    elif c_idx==3:
                        target['A8']="Father's Name : "+str(value)
                    elif c_idx==4:
                        target['A6']="Registration No. :- "+str(value)

                    elif is_abs_num==0 and value==absent_label:
                        is_abs_num=1
                        start=columns[c_idx-1]
                        end=columns[c_idx-1]
                    elif value==absent_label:
                        is_abs_num+=1
                        end=columns[c_idx-1]
                    elif is_abs_num:
                        #target.cell(row=row_index+13, column=1+column_offset, value=is_abs_num)
                     #   print("write",row_index,row_offset,row_index+row_offset[target.title])
                        start_date=start.split("\n")[1].replace("/","-")+"-"+str(year)
                        end_date=end.split("\n")[1].replace("/","-")+"-"+str(year)
                        form_write[absent_label](row_index+row_offset[target.title],target,start_date,end_date,is_abs_num,name)
                        target.insert_rows(row_index+row_offset[target.title]+1)
                        is_abs_num=0
                        row_index+=1
                        added[target.title]+=1

            return added
        offset={}
        initial_offset=14
            
        
        offset=Counter(offset)+Counter(start_end_date_attendance(dataframe_to_rows(formXII_data, index=False, header=False),"PL",offset,initial_offset))
        
        for sheet in formXIIfile.sheetnames:
            offset[sheet]+=20
            initial_offset+=20
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-2)+":A"+str(offset[sheet]-1))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-3)+":C"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-3)+":E"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-2)+":C"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-2)+":E"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("F"+str(offset[sheet]-2)+":F"+str(offset[sheet]-1))
            formXIIfile[sheet].merge_cells("G"+str(offset[sheet]-3)+":H"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("G"+str(offset[sheet]-2)+":H"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("I"+str(offset[sheet]-3)+":J"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("I"+str(offset[sheet]-2)+":J"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-4)+":J"+str(offset[sheet]-4))
            cell_write(sheet=formXIIfile[sheet],r_idx=offset[sheet]-4,c_idx=1,value="Sick Leave")
            
        offset+=Counter(start_end_date_attendance(dataframe_to_rows(formXII_data, index=False, header=False),"SL",offset,initial_offset))
        
        for sheet in formXIIfile.sheetnames:
            offset[sheet]+=6
            initial_offset+=6
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-2)+":A"+str(offset[sheet]-1))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-3)+":C"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-3)+":E"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-2)+":C"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-2)+":E"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("F"+str(offset[sheet]-2)+":F"+str(offset[sheet]-1))
            formXIIfile[sheet].merge_cells("G"+str(offset[sheet]-3)+":H"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("G"+str(offset[sheet]-2)+":H"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("I"+str(offset[sheet]-3)+":J"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("I"+str(offset[sheet]-2)+":J"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-4)+":J"+str(offset[sheet]-4))
            cell_write(sheet=formXIIfile[sheet],r_idx=offset[sheet]-4,c_idx=1,value="Casual Leave")
        
        
        offset+=Counter(start_end_date_attendance(dataframe_to_rows(formXII_data, index=False, header=False),"CL",offset,initial_offset))
        
        for sheet in formXIIfile.sheetnames:
            offset[sheet]+=6
            initial_offset+=6
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-2)+":A"+str(offset[sheet]-1))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-3)+":C"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-2)+":C"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-3)+":E"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-2)+":E"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("F"+str(offset[sheet]-3)+":G"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("F"+str(offset[sheet]-2)+":G"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-4)+":G"+str(offset[sheet]-4))
            cell_write(sheet=formXIIfile[sheet],r_idx=offset[sheet]-4,c_idx=1,value="Maternity Leave")
        offset+=Counter(start_end_date_attendance(dataframe_to_rows(formXII_data, index=False, header=False),"ML",offset,initial_offset))
        formXIIfile.remove(formXIIfile["Sheet1"])
        formXIIfile.remove(formXIIfile["Sheet2"])
        formXIIfile.remove(formXIIfile["Sheet3"])
        formXIIfinalfile = os.path.join(filelocation,'Form XII Register of leave.xlsx')
        formXIIfile.save(filename=formXIIfinalfile)
        
       
    def Form_XXI():
        formXXIfilepath = os.path.join(Goafilespath,'Form XXI Register of Employment.xlsx')
        formXXIfile = load_workbook(filename=formXXIfilepath)
        logging.info('Form XXI file has sheet: '+str(formXXIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formXXI = data.copy(deep=True)
        data_formXXI=data_formXXI.drop_duplicates(subset="Employee Name", keep="last")

        
        columns=['S.no',"Employee Name","Father's Name","Gender","Designation","Date_of_appoinment"]
        
        interval_for_reset_to=data_formXXI.rest_interval.str.split("-",expand=True)[1].unique()[0]
        interval_for_reset_from=data_formXXI.rest_interval.str.split("-",expand=True)[0].unique()[0]
        start_time=data_formXXI["start_time"].unique()[0]
        end_time=data_formXXI["end_time"].unique()[0]

        data_formXXI_columns=list(data_formXXI.columns)
        start=data_formXXI_columns.index('Emp Code')
        end=data_formXXI_columns.index('Total\r\nDP')
        columns.extend(data_formXXI_columns[start+1:end])
        
        less=31-len(data_formXXI_columns[start+1:end])
        for i in range(less):
            columns.extend(["less"+str(i+1)])
            data_formXXI["less"+str(i+1)]=""

        columns.extend(["normal_hours",'Overtime_hrs',"remarks"])
        data_formXXI["Date_of_appoinment"]=data_formXXI['Date Joined']
        data_formXXI["normal_hours"]=len(data_formXXI_columns[start+1:end])-data_formXXI['Total\r\nDP'].astype(float)
        data_formXXI['Overtime_hrs']=data_formXXI['Total\r\nOT Hrs']
        data_formXXI["remarks"]=""
        data_formXXI['S.no'] = list(range(1,len(data_formXXI)+1))

        formXXI_data=data_formXXI[columns]
        formXXIsheet = formXXIfile['Sheet1']
        formXXIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form XXI is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formXXI_data, index=False, header=False)

        logging.info('rows taken out from data')
        formXXIsheet.unmerge_cells('A23:E23')
        row_num=0
        for r_idx, row in enumerate(rows, 14):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formXXIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formXXI):
                    formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formXXIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formXXIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formXXI):
                    formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formXXIsheet.row_dimensions[r_idx].height = 20
                else:
                    formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formXXIsheet.row_dimensions[r_idx].height = 20
        
        formXXIsheet['AE4']=formXXIsheet['AE4'].value+"   "+str(data_formXXI['Registration_no'].unique()[0])
        formXXIsheet['AG5']=start_time
        formXXIsheet['AK5']=end_time
        formXXIsheet['AG6']="8 hrs"
        #formXXIsheet['AK6']=""
        formXXIsheet['AG7']=interval_for_reset_from
        formXXIsheet['AK7']=interval_for_reset_to
        formXXIsheet['A4']=formXXIsheet['A4'].value+" "+str(data_formXXI['Company Name'].unique()[0])+", "+str(data_formXXI['Company Address'].unique()[0])
        #formXXIsheet['A5']=formXXIsheet['A5'].value+" "+str(data_formXXI['Unit'].unique()[0])+", "+str(data_formXXI['Location'].unique()[0])
        if data_formXXI['PE_or_contract'].unique()[0]=="Contractor":
            formXXIsheet["A5"]="Name of Employer and address:-  "+str(data_formXXI['UnitName'].unique()[0])+", "+str(data_formXXI['Address'].unique()[0])
        else:
            formXXIsheet["A5"]="Name of Employer and address:-  "+"---"
        formXXIfinalfile = os.path.join(filelocation,'Form XXI register of Over time.xlsx')
        formXXIfile.save(filename=formXXIfinalfile)



    def Form_XXIII():
        formXXIIIfilepath = os.path.join(Goafilespath,'Form XXIII Register of wages.xlsx')
        formXXIIIfile = load_workbook(filename=formXXIIIfilepath)
        logging.info('Form XXIII file has sheet: '+str(formXXIIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formXXIII = data.copy()
        
        columns=['S.no',"Employee Name","Father's Name","Designation",'Basic','DA',
                                'Earned Basic','Dearness_Allowance','all_Other_Allowance','Overtime',
                                 'Total Earning','Salary Advance','PF', 'Other_auth_Deduction',
                                 'Total Deductions','Net Paid',"sign",'Date of payment']
        
        data_formXXIII[["sign","remarks"]]=""
        data_formXXIII['Dearness_Allowance']=data_formXXIII['DA']
        
        data_formXXIII["Basic"]=min_wages_goa
        all_other_allowance_columns=['Other Allowance','OtherAllowance1','OtherAllowance2', 'OtherAllowance3', 'OtherAllowance4', 'OtherAllowance5']
        
        data_formXXIII[all_other_allowance_columns]=data_formXXIII[all_other_allowance_columns].astype(float)
        data_formXXIII['all_Other_Allowance']= data_formXXIII.loc[:,all_other_allowance_columns].sum(axis=1)

        Other_auth_Deduction_columns=['Insurance','CSR','ESIC','P.Tax','LWF EE','Loan Deduction','Loan Interest','Other Deduction','TDS',
                                            'OtherDeduction1', 'OtherDeduction2',
                                                    'OtherDeduction3', 'OtherDeduction4', 'OtherDeduction5']
        data_formXXIII[Other_auth_Deduction_columns]=data_formXXIII[Other_auth_Deduction_columns].astype(float)
        data_formXXIII['Other_auth_Deduction']= data_formXXIII.loc[:,Other_auth_Deduction_columns].sum(axis=1)


        data_formXXIII['S.no'] = list(range(1,len(data_formXXIII)+1))

        formXXIII_data=data_formXXIII[columns]
        formXXIIIsheet = formXXIIIfile['Sheet1']
        formXXIIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form XXIII is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formXXIII_data, index=False, header=False)
        rows_copy = list(dataframe_to_rows(formXXIII_data, index=False, header=False))
        logging.info('rows taken out from data')
        formXXIIIsheet.unmerge_cells('P15:R15')
        formXXIIIsheet["P15"]=""
        row_num=0
        for r_idx, row in enumerate(rows, 10):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                #formXXIIIsheet.cell(row=r_idx, column=c_idx).value=value
                formXXIIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formXXIII):
                    formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formXXIIIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formXXIIIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formXXIII):
                    formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formXXIIIsheet.row_dimensions[r_idx].height = 20
                else:
                    formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formXXIIIsheet.row_dimensions[r_idx].height = 20
        
        
        formXXIIIsheet['P'+str(len(list(rows_copy))+10+5)].value="Signature of Employer"
        
        formXXIIIsheet.merge_cells('P'+str(len(list(rows_copy))+10+5)+':R'+str(len(list(rows_copy))+10+5))
        
        formXXIIIsheet['P4']=formXXIIIsheet['P4'].value+"   "+str(data_formXXIII['Registration_no'].unique()[0])
        formXXIIIsheet['P5']=formXXIIIsheet['P5'].value+"   "+month

        if data["PE_or_contract"].unique()[0].upper()=="PE":
            formXXIIIsheet['A4']=" Name of Establishment:-   "+str(data_formXXIII['Company Name'][0])+" "+str(data_formXXIII['Company Address'][0])
        else:
            formXXIIIsheet['A4']=" Name of Establishment:-   "+str(data_formXXIII['UnitName'][0])+" "+str(data_formXXIII['Address'][0])
            formXXIIIsheet['A5']="Name of Employer and address:-   "+str(data_formXXIII['Contractor_name'][0])+","+str(data_formXXIII['Contractor_Address'][0])
        
        
        formXXIIIfinalfile = os.path.join(filelocation,'Form XXIII Register of wages.xlsx')
        formXXIIIfile.save(filename=formXXIIIfinalfile)
        
    Form_I()
    Form_II()
    Form_VIII()
    From_XII()
    Form_XXI()
    Form_XXIII()




def Maharashtra(data,contractor_name,contractor_address,filelocation,month,year):
    logging.info('Maharashtra forms')

    Maharashtrafilespath = os.path.join(Statefolder,'Maharashtra')
    logging.info('Maharashtra files path is :'+str(Maharashtrafilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]
    #Min wages xl
    input_filelocation=filelocation.split("Registers")[0]
    min_wages_maharashtra=read_min_wages_file("MAHARASHTRA","SEMI-SKILLED",input_filelocation)
    
    def Read_Holiday_file():

        inputfolder = filelocation.split("Registers")[0]
        file_list = os.listdir(inputfolder)
        logging.info('input folder is '+str(inputfolder))
        for f in file_list:
            if f[0:12].upper()=='HOLIDAY LIST':
                holidayfilename = f
                logging.info('holidayfilename is :'+f)
        if 'holidayfilename' in locals():
            holidayfile = os.path.join(inputfolder,holidayfilename)
            holiday=pd.read_excel(holidayfile).dropna()
            holiday.columns=[ "SN.", "Date"," Day"," Occasion"]
            convert=lambda variable: datetime.datetime.strptime(variable,'%d%m%Y')
            from dateutil import parser
            holiday=holiday[1:]
            holiday["Date"]=holiday["Date"].apply(str)
            holiday["Date"]=holiday["Date"].apply(parser.parse)
            holiday=holiday.set_index(pd.PeriodIndex(data=holiday.Date,freq='D'))
        else:
            holiday_columns = [ "SN.", "Date"," Day"," Occasion"]
            holiday = pd.DataFrame(columns = holiday_columns)
            holiday=holiday.set_index(pd.PeriodIndex(data=holiday.Date,freq='D'))
        return holiday.sort_index()
    # print(sorted(list(data.columns)))
    #print("------------")
    #print(data["Opening"])
    #print("-------------------")
    #print(data["Employee Name"])
    
        
    def Form_I():
        formIfilepath = os.path.join(Maharashtrafilespath,'Form I register of fine.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')
        data_formI = data.copy(deep=True)
        data_formI=data_formI.drop_duplicates(subset="Employee Name", keep="last")

        data_formI.fillna(value=0, inplace=True)
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","name&date_of_offence","cause_against_fine",
                                        "FIXED MONTHLY GROSS","Date of payment","Date of Fine","remarks"]

        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        data_formI[["name&date_of_offence","cause_against_fine","FIXED MONTHLY GROSS","Date of payment","Date of Fine","remarks"]]="NIL"
        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')
        border_sides_thick = Side(style='thick')       
        border_sides_thin = Side(style='thin')            
        for r_idx, row in enumerate(rows, 8):
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                if len(row)==c_idx and int(row[0])==len(data_formI):
                   formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                elif len(row)==c_idx:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                elif int(row[0])==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                else:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
        formIsheet['A5']=formIsheet['A5'].value+" : "+str(data_formI['Company Name'].unique()[0])
        formIsheet['A6']=formIsheet['A6'].value+" : "+str(month)+" "+str(year)
        formIfinalfile = os.path.join(filelocation,'Form I register of fine.xlsx')
        formIfile.save(filename=formIfinalfile)
    
    def Form_II_Muster_Roll():
        formIIfilepath = os.path.join(Maharashtrafilespath,'Form II muster roll.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy(deep=True)
        data_formII=data_formII.drop_duplicates(subset="Employee Name", keep="last")

        data_formII.fillna(value=0, inplace=True)
        columns=['S.no',"Employee Code","Employee Name","start_time","end_time",
                                        "interval_for_reset_from","interval_for_reset_to"]
        
        data_formII_columns=list(data_formII.columns)
        start=data_formII_columns.index('Emp Code')
        end=data_formII_columns.index('Total\r\nDP')
        columns.extend(data_formII_columns[start+1:end])
        less=31-len(data_formII_columns[start+1:end])
        for i in range(less):
            columns.extend(["less"+str(i+1)])
            data_formII["less"+str(i+1)]=""

        columns.extend(["Total\r\nDP"])
        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII['interval_for_reset_to']="2:00 PM"
        data_formII['interval_for_reset_from']="1:00 PM"
        data_formII["start_time"]="9:30 AM"
        data_formII["end_time"]="6:30 PM"
        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formII_data, index=False, header=False)

        logging.info('rows taken out from data')
        border_sides_thick = Side(style='thick')       
        border_sides_thin = Side(style='thin')
        for r_idx, row in enumerate(rows, 8):
            for c_idx, value in enumerate(row, 1):
                formIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                if len(row)==c_idx and int(row[0])==len(data_formII):
                       formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                elif len(row)==c_idx:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                elif int(row[0])==len(data_formII):
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                else:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                #border_sides = Side(style='thin')
                #formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        formIIsheet['A2']=formIIsheet['A2'].value+"   "+month
        if not data["PE_or_contract"].unique()[0].upper()=="PE":
            formIIsheet['A3']=formIIsheet['A3'].value+"   "+str(data_formII['Contractor_name'].unique()[0])+","+str(data_formII['Contractor_Address'].unique()[0])
            formIIsheet['A4']=formIIsheet['A4'].value+" "+str(data_formII['Unit'].unique()[0])+","+str(data_formII['Address'].unique()[0])
        formIIfinalfile = os.path.join(filelocation,'Form II muster roll.xlsx')
        formIIfile.save(filename=formIIfinalfile)

    def Form_II_reg_damage_loss():
        formIIfilepath = os.path.join(Maharashtrafilespath,'Form II register of damage or losses.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy(deep=True)
        data_formII=data_formII.drop_duplicates(subset="Employee Name", keep="last")

        data_formII.fillna(value=0, inplace=True)
        #print(sorted(data_formII.columns))
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","Damage or Loss","whether_work_showed_cause",
                                        "Date of payment & amount of deduction","num_instalments","Date of payment","remarks"]
        
        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII[["Damage or Loss","whether_work_showed_cause","Date of payment & amount of deduction","num_instalments","Date of payment","remarks"]]="NIL"
        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formII_data, index=False, header=False)
        border_sides_thick = Side(style='thick')       
        border_sides_thin = Side(style='thin')
        
        logging.info('rows taken out from data')
        for r_idx, row in enumerate(rows, 9):
            for c_idx, value in enumerate(row, 1):
                formIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                if len(row)==c_idx and int(row[0])==len(data_formII):
                       formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                elif len(row)==c_idx:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                elif int(row[0])==len(data_formII):
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                else:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                #border_sides = Side(style='thin')
                #formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        
        formIIsheet['A5']="Name and Address of the Establishment "+str(data_formII['Company Name'].unique()[0])+","+str(data_formII['Address'].unique()[0])
        formIIsheet['A6']="PERIOD "+str(month)+" "+str(year)
        formIIfinalfile = os.path.join(filelocation,'Form II register of damage or losses.xlsx')
        formIIfile.save(filename=formIIfinalfile)

    def Form_II_wages_reg():
        # print("----------------------------------------------")
        # print(filelocation)
        formIIfilepath = os.path.join(Maharashtrafilespath,'Form II wages register.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy(deep=True)
        leave_file_data=data_formII[["Employee Code","Employee Name","Leave Type","Opening","Monthly Increment","Leave Accrued","Used","Encash","Closing"]]
        data_formII=data_formII.drop_duplicates(subset="Employee Name", keep="last")
        data_formII.fillna(value=0, inplace=True)
        #print(sorted(data_formII.columns))
        columns=['S.no',"Employee Code","Employee Name",'Age',"Gender","Designation","Date Joined","Days Paid",
                                    "min_wages","FIXED MONTHLY GROSS","Total_Production_Piece_Rate",'Total\r\nOT Hrs',
                                    "FIXED MONTHLY GROSS","Earned Basic","HRA/Earned_basic","HRA","Tel and Int Reimb",
                                    "Bonus","Fuel Reimb","Corp Attire Reimb","CCA","Overtime","Total Earning",
                                    "PF","P.Tax","Insurance","sal_fine_damage","Total Deductions","Net Paid",
                                    "Prev_balance","Earned_during_month","Availed","colsing_bal","Date of payment",
                                    "Bank A/c Number",'Cheque No - NEFT date',"Net Paid","sign"]
        # print(leave_file_data)
        data_formII[["Prev_balance","Earned_during_month","Availed","colsing_bal"]]=""
        data_formII["Designation"] = data_formII["Designation"].astype(str)
        def date_format_change(val):
            return val.strftime('%d-%m-%y')

        data_formII["Date of payment"]=data_formII["Date of payment"].apply(date_format_change)
        for employee_name_leave_file in data_formII["Employee Name"]:
            #opening
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==employee_name_leave_file,:]
            opening_pl=emp_details["Opening"].loc[emp_details["Leave Type"]=="PL"].astype(float)
            opening_cl=emp_details["Opening"].loc[emp_details["Leave Type"]=="CL"].astype(float)
            opening_sl=emp_details["Opening"].loc[emp_details["Leave Type"]=="SL"].astype(float)
            prev_bal=opening_pl.add(opening_cl.add(opening_sl,fill_value=0), fill_value=0).sum()
            
            data_formII.loc[data_formII["Employee Name"]==employee_name_leave_file,"Prev_balance"]=prev_bal
            
            #monthly_inr
            mon_inr_pl=emp_details["Monthly Increment"].loc[emp_details["Leave Type"]=="PL"].astype(float)
            mon_inr_cl=emp_details["Monthly Increment"].loc[emp_details["Leave Type"]=="CL"].astype(float)
            mon_inr_sl=emp_details["Monthly Increment"].loc[emp_details["Leave Type"]=="SL"].astype(float)
            earned=mon_inr_cl.add(mon_inr_pl.add(mon_inr_sl,fill_value=0), fill_value=0).sum()
            data_formII.loc[data_formII["Employee Name"]==employee_name_leave_file,"Earned_during_month"]=earned
            #availed during month
            Used_pl=emp_details["Used"].loc[emp_details["Leave Type"]=="PL"].astype(float)
            Used_cl=emp_details["Used"].loc[emp_details["Leave Type"]=="CL"].astype(float)
            Used_sl=emp_details["Used"].loc[emp_details["Leave Type"]=="SL"].astype(float)
            availed=Used_cl.add(Used_pl.add(Used_sl,fill_value=0), fill_value=0).sum()
            
            data_formII.loc[data_formII["Employee Name"]==employee_name_leave_file,"Availed"]=availed
            #closing
            Closing_pl=emp_details["Closing"].loc[emp_details["Leave Type"]=="PL"].astype(float)
            Closing_cl=emp_details["Closing"].loc[emp_details["Leave Type"]=="CL"].astype(float)
            Closing_sl=emp_details["Closing"].loc[emp_details["Leave Type"]=="SL"].astype(float)
            closing=Closing_cl.add(Closing_pl.add(Closing_sl,fill_value=0), fill_value=0).sum()
            
            data_formII.loc[data_formII["Employee Name"]==employee_name_leave_file,"colsing_bal"]=closing
     
        def convert(input_str):
            if input_str=="nan":
                return ""
            else:
                return input_str.split(".")[0]+"."+input_str.split(".")[1][:2]
        data_formII["HRA/Earned_basic"]=((data_formII["HRA"].apply(float)/data_formII["Earned Basic"].apply(float))*100.0).apply(str).apply(convert)
        
        data_formII["Fine"]=data_formII["Fine"].fillna(0)
        data_formII["Damage or Loss"]=data_formII["Damage or Loss"].fillna(0)
        
        data_formII["sal_fine_damage"]=data_formII["Fine"].apply(float)+data_formII["Damage or Loss"].apply(float)
        remove_point=lambda input_str: input_str.split(".")[0]
        data_formII["Bank A/c Number"]=data_formII["Bank A/c Number"].apply(str).apply(remove_point)
        data_formII['S.no'] = list(range(1,len(data_formII)+1))

        data_formII[["Total_Production_Piece_Rate"]]="----"
        data_formII["min_wages"]=min_wages_maharashtra
        data_formII[["sign"]]=""
        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formII_data, index=False, header=False)
        border_sides_thick = Side(style='thick')       
        border_sides_thin = Side(style='thin')
        logging.info('rows taken out from data')
        for r_idx, row in enumerate(rows, 7):
            for c_idx, value in enumerate(row, 1):
                # if data_formII.loc[data_formII["Employee Name"]=="Nilesh Tanaji Patil","HRA"].apply(float):
                if str(value)=="nan":
                    value=""
                formIIsheet.cell(row=r_idx, column=c_idx, value=value) 
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                if len(row)==c_idx and int(row[0])==len(data_formII):
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                elif len(row)==c_idx:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                elif int(row[0])==len(data_formII):
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                else:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                #formIIsheet.column.format("", str)
                #border_sides = Side(style='thin')
                #formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        formIIsheet['A2']=formIIsheet['A2'].value+"   "+str(month)
        #formIIsheet['A3']="Name and address of Contractor :- "+str(data_formII['Contractor_name'].unique()[0])+","+str(data_formII['Contractor_Address'].unique()[0])
        formIIsheet['A4']="Name and   address of Principal Employer :- "+str(data_formII['Company Name'].unique()[0])#+","+str(data_formII['Address'].unique()[0])
        formIIfinalfile = os.path.join(filelocation,'Form II wages register.xlsx')
        formIIfile.save(filename=formIIfinalfile)

    def Form_VI_Overtime():
        formIVfilepath = os.path.join(Maharashtrafilespath,'Form IV Overtime register.xlsx')
        formIVfile = load_workbook(filename=formIVfilepath)
        logging.info('Form IV file has sheet: '+str(formIVfile.sheetnames))
        logging.info('create columns which are now available')

        data_formIV = data.copy(deep=True)
        data_formIV=data_formIV.drop_duplicates(subset="Employee Name", keep="last")


        if str(data_formIV['Designation'].dtype)[0:3] != 'obj':
            data_formIV["Designation"] = data_formIV["Designation"].astype(str)
        if str(data_formIV['Department'].dtype)[0:3] != 'obj':
            data_formIV["Department"] = data_formIV["Department"].astype(str)

        u = data_formIV.select_dtypes(exclude=['object'])
        data_formIV[u.columns] = u.fillna(value=0)
        columns=['S.no',"Employee Name","Father's Name","Gender","Designation_Dept","Date_overtime_worked",
                                        "Extent of over-time",'Total\r\nOT Hrs','Normal hrs ',
                                        "FIXED MONTHLY GROSS","overtime rate","Total Earning-Overtime","Overtime",'Total Earning',"Date of payment"]
        
        data_formIV['S.no'] = list(range(1,len(data_formIV)+1))
        data_formIV["Overtime"]=data_formIV["Overtime"].astype(str)
        
        data_formIV["Overtime"]=data_formIV["Overtime"].str.replace("","0")
        data_formIV["Overtime"]=data_formIV["Overtime"].astype(float)
        data_formIV.loc[data_formIV["Overtime"]==0,"Date of payment"]="---"
        data_formIV["Date of payment"]=data_formIV["Date of payment"].replace(0,"---")
        data_formIV['Designation_Dept']=data_formIV["Designation"]+"_"+data_formIV["Department"]
        data_formIV["Total Earning-Overtime"]=data_formIV['Total Earning'].astype(float)-data_formIV["Overtime"].astype(float)
        data_formIV[["Date_overtime_worked","Extent of over-time"]]="NIL"
        formIV_data=data_formIV[columns]
        formIVsheet = formIVfile['Sheet1']
        formIVsheet.sheet_properties.pageSetUpPr.fitToPage = True
        
        logging.info('data for form IV is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formIV_data, index=False, header=False)

        logging.info('rows taken out from data')
        border_sides_thick = Side(style='thick')       
        border_sides_thin = Side(style='thin')
        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 1):
                formIVsheet.cell(row=r_idx, column=c_idx, value=value)
                formIVsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIVsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                if len(row)==c_idx and int(row[0])==len(data_formIV):
                       formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                elif len(row)==c_idx:
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                elif int(row[0])==len(data_formIV):
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                else:
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                
        #formIVsheet['A4']=formIVsheet['A4'].value+" : "+data_formIV['Unit'][0]
        formIVsheet['A7']="Name of the Establishment : "+str(data_formIV['Contractor_name'].unique()[0])
        for i in range(1,16):
            formIVsheet.cell(row=7, column=i).border = Border(outline= True,bottom=border_sides_thick)
        formIVsheet['A5']=formIVsheet['A5'].value+" "+str(month)+" "+str(year)
        formIVfinalfile = os.path.join(filelocation,'Form IV Overtime register.xlsx')
        formIVfile.save(filename=formIVfinalfile)

    def Form_VI_reg_advance():
        formIVfilepath = os.path.join(Maharashtrafilespath,'Form IV register of advance.xlsx')
        formIVfile = load_workbook(filename=formIVfilepath)
        logging.info('Form IV file has sheet: '+str(formIVfile.sheetnames))
        logging.info('create columns which are now available')

        data_formIV = data.copy(deep=True)
        data_formIV=data_formIV.drop_duplicates(subset="Employee Name", keep="last")

        data_formIV.fillna(value=0, inplace=True)
        columns=['S.no',"Employee Name","Father's Name","Department","Salary Advance","purpose_advance",
                                        "num_installments_advance","Postponement_granted",
                                        "Date repaid","remarks"]
                                        
                                        
        data_formIV['S.no'] = list(range(1,len(data_formIV)+1))
        data_formIV["Salary Advance"]=data_formIV["Salary Advance"].astype(str)
        data_formIV=data_formIV.replace({"Salary Advance":{"":"NIL","0.":"NIL","0":"NIL","0.0":"NIL"}})
        
        data_formIV[["purpose_advance","num_installments_advance","Postponement_granted","Date repaid","remarks"]]="NIL"
        formIV_data=data_formIV[columns]
        formIVsheet = formIVfile['Sheet1']
        formIVsheet.sheet_properties.pageSetUpPr.fitToPage = True
        
        logging.info('data for form IV is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formIV_data, index=False, header=False)

        logging.info('rows taken out from data')
        border_sides_thick = Side(style='thick')       
        border_sides_thin = Side(style='thin')
        for r_idx, row in enumerate(rows, 13):
            for c_idx, value in enumerate(row, 1):
                formIVsheet.cell(row=r_idx, column=c_idx, value=value)
                formIVsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIVsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                if len(row)==c_idx and int(row[0])==len(data_formIV):
                       formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                       formIVsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIVsheet.row_dimensions[r_idx].height = 20
                elif int(row[0])==len(data_formIV):
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIVsheet.row_dimensions[r_idx].height = 20
                else:
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIVsheet.row_dimensions[r_idx].height = 20
                #border_sides = Side(style='thin')
                #formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        #formIVsheet['A4']=formIVsheet['A4'].value+" : "+data_formIV['Unit'][0]
        formIVsheet['A6']="Name of Factory or Industrial Establishment. : "+str(data_formIV['Company Name'].unique()[0])
        formIVsheet['A7']="PERIOD "+str(month)+" "+str(year)

        formIVfinalfile = os.path.join(filelocation,'Form IV register of advance.xlsx')
        formIVfile.save(filename=formIVfinalfile)



    def From_O():
        formOfilepath = os.path.join(Maharashtrafilespath,'Form O leave book.xlsx')
        formOfile = load_workbook(filename=formOfilepath)
        logging.info('Form O file has sheet: '+str(formOfile.sheetnames))
        #print(formOfile.sheetnames)
        logging.info('create columns which are now available')

        data_formO = data.copy(deep=True)
        leave_file_data=data_formO[["Employee Code","Employee Name","Leave Type","Opening","Monthly Increment","Leave Accrued","Used","Encash","Closing"]]
        data_formO=data_formO.drop_duplicates(subset="Employee Name", keep="last")

        data_formO.fillna(value=0, inplace=True)
        columns=["Employee Name & Code","Date Joined","Department","Registration_no"]
        data_formO["Employee Name & Code"]=data_formO["Employee Name"].astype(str)+"||"+data_formO["Employee Code"].astype(str)

        data_formO[["num_days","Earned_during_month","Availed","colsing_bal",'Cheque No - NEFT date']]=""
        for employee_name_leave_file in data_formO["Employee Name"]:
            #opening
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==employee_name_leave_file,:]
            opening_pl=emp_details["Opening"].loc[emp_details["Leave Type"]=="PL"]
            if opening_pl.empty:
                opening_pl="0"
            else:
                opening_pl=opening_pl.to_string(index=False)
            if opening_pl=="Nan" or opening_pl=="nan":
                opening_pl="0"
            
            data_formO.loc[data_formO["Employee Name"]==employee_name_leave_file,"num_days"]=opening_pl
            
           
     
        data_formO_columns=list(data_formO.columns)
        start_col=data_formO_columns.index('Emp Code')
        end=data_formO_columns.index('Total\r\nDP')
        num_days=len(data_formO_columns[start_col+1:end])
        start_month=data_formO_columns[start_col+1]
        end_month=data_formO_columns[end-1]
        
        columns.extend(data_formO_columns[start_col+1:end])


        formO_data=data_formO[columns]
        formOsheet = formOfile['Sheet1']

        formOsheet.sheet_properties.pageSetUpPr.fitToPage = True

        #for column in  range(ord('A'), ord('G') + 1):
        #    formOsheet.unmerge_cells(chr(column)+"11:"+chr(column)+"15")
        formOsheet.unmerge_cells("A22:H22")
        formOsheet.unmerge_cells("A23:B23")
        formOsheet.unmerge_cells("C23:C24")
        formOsheet.unmerge_cells("D23:D24")
        formOsheet.unmerge_cells("E23:E24")
        formOsheet.unmerge_cells("F23:G24")
        formOsheet.unmerge_cells("H23:H24")
        formOsheet.unmerge_cells("F25:G25")
        formOsheet.unmerge_cells("F26:G26")
        formOsheet.unmerge_cells("F27:G27")
        
        formOsheet.unmerge_cells("A28:F28")
        formOsheet.unmerge_cells("A29:B30")
        formOsheet.unmerge_cells("C29:C31")
        formOsheet.unmerge_cells("D29:D31")
        formOsheet.unmerge_cells("E29:E31")
        formOsheet.unmerge_cells("F29:F31")
        
        
        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        #rows_copy = list(dataframe_to_rows(formO_data, index=False, header=False))
        def cell_write(sheet,r_idx,c_idx,value):
                if not (str(value)=="nan" or str(value)=="NaN"):
                    sheet.cell(row=r_idx, column=c_idx, value=value)
                    sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                    sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                
        def PL_write(row_index,target,start,end,is_abs_num):

            cell_write(target,row_index,3,start+"--"+end)
            target.row_dimensions[row_index].height = 50
            cell_write(target,row_index , 1,data_formO_columns[start_col+1])
            cell_write(target,row_index , 4,"----")
            cell_write(target,row_index , 5,"----")
            # cell_write(target,row_index , 6,"----")
            # cell_write(target,row_index , 7,"----")
            # cell_write(target,row_index , 8,"----")
            # #print(data_formO.loc[data_formO[columns[0]]],emp_name)
            def get_emp_name(var):
                return var.split("||")[0]
            temp=str(data_formO.loc[data_formO[columns[0]].apply(get_emp_name)==emp_name,"Date Left"].tolist()[0])
            if not (temp=="nan" or temp=="0"):
                cell_write(target,row_index , 9,temp)
            else:
                cell_write(target,row_index , 9,"---")
            cell_write(target,row_index ,10,data_formO.loc[data_formO[columns[0]].apply(get_emp_name)==emp_name,"Leave Encashment"].to_string(index=False))
            #cell_write(target,row_index,4,is_abs_num)
            #cell_write(target,row_index,5,start)
            #cell_write(target,row_index,6,end)

        # def FL_write(row_index,target,start,end,is_abs_num):
        #     cell_write(target,row_index,1,start)
        #     cell_write(target,row_index,2,end)
        #     cell_write(target,row_index, 6, "-----")
        #     formOfile[sheet].merge_cells("F"+str(row_index)+":G"+str(row_index))
        #     #print("F"+str(row_index)+":G"+str(row_index))
        #     #cell_write(target,row_index,4,is_abs_num)
        #     #cell_write(target,row_index,5,start)
        #     #cell_write(target,row_index,6,end)
        
        # def CL_write(row_index,target,start,end,is_abs_num):
        #     cell_write(target,row_index,2,start)
        #     cell_write(target,row_index,3,end)
        #     #cell_write(target,row_index,5,start)
        #     #cell_write(target,row_index,6,end)

        form_write={'PL':PL_write}#,'FL':FL_write,'CL':CL_write}
        
        def start_end_date_attendance(rows,absent_label,row_offset,initial_offset):  
           # print("infunction",row_offset)
            is_abs_num=0
            row_index=0
            added={}
            for sheet_idx, row in enumerate(rows, 10):
                row_index=0
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        if code =="nan":
                            code=name
                        try:
                            target=formOfile[code]
                        except:
                            target = formOfile.copy_worksheet(formOsheet)
                            target.title=code
                            #initial offset
                            row_offset[code]=initial_offset
                        
                        target['A4']="Name and address of the Establishment:- "" "+str(data_formO['Company Name'].unique()[0])#+","+str(data_formO['Address'].unique()[0])
                        #target['A5']="Name of Employer:- "" "+str(data_formO['Unit'].unique()[0])
                        target["H4"]="Name of the employee:- "+str(name)+"\n"+" Receipt of leave book - "
                        target['A7']="Name of worker : "+str(name)
                        global emp_name
                        emp_name=str(name)
                        added[target.title]=0
                        form_write[absent_label](row_index+row_offset[target.title],target,"","","")
                        cell_write(target,row_index+row_offset[target.title] , 1, str("01"+"-"+str(month_num)+"-"+str(year)))
                        num=data_formO.loc[data_formO["Employee Name"]==emp_name,"num_days"]
                        if num.empty:
                            cell_write(target,row_index+row_offset[target.title] , 2,0)
                        else:
                            cell_write(target,row_index+row_offset[target.title] , 2,num.to_string(index=False))
                        #cell_write(target,row_index+row_offset[target.title] , 2,data_formO.loc[data_formO["Employee Name"]==emp_name,"Opening"].to_string(index=False))
                        
                    elif c_idx==2:
                        target['H8']="Date of entry into service :- "+str(value)
                    elif c_idx==3:
                        target['A8']="Description of the Department (If Applicable) :-  "+str(value)
                    elif c_idx==4:
                        target['A6']="Registration No. :- "+str(value)
                    elif is_abs_num==0 and value==absent_label:
                        is_abs_num=1
                        start=columns[c_idx-1]
                        end=columns[c_idx-1]
                    elif value==absent_label:
                        is_abs_num+=1
                        end=columns[c_idx-1]
                    elif is_abs_num:
                        start=start.split("\n")[1].replace("/","-")+"-"+str(year)
                        end=end.split("\n")[1].replace("/","-")+"-"+str(year)
                        
                        form_write[absent_label](row_index+row_offset[target.title],target,start,end,is_abs_num)
                        
                        #Uncomment these lines if there are too many lines in the first part of the form(This will cause border problems)
                        #target.insert_rows(row_index+row_offset[target.title]+1)
                        #added[target.title]+=1
                        
                        is_abs_num=0
                        num=data_formO.loc[data_formO["Employee Name"]==emp_name,"num_days"]
                        if num.empty:
                            cell_write(target,row_index+row_offset[target.title] , 2,0)
                        else:
                            cell_write(target,row_index+row_offset[target.title] , 2,num.to_string(index=False))
                        cell_write(target,row_index+row_offset[target.title] , 1, str("01"+"-"+str(month_num)+"-"+str(year)))
                        row_index+=1
                    
            # print(added)
            return added
        offset={}
        initial_offset=13
        #for sheet in formOfile.sheetnames:
        #    offset[sheet]=initial_offset
        offset=Counter(offset)+Counter(start_end_date_attendance(dataframe_to_rows(formO_data, index=False, header=False),"PL",offset,initial_offset))
        
        for sheet in formOfile.sheetnames:
            offset[sheet]+=25
            initial_offset+=25
            formOfile[sheet].merge_cells("A"+str(offset[sheet]-3)+":H"+str(offset[sheet]-3))
            formOfile[sheet].merge_cells("A"+str(offset[sheet]-2)+":B"+str(offset[sheet]-2))
            formOfile[sheet].merge_cells("C"+str(offset[sheet]-2)+":C"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("D"+str(offset[sheet]-2)+":D"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("E"+str(offset[sheet]-2)+":E"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("F"+str(offset[sheet]-2)+":G"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("H"+str(offset[sheet]-2)+":H"+str(offset[sheet]-1))
        columns=["Employee Name & Code"]
        data_formO["Employee Name & Code"]=data_formO["Employee Name"].astype(str)+"||"+data_formO["Employee Code"].astype(str)
        formO_data=data_formO[columns]
        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formO_data, index=False, header=False)
        logging.info('rows taken out from data')
        holidays=Read_Holiday_file()
        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 1):
                    name=value.split("||")[0]
                    code=value.split("||")[1]
                    if code =="nan":
                        code=name
                    target=formOfile[code]
                    import calendar
                    last_day=calendar.monthrange(int(year),month_num)[1]
                    start_date = str(year)+"-"+str(month_num)+"-01"
                    end_date = str(year)+"-"+str(month_num)+"-"+str(last_day)

                    after_start_date = holidays.index >= start_date
                    before_end_date = holidays.index <= end_date
                    between_two_dates = after_start_date & before_end_date

                    filtered_dates = holidays.loc[after_start_date,"Date"]
                    start_date=datetime.datetime.strptime(start_date,'%Y-%m-%d')
                    end_date=datetime.datetime.strptime(end_date,'%Y-%m-%d')
                    
                    #for index,date in enumerate(filtered_dates):
                    index=0
                    target.cell(row=offset[code]+index, column=1, value=start_date.date().strftime('%d-%m-%y'))
                    target.cell(row=offset[code]+index, column=1).font =Font(name ='Verdana', size =8)
                    target.cell(row=offset[code]+index, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=offset[code]+index, column=1).border = Border(outline= True, right=border_sides, bottom=border_sides)
                    
                    target.cell(row=offset[code]+index, column=2, value=end_date.date().strftime('%d-%m-%y'))
                    target.cell(row=offset[code]+index, column=2).font =Font(name ='Verdana', size =8)
                    target.cell(row=offset[code]+index, column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=offset[code]+index, column=2).border = Border(outline= True, right=border_sides, bottom=border_sides)

                    target.cell(row=offset[code]+index, column=3, value=len(holidays.loc[holidays.index >=str(start_date)]))
                    target.cell(row=offset[code]+index, column=3).font =Font(name ='Verdana', size =8)
                    target.cell(row=offset[code]+index, column=3).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=offset[code]+index, column=3).border = Border(outline= True, right=border_sides, bottom=border_sides)    

                    target.cell(row=offset[code]+index, column=4, value=between_two_dates.sum())
                    target.cell(row=offset[code]+index, column=4).font =Font(name ='Verdana', size =8)
                    target.cell(row=offset[code]+index, column=4).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=offset[code]+index, column=4).border = Border(outline= True, right=border_sides, bottom=border_sides)    

                    target.cell(row=offset[code]+index, column=5, value=len(holidays.loc[holidays.index >=str(start_date)])-between_two_dates.sum())
                    target.cell(row=offset[code]+index, column=5).font =Font(name ='Verdana', size =8)
                    target.cell(row=offset[code]+index, column=5).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=offset[code]+index, column=5).border = Border(outline= True, right=border_sides, bottom=border_sides)    
                
                    #offset[code]+=1
                    
        #offset+=Counter(start_end_date_attendance(dataframe_to_rows(formO_data, index=False, header=False),"FL",offset,initial_offset))
        
        for sheet in formOfile.sheetnames:
            offset[sheet]+=7
            initial_offset+=7
            formOfile[sheet].merge_cells("A"+str(offset[sheet]-4)+":F"+str(offset[sheet]-4))
            formOfile[sheet].merge_cells("A"+str(offset[sheet]-3)+":B"+str(offset[sheet]-2))
            formOfile[sheet].merge_cells("C"+str(offset[sheet]-3)+":C"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("D"+str(offset[sheet]-3)+":D"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("E"+str(offset[sheet]-3)+":E"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("F"+str(offset[sheet]-3)+":F"+str(offset[sheet]-1))

        columns=["Employee Name & Code","total_leave","availed","balance","remarks"]
        data_formO[["total_leave","availed","balance"]]=""
        for employee_name_leave_file in data_formO["Employee Name"]:
            #opening
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==employee_name_leave_file,:]
            opening_cl=emp_details["Opening"].loc[emp_details["Leave Type"]=="CL"]
            if opening_cl.empty:
                data_formO.loc[data_formO["Employee Name"]==employee_name_leave_file,"total_leave"]="0"
            else:
                opening_cl=opening_cl.to_string(index=False)
                data_formO.loc[data_formO["Employee Name"]==employee_name_leave_file,"total_leave"]=opening_cl if not opening_cl=="" else "0"

            availed=emp_details["Used"].loc[emp_details["Leave Type"]=="CL"]
            
            if availed.empty:
                data_formO.loc[data_formO["Employee Name"]==employee_name_leave_file,"availed"]="0"
            else:
                availed=availed.to_string(index=False)
                data_formO.loc[data_formO["Employee Name"]==employee_name_leave_file,"availed"]=availed if not availed=="" else "0"
                # print("------------------------------------------------------------------------------------------------------")
                # print("availed")
                # print(availed)

            balance=emp_details["Closing"].loc[emp_details["Leave Type"]=="CL"]
           
            if balance.empty:
                data_formO.loc[data_formO["Employee Name"]==employee_name_leave_file,"balance"]="0"
            else:
                balance=balance.to_string(index=False)
                data_formO.loc[data_formO["Employee Name"]==employee_name_leave_file,"balance"]=balance if not balance=="" else "0"
                # print("balance")
                # print(balance)

            
            
        data_formO[["remarks"]]=""
        data_formO["Employee Name & Code"]=data_formO["Employee Name"].astype(str)+"||"+data_formO["Employee Code"].astype(str)
        formO_data=data_formO[columns]
        
        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formO_data, index=False, header=False)
        logging.info('rows taken out from data')
        offset[code]+=1
        border_sides_thin = Side(style='thin')
        border_sides_thick = Side(style='thick')
        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 1):
                if c_idx==1:
                    name=value.split("||")[0]
                    code=value.split("||")[1]
                    if code =="nan":
                        code=name
                    target=formOfile[code]
                    start_date = "01"+"-"+str(month_num)+"-"+str(year)
                    end_date = str(last_day)+"-"+str(month_num)+"-"+str(year)
                    target.cell(row=offset[code], column=1, value=start_date)
                    target.cell(row=offset[code], column=1).font =Font(name ='Verdana', size =8)
                    target.cell(row=offset[code], column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=offset[code], column=1).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    
                    target.cell(row=offset[code], column=2, value=end_date)
                    target.cell(row=offset[code], column=2).font =Font(name ='Verdana', size =8)
                    target.cell(row=offset[code], column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=offset[code], column=2).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    target.row_dimensions[offset[code]].height = 20
                else:
                   # print("--------------------------------------------------------")
                   # print(offset[code],c_idx+1)
                    target.cell(row=offset[code], column=c_idx+1, value=str(value))
                    target.cell(row=offset[code], column=c_idx+1).font =Font(name ='Verdana', size =8)
                    target.cell(row=offset[code], column=c_idx+1).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=offset[code], column=c_idx+1).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    target.row_dimensions[offset[code]].height = 20
            #offset[code]+=1


        #offset+=Counter(start_end_date_attendance(dataframe_to_rows(formO_data, index=False, header=False),"CL",offset,initial_offset))
        formOfile.remove(formOfile["Sheet1"])
        formOfile.remove(formOfile["Sheet2"])
        formOfile.remove(formOfile["Sheet3"])
        formOfinalfile = os.path.join(filelocation,'Form O leave book.xlsx')
        formOfile.save(filename=formOfinalfile)

    Form_I()
    Form_II_Muster_Roll()
    Form_II_reg_damage_loss()
    Form_II_wages_reg()
    Form_VI_Overtime()
    Form_VI_reg_advance()
    From_O()


def Delhi(data,contractor_name,contractor_address,filelocation,month,year):
    Delhifilespath = os.path.join(Statefolder,'Delhi')
    logging.info('Goa files path is :'+str(Delhifilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]

    #print(sorted(list(data.columns)))
    def Form_G():
        formGfilepath = os.path.join(Delhifilespath,'Form G.xlsx')
        formGfile = load_workbook(filename=formGfilepath)
        logging.info('Form G file has sheet: '+str(formGfile.sheetnames))
        logging.info('create columns which are now available')

        data_formG = data.copy(deep=True)
        
        leave_file_data=data_formG[["Employee Code","Employee Name","Leave Type","Opening","Monthly Increment","Leave Accrued","Used","Encash","Closing"]]
        data_formG=data_formG.drop_duplicates(subset="Employee Name", keep="last")
        

        #Part 1 form
        data_formG["Employee Name & Code"]=data_formG["Employee Name"].astype(str)+"||"+data_formG["Employee Code"].astype(str)
        
        columns=["Employee Name & Code",'Nature of work',"Date","start_time","end_time","interval_for_reset_from","interval_for_reset_to","Total_hrs_worked",
                                            'Total\r\nOT Hrs','Overtime',"CL_Sl","leave_due","leave_availed","Balance","sign","remarks"]

        data_formG["leave_due"]=""
        data_formG["leave_availed"]=""
        data_formG["Balance"]=""
        data_formG["remarks"]="---"
        data_formG["sign"]=""
        for employee_name_leave_file in data_formG["Employee Name"]:
            #opening+monthly increment
            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==employee_name_leave_file,:]
            opening_pl=emp_details["Opening"].loc[emp_details["Leave Type"]=="PL"].astype(float)
            mon_inr_pl=emp_details["Monthly Increment"].loc[emp_details["Leave Type"]=="PL"].astype(float)
            leave_due=mon_inr_pl.add(opening_pl,fill_value=0).sum()
            data_formG.loc[data_formG["Employee Name"]==employee_name_leave_file,"leave_due"]=leave_due
            ##############################################################################################################################
            #used
            used_pl=emp_details["Used"].loc[emp_details["Leave Type"]=="PL"].astype(float)
            data_formG.loc[data_formG["Employee Name"]==employee_name_leave_file,"leave_availed"]=used_pl
            #closing
            balance_pl=emp_details["Closing"].loc[emp_details["Leave Type"]=="PL"].astype(float)
            data_formG.loc[data_formG["Employee Name"]==employee_name_leave_file,"Balance"]=balance_pl
            ###############################################################################################################################

            
        
        data_formG["Date"]="01"+"-"+str(month)+"-"+str(year)
        #print(data_formG["Date"])
        
        data_formG['Total_hrs_worked']="8 Hours"
        data_formG["CL_Sl"]=data_formG['Total\r\nCL'].astype(float)+data_formG['Total\r\nSL'].astype(float)

        data_formG["Fine_damage_loss"]=data_formG["Fine"].astype(str)+"\n"+data_formG["Damage or Loss"].astype(str)
        data_formG['interval_for_reset_to']=data_formG.rest_interval.str.split("-",expand=True)[1]
        data_formG['interval_for_reset_from']=data_formG.rest_interval.str.split("-",expand=True)[0]

        data_formG_columns=list(data_formG.columns)
        start=data_formG_columns.index('Emp Code')
        end=data_formG_columns.index('Total\r\nDP')
        start_date=data_formG_columns[start+1]
        end_date=data_formG_columns[end-1]
        start_date=start_date.split("\n")[1].replace("/","-")+"-"+str(year)
        end_date=end_date.split("\n")[1].replace("/","-")+"-"+str(year)

        formG_data=data_formG[columns]
        formGsheet = formGfile['Sheet1']
        formGsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form G is ready')



        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formG_data, index=False, header=False)

        logging.info('rows taken out from data')
        added=0
        #print("--------------------------------")
        for r_idx, row in enumerate(rows, 15):
            for c_idx, value in enumerate(row, 1):
                if c_idx==1:
                    try:
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        
                        target=formGfile[code]
                    except:
                        target = formGfile.copy_worksheet(formGsheet)
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        target.title=code
                        target["A8"]="Name of Employee "+name
                        target['A7']="Name of Establishment : "+data_formG['Company Name'].unique()[0]
                        target['A4']="Year:- "+str(year)+"  Month:- "+month
                        target['A5']="Wage Period:- "+start_date+" to  "+end_date
                        target["A10"]="Date of Employment   {}".format(data_formG.loc[data_formG["Employee Name"]==name,'Date Joined'].to_string(index=False))
                        
                elif c_idx==2:
                    target["A9"]="Nature of Work:- "+str(value)
                else:
                   # print(value)
                    target.cell(row=15+added, column=c_idx-2, value=value)
                    target.cell(row=15+added, column=c_idx-2).font =Font(name ='Verdana', size =8)
                    target.cell(row=15+added, column=c_idx-2).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=15+added, column=c_idx-2).border = Border(outline= True, right=border_sides, bottom=border_sides)
                    ###
                    border_sides_thick = Side(style='thick')       
                    border_sides_thin = Side(style='thin')

        #print("--------------------------")

        #Part 2 form
        data_formG["Employee Name & Code"]=data_formG["Employee Name"].astype(str)+"||"+data_formG["Employee Code"].astype(str)
        
        columns=["Employee Name & Code","Earned Basic","Overtime","All_Allowance_sum","Total Earning",
                                            "Fine_damage_loss","all_Other_Deduction_sum","date_of_payment",'Salary Advance',
                                            "Total_ded","Net Paid",'Date of payment',"sign"
                                            ]
        
        data_formG['Salary Advance']=data_formG['Salary Advance'].astype(str)
        data_formG=data_formG.replace({'Salary Advance':{"":"0","0.":"0","0.0":"0","nan":"0"}})
        
        
        get_date_of_payment=data_formG['Salary Advance']!="0"
        data_formG["date_of_payment"]=""
        data_formG["date_of_payment"]=data_formG.loc[get_date_of_payment,'Date of payment']
        all_deductions_columns_name=['HRA','Conveyance','Medical Allowance','Telephone Reimb','Tel and Int Reimb',
                                            'Bonus','Other Allowance', 'Fuel Reimb','Prof Dev Reimb','Corp Attire Reimb',
                                            'Meal Allowance','Special Allowance','Personal Allowance','CCA','Other Reimb',
                                            'Arrears','Other Earning',"Retention Pay",'Variable Pay','Leave Encashment',
                                            'Stipend','Consultancy Fees','Covid Deduction','OtherAllowance1', 
                                            'OtherAllowance2', 'OtherAllowance3', 'OtherAllowance4', 'OtherAllowance5'
                                            ]
        if "Covid Deduction" not in data_formG.columns:
            data_formG["Covid Deduction"]=0
        if "Retention Pay" not in data_formG.columns:
            data_formG["Retention Pay"]=0
            
        data_formG[all_deductions_columns_name]=data_formG[all_deductions_columns_name].astype(float)
        data_formG['All_Allowance_sum']= data_formG.loc[:,all_deductions_columns_name].sum(axis=1)

        data_formG["Fine_damage_loss"]=data_formG['Fine'].astype(float)+data_formG['Damage or Loss'].astype(float)

        other_deductions_columns_name=['Other Deduction','OtherDeduction1', 'OtherDeduction2',
                                                        'OtherDeduction3', 'OtherDeduction4', 'OtherDeduction5']

        data_formG[other_deductions_columns_name]=data_formG[other_deductions_columns_name].astype(float)
        data_formG["all_Other_Deduction_sum"]= data_formG.loc[:,other_deductions_columns_name].sum(axis=1)

        data_formG["Total_ded"]=data_formG["all_Other_Deduction_sum"]-data_formG['Salary Advance'].astype(float)
        
        data_formG["sign"]=""

        formG_data=data_formG[columns]
        formGsheet = formGfile['Sheet1']
        formGfile.remove(formGfile["Sheet1"])
        formGsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form G is ready')



        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formG_data, index=False, header=False)

        added=0
        for r_idx, row in enumerate(rows, 28):
            for c_idx, value in enumerate(row, 1):
                if c_idx==1:
                    try:
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        target=formGfile[code]
                    except:
                        target = formGfile.copy_worksheet(formGsheet)
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        
                        target.title=code
                        target["A8"]=target["A8"].value+" "+name
                        target['A7']=target['A7'].value+" : "+data_formG['Company Name'][0]
                        target['A4']="Year "+str(year)+"Month "+month
                        target['A5']="Wage Period:- "+start_date+"-"+end_date
                        target["A10"]="" #"Date of Employment   {}".format(data_formG.loc[data_formG["Employee Name"]==name,'Date Joined'])
                else:
                    target.cell(row=28+added, column=c_idx-1, value=value)
                    target.cell(row=28+added, column=c_idx-1).font =Font(name ='Verdana', size =8)
                    target.cell(row=28+added, column=c_idx-1).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=28+added, column=c_idx-1).border = Border(outline= True, right=border_sides, bottom=border_sides)
                    border_sides_thick = Side(style='thick')       
                    border_sides_thin = Side(style='thin')


        formGfinalfile = os.path.join(filelocation,'Form G.xlsx')
        formGfile.save(filename=formGfinalfile)



    def Form_H():
        formHfilepath = os.path.join(Delhifilespath,'Form H.xlsx')
        formHfile = load_workbook(filename=formHfilepath)
        logging.info('Form H file has sheet: '+str(formHfile.sheetnames))
        logging.info('create columns which are now available')

        def Part_I():
            data_formH = data.copy(deep=True)
            data_formH=data_formH.drop_duplicates(subset="Employee Name", keep="last")
            columns=['S.no',"Employee Name",'Nature of work']

            data_formH_columns=list(data_formH.columns)
            start=data_formH_columns.index('Emp Code')
            end=data_formH_columns.index('Total\r\nDP')
            columns.extend(data_formH_columns[start+1:end])
            
            less=31-len(data_formH_columns[start+1:end])
            for i in range(less):
                columns.extend(["less"+str(i+1)])
                data_formH["less"+str(i+1)]=""
            columns.extend(["remarks"])


            data_formH['S.no'] = list(range(1,len(data_formH)+1))
            data_formH[["remarks"]]=""
            formH_data=data_formH[columns]
            formHsheet = formHfile['Sheet1']
            formHsheet.sheet_properties.pageSetUpPr.fitToPage = True
            logging.info('data for form H is ready')

            from openpyxl.utils.dataframe import dataframe_to_rows
            rows = dataframe_to_rows(formH_data, index=False, header=False)
            rows_copy = list(dataframe_to_rows(formH_data, index=False, header=False))
            

            logging.info('rows taken out from data')
            formHsheet.unmerge_cells("A15:N15")
            formHsheet.unmerge_cells("A18:A19")
            formHsheet.unmerge_cells("B18:B19")
            
            formHsheet.unmerge_cells("C18:G18")
            formHsheet.unmerge_cells("H18:K18")
            formHsheet.unmerge_cells("L18:L19")
            formHsheet.unmerge_cells("M18:M19")
            formHsheet.unmerge_cells("N18:N19")
            
            from string import ascii_uppercase
            for char in ascii_uppercase[:14]:
                formHsheet.unmerge_cells(char+str(20)+':'+char+str(22))
            formHsheet.insert_rows(10,len(rows_copy))
            #formHsheet.delete_rows(18,2)
            formHsheet.merge_cells("C"+str(len(rows_copy)+18)+":G"+str(len(rows_copy)+18))
            formHsheet.merge_cells("H"+str(len(rows_copy)+18)+":K"+str(len(rows_copy)+18))
            formHsheet.merge_cells("A"+str(len(rows_copy)+18)+":A"+str(len(rows_copy)+19))
            formHsheet.merge_cells("B"+str(len(rows_copy)+18)+":B"+str(len(rows_copy)+19))
            formHsheet.merge_cells("L"+str(len(rows_copy)+18)+":L"+str(len(rows_copy)+19))
            formHsheet.merge_cells("M"+str(len(rows_copy)+18)+":M"+str(len(rows_copy)+19))
            formHsheet.merge_cells("N"+str(len(rows_copy)+18)+":N"+str(len(rows_copy)+19))
            formHsheet.merge_cells("A"+str(len(rows_copy)+15)+":N"+str(len(rows_copy)+15))

            row_num=0
            for r_idx, row in enumerate(rows, 10):
                row_num+=1
                for c_idx, value in enumerate(row, 1):
                    formHsheet.cell(row=r_idx, column=c_idx, value=value)
                    formHsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                    formHsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                    border_sides_thick = Side(style='thick')       
                    border_sides_thin = Side(style='thin')
                    if len(row)==c_idx and row_num==len(data_formH):
                        formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                        formHsheet.row_dimensions[r_idx].height = 20
                    elif len(row)==c_idx:
                        formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                        formHsheet.row_dimensions[r_idx].height = 20
                    elif row_num==len(data_formH):
                        formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                        formHsheet.row_dimensions[r_idx].height = 20
                    else:
                        formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                        formHsheet.row_dimensions[r_idx].height = 20
                    
        def Part_II():
            data_formH = data.copy(deep=True)
            data_formH=data_formH.drop_duplicates(subset="Employee Name", keep="last")
            columns=["Employee Name","Designation",'Earned Basic','DA','All_Other_Allowance',
                                'Total Earning','Overtime','Salary Advance',"sal_fine_damage",
                                'All_other_deductions','Total Deductions','Net Paid','sign','Date of payment']

            all_other_allowance_columns=['Other Allowance','OtherAllowance1', 'OtherAllowance2', 'OtherAllowance3', 'OtherAllowance4', 'OtherAllowance5']
            
            data_formH[all_other_allowance_columns]=data_formH[all_other_allowance_columns].astype(float)
            data_formH['All_Other_Allowance']= data_formH.loc[:,all_other_allowance_columns].sum(axis=1)
            data_formH["sal_fine_damage"]=data_formH["Fine"].apply(float)+data_formH["Damage or Loss"].apply(float)


            other_deductions_columns=['Insurance','CSR','PF','ESIC','P.Tax','LWF EE','Loan Deduction','Loan Interest','Other Deduction','TDS']
            data_formH[other_deductions_columns]=data_formH[other_deductions_columns].astype(float)
            data_formH['All_other_deductions']= data_formH.loc[:,other_deductions_columns].sum(axis=1)
            
            data_formH[["remarks",'Amount_Due','sign','Dearness_Allowance']]=""
            formH_data=data_formH[columns]
            formHsheet = formHfile['Sheet1']
            formHsheet.sheet_properties.pageSetUpPr.fitToPage = True
            logging.info('data for form H is ready')

            from openpyxl.utils.dataframe import dataframe_to_rows
            rows = dataframe_to_rows(formH_data, index=False, header=False)
            rows_copy = list(dataframe_to_rows(formH_data, index=False, header=False))
            

            logging.info('rows taken out from data')
            formHsheet.insert_rows(len(rows_copy)+20,len(rows_copy))
            row_num=0
            for r_idx, row in enumerate(rows, len(rows_copy)+20):
                row_num+=1
                for c_idx, value in enumerate(row, 1):
                    formHsheet.cell(row=r_idx, column=c_idx, value=value)
                    formHsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                    formHsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                    border_sides_thick = Side(style='thick')       
                    border_sides_thin = Side(style='thin')
                    if len(row)==c_idx and row_num==len(data_formH):
                        formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                        formHsheet.row_dimensions[r_idx].height = 20
                    elif len(row)==c_idx:
                        formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                        formHsheet.row_dimensions[r_idx].height = 20
                    elif row_num==len(data_formH):
                        formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                        formHsheet.row_dimensions[r_idx].height = 20
                    else:
                        formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                        formHsheet.row_dimensions[r_idx].height = 20

            formHsheet['A5']="Name of Establishment   "+str(data_formH['Company Name'].unique()[0])
            formHsheet['H5']=str(data_formH['start_time'].unique()[0])
            
            formHsheet['A6']="Registration No   "+str(data_formH['Registration_no'].unique()[0])
            formHsheet['H6']=str(data_formH['end_time'].unique()[0])
            formHsheet['Q7']=str(data_formH.rest_interval.str.split("-",expand=True)[0].unique()[0])
            formHsheet['U7']=str(data_formH.rest_interval.str.split("-",expand=True)[1].unique()[0])
            

            formHsheet['A'+str(len(rows_copy)+16)]="Name of Establishment   "+str(data_formH['Company Name'].unique()[0])
            formHsheet['A'+str(len(rows_copy)+17)]="Registration No   "+str(data_formH['Registration_no'].unique()[0])
            formHsheet['E'+str(len(rows_copy)+17)]="Wage Period :  "+str(month)+"  "+str(year)
        Part_I()
        Part_II()
        
        formHfinalfile = os.path.join(filelocation,'Form H.xlsx')
        formHfile.save(filename=formHfinalfile)
        

    def Form_I_reg():
        formIfilepath = os.path.join(Delhifilespath,'Form I register of Fine.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formI = data.copy(deep=True)
        data_formI=data_formI.drop_duplicates(subset="Employee Name", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","nature_of_offence","cause_against_fine","FIXED MONTHLY GROSS",
                                        "Date of payment&Fine",'Date_fine',"remarks"]

        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        data_formI[["nature_of_offence","cause_against_fine",'Date_fine']]="-----"
        data_formI["remarks"]=""
        
        data_formI['Fine']=data_formI['Fine'].astype(float)
        data_formI['Fine']=data_formI['Fine'].fillna(0)
        data_formI["Date of payment&Fine"]=data_formI['Date of payment']
        data_formI.loc[data_formI['Fine']==0,"Date of payment&Fine"]="---"
        


        #data_formI['Date of payment']+"\n"+data_formI["Fine"]
        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 7):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIsheet.row_dimensions[r_idx].height = 20
                else:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIsheet.row_dimensions[r_idx].height = 20

        print(data_formI['Company Name'].unique()[0])
        formIsheet['A4']=formIsheet['A4'].value+" : "+data_formI['Company Name'].unique()[0]
        formIfinalfile = os.path.join(filelocation,'Form I register of Fine.xlsx')
        formIfile.save(filename=formIfinalfile)

    def Form_I():
        formIfilepath = os.path.join(Delhifilespath,'Form I.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formI = data.copy(deep=True)
        leave_file_data=data_formI[["Employee Code","Employee Name","Leave Type","Opening","Monthly Increment","Leave Accrued","Used","Encash","Closing"]]
        
        data_formI=data_formI.drop_duplicates(subset="Employee Name", keep="last")
        columns=["Employee Name & Code","Date Joined"]
        data_formI["Employee Name & Code"]=data_formI["Employee Name"].astype(str)+"||"+data_formI["Employee Code"].astype(str)
        data_formI_columns=list(data_formI.columns)
        start=data_formI_columns.index('Emp Code')
        end=data_formI_columns.index('Total\r\nDP')
        columns.extend(data_formI_columns[start+1:end])


        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']

        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        # for column in  range(ord('A'), ord('G') + 1):
        #     formIsheet.unmerge_cells(chr(column)+"11:"+chr(column)+"15")
        # formIsheet.unmerge_cells("H11:I15")
        # formIsheet.unmerge_cells("J11:J15")
        # formIsheet.unmerge_cells("K11:K15")
        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')

        def cell_write(sheet,r_idx,c_idx,value):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            
        def start_end_date_attendance(absent_label,column_offset,row_offset,initial_offset):  
            is_abs_num=0
            row_index=0
            all_start_dates={}
            all_end_dates={}
            added={}
            for sheet_idx, row in enumerate(dataframe_to_rows(formI_data, index=False, header=False), 12):
                row_index=0
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        if code =="nan":
                            code=name
                        try:
                            target=formIfile[code]
                            added[target.title]=0
                        except:
                            target = formIfile.copy_worksheet(formIsheet)
                            target.title=code
                            #initial offset
                            row_offset[code]=initial_offset
                            added[target.title]=0
                        target['A6']="Name of Employee : "+name
                        target['A4']="Name of Establishment : "+data_formI['Company Name'].unique()[0]
                        target['A7']="Period "+str(month)+" "+str(year)
                    elif c_idx==2:
                        target['A5']="Date of Employment : "+value
                    elif is_abs_num==0 and value==absent_label:
                        is_abs_num=1
                        start=columns[c_idx-1]
                        end=columns[c_idx-1]
                    elif value==absent_label:
                        is_abs_num+=1
                        end=columns[c_idx-1]
                    elif is_abs_num:
                        #target.cell(row=row_index+13, column=1+column_offset, value=is_abs_num)
                        start=start.split("\n")[1].replace("/","-")+"-"+str(year)
                        end=end.split("\n")[1].replace("/","-")+"-"+str(year)
                        
                        cell_write(target,row_index+row_offset[target.title],3+column_offset,start)
                        cell_write(target,row_index+row_offset[target.title],4+column_offset,end)
                        emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==name,:]
                        if absent_label=="PL":
                            total=emp_details["Used"].loc[emp_details["Leave Type"]=="PL"]
                            Closing_pl=emp_details["Closing"].loc[emp_details["Leave Type"]=="PL"].astype(float)
                            Closing_cl=emp_details["Closing"].loc[emp_details["Leave Type"]=="CL"].astype(float)
                            Closing_sl=emp_details["Closing"].loc[emp_details["Leave Type"]=="SL"].astype(float)
                            balance=Closing_cl.add(Closing_pl.add(Closing_sl,fill_value=0), fill_value=0).sum()
                            cell_write(target,row_index+row_offset[target.title],6,"----")
                            cell_write(target,row_index+row_offset[target.title],7,"----")
                            cell_write(target,row_index+row_offset[target.title],10,total.to_string(index=False))
                            cell_write(target,row_index+row_offset[target.title],11,balance)
                        else:
                            temp=data_formI.loc[data_formI["Employee Name"]==name,'Total\r\nCL'].astype(float)
                            amt_leave_requested=data_formI.loc[data_formI["Employee Name"]==name,'Total\r\nSL'].astype(float)+temp
                            #print(amt_leave_requested)
                            Used_cl=emp_details["Used"].loc[emp_details["Leave Type"]=="CL"].astype(float)
                            Used_sl=emp_details["Used"].loc[emp_details["Leave Type"]=="SL"].astype(float)
                            availed=Used_cl.add(Used_sl,fill_value=0).sum()
                            cell_write(target,row_index+row_offset[target.title],1,amt_leave_requested.to_string(index=False))
                            cell_write(target,row_index+row_offset[target.title],2,"----")
                            cell_write(target,row_index+row_offset[target.title],5,availed)
                            
                        # cell_write(target,row_index+11,5+offset,is_abs_num)
                        is_abs_num=0
                        row_index+=1
                        added[target.title]+=1
                        # border_sides_thick = Side(style='thick')       
                        # border_sides_thin = Side(style='thin')
                        # if len(row)==c_idx and len(row)==len(data_formI):
                        #     formIsheet.cell(row=row_index+row_offset[target.title], column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                        #     formIsheet.row_dimensions[row_index+row_offset[target.title]].height = 20
                        # elif len(row)==c_idx:
                        #     formIsheet.cell(row=row_index+row_offset[target.title], column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                        #     formIsheet.row_dimensions[row_index+row_offset[target.title]].height = 20
                        # elif len(row)==len(data_formI):
                        #     formIsheet.cell(row=row_index+row_offset[target.title], column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                        #     formIsheet.row_dimensions[row_index+row_offset[target.title]].height = 20
                        # else:
                        #     formIsheet.cell(row=row_index+row_offset[target.title], column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                        #     formIsheet.row_dimensions[row_index+row_offset[target.title]].height = 20
            return added
                    
        absent_label="PL"
        column_offset=5
        initial_offset=12
        row_offset={}          
        row_offset=start_end_date_attendance(absent_label,column_offset,row_offset,initial_offset)
        #reset row_offset since it was for PL
        
        row_offset = {x: initial_offset for x in row_offset}
        
        absent_label="CL"
        column_offset=0
        from collections import Counter
        #increment = {x: initial_offset for x in row_offset}
        row_offset=Counter(start_end_date_attendance(absent_label,column_offset,row_offset,initial_offset))+Counter(row_offset)
        absent_label="SL"
        column_offset=0
        row_offset=start_end_date_attendance(absent_label,column_offset,row_offset,initial_offset)
        

        formIfile.remove(formIfile["Sheet1"])
        formIfinalfile = os.path.join(filelocation,'Form I.xlsx')
        formIfile.save(filename=formIfinalfile)


    def Form_II():
        formIIfilepath = os.path.join(Delhifilespath,'Form II.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy(deep=True)
        data_formII=data_formII.drop_duplicates(subset="Employee Name", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Gender","Department",
                                        "Damage_loss_with_date","whether_work_showed_cause",
                                        "Date of payment & amount of deduction","num_instalments",'Date on which total amount realised',"remarks"]
        
        data_formII["Damage or Loss"]=data_formII["Damage or Loss"].astype(float)
        data_formII["Damage or Loss"]=data_formII["Damage or Loss"].fillna(0)
        
        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII['Date of payment']=data_formII['Date of payment'].apply(lambda x: x.strftime('%d-%m-%Y'))
        data_formII["Damage_loss_with_date"]=data_formII['Date of payment']+" & "+data_formII["Damage or Loss"].astype(str)
        data_formII.loc[data_formII["Damage or Loss"]==0,"Damage_loss_with_date"]="---"

        data_formII["Date of payment & amount of deduction"]=data_formII["Damage_loss_with_date"]
        data_formII["num_instalments"]="1"
        data_formII.loc[data_formII["Damage or Loss"]==0,"num_instalments"]="---"

        data_formII["Date on which total amount realised"]=data_formII['Date of payment']
        data_formII.loc[data_formII["Damage or Loss"]==0,"Date on which total amount realised"]="---"

        data_formII[["remarks","whether_work_showed_cause"]]="-----"
        

        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formII_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 7):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                
                if len(row)==c_idx and row_num==len(data_formII):
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formII):
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIIsheet.row_dimensions[r_idx].height = 20
                else:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIIsheet.row_dimensions[r_idx].height = 20

        formIIsheet['A4']=formIIsheet['A4'].value+" : "+data_formII['Company Name'].unique()[0]
        formIIfinalfile = os.path.join(filelocation,'Form II.xlsx')
        formIIfile.save(filename=formIIfinalfile)

    def Form_IV():
        formIVfilepath = os.path.join(Delhifilespath,'Form IV.xlsx')
        formIVfile = load_workbook(filename=formIVfilepath)
        logging.info('Form IV file has sheet: '+str(formIVfile.sheetnames))
        logging.info('create columns which are now available')

        data_formIV = data.copy(deep=True)
        data_formIV=data_formIV.drop_duplicates(subset="Employee Name", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Gender","Designation_Dept","Date_overtime_worked",
                                        "Extent of over-time","Total over-time","Normal hrs ",
                                        "FIXED MONTHLY GROSS","overtime rate",
                                        "normal_earning","Overtime",'Total Earning',"date_overtime_paid"]
                                        
        data_formIV['Total\r\nOT Hrs']=data_formIV[['Total\r\nOT Hrs',"Overtime",'Total Earning']].astype(float)
        data_formIV["Total over-time"]=data_formIV['Total\r\nOT Hrs']
        data_formIV["normal_earning"]=data_formIV['Total Earning']-data_formIV["Overtime"]
        data_formIV.loc[data_formIV['Total\r\nOT Hrs']==0,["Total over-time","Normal hrs ",
                                        "FIXED MONTHLY GROSS","overtime rate",
                                        "normal_earning","Overtime",'Total Earning']]="---"

        data_formIV["date_overtime_paid"]=data_formIV['Date of payment']
        data_formIV.loc[data_formIV["Overtime"]==0,"date_overtime_paid"]="---"
        data_formIV.loc[data_formIV['Total\r\nOT Hrs']==0,"date_overtime_paid"]="---"

        data_formIV['S.no'] = list(range(1,len(data_formIV)+1))
        data_formIV['Designation_Dept']=data_formIV["Designation"]+"_"+data_formIV["Department"]
        data_formIV["Extent of over-time"]="-----"

        data_formIV["Date_overtime_worked"]="-----"
        # data_formIV["Date of payment & amount of deduction"]=data_formIV['Date of payment']+"\n"+data_formIV["Total Deductions"]
        formIV_data=data_formIV[columns]
        formIVsheet = formIVfile['Sheet1']
        formIVsheet.sheet_properties.pageSetUpPr.fitToPage = True
        for column in  range(ord('A'), ord('O') + 1):
            formIVsheet.unmerge_cells(chr(column)+"7:"+chr(column)+"14")

        logging.info('data for form IV is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formIV_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 7):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formIVsheet.cell(row=r_idx, column=c_idx, value=value)
                formIVsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIVsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formIV):
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIVsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIVsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formIV):
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIVsheet.row_dimensions[r_idx].height = 20
                else:
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIVsheet.row_dimensions[r_idx].height = 20

        formIVsheet['A4']=formIVsheet['A4'].value+"  "+data_formIV['Company Name'].unique()[0]+"  "+data_formIV['Company Address'].unique()[0]+"                                Month Ending: "+month+" "+str(year)
        # formIVsheet.cell(row=4, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
        #formIVsheet['A4']="Month Ending: "+month+" "+str(year)
        formIVfinalfile = os.path.join(filelocation,'Form IV.xlsx')
        formIVfile.save(filename=formIVfinalfile)
        
    Form_H()
    Form_I_reg()
    Form_I()
    Form_II()
    Form_IV()
    Form_G()


def Tamilnadu(data,contractor_name,contractor_address,filelocation,month,year):
    logging.info('Tamilnadu forms')
    
    

def Telangana(data,contractor_name,contractor_address,filelocation,month,year):
    logging.info('Telangana forms')

def Uttar_Pradesh(data,contractor_name,contractor_address,filelocation,month,year):
    logging.info('Uttar Pradesh forms')


def Karnataka(data,contractor_name,contractor_address,filelocation,month,year):
    karnatakafilespath = os.path.join(Statefolder,'Karnataka')
    logging.info('karnataka files path is :'+str(karnatakafilespath))
    data.reset_index(drop=True, inplace=True)

    month_num = monthdict[month]

    def create_form_A():

        formAfilepath = os.path.join(karnatakafilespath,'FormA.xlsx')
        formAfile = load_workbook(filename=formAfilepath)
        logging.info('Form A file has sheet: '+str(formAfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formA = data.drop_duplicates(subset=['Employee Code']).copy()

        data_formA.fillna(value=0, inplace=True)

        data_formA['S.no'] = list(range(1,len(data_formA)+1))

        #data_formA["Nationality"] =''
        #data_formA['education level'] = ''
        #data_formA['Category address'] = ''
        #data_formA['type of employment'] = ''
        #data_formA['lwf'] = ''
        #data_formA['Service Book No'] = ''
        data_formA["a"] = ''
        data_formA["b"] = ''
        data_formA["c"] = ''
        formA_columns = ["S.no",'Employee Code','Employee Name','Unit','Location',"Gender","Father's Name",'Date of Birth',"Nationality","Education Level",'Date Joined','Designation','Category Address',"Type of Employment",'Mobile Tel No.','UAN Number',"PAN Number","ESIC Number","LWF EE","Aadhar Number","Bank A/c Number","Bank Name","Account Code","P","L","Service Book No","Date Left","Reason for Leaving","Identification mark","a","b","c"]
        formA_data = data_formA[formA_columns]


        formAsheet = formAfile['FORM A']

        formAsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form A is ready')

        
        rows = dataframe_to_rows(formA_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 13):
            for c_idx, value in enumerate(row, 1):
                formAsheet.cell(row=r_idx, column=c_idx, value=value)
                formAsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formAsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formAsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                if c_idx==15 or c_idx==16 or c_idx==20 or c_idx==21:
                    formAsheet.cell(row=r_idx, column=c_idx).number_format= numbers.FORMAT_NUMBER

        logging.info('')


        if str(data_formA['Company Name'].dtype)[0:3] != 'obj':
            data_formA['Company Name'] = data_formA['Company Name'].astype(str)

        if str(data_formA['Company Address'].dtype)[0:3] != 'obj':
            data_formA['Company Address'] = data_formA['Company Address'].astype(str)

        if str(data_formA['Contractor_name'].dtype)[0:3] != 'obj':
            data_formA['Contractor_name'] = data_formA['Contractor_name'].astype(str)

        if str(data_formA['Contractor_Address'].dtype)[0:3] != 'obj':
            data_formA['Contractor_Address'] = data_formA['Contractor_Address'].astype(str)

        if str(data_formA['Unit'].dtype)[0:3] != 'obj':
            data_formA['Unit'] = data_formA['Unit'].astype(str)

        if str(data_formA['Branch'].dtype)[0:3] != 'obj':
            data_formA['Branch'] = data_formA['Branch'].astype(str)

        establishment = formAsheet['L6'].value
        if data_formA['PE_or_contract'][0] == 'PE':
            L6_data = establishment+' '+data_formA['Company Name'][0] +', '+data_formA['Company Address'][0]  
        else:    
            L6_data = establishment+' '+data_formA['Contractor_name'][0]+', '+data_formA['Contractor_Address'][0]
        formAsheet['L6'] = L6_data


        company = formAsheet['A10'].value
        A10_data = company+' '+data_formA['Unit'][0]+', '+data_formA['Branch'][0]
        formAsheet['A10'] = A10_data

        
        formAfinalfile = os.path.join(filelocation,'FormA.xlsx')
        logging.info('Form A file is' +str(formAfinalfile))
        formAfile.save(filename=formAfinalfile)
        

    def create_form_B():
        formBfilepath = os.path.join(karnatakafilespath,'FormB.xlsx')
        formBfile = load_workbook(filename=formBfilepath)
        logging.info('Form B file has sheet: '+str(formBfile.sheetnames))

        
        logging.info('create columns which are now available')


        data_formB = data.drop_duplicates(subset=['Employee Code']).copy()
        data_formB.fillna(value=0, inplace=True)

        #data_formB['OT hours'] = 0
        #data_formB['Pay OT'] = 0
        if str(data_formB['Earned Basic'].dtype)[0:3] != 'int':
            data_formB['Earned Basic']= data_formB['Earned Basic'].astype(int)
        if str(data_formB['DA'].dtype)[0:3] != 'int':
            data_formB['DA']= data_formB['DA'].astype(int)
        data_formB['basic_and_allo'] = data_formB['Earned Basic']+ data_formB['DA']
        #data_formB['Other EAR'] = data_formB['Other Reimb']+data_formB['Arrears']+data_formB['Other Earning']+data_formB['Variable Pay']+data_formB['Stipend'] +data_formB['Consultancy Fees']
        #data_formB['VPF']=0
        data_formB['Society']="---"
        data_formB['Income Tax']="---"
        if str(data_formB['Other Deduction'].dtype)[0:3] != 'int':
            data_formB['Other Deduction']= data_formB['Other Deduction'].astype(int)
        if str(data_formB['Other Deduction'].dtype)[0:3] != 'int':
            data_formB['Other Deduction']= data_formB['Other Deduction'].astype(int)
        data_formB['Other Deduc']= data_formB['Other Deduction']+ data_formB['Salary Advance']
        data_formB['EMP PF'] = data_formB['PF']
        #data_formB['BankID'] = ''
        #data_formB['Pay Date'] = ''
        data_formB['Remarks'] =''

        formB_columns = ['Employee Code','Employee Name','FIXED MONTHLY GROSS',	'Days Paid','Total\r\nOT Hrs',	'basic_and_allo', 'Overtime',	'HRA',	'Tel and Int Reimb', 'Bonus', 'Fuel Reimb',	'Prof Dev Reimb', 'Corp Attire Reimb',	'CCA',	'Leave Encashment','Conveyance', 'Medical Allowance', 'Telephone Reimb', 'Other Allowance', 'Meal Allowance',
       'Special Allowance', 'Personal Allowance','Other Reimb', 'Arrears', 'Other Earning', 'Variable Pay','Stipend' ,'Total Earning', 'PF',	'ESIC',	'VPF', 'Loan Deduction', 'Loan Interest', 'P.Tax',	'Society',	'Income Tax', 'Insurance',	'LWF EE',	'Other Deduction',	'TDS',	'Total Deductions',	'Net Paid',	'EMP PF','Bank A/c Number','Date of payment','Remarks']

        formB_data = data_formB[formB_columns]

        formBsheet = formBfile['FORM B']

        formBsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form B is ready')

        
        rows = dataframe_to_rows(formB_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 21):
            for c_idx, value in enumerate(row, 2):
                formBsheet.cell(row=r_idx, column=c_idx, value=value)
                formBsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formBsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formBsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                if c_idx==45:
                    formBsheet.cell(row=r_idx, column=c_idx).number_format= numbers.FORMAT_NUMBER

        contractline = formBsheet['B10'].value
        B10_data = contractline+' '+contractor_name+', '+contractor_address
        formBsheet['B10'] = B10_data

        if str(data_formB['Nature of work'].dtype)[0:3] != 'obj':
            data_formB['Nature of work'] = data_formB['Nature of work'].astype(str)

        if str(data_formB['Location'].dtype)[0:3] != 'obj':
            data_formB['Location'] = data_formB['Location'].astype(str)

        if str(data_formB['Company Name'].dtype)[0:3] != 'obj':
            data_formB['Company Name'] = data_formB['Company Name'].astype(str)

        if str(data_formB['Company Address'].dtype)[0:3] != 'obj':
            data_formB['Company Address'] = data_formB['Company Address'].astype(str)

        if str(data_formB['Unit'].dtype)[0:3] != 'obj':
            data_formB['Unit'] = data_formB['Unit'].astype(str)

        if str(data_formB['Address'].dtype)[0:3] != 'obj':
            data_formB['Address'] = data_formB['Address'].astype(str)

        locationline = formBsheet['B11'].value
        B11_data = locationline+' '+data_formB['Nature of work'][0]+', '+data_formB['Location'][0]
        formBsheet['B11'] = B11_data

        establine = formBsheet['B12'].value
        if data_formB['PE_or_contract'][0]== 'PE':
            B12_data = establine+' '+data_formB['Company Name'][0]+', '+data_formB['Company Address'][0]  
        else:    
            B12_data = establine+' '+data_formB['Unit'][0]+', '+data_formB['Address'][0]
        formBsheet['B12'] = B12_data

        peline = formBsheet['B13'].value
        if data_formB['PE_or_contract'][0]== 'PE':
            B13_data = peline+' '+data_formB['Company Name'][0]+', '+data_formB['Company Address'][0]  
        else:    
            B13_data = peline+' '+data_formB['Unit'][0]+', '+data_formB['Address'][0]
        formBsheet['B13'] = B13_data

        monthstart = datetime.date(year,month_num,1)
        monthend = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
        formBsheet['B16'] = 'Wage period From: '+str(monthstart)+' to '+str(monthend)

        formBfinalfile = os.path.join(filelocation,'FormB.xlsx')
        formBfile.save(filename=formBfinalfile)

    def create_form_XXI():
        formXXIfilepath = os.path.join(karnatakafilespath,'FormXXI.xlsx')
        formXXIfile = load_workbook(filename=formXXIfilepath)
        logging.info('Form XXI file has sheet: '+str(formXXIfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXXI = data.drop_duplicates(subset=['Employee Code']).copy()

        data_formXXI.fillna(value=0, inplace=True)

        data_formXXI['S.no'] = list(range(1,len(data_formXXI)+1))

        data_formXXI['a'] ='---'
        data_formXXI['b'] ='---'
        data_formXXI['c'] ='---'
        #data_formXXI['e'] ='---'
        data_formXXI['f'] ='---'
        data_formXXI['g'] =''

        formXXI_columns = ['S.no','Employee Name',"Father's Name",'Designation','a','b','c','FIXED MONTHLY GROSS','Fine','f','g']

        formXXI_data = data_formXXI[formXXI_columns]

        formXXIsheet = formXXIfile['FORM XXI']

        formXXIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XXI is ready')

        
        rows = dataframe_to_rows(formXXI_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 14):
            for c_idx, value in enumerate(row, 3):
                formXXIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXIsheet['C7'].value
        C7_data = contractline+' '+contractor_name+', '+contractor_address
        formXXIsheet['C7'] = C7_data

        if str(data_formXXI['Nature of work'].dtype)[0:3] != 'obj':
            data_formXXI['Nature of work'] = data_formXXI['Nature of work'].astype(str)

        if str(data_formXXI['Location'].dtype)[0:3] != 'obj':
            data_formXXI['Location'] = data_formXXI['Location'].astype(str)

        if str(data_formXXI['Company Name'].dtype)[0:3] != 'obj':
            data_formXXI['Company Name'] = data_formXXI['Company Name'].astype(str)

        if str(data_formXXI['Company Address'].dtype)[0:3] != 'obj':
            data_formXXI['Company Address'] = data_formXXI['Company Address'].astype(str)

        if str(data_formXXI['Unit'].dtype)[0:3] != 'obj':
            data_formXXI['Unit'] = data_formXXI['Unit'].astype(str)

        if str(data_formXXI['Address'].dtype)[0:3] != 'obj':
            data_formXXI['Address'] = data_formXXI['Address'].astype(str)

        locationline = formXXIsheet['C8'].value
        C8_data = locationline+' '+data_formXXI['Nature of work'][0]+', '+data_formXXI['Location'][0]
        formXXIsheet['C8'] = C8_data

        establine = formXXIsheet['C9'].value
        if data_formXXI['PE_or_contract'][0]== 'PE':
            C9_data = establine+' '+data_formXXI['Company Name'][0]+', '+data_formXXI['Company Address'][0]
        else:
            C9_data = establine+' '+data_formXXI['Unit'][0]+', '+data_formXXI['Address'][0]
        formXXIsheet['C9'] = C9_data

        peline = formXXIsheet['C10'].value
        if data_formXXI['PE_or_contract'][0]== 'PE':
            C10_data = peline+' '+data_formXXI['Company Name'][0]+', '+data_formXXI['Company Address'][0]
        else:
            C10_data = peline+' '+data_formXXI['Unit'][0]+', '+data_formXXI['Address'][0]
        formXXIsheet['C10'] = C10_data

        #border the region
        count1 = len(data_formXXI)
        border_1 = Side(style='thick')
        for i in range(2,15):
            formXXIsheet.cell(row=2, column=i).border = Border(outline= True, top=border_1)
            formXXIsheet.cell(row=count1+15, column=i).border = Border(outline= True, bottom=border_1)
        for i in range(2,count1+16):
            formXXIsheet.cell(row=i, column=2).border = Border(outline= True, left=border_1)
            formXXIsheet.cell(row=i, column=14).border = Border(outline= True, right=border_1)

        formXXIfinalfile = os.path.join(filelocation,'FormXXI.xlsx')
        formXXIfile.save(filename=formXXIfinalfile)



    def create_form_XXII():
        formXXIIfilepath = os.path.join(karnatakafilespath,'FormXXII.xlsx')
        formXXIIfile = load_workbook(filename=formXXIIfilepath)
        logging.info('Form XXII file has sheet: '+str(formXXIIfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXXII = data.drop_duplicates(subset=['Employee Code']).copy()

        data_formXXII.fillna(value=0, inplace=True)

        data_formXXII['S.no'] = list(range(1,len(data_formXXII)+1))

        data_formXXII['b'] ='---'
        data_formXXII['c'] ='---'
        data_formXXII['d'] ='---'
        data_formXXII['e'] ='---'
        data_formXXII['f'] ='---'
        data_formXXII['g'] =''

        formXXII_columns = ['S.no','Employee Name',"Father's Name",'Designation','FIXED MONTHLY GROSS','b','c','d','e','f','g']

        formXXII_data = data_formXXII[formXXII_columns]

        formXXIIsheet = formXXIIfile['FORM XXII']

        formXXIIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XXII is ready')

        
        rows = dataframe_to_rows(formXXII_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 15):
            for c_idx, value in enumerate(row, 3):
                formXXIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXIIsheet['C7'].value
        C7_data = contractline+' '+contractor_name+', '+contractor_address
        formXXIIsheet['C7'] = C7_data


        if str(data_formXXII['Nature of work'].dtype)[0:3] != 'obj':
            data_formXXII['Nature of work'] = data_formXXII['Nature of work'].astype(str)

        if str(data_formXXII['Location'].dtype)[0:3] != 'obj':
            data_formXXII['Location'] = data_formXXII['Location'].astype(str)

        if str(data_formXXII['Company Name'].dtype)[0:3] != 'obj':
            data_formXXII['Company Name'] = data_formXXII['Company Name'].astype(str)

        if str(data_formXXII['Company Address'].dtype)[0:3] != 'obj':
            data_formXXII['Company Address'] = data_formXXII['Company Address'].astype(str)

        if str(data_formXXII['Unit'].dtype)[0:3] != 'obj':
            data_formXXII['Unit'] = data_formXXII['Unit'].astype(str)

        if str(data_formXXII['Address'].dtype)[0:3] != 'obj':
            data_formXXII['Address'] = data_formXXII['Address'].astype(str)

        locationline = formXXIIsheet['C8'].value
        C8_data = locationline+' '+data_formXXII['Nature of work'][0]+', '+data_formXXII['Location'][0]
        formXXIIsheet['C8'] = C8_data

        establine = formXXIIsheet['C9'].value
        if data_formXXII['PE_or_contract'][0]== 'PE':
            C9_data = establine+' '+data_formXXII['Company Name'][0]+', '+data_formXXII['Company Address'][0]
        else:
            C9_data = establine+' '+data_formXXII['Unit'][0]+', '+data_formXXII['Address'][0]
        formXXIIsheet['C9'] = C9_data

        peline = formXXIIsheet['C10'].value
        if data_formXXII['PE_or_contract'][0]== 'PE':
            C10_data = peline+' '+data_formXXII['Company Name'][0]+', '+data_formXXII['Company Address'][0]
        else:
            C10_data = peline+' '+data_formXXII['Unit'][0]+', '+data_formXXII['Address'][0]
        formXXIIsheet['C10'] = C10_data

        #border the region
        count1 = len(data_formXXII)
        border_1 = Side(style='thick')
        for i in range(2,15):
            formXXIIsheet.cell(row=2, column=i).border = Border(outline= True, top=border_1)
            formXXIIsheet.cell(row=count1+16, column=i).border = Border(outline= True, bottom=border_1)
        for i in range(2,count1+17):
            formXXIIsheet.cell(row=i, column=2).border = Border(outline= True, left=border_1)
            formXXIIsheet.cell(row=i, column=14).border = Border(outline= True, right=border_1)

        formXXIIfinalfile = os.path.join(filelocation,'FormXXII.xlsx')
        formXXIIfile.save(filename=formXXIIfinalfile)


    def create_form_XXIII():
        formXXIIIfilepath = os.path.join(karnatakafilespath,'FormXXIII.xlsx')
        formXXIIIfile = load_workbook(filename=formXXIIIfilepath)
        logging.info('Form XXIII file has sheet: '+str(formXXIIIfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXXIII = data.drop_duplicates(subset=['Employee Code']).copy()
        data_formXXIII.fillna(value=0, inplace=True)

        data_formXXIII['S.no'] = list(range(1,len(data_formXXIII)+1))

        data_formXXIII['b'] ='---'
        data_formXXIII['c'] ='---'
        data_formXXIII['d'] ='---'
        data_formXXIII['e'] ='---'
        data_formXXIII['f'] ='---'
        data_formXXIII['g'] =''

        formXXIII_columns = ['S.no','Employee Name',"Father's Name",'Designation','FIXED MONTHLY GROSS','b','c','d','e','f','g']

        formXXIII_data = data_formXXIII[formXXIII_columns]

        formXXIIIsheet = formXXIIIfile['FORM XXIII']

        formXXIIIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XXIII is ready')

        
        rows = dataframe_to_rows(formXXIII_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 12):
            for c_idx, value in enumerate(row, 3):
                formXXIIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXIIIsheet['C5'].value
        C5_data = contractline+' '+contractor_name+', '+contractor_address
        formXXIIIsheet['C5'] = C5_data

        if str(data_formXXIII['Nature of work'].dtype)[0:3] != 'obj':
            data_formXXIII['Nature of work'] = data_formXXIII['Nature of work'].astype(str)

        if str(data_formXXIII['Location'].dtype)[0:3] != 'obj':
            data_formXXIII['Location'] = data_formXXIII['Location'].astype(str)

        if str(data_formXXIII['Company Name'].dtype)[0:3] != 'obj':
            data_formXXIII['Company Name'] = data_formXXIII['Company Name'].astype(str)

        if str(data_formXXIII['Company Address'].dtype)[0:3] != 'obj':
            data_formXXIII['Company Address'] = data_formXXIII['Company Address'].astype(str)

        if str(data_formXXIII['Unit'].dtype)[0:3] != 'obj':
            data_formXXIII['Unit'] = data_formXXIII['Unit'].astype(str)

        if str(data_formXXIII['Address'].dtype)[0:3] != 'obj':
            data_formXXIII['Address'] = data_formXXIII['Address'].astype(str)

        locationline = formXXIIIsheet['C6'].value
        C6_data = locationline+' '+data_formXXIII['Nature of work'][0]+', '+data_formXXIII['Location'][0]
        formXXIIIsheet['C6'] = C6_data

        establine = formXXIIIsheet['C7'].value
        if data_formXXIII['PE_or_contract'][0]== 'PE':
            C7_data = establine+' '+data_formXXIII['Company Name'][0]+', '+data_formXXIII['Company Address'][0]
        else:
            C7_data = establine+' '+data_formXXIII['Unit'][0]+', '+data_formXXIII['Address'][0]
        formXXIIIsheet['C7'] = C7_data

        peline = formXXIIIsheet['C8'].value
        if data_formXXIII['PE_or_contract'][0]== 'PE':
            C8_data = peline+' '+data_formXXIII['Company Name'][0]+', '+data_formXXIII['Company Address'][0]
        else:
            C8_data = peline+' '+data_formXXIII['Unit'][0]+', '+data_formXXIII['Address'][0]
        formXXIIIsheet['C8'] = C8_data


        #border the region
        count1 = len(data_formXXIII)
        border_1 = Side(style='thick')
        for i in range(2,15):
            formXXIIIsheet.cell(row=2, column=i).border = Border(outline= True, top=border_1)
            formXXIIIsheet.cell(row=count1+13, column=i).border = Border(outline= True, bottom=border_1)
        for i in range(2,count1+14):
            formXXIIIsheet.cell(row=i, column=2).border = Border(outline= True, left=border_1)
            formXXIIIsheet.cell(row=i, column=14).border = Border(outline= True, right=border_1)


        formXXIIIfinalfile = os.path.join(filelocation,'FormXXIII.xlsx')
        formXXIIIfile.save(filename=formXXIIIfinalfile)


    def create_form_XX():
        formXXfilepath = os.path.join(karnatakafilespath,'FormXX.xlsx')
        formXXfile = load_workbook(filename=formXXfilepath)
        logging.info('Form XX file has sheet: '+str(formXXfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXX = data.drop_duplicates(subset=['Employee Code']).copy()
        data_formXX.fillna(value=0, inplace=True)

        data_formXX['S.no'] = list(range(1,len(data_formXX)+1))

        data_formXX['a'] ='---'
        data_formXX['b'] ='---'
        data_formXX['c'] ='---'
        data_formXX['d'] ='---'
        data_formXX['e'] ='---'
        data_formXX['f'] ='---'
        data_formXX['g'] ='---'
        data_formXX['h'] ='---'
        data_formXX['i'] =''

        formXX_columns = ['S.no','Employee Name',"Father's Name",'Designation','a','b','c','d','e','f','g','h','i']

        formXX_data = data_formXX[formXX_columns]

        formXXsheet = formXXfile['FORM XX']

        formXXsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XX is ready')

        
        rows = dataframe_to_rows(formXX_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 15):
            for c_idx, value in enumerate(row, 3):
                formXXsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXsheet['C6'].value
        C6_data = contractline+' '+contractor_name+', '+contractor_address
        formXXsheet['C6'] = C6_data

        if str(data_formXX['Nature of work'].dtype)[0:3] != 'obj':
            data_formXX['Nature of work'] = data_formXX['Nature of work'].astype(str)

        if str(data_formXX['Location'].dtype)[0:3] != 'obj':
            data_formXX['Location'] = data_formXX['Location'].astype(str)

        if str(data_formXX['Company Name'].dtype)[0:3] != 'obj':
            data_formXX['Company Name'] = data_formXX['Company Name'].astype(str)

        if str(data_formXX['Company Address'].dtype)[0:3] != 'obj':
            data_formXX['Company Address'] = data_formXX['Company Address'].astype(str)

        if str(data_formXX['Unit'].dtype)[0:3] != 'obj':
            data_formXX['Unit'] = data_formXX['Unit'].astype(str)

        if str(data_formXX['Address'].dtype)[0:3] != 'obj':
            data_formXX['Address'] = data_formXX['Address'].astype(str)

        locationline = formXXsheet['C7'].value
        C7_data = locationline+' '+data_formXX['Nature of work'][0]+', '+data_formXX['Location'][0]
        formXXsheet['C7'] = C7_data

        establine = formXXsheet['C8'].value
        if data_formXX['PE_or_contract'][0]== 'PE':
            C8_data = establine+' '+data_formXX['Company Name'][0]+', '+data_formXX['Company Address'][0]
        else:
            C8_data = establine+' '+data_formXX['Unit'][0]+', '+data_formXX['Address'][0]
        formXXsheet['C8'] = C8_data

        peline = formXXsheet['C9'].value
        if data_formXX['PE_or_contract'][0]== 'PE':
            C9_data = peline+' '+data_formXX['Company Name'][0]+', '+data_formXX['Company Address'][0]
        else:
            C9_data = peline+' '+data_formXX['Unit'][0]+', '+data_formXX['Address'][0]
        formXXsheet['C9'] = C9_data

        #border the region
        count1 = len(data_formXX)
        border_1 = Side(style='thick')
        for i in range(2,17):
            formXXsheet.cell(row=2, column=i).border = Border(outline= True, top=border_1)
            formXXsheet.cell(row=count1+16, column=i).border = Border(outline= True, bottom=border_1)
        for i in range(2,count1+17):
            formXXsheet.cell(row=i, column=2).border = Border(outline= True, left=border_1)
            formXXsheet.cell(row=i, column=16).border = Border(outline= True, right=border_1)

        formXXfinalfile = os.path.join(filelocation,'FormXX.xlsx')
        formXXfile.save(filename=formXXfinalfile)

    def create_wages():
        wagesfilepath = os.path.join(karnatakafilespath,'Wages.xlsx')
        wagesfile = load_workbook(filename=wagesfilepath)
        logging.info('wages file has sheet: '+str(wagesfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_wages = data.drop_duplicates(subset=['Employee Code']).copy()
        data_wages.fillna(value=0, inplace=True)

        data_wages['S.no'] = list(range(1,len(data_wages)+1))


        data_wages['fixed_wage'] = '---'
        #data_wages['OT hours'] = 0

        if str(data_wages['Earned Basic'].dtype)[0:3] != 'int':
            data_wages['Earned Basic']= data_wages['Earned Basic'].astype(int)
        if str(data_wages['DA'].dtype)[0:3] != 'int':
            data_wages['DA']= data_wages['DA'].astype(int)
        data_wages['basic_and_allo'] = data_wages['Earned Basic']+ data_wages['DA']
        data_wages['NFH'] = '---'
        data_wages['maturity'] = '---'
        data_wages['Sub Allow'] = '---'
        data_wages['Society'] = '---'
        data_wages['Fines']= '---'
        data_wages['Damages']= '---'
        data_wages['Pay mode'] = 'Bank Transfer'
        data_wages['Remarks'] =''

        wages_columns = ['S.no','Employee Code','Employee Name',"Father's Name",'Gender','Designation','Department','Address','Date Joined','ESIC Number','PF Number','fixed_wage','Days Paid','Total\r\nOT Hrs','basic_and_allo','HRA','Conveyance','Medical Allowance','Tel and Int Reimb', 'Bonus', 'Fuel Reimb',	'Prof Dev Reimb', 'Corp Attire Reimb','Special Allowance','Overtime','NFH','maturity','Other Reimb', 'CCA', 'Medical Allowance', 'Telephone Reimb', 'Other Allowance', 'Meal Allowance',
       'Special Allowance', 'Personal Allowance', 'Arrears', 'Other Earning', 'Variable Pay','Stipend','Sub Allow','Leave Encashment', 'Total Earning','ESIC', 'PF','P.Tax','TDS','Society','Insurance','Salary Advance','Fines','Damages','Other Deduction',	'Total Deductions',	'Net Paid','Pay mode','Bank A/c Number','Remarks']

        wages_data = data_wages[wages_columns]

        wagessheet = wagesfile['Wages']

        wagessheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for wages is ready')

        
        rows = dataframe_to_rows(wages_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 18):
            for c_idx, value in enumerate(row, 1):
                wagessheet.cell(row=r_idx, column=c_idx, value=value)
                wagessheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                wagessheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                wagessheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                if c_idx==56:
                    wagessheet.cell(row=r_idx, column=c_idx).number_format= numbers.FORMAT_NUMBER

        contractline = wagessheet['A10'].value
        A10_data = contractline+' '+contractor_name+', '+contractor_address
        wagessheet['A10'] = A10_data

        if str(data_wages['Nature of work'].dtype)[0:3] != 'obj':
            data_wages['Nature of work'] = data_wages['Nature of work'].astype(str)

        if str(data_wages['Location'].dtype)[0:3] != 'obj':
            data_wages['Location'] = data_wages['Location'].astype(str)

        if str(data_wages['Company Name'].dtype)[0:3] != 'obj':
            data_wages['Company Name'] = data_wages['Company Name'].astype(str)

        if str(data_wages['Company Address'].dtype)[0:3] != 'obj':
            data_wages['Company Address'] = data_wages['Company Address'].astype(str)

        if str(data_wages['Unit'].dtype)[0:3] != 'obj':
            data_wages['Unit'] = data_wages['Unit'].astype(str)

        if str(data_wages['Address'].dtype)[0:3] != 'obj':
            data_wages['Address'] = data_wages['Address'].astype(str)

        locationline = wagessheet['A11'].value
        A11_data = locationline+' '+data_wages['Nature of work'][0]+', '+data_wages['Location'][0]
        wagessheet['A11'] = A11_data

        establine = wagessheet['A12'].value
        if data_wages['PE_or_contract'][0]== 'PE':
            A12_data = establine+' '+data_wages['Company Name'][0]+', '+data_wages['Company Address'][0]
        else:
            A12_data = establine+' '+data_wages['Unit'][0]+', '+data_wages['Address'][0]
        wagessheet['A12'] = A12_data

        peline = wagessheet['A13'].value
        if data_wages['PE_or_contract'][0]== 'PE':
            A13_data = peline+' '+data_wages['Company Name'][0]+', '+data_wages['Company Address'][0]
        else:
            A13_data = peline+' '+data_wages['Unit'][0]+', '+data_wages['Address'][0]
        wagessheet['A13'] = A13_data

        wagessheet['F4'] = 'Combined Muster Roll-cum-Register of Wages in lieu of '+month+' '+str(year)

        wagesfinalfile = os.path.join(filelocation,'Wages.xlsx')
        wagesfile.save(filename=wagesfinalfile)

    
    def create_form_H_F(form):
        if form=='FORM H':
            formHfilepath = os.path.join(karnatakafilespath,'FormH.xlsx')
        if form=='FORM F':
            formHfilepath = os.path.join(karnatakafilespath,'FormF.xlsx')
        formHfile = load_workbook(filename=formHfilepath)
        logging.info('file has sheet: '+str(formHfile.sheetnames))
        sheetformh = formHfile[form]

        
        logging.info('create columns which are now available')

        data_formH = data[data['Leave Type']=='PL'].copy()
        data_formH.fillna(value=0, inplace=True)

        def attandance_data(employee_attendance,i):

            leavelist = list(employee_attendance.columns[(employee_attendance=='PL').iloc[i]])
            empcodeis = employee_attendance.iloc[i]['Employee Code']
            logging.info(empcodeis)
            if 'Leave Type' in leavelist:
                leavelist.remove('Leave Type')
            emp1 = pd.DataFrame(leavelist)
            
            
            if len(emp1.index)==0:
                defaultemp = {'emp':(employee_attendance).iloc[i]['Employee Code'],'startdate':0,'enddate':0,'days':0,'start_date':'-------','end_date':'-------'}
                emp1 = pd.DataFrame(defaultemp, index=[0])
                emp1.index = np.arange(1, len(emp1) + 1)
                emp1['s.no'] = emp1.index
                emp1.reset_index(drop=True, inplace=True)
                emp1['from'] = datetime.date(year,month_num,1)
                emp1['to'] = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
                emp1['totaldays'] = (employee_attendance).iloc[i]['Days Paid']
                emp1['leavesearned'] = float(employee_attendance.iloc[i]['Monthly Increment'])
                emp1['leavesstart'] = float(employee_attendance.iloc[i]['Opening'])
                emp1['leavesend'] = float(employee_attendance.iloc[i]['Closing'])
                emp1['Date of Payment and fixed'] = str(employee_attendance.iloc[i]['Date of payment'])+' and '+str(employee_attendance.iloc[i]['FIXED MONTHLY GROSS'])
                emp1['a']='---'
                emp1['b']='---'
                emp1['c']='---'
                emp1['d']='---'
                emp1['e']='---'
                emp1['f']='---'
                emp1 = emp1[["s.no","from","to","totaldays","leavesearned","leavesstart","start_date","end_date","days","leavesend","Date of Payment and fixed",'a','b','c','d','e','f']]
            else:
                logging.info(emp1)
                emp1.columns = ['Leaves']
                emp1['emp'] = (employee_attendance).iloc[i]['Employee Code']
                emp1['Leavesdays'] = emp1.Leaves.str[5:7].astype(int)
                emp1['daysdiff'] = (emp1.Leavesdays.shift(-1) - emp1.Leavesdays).fillna(0).astype(int)
                emp1['startdate'] = np.where(emp1.daysdiff.shift() != 1, emp1.Leavesdays, 0)
                emp1['enddate'] = np.where(emp1.daysdiff!=1, emp1.Leavesdays, 0)
                emp1.drop(emp1[(emp1.startdate==0) & (emp1.enddate==0)].index, inplace=True)
                emp1['startdate'] = np.where(emp1.startdate ==0, emp1.startdate.shift(), emp1.startdate).astype(int)
                emp1['enddate'] = np.where(emp1.enddate ==0, emp1.enddate.shift(-1), emp1.enddate).astype(int)
                emp1 = emp1[['emp','startdate','enddate']]
                emp1.drop_duplicates(subset='startdate', inplace=True)
                emp1['days'] = emp1.enddate -emp1.startdate +1
                emp1['start_date'] = [datetime.date(year,month_num,x) for x in emp1.startdate]
                emp1['end_date'] = [datetime.date(year,month_num,x) for x in emp1.enddate]
                emp1.index = np.arange(1, len(emp1) + 1)
                emp1['s.no'] = emp1.index
                emp1.reset_index(drop=True, inplace=True)
                emp1['from'] = datetime.date(year,month_num,1)
                emp1['to'] = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
                emp1['totaldays'] = (employee_attendance).iloc[i]['Days Paid']
                emp1['leavesearned'] = float((employee_attendance).iloc[i]['Monthly Increment'])
                emp1['totalleaves']= float(employee_attendance.iloc[i]['Opening'])
                emp1['leavesend'] = float(employee_attendance.iloc[i]['Closing'])
                emp1['leavesstart'] =emp1['totalleaves']
                emp1['Date of Payment and fixed'] = str(employee_attendance.iloc[i]['Date of payment'])+' and '+str(employee_attendance.iloc[i]['FIXED MONTHLY GROSS'])
                emp1['a']='---'
                emp1['b']='---'
                emp1['c']='---'
                emp1['d']='---'
                emp1['e']='---'
                emp1['f']='---'
                emp1 = emp1[["s.no","from","to","totaldays","leavesearned","leavesstart","start_date","end_date","days","leavesend","Date of Payment and fixed",'a','b','c','d','e','f']]
                
            
            return emp1

        def prepare_emp_sheet(emp1,sheet_key,key,name,fathername):
            
            sheet1 = formHfile.copy_worksheet(sheetformh)
            sheet1.title = sheet_key
            lastline = sheet1['B18'].value
            sheet1['B18'] =''

            if len(emp1)>3:
                lastlinerow = 'B'+str(18+len(emp1))
            else:
                lastlinerow = 'B18'

            
            logging.info(lastlinerow)
            sheet1[lastlinerow] = lastline

            
            
            rows = dataframe_to_rows(emp1, index=False, header=False)

            for r_idx, row in enumerate(rows, 14):
                for c_idx, value in enumerate(row, 2):
                    sheet1.cell(row=r_idx, column=c_idx, value=value)
                    sheet1.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                    sheet1.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    sheet1.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            sheet1['H5']=key
            sheet1['F7']=name
            sheet1['F8']=fathername

            sheet1.sheet_properties.pageSetUpPr.fitToPage = True


        emp_count = len(data_formH.index)
        emp_dic = dict()
        for i in range(0,emp_count):
            key = (data_formH).iloc[i]['Employee Code']
            emp_dic[key] = attandance_data(data_formH,i)
            sheet_key = form+'_'+str(key)
            name= data_formH[data_formH['Employee Code']==key]['Employee Name'].values[0]
            fathername= data_formH[data_formH['Employee Code']==key]["Father's Name"].values[0]
            logging.info(name)
            logging.info(fathername)
            prepare_emp_sheet(emp_dic[key],sheet_key,key,name,fathername)
            logging.info(key)
            logging.info(sheet_key)
        if form=='FORM H':
            formHfinalfile = os.path.join(filelocation,'FormH.xlsx')
        if form=='FORM F':
            formHfinalfile = os.path.join(filelocation,'FormF.xlsx')
        
        formHfile.remove(sheetformh)
        formHfile.save(filename=formHfinalfile)

    
    def create_muster():

        musterfilepath = os.path.join(karnatakafilespath,'Muster.xlsx')
        musterfile = load_workbook(filename=musterfilepath)
        logging.info('muster file has sheet: '+str(musterfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_muster = data.drop_duplicates(subset=['Employee Code']).copy()
        data_muster.fillna(value=0, inplace=True)

        data_muster['S.no'] = list(range(1,len(data_muster)+1))

        data_muster['datelast'] ='---'

        first3columns = ["S.no",'Employee Code','Employee Name']
        last2columns = ["datelast","Days Paid"]

        columnstotake =[]
        days = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31']
        for day in days:
            for col in data_muster.columns:
                if col[5:7]==day:
                    columnstotake.append(col)
        if len(columnstotake)==28:

            columnstotake.append('29')
            columnstotake.append('30')
            columnstotake.append('31')
            data_muster['29'] = ''
            data_muster['30'] = ''
            data_muster['31'] = ''
            
        elif len(columnstotake)==29:
            columnstotake.append('30')
            columnstotake.append('31')
            data_muster['30'] = ''
            data_muster['31'] = ''

        elif len(columnstotake)==30:
            columnstotake.append('31')
            data_muster['31'] = ''

        muster_columns = first3columns+columnstotake+last2columns

        muster_data = data_muster[muster_columns]

        mustersheet = musterfile['Muster']

        mustersheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for muster is ready')

        
        rows = dataframe_to_rows(muster_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 18):
            for c_idx, value in enumerate(row, 2):
                mustersheet.cell(row=r_idx, column=c_idx, value=value)
                mustersheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                mustersheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                mustersheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        logging.info('')

        contractline = mustersheet['B10'].value
        B10_data = contractline+' '+contractor_name+', '+contractor_address
        mustersheet['B10'] = B10_data

        if str(data_muster['Nature of work'].dtype)[0:3] != 'obj':
            data_muster['Nature of work'] = data_muster['Nature of work'].astype(str)

        if str(data_muster['Location'].dtype)[0:3] != 'obj':
            data_muster['Location'] = data_muster['Location'].astype(str)

        if str(data_muster['Company Name'].dtype)[0:3] != 'obj':
            data_muster['Company Name'] = data_muster['Company Name'].astype(str)

        if str(data_muster['Company Address'].dtype)[0:3] != 'obj':
            data_muster['Company Address'] = data_muster['Company Address'].astype(str)

        if str(data_muster['Unit'].dtype)[0:3] != 'obj':
            data_muster['Unit'] = data_muster['Unit'].astype(str)

        if str(data_muster['Address'].dtype)[0:3] != 'obj':
            data_muster['Address'] = data_muster['Address'].astype(str)

        locationline = mustersheet['B11'].value
        B11_data = locationline+' '+data_muster['Nature of work'][0]+', '+data_muster['Location'][0]
        mustersheet['B11'] = B11_data

        establine = mustersheet['B12'].value
        if data_muster['PE_or_contract'][0]== 'PE':
            B12_data = establine+' '+data_muster['Company Name'][0]+', '+data_muster['Company Address'][0]
        else:
            B12_data = establine+' '+data_muster['Unit'][0]+', '+data_muster['Address'][0]
        mustersheet['B12'] = B12_data

        peline = mustersheet['B13'].value
        if data_muster['PE_or_contract'][0]== 'PE':
            B13_data = peline+' '+data_muster['Company Name'][0]+', '+data_muster['Company Address'][0]
        else:
            B13_data = peline+' '+data_muster['Unit'][0]+', '+data_muster['Address'][0]
        mustersheet['B13'] = B13_data

        mustersheet['B4'] = 'Combined Muster Roll-cum-Register of Wages in lieu of '+month+' '+str(year)

        musterfinalfile = os.path.join(filelocation,'Muster.xlsx')
        musterfile.save(filename=musterfinalfile)

    def create_formXIX():

        formXIXfilepath = os.path.join(karnatakafilespath,'FormXIX.xlsx')
        formXIXfile = load_workbook(filename=formXIXfilepath)
        logging.info('Form XIX file has sheet: '+str(formXIXfile.sheetnames))
        sheetformXIX = formXIXfile['FORM XIX']

        
        logging.info('create columns which are now available')

        data_formXIX = data.drop_duplicates(subset=['Employee Code']).copy()

        data_formXIX.fillna(value=0, inplace=True)

        emp_count = len(data_formXIX.index)
        
        for i in range(0,emp_count):
            key = (data_formXIX).iloc[i]['Employee Code']
            sheet_key = 'FORM XIX_'+str(key)

            emp_data = (data_formXIX).iloc[i]

            sheet1 = formXIXfile.copy_worksheet(sheetformXIX)
            sheet1.title = sheet_key
            sheet1['D7'] = contractor_name+', '+contractor_address
            sheet1['D8'] = emp_data['Nature of work']+', '+emp_data['Location']
            if emp_data['PE_or_contract'][0]== 'PE':
                sheet1['D9'] = emp_data['Company Name']+', '+emp_data['Company Address']
                sheet1['D10'] = emp_data['Company Name']+', '+emp_data['Company Address']
            else:
                sheet1['D9'] = emp_data['Unit']+', '+emp_data['Address']
                sheet1['D10'] = emp_data['Unit']+', '+emp_data['Address']
            sheet1['D11'] = emp_data['Employee Name']
            sheet1['D12'] = emp_data['Gender']
            sheet1['D13'] = month+'-'+str(year)
            sheet1['D14'] = key
            sheet1['D15'] = emp_data['Days Paid']
            sheet1['D16'] = emp_data['Earned Basic']
            sheet1['D17'] = emp_data['DA']
            sheet1['D18'] = emp_data['HRA']
            sheet1['D19'] = emp_data['Tel and Int Reimb']
            sheet1['D20'] = emp_data['Bonus']
            sheet1['D21'] = emp_data['Fuel Reimb']
            sheet1['D22'] = emp_data['Corp Attire Reimb']
            sheet1['D23'] = emp_data['CCA']
            sheet1['D24'] = emp_data['Conveyance']+emp_data['Medical Allowance']+emp_data['Telephone Reimb']+emp_data['Other Allowance']+emp_data['Prof Dev Reimb']+emp_data['Meal Allowance']+emp_data['Special Allowance']+emp_data['Personal Allowance']+emp_data['Other Reimb']+emp_data['Arrears']+emp_data['Variable Pay']+emp_data['Other Earning']+emp_data['Leave Encashment']+emp_data['Stipend']
            sheet1['D25'] = emp_data['Total Earning']
            sheet1['D26'] = emp_data['Insurance']
            sheet1['D27'] = emp_data['PF']
            sheet1['D28'] = emp_data['ESIC']
            sheet1['D29'] = emp_data['P.Tax']
            sheet1['D30'] = emp_data['TDS']
            sheet1['D31'] = emp_data['CSR']+emp_data['VPF']+emp_data['LWF EE']+emp_data['Salary Advance']+emp_data['Loan Deduction']+emp_data['Loan Interest']+emp_data['Other Deduction']
            sheet1['D32'] = emp_data['Total Deductions']
            sheet1['D33'] = emp_data['Net Paid']

        formXIXfinalfile = os.path.join(filelocation,'FormXIX.xlsx')
        formXIXfile.remove(sheetformXIX)
        formXIXfile.save(filename=formXIXfinalfile)

    def create_ecard():

        ecardfilepath = os.path.join(karnatakafilespath,'Employment card.xlsx')
        ecardfile = load_workbook(filename=ecardfilepath)
        logging.info('Employment card file has sheet: '+str(ecardfile.sheetnames))
        sheetecard = ecardfile['Employment card']

        
        logging.info('create columns which are now available')

        data_ecard = data.drop_duplicates(subset=['Employee Code']).copy()
        data_ecard.fillna(value=0, inplace=True)

        emp_count = len(data_ecard.index)
        
        for i in range(0,emp_count):
            key = (data_ecard).iloc[i]['Employee Code']
            sheet_key = 'Employment card_'+str(key)

            emp_data = (data_ecard).iloc[i]
            emp_data.fillna(value='', inplace=True)

            sheet1 = ecardfile.copy_worksheet(sheetecard)
            sheet1.title = sheet_key
            sheet1['B4'] = contractor_name
            sheet1['B5'] = str(emp_data['Contractor_LIN'])+' / '+str(emp_data['Contractor_PAN'])
            sheet1['B6'] = emp_data['Contractor_email']
            sheet1['B7'] = emp_data['Contractor_mobile']
            sheet1['B7'].number_format= numbers.FORMAT_NUMBER
            sheet1['B8'] = emp_data['Nature of work']
            sheet1['B9'] = contractor_address
            if emp_data['PE_or_contract'][0]== 'PE':
                sheet1['B10'] = emp_data['Company Name']
            else:
                sheet1['B10'] = emp_data['Unit']
            sheet1['B11'] = str(emp_data['Unit_LIN'])+' / '+str(emp_data['Unit_PAN'])
            sheet1['B12'] = emp_data['Unit_email']
            sheet1['B13'] = emp_data['Unit_mobile']
            sheet1['B13'].number_format= numbers.FORMAT_NUMBER
            sheet1['B14'] = emp_data['Employee Name']
            sheet1['B15'] = emp_data['Aadhar Number']
            sheet1['B15'].number_format= numbers.FORMAT_NUMBER
            sheet1['B16'] = emp_data['Mobile Tel No.']
            sheet1['B16'].number_format= numbers.FORMAT_NUMBER
            sheet1['B17'] = key
            sheet1['B18'] = emp_data['Designation']
            sheet1['B19'] = emp_data['FIXED MONTHLY GROSS']
            sheet1['B20'] = emp_data['Date Joined']
            sheet1['B21'] = '-'
            

        ecardfinalfile = os.path.join(filelocation,'Employment card.xlsx')
        ecardfile.remove(sheetecard)
        ecardfile.save(filename=ecardfinalfile)
            

    create_form_A()
    create_form_B()
    create_form_XXI()
    create_form_XXII()
    create_form_XXIII()
    create_form_XX()
    create_wages()
    create_form_H_F('FORM H')
    create_form_H_F('FORM F')
    create_muster()
    create_formXIX()
    create_ecard()
    

def West_Bengal(data,contractor_name,contractor_address,filelocation,month,year):
    logging.info("West_Bengal form creation")

def Uttarakhand(data,contractor_name,contractor_address,filelocation,month,year):
    logging.info("Uttarakhand form creation")

def Contractor_Process(data,filelocation,month,year):
    Contractorfilespath = os.path.join(Statefolder,'CLRA')
    logging.info('Contractor files path is :'+str(Contractorfilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]

    def create_form_A():
    
        formAfilepath = os.path.join(Contractorfilespath,'Form A Employee register.xlsx')
        formAfile = load_workbook(filename=formAfilepath)
        logging.info('Form A file has sheet: '+str(formAfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formA = data.drop_duplicates(subset=['Employee Code']).copy()

        data_formA.fillna(value=0, inplace=True)

        data_formA['S.no'] = list(range(1,len(data_formA)+1))

        formA_columns = ["S.no",'Employee Code','Employee Name',"Gender","Father's Name",'Date of Birth',"Nationality","Education Level",'Date Joined',
                        'Designation','Category Address',"Type of Employment",'Mobile Tel No.','UAN Number',"PAN Number","ESIC Number","LWF EE","Aadhar Number",
                        "Bank A/c Number","Bank Name",'Branch',"Present_Address","Permanent_Address",'Service Book No',"Date Left","Reason for Leaving","Identification mark"
                        "photo","sign","remarks"]
        data_formA['Local Address 1', 'Local Address 2','Local Address 3', 'Local Address 4'].astype(str,inplace=True)
        data_formA['Permanent Address 1', 'Permanent Address 2','Permanent Address 3', 'Permanent Address 4'].astype(str,inplace=True)
        data_formA["Present_Address"]=data_formA['Local Address 1']+data_formA['Local Address 2']+data_formA['Local Address 3']+ data_formA['Local Address 4']
        data_formA["Permanent_Address"]=data_formA['Permanent Address 1']+data_formA['Permanent Address 2']+data_formA['Permanent Address 3']+ data_formA['Permanent Address 4']
        
        formA_data = data_formA[formA_columns]


        formAsheet = formAfile['Sheet1']

        formAsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form A is ready')

        
        rows = dataframe_to_rows(formA_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 11):
            for c_idx, value in enumerate(row, 1):
                formAsheet.cell(row=r_idx, column=c_idx, value=value)
                formAsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formAsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formAsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                if c_idx==15 or c_idx==16 or c_idx==20 or c_idx==21:
                    formAsheet.cell(row=r_idx, column=c_idx).number_format= numbers.FORMAT_NUMBER

        logging.info('')


        if str(data_formA['Company Name'].dtype)[0:3] != 'obj':
            data_formA['Company Name'] = data_formA['Company Name'].astype(str)

        if str(data_formA['Company Address'].dtype)[0:3] != 'obj':
            data_formA['Company Address'] = data_formA['Company Address'].astype(str)

        if str(data_formA['Contractor_name'].dtype)[0:3] != 'obj':
            data_formA['Contractor_name'] = data_formA['Contractor_name'].astype(str)

        if str(data_formA['Contractor_Address'].dtype)[0:3] != 'obj':
            data_formA['Contractor_Address'] = data_formA['Contractor_Address'].astype(str)

        if str(data_formA['Unit'].dtype)[0:3] != 'obj':
            data_formA['Unit'] = data_formA['Unit'].astype(str)

        if str(data_formA['Branch'].dtype)[0:3] != 'obj':
            data_formA['Branch'] = data_formA['Branch'].astype(str)

        establishment = formAsheet['A5'].value
        # print(establishment)
        # print(data_formA['Contractor_name'][0])
        # print(data_formA['Contractor_Address'][0])
        if data_formA['PE_or_contract'][0] == 'PE':
            A5_data = establishment+' '+data_formA['Company Name'][0] +', '+data_formA['Company Address'][0]  
        else:    
            A5_data = establishment+' '+data_formA['Contractor_name'][0]+', '+data_formA['Contractor_Address'][0]
        formAsheet['A5'] = A5_data


        # company = formAsheet['A10'].value
        # A10_data = company+' '+data_formA['Unit'][0]+', '+data_formA['Branch'][0]
        # formAsheet['A10'] = A10_data

        
        formAfinalfile = os.path.join(filelocation,'Form A Employee register.xlsx')
        logging.info('Form A file is' +str(formAfinalfile))
        formAfile.save(filename=formAfinalfile)
        

    def create_form_B():
        formBfilepath = os.path.join(karnatakafilespath,'Form B wage register equal remuniration.xlsx')
        formBfile = load_workbook(filename=formBfilepath)
        logging.info('Form B file has sheet: '+str(formBfile.sheetnames))

        
        logging.info('create columns which are now available')


        data_formB = data.drop_duplicates(subset=['Employee Code']).copy()
        data_formB.fillna(value=0, inplace=True)

        #data_formB['OT hours'] = 0
        #data_formB['Pay OT'] = 0
        if str(data_formB['Earned Basic'].dtype)[0:3] != 'int':
            data_formB['Earned Basic']= data_formB['Earned Basic'].astype(int)
        if str(data_formB['DA'].dtype)[0:3] != 'int':
            data_formB['DA']= data_formB['DA'].astype(int)
        data_formB['basic_and_allo'] = data_formB['Earned Basic']+ data_formB['DA']
        #data_formB['Other EAR'] = data_formB['Other Reimb']+data_formB['Arrears']+data_formB['Other Earning']+data_formB['Variable Pay']+data_formB['Stipend'] +data_formB['Consultancy Fees']
        #data_formB['VPF']=0
        data_formB['Society']="---"
        data_formB['Income Tax']="---"
        if str(data_formB['Other Deduction'].dtype)[0:3] != 'int':
            data_formB['Other Deduction']= data_formB['Other Deduction'].astype(int)
        if str(data_formB['Other Deduction'].dtype)[0:3] != 'int':
            data_formB['Other Deduction']= data_formB['Other Deduction'].astype(int)
        data_formB['Other Deduc']= data_formB['Other Deduction']+ data_formB['Salary Advance']
        data_formB['EMP PF'] = data_formB['PF']
        #data_formB['BankID'] = ''
        #data_formB['Pay Date'] = ''
        data_formB['Remarks'] =''

        formB_columns = ['Employee Code','Employee Name','FIXED MONTHLY GROSS',	'Days Paid','Total\r\nOT Hrs',	'basic_and_allo', 'Overtime',	
                        'HRA',	'Tel and Int Reimb', 'Bonus', 'Fuel Reimb',	'Prof Dev Reimb', 'Corp Attire Reimb',	'CCA',	'Leave Encashment',
                        'Conveyance', 'Medical Allowance', 'Telephone Reimb', 'Other Allowance', 'Meal Allowance',
                        'Special Allowance', 'Personal Allowance','Other Reimb', 'Arrears', 'Other Earning', 'Variable Pay','Stipend' ,
                        'Total Earning', 'PF',	'ESIC',	'VPF', 'Loan Deduction', 'Loan Interest', 'P.Tax',	'Society',	'Income Tax', 
                        'Insurance',	'LWF EE',	'Other Deduction',	'TDS',	'Total Deductions',	'Net Paid',	'EMP PF','Bank A/c Number','Date of payment','Remarks']

        formB_data = data_formB[formB_columns]

        formBsheet = formBfile['FORM B']

        formBsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form B is ready')

        
        rows = dataframe_to_rows(formB_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 21):
            for c_idx, value in enumerate(row, 2):
                formBsheet.cell(row=r_idx, column=c_idx, value=value)
                formBsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formBsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formBsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                if c_idx==45:
                    formBsheet.cell(row=r_idx, column=c_idx).number_format= numbers.FORMAT_NUMBER

        contractline = formBsheet['B10'].value
        B10_data = contractline+' '+contractor_name+', '+contractor_address
        formBsheet['B10'] = B10_data

        if str(data_formB['Nature of work'].dtype)[0:3] != 'obj':
            data_formB['Nature of work'] = data_formB['Nature of work'].astype(str)

        if str(data_formB['Location'].dtype)[0:3] != 'obj':
            data_formB['Location'] = data_formB['Location'].astype(str)

        if str(data_formB['Company Name'].dtype)[0:3] != 'obj':
            data_formB['Company Name'] = data_formB['Company Name'].astype(str)

        if str(data_formB['Company Address'].dtype)[0:3] != 'obj':
            data_formB['Company Address'] = data_formB['Company Address'].astype(str)

        if str(data_formB['Unit'].dtype)[0:3] != 'obj':
            data_formB['Unit'] = data_formB['Unit'].astype(str)

        if str(data_formB['Address'].dtype)[0:3] != 'obj':
            data_formB['Address'] = data_formB['Address'].astype(str)

        locationline = formBsheet['B11'].value
        B11_data = locationline+' '+data_formB['Nature of work'][0]+', '+data_formB['Location'][0]
        formBsheet['B11'] = B11_data

        establine = formBsheet['B12'].value
        if data_formB['PE_or_contract'][0]== 'PE':
            B12_data = establine+' '+data_formB['Company Name'][0]+', '+data_formB['Company Address'][0]  
        else:    
            B12_data = establine+' '+data_formB['Unit'][0]+', '+data_formB['Address'][0]
        formBsheet['B12'] = B12_data

        peline = formBsheet['B13'].value
        if data_formB['PE_or_contract'][0]== 'PE':
            B13_data = peline+' '+data_formB['Company Name'][0]+', '+data_formB['Company Address'][0]  
        else:    
            B13_data = peline+' '+data_formB['Unit'][0]+', '+data_formB['Address'][0]
        formBsheet['B13'] = B13_data

        monthstart = datetime.date(year,month_num,1)
        monthend = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
        formBsheet['B16'] = 'Wage period From: '+str(monthstart)+' to '+str(monthend)

        formBfinalfile = os.path.join(filelocation,'Form B wage register equal remuniration.xlsx')
        formBfile.save(filename=formBfinalfile)

    def create_form_XXI():
        formXXIfilepath = os.path.join(karnatakafilespath,'FormXXI.xlsx')
        formXXIfile = load_workbook(filename=formXXIfilepath)
        logging.info('Form XXI file has sheet: '+str(formXXIfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXXI = data.drop_duplicates(subset=['Employee Code']).copy()

        data_formXXI.fillna(value=0, inplace=True)

        data_formXXI['S.no'] = list(range(1,len(data_formXXI)+1))

        data_formXXI['a'] ='---'
        data_formXXI['b'] ='---'
        data_formXXI['c'] ='---'
        #data_formXXI['e'] ='---'
        data_formXXI['f'] ='---'
        data_formXXI['g'] =''

        formXXI_columns = ['S.no','Employee Name',"Father's Name",'Designation','a','b','c','FIXED MONTHLY GROSS','Fine','f','g']

        formXXI_data = data_formXXI[formXXI_columns]

        formXXIsheet = formXXIfile['FORM XXI']

        formXXIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XXI is ready')

        
        rows = dataframe_to_rows(formXXI_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 14):
            for c_idx, value in enumerate(row, 3):
                formXXIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXIsheet['C7'].value
        C7_data = contractline+' '+contractor_name+', '+contractor_address
        formXXIsheet['C7'] = C7_data

        if str(data_formXXI['Nature of work'].dtype)[0:3] != 'obj':
            data_formXXI['Nature of work'] = data_formXXI['Nature of work'].astype(str)

        if str(data_formXXI['Location'].dtype)[0:3] != 'obj':
            data_formXXI['Location'] = data_formXXI['Location'].astype(str)

        if str(data_formXXI['Company Name'].dtype)[0:3] != 'obj':
            data_formXXI['Company Name'] = data_formXXI['Company Name'].astype(str)

        if str(data_formXXI['Company Address'].dtype)[0:3] != 'obj':
            data_formXXI['Company Address'] = data_formXXI['Company Address'].astype(str)

        if str(data_formXXI['Unit'].dtype)[0:3] != 'obj':
            data_formXXI['Unit'] = data_formXXI['Unit'].astype(str)

        if str(data_formXXI['Address'].dtype)[0:3] != 'obj':
            data_formXXI['Address'] = data_formXXI['Address'].astype(str)

        locationline = formXXIsheet['C8'].value
        C8_data = locationline+' '+data_formXXI['Nature of work'][0]+', '+data_formXXI['Location'][0]
        formXXIsheet['C8'] = C8_data

        establine = formXXIsheet['C9'].value
        if data_formXXI['PE_or_contract'][0]== 'PE':
            C9_data = establine+' '+data_formXXI['Company Name'][0]+', '+data_formXXI['Company Address'][0]
        else:
            C9_data = establine+' '+data_formXXI['Unit'][0]+', '+data_formXXI['Address'][0]
        formXXIsheet['C9'] = C9_data

        peline = formXXIsheet['C10'].value
        if data_formXXI['PE_or_contract'][0]== 'PE':
            C10_data = peline+' '+data_formXXI['Company Name'][0]+', '+data_formXXI['Company Address'][0]
        else:
            C10_data = peline+' '+data_formXXI['Unit'][0]+', '+data_formXXI['Address'][0]
        formXXIsheet['C10'] = C10_data

        #border the region
        count1 = len(data_formXXI)
        border_1 = Side(style='thick')
        for i in range(2,15):
            formXXIsheet.cell(row=2, column=i).border = Border(outline= True, top=border_1)
            formXXIsheet.cell(row=count1+15, column=i).border = Border(outline= True, bottom=border_1)
        for i in range(2,count1+16):
            formXXIsheet.cell(row=i, column=2).border = Border(outline= True, left=border_1)
            formXXIsheet.cell(row=i, column=14).border = Border(outline= True, right=border_1)

        formXXIfinalfile = os.path.join(filelocation,'FormXXI.xlsx')
        formXXIfile.save(filename=formXXIfinalfile)



    def create_form_XXII():
        formXXIIfilepath = os.path.join(karnatakafilespath,'FormXXII.xlsx')
        formXXIIfile = load_workbook(filename=formXXIIfilepath)
        logging.info('Form XXII file has sheet: '+str(formXXIIfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXXII = data.drop_duplicates(subset=['Employee Code']).copy()

        data_formXXII.fillna(value=0, inplace=True)

        data_formXXII['S.no'] = list(range(1,len(data_formXXII)+1))

        data_formXXII['b'] ='---'
        data_formXXII['c'] ='---'
        data_formXXII['d'] ='---'
        data_formXXII['e'] ='---'
        data_formXXII['f'] ='---'
        data_formXXII['g'] =''

        formXXII_columns = ['S.no','Employee Name',"Father's Name",'Designation','FIXED MONTHLY GROSS','b','c','d','e','f','g']

        formXXII_data = data_formXXII[formXXII_columns]

        formXXIIsheet = formXXIIfile['FORM XXII']

        formXXIIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XXII is ready')

        
        rows = dataframe_to_rows(formXXII_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 15):
            for c_idx, value in enumerate(row, 3):
                formXXIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXIIsheet['C7'].value
        C7_data = contractline+' '+contractor_name+', '+contractor_address
        formXXIIsheet['C7'] = C7_data


        if str(data_formXXII['Nature of work'].dtype)[0:3] != 'obj':
            data_formXXII['Nature of work'] = data_formXXII['Nature of work'].astype(str)

        if str(data_formXXII['Location'].dtype)[0:3] != 'obj':
            data_formXXII['Location'] = data_formXXII['Location'].astype(str)

        if str(data_formXXII['Company Name'].dtype)[0:3] != 'obj':
            data_formXXII['Company Name'] = data_formXXII['Company Name'].astype(str)

        if str(data_formXXII['Company Address'].dtype)[0:3] != 'obj':
            data_formXXII['Company Address'] = data_formXXII['Company Address'].astype(str)

        if str(data_formXXII['Unit'].dtype)[0:3] != 'obj':
            data_formXXII['Unit'] = data_formXXII['Unit'].astype(str)

        if str(data_formXXII['Address'].dtype)[0:3] != 'obj':
            data_formXXII['Address'] = data_formXXII['Address'].astype(str)

        locationline = formXXIIsheet['C8'].value
        C8_data = locationline+' '+data_formXXII['Nature of work'][0]+', '+data_formXXII['Location'][0]
        formXXIIsheet['C8'] = C8_data

        establine = formXXIIsheet['C9'].value
        if data_formXXII['PE_or_contract'][0]== 'PE':
            C9_data = establine+' '+data_formXXII['Company Name'][0]+', '+data_formXXII['Company Address'][0]
        else:
            C9_data = establine+' '+data_formXXII['Unit'][0]+', '+data_formXXII['Address'][0]
        formXXIIsheet['C9'] = C9_data

        peline = formXXIIsheet['C10'].value
        if data_formXXII['PE_or_contract'][0]== 'PE':
            C10_data = peline+' '+data_formXXII['Company Name'][0]+', '+data_formXXII['Company Address'][0]
        else:
            C10_data = peline+' '+data_formXXII['Unit'][0]+', '+data_formXXII['Address'][0]
        formXXIIsheet['C10'] = C10_data

        #border the region
        count1 = len(data_formXXII)
        border_1 = Side(style='thick')
        for i in range(2,15):
            formXXIIsheet.cell(row=2, column=i).border = Border(outline= True, top=border_1)
            formXXIIsheet.cell(row=count1+16, column=i).border = Border(outline= True, bottom=border_1)
        for i in range(2,count1+17):
            formXXIIsheet.cell(row=i, column=2).border = Border(outline= True, left=border_1)
            formXXIIsheet.cell(row=i, column=14).border = Border(outline= True, right=border_1)

        formXXIIfinalfile = os.path.join(filelocation,'FormXXII.xlsx')
        formXXIIfile.save(filename=formXXIIfinalfile)


    def create_form_XXIII():
        formXXIIIfilepath = os.path.join(karnatakafilespath,'FormXXIII.xlsx')
        formXXIIIfile = load_workbook(filename=formXXIIIfilepath)
        logging.info('Form XXIII file has sheet: '+str(formXXIIIfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXXIII = data.drop_duplicates(subset=['Employee Code']).copy()
        data_formXXIII.fillna(value=0, inplace=True)

        data_formXXIII['S.no'] = list(range(1,len(data_formXXIII)+1))

        data_formXXIII['b'] ='---'
        data_formXXIII['c'] ='---'
        data_formXXIII['d'] ='---'
        data_formXXIII['e'] ='---'
        data_formXXIII['f'] ='---'
        data_formXXIII['g'] =''

        formXXIII_columns = ['S.no','Employee Name',"Father's Name",'Designation','FIXED MONTHLY GROSS','b','c','d','e','f','g']

        formXXIII_data = data_formXXIII[formXXIII_columns]

        formXXIIIsheet = formXXIIIfile['FORM XXIII']

        formXXIIIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XXIII is ready')

        
        rows = dataframe_to_rows(formXXIII_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 12):
            for c_idx, value in enumerate(row, 3):
                formXXIIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXIIIsheet['C5'].value
        C5_data = contractline+' '+contractor_name+', '+contractor_address
        formXXIIIsheet['C5'] = C5_data

        if str(data_formXXIII['Nature of work'].dtype)[0:3] != 'obj':
            data_formXXIII['Nature of work'] = data_formXXIII['Nature of work'].astype(str)

        if str(data_formXXIII['Location'].dtype)[0:3] != 'obj':
            data_formXXIII['Location'] = data_formXXIII['Location'].astype(str)

        if str(data_formXXIII['Company Name'].dtype)[0:3] != 'obj':
            data_formXXIII['Company Name'] = data_formXXIII['Company Name'].astype(str)

        if str(data_formXXIII['Company Address'].dtype)[0:3] != 'obj':
            data_formXXIII['Company Address'] = data_formXXIII['Company Address'].astype(str)

        if str(data_formXXIII['Unit'].dtype)[0:3] != 'obj':
            data_formXXIII['Unit'] = data_formXXIII['Unit'].astype(str)

        if str(data_formXXIII['Address'].dtype)[0:3] != 'obj':
            data_formXXIII['Address'] = data_formXXIII['Address'].astype(str)

        locationline = formXXIIIsheet['C6'].value
        C6_data = locationline+' '+data_formXXIII['Nature of work'][0]+', '+data_formXXIII['Location'][0]
        formXXIIIsheet['C6'] = C6_data

        establine = formXXIIIsheet['C7'].value
        if data_formXXIII['PE_or_contract'][0]== 'PE':
            C7_data = establine+' '+data_formXXIII['Company Name'][0]+', '+data_formXXIII['Company Address'][0]
        else:
            C7_data = establine+' '+data_formXXIII['Unit'][0]+', '+data_formXXIII['Address'][0]
        formXXIIIsheet['C7'] = C7_data

        peline = formXXIIIsheet['C8'].value
        if data_formXXIII['PE_or_contract'][0]== 'PE':
            C8_data = peline+' '+data_formXXIII['Company Name'][0]+', '+data_formXXIII['Company Address'][0]
        else:
            C8_data = peline+' '+data_formXXIII['Unit'][0]+', '+data_formXXIII['Address'][0]
        formXXIIIsheet['C8'] = C8_data


        #border the region
        count1 = len(data_formXXIII)
        border_1 = Side(style='thick')
        for i in range(2,15):
            formXXIIIsheet.cell(row=2, column=i).border = Border(outline= True, top=border_1)
            formXXIIIsheet.cell(row=count1+13, column=i).border = Border(outline= True, bottom=border_1)
        for i in range(2,count1+14):
            formXXIIIsheet.cell(row=i, column=2).border = Border(outline= True, left=border_1)
            formXXIIIsheet.cell(row=i, column=14).border = Border(outline= True, right=border_1)


        formXXIIIfinalfile = os.path.join(filelocation,'FormXXIII.xlsx')
        formXXIIIfile.save(filename=formXXIIIfinalfile)


    def create_form_XX():
        formXXfilepath = os.path.join(karnatakafilespath,'FormXX.xlsx')
        formXXfile = load_workbook(filename=formXXfilepath)
        logging.info('Form XX file has sheet: '+str(formXXfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXX = data.drop_duplicates(subset=['Employee Code']).copy()
        data_formXX.fillna(value=0, inplace=True)

        data_formXX['S.no'] = list(range(1,len(data_formXX)+1))

        data_formXX['a'] ='---'
        data_formXX['b'] ='---'
        data_formXX['c'] ='---'
        data_formXX['d'] ='---'
        data_formXX['e'] ='---'
        data_formXX['f'] ='---'
        data_formXX['g'] ='---'
        data_formXX['h'] ='---'
        data_formXX['i'] =''

        formXX_columns = ['S.no','Employee Name',"Father's Name",'Designation','a','b','c','d','e','f','g','h','i']

        formXX_data = data_formXX[formXX_columns]

        formXXsheet = formXXfile['FORM XX']

        formXXsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XX is ready')

        
        rows = dataframe_to_rows(formXX_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 15):
            for c_idx, value in enumerate(row, 3):
                formXXsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXsheet['C6'].value
        C6_data = contractline+' '+contractor_name+', '+contractor_address
        formXXsheet['C6'] = C6_data

        if str(data_formXX['Nature of work'].dtype)[0:3] != 'obj':
            data_formXX['Nature of work'] = data_formXX['Nature of work'].astype(str)

        if str(data_formXX['Location'].dtype)[0:3] != 'obj':
            data_formXX['Location'] = data_formXX['Location'].astype(str)

        if str(data_formXX['Company Name'].dtype)[0:3] != 'obj':
            data_formXX['Company Name'] = data_formXX['Company Name'].astype(str)

        if str(data_formXX['Company Address'].dtype)[0:3] != 'obj':
            data_formXX['Company Address'] = data_formXX['Company Address'].astype(str)

        if str(data_formXX['Unit'].dtype)[0:3] != 'obj':
            data_formXX['Unit'] = data_formXX['Unit'].astype(str)

        if str(data_formXX['Address'].dtype)[0:3] != 'obj':
            data_formXX['Address'] = data_formXX['Address'].astype(str)

        locationline = formXXsheet['C7'].value
        C7_data = locationline+' '+data_formXX['Nature of work'][0]+', '+data_formXX['Location'][0]
        formXXsheet['C7'] = C7_data

        establine = formXXsheet['C8'].value
        if data_formXX['PE_or_contract'][0]== 'PE':
            C8_data = establine+' '+data_formXX['Company Name'][0]+', '+data_formXX['Company Address'][0]
        else:
            C8_data = establine+' '+data_formXX['Unit'][0]+', '+data_formXX['Address'][0]
        formXXsheet['C8'] = C8_data

        peline = formXXsheet['C9'].value
        if data_formXX['PE_or_contract'][0]== 'PE':
            C9_data = peline+' '+data_formXX['Company Name'][0]+', '+data_formXX['Company Address'][0]
        else:
            C9_data = peline+' '+data_formXX['Unit'][0]+', '+data_formXX['Address'][0]
        formXXsheet['C9'] = C9_data

        #border the region
        count1 = len(data_formXX)
        border_1 = Side(style='thick')
        for i in range(2,17):
            formXXsheet.cell(row=2, column=i).border = Border(outline= True, top=border_1)
            formXXsheet.cell(row=count1+16, column=i).border = Border(outline= True, bottom=border_1)
        for i in range(2,count1+17):
            formXXsheet.cell(row=i, column=2).border = Border(outline= True, left=border_1)
            formXXsheet.cell(row=i, column=16).border = Border(outline= True, right=border_1)

        formXXfinalfile = os.path.join(filelocation,'FormXX.xlsx')
        formXXfile.save(filename=formXXfinalfile)

    def create_wages():
        wagesfilepath = os.path.join(karnatakafilespath,'Wages.xlsx')
        wagesfile = load_workbook(filename=wagesfilepath)
        logging.info('wages file has sheet: '+str(wagesfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_wages = data.drop_duplicates(subset=['Employee Code']).copy()
        data_wages.fillna(value=0, inplace=True)

        data_wages['S.no'] = list(range(1,len(data_wages)+1))


        data_wages['fixed_wage'] = '---'
        #data_wages['OT hours'] = 0

        if str(data_wages['Earned Basic'].dtype)[0:3] != 'int':
            data_wages['Earned Basic']= data_wages['Earned Basic'].astype(int)
        if str(data_wages['DA'].dtype)[0:3] != 'int':
            data_wages['DA']= data_wages['DA'].astype(int)
        data_wages['basic_and_allo'] = data_wages['Earned Basic']+ data_wages['DA']
        data_wages['NFH'] = '---'
        data_wages['maturity'] = '---'
        data_wages['Sub Allow'] = '---'
        data_wages['Society'] = '---'
        data_wages['Fines']= '---'
        data_wages['Damages']= '---'
        data_wages['Pay mode'] = 'Bank Transfer'
        data_wages['Remarks'] =''

        wages_columns = ['S.no','Employee Code','Employee Name',"Father's Name",'Gender','Designation','Department','Address','Date Joined','ESIC Number','PF Number','fixed_wage','Days Paid','Total\r\nOT Hrs','basic_and_allo','HRA','Conveyance','Medical Allowance','Tel and Int Reimb', 'Bonus', 'Fuel Reimb',	'Prof Dev Reimb', 'Corp Attire Reimb','Special Allowance','Overtime','NFH','maturity','Other Reimb', 'CCA', 'Medical Allowance', 'Telephone Reimb', 'Other Allowance', 'Meal Allowance',
       'Special Allowance', 'Personal Allowance', 'Arrears', 'Other Earning', 'Variable Pay','Stipend','Sub Allow','Leave Encashment', 'Total Earning','ESIC', 'PF','P.Tax','TDS','Society','Insurance','Salary Advance','Fines','Damages','Other Deduction',	'Total Deductions',	'Net Paid','Pay mode','Bank A/c Number','Remarks']

        wages_data = data_wages[wages_columns]

        wagessheet = wagesfile['Wages']

        wagessheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for wages is ready')

        
        rows = dataframe_to_rows(wages_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 18):
            for c_idx, value in enumerate(row, 1):
                wagessheet.cell(row=r_idx, column=c_idx, value=value)
                wagessheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                wagessheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                wagessheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                if c_idx==56:
                    wagessheet.cell(row=r_idx, column=c_idx).number_format= numbers.FORMAT_NUMBER

        contractline = wagessheet['A10'].value
        A10_data = contractline+' '+contractor_name+', '+contractor_address
        wagessheet['A10'] = A10_data

        if str(data_wages['Nature of work'].dtype)[0:3] != 'obj':
            data_wages['Nature of work'] = data_wages['Nature of work'].astype(str)

        if str(data_wages['Location'].dtype)[0:3] != 'obj':
            data_wages['Location'] = data_wages['Location'].astype(str)

        if str(data_wages['Company Name'].dtype)[0:3] != 'obj':
            data_wages['Company Name'] = data_wages['Company Name'].astype(str)

        if str(data_wages['Company Address'].dtype)[0:3] != 'obj':
            data_wages['Company Address'] = data_wages['Company Address'].astype(str)

        if str(data_wages['Unit'].dtype)[0:3] != 'obj':
            data_wages['Unit'] = data_wages['Unit'].astype(str)

        if str(data_wages['Address'].dtype)[0:3] != 'obj':
            data_wages['Address'] = data_wages['Address'].astype(str)

        locationline = wagessheet['A11'].value
        A11_data = locationline+' '+data_wages['Nature of work'][0]+', '+data_wages['Location'][0]
        wagessheet['A11'] = A11_data

        establine = wagessheet['A12'].value
        if data_wages['PE_or_contract'][0]== 'PE':
            A12_data = establine+' '+data_wages['Company Name'][0]+', '+data_wages['Company Address'][0]
        else:
            A12_data = establine+' '+data_wages['Unit'][0]+', '+data_wages['Address'][0]
        wagessheet['A12'] = A12_data

        peline = wagessheet['A13'].value
        if data_wages['PE_or_contract'][0]== 'PE':
            A13_data = peline+' '+data_wages['Company Name'][0]+', '+data_wages['Company Address'][0]
        else:
            A13_data = peline+' '+data_wages['Unit'][0]+', '+data_wages['Address'][0]
        wagessheet['A13'] = A13_data

        wagessheet['F4'] = 'Combined Muster Roll-cum-Register of Wages in lieu of '+month+' '+str(year)

        wagesfinalfile = os.path.join(filelocation,'Wages.xlsx')
        wagesfile.save(filename=wagesfinalfile)

    
    def create_form_H_F(form):
        if form=='FORM H':
            formHfilepath = os.path.join(karnatakafilespath,'FormH.xlsx')
        if form=='FORM F':
            formHfilepath = os.path.join(karnatakafilespath,'FormF.xlsx')
        formHfile = load_workbook(filename=formHfilepath)
        logging.info('file has sheet: '+str(formHfile.sheetnames))
        sheetformh = formHfile[form]

        
        logging.info('create columns which are now available')

        data_formH = data[data['Leave Type']=='PL'].copy()
        data_formH.fillna(value=0, inplace=True)

        def attandance_data(employee_attendance,i):

            leavelist = list(employee_attendance.columns[(employee_attendance=='PL').iloc[i]])
            empcodeis = employee_attendance.iloc[i]['Employee Code']
            logging.info(empcodeis)
            if 'Leave Type' in leavelist:
                leavelist.remove('Leave Type')
            emp1 = pd.DataFrame(leavelist)
            
            
            if len(emp1.index)==0:
                defaultemp = {'emp':(employee_attendance).iloc[i]['Employee Code'],'startdate':0,'enddate':0,'days':0,'start_date':'-------','end_date':'-------'}
                emp1 = pd.DataFrame(defaultemp, index=[0])
                emp1.index = np.arange(1, len(emp1) + 1)
                emp1['s.no'] = emp1.index
                emp1.reset_index(drop=True, inplace=True)
                emp1['from'] = datetime.date(year,month_num,1)
                emp1['to'] = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
                emp1['totaldays'] = (employee_attendance).iloc[i]['Days Paid']
                emp1['leavesearned'] = float(employee_attendance.iloc[i]['Monthly Increment'])
                emp1['leavesstart'] = float(employee_attendance.iloc[i]['Opening'])
                emp1['leavesend'] = float(employee_attendance.iloc[i]['Closing'])
                emp1['Date of Payment and fixed'] = str(employee_attendance.iloc[i]['Date of payment'])+' and '+str(employee_attendance.iloc[i]['FIXED MONTHLY GROSS'])
                emp1['a']='---'
                emp1['b']='---'
                emp1['c']='---'
                emp1['d']='---'
                emp1['e']='---'
                emp1['f']='---'
                emp1 = emp1[["s.no","from","to","totaldays","leavesearned","leavesstart","start_date","end_date","days","leavesend","Date of Payment and fixed",'a','b','c','d','e','f']]
            else:
                logging.info(emp1)
                emp1.columns = ['Leaves']
                emp1['emp'] = (employee_attendance).iloc[i]['Employee Code']
                emp1['Leavesdays'] = emp1.Leaves.str[5:7].astype(int)
                emp1['daysdiff'] = (emp1.Leavesdays.shift(-1) - emp1.Leavesdays).fillna(0).astype(int)
                emp1['startdate'] = np.where(emp1.daysdiff.shift() != 1, emp1.Leavesdays, 0)
                emp1['enddate'] = np.where(emp1.daysdiff!=1, emp1.Leavesdays, 0)
                emp1.drop(emp1[(emp1.startdate==0) & (emp1.enddate==0)].index, inplace=True)
                emp1['startdate'] = np.where(emp1.startdate ==0, emp1.startdate.shift(), emp1.startdate).astype(int)
                emp1['enddate'] = np.where(emp1.enddate ==0, emp1.enddate.shift(-1), emp1.enddate).astype(int)
                emp1 = emp1[['emp','startdate','enddate']]
                emp1.drop_duplicates(subset='startdate', inplace=True)
                emp1['days'] = emp1.enddate -emp1.startdate +1
                emp1['start_date'] = [datetime.date(year,month_num,x) for x in emp1.startdate]
                emp1['end_date'] = [datetime.date(year,month_num,x) for x in emp1.enddate]
                emp1.index = np.arange(1, len(emp1) + 1)
                emp1['s.no'] = emp1.index
                emp1.reset_index(drop=True, inplace=True)
                emp1['from'] = datetime.date(year,month_num,1)
                emp1['to'] = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
                emp1['totaldays'] = (employee_attendance).iloc[i]['Days Paid']
                emp1['leavesearned'] = float((employee_attendance).iloc[i]['Monthly Increment'])
                emp1['totalleaves']= float(employee_attendance.iloc[i]['Opening'])
                emp1['leavesend'] = float(employee_attendance.iloc[i]['Closing'])
                emp1['leavesstart'] =emp1['totalleaves']
                emp1['Date of Payment and fixed'] = str(employee_attendance.iloc[i]['Date of payment'])+' and '+str(employee_attendance.iloc[i]['FIXED MONTHLY GROSS'])
                emp1['a']='---'
                emp1['b']='---'
                emp1['c']='---'
                emp1['d']='---'
                emp1['e']='---'
                emp1['f']='---'
                emp1 = emp1[["s.no","from","to","totaldays","leavesearned","leavesstart","start_date","end_date","days","leavesend","Date of Payment and fixed",'a','b','c','d','e','f']]
                
            
            return emp1

        def prepare_emp_sheet(emp1,sheet_key,key,name,fathername):
            
            sheet1 = formHfile.copy_worksheet(sheetformh)
            sheet1.title = sheet_key
            lastline = sheet1['B18'].value
            sheet1['B18'] =''

            if len(emp1)>3:
                lastlinerow = 'B'+str(18+len(emp1))
            else:
                lastlinerow = 'B18'

            
            logging.info(lastlinerow)
            sheet1[lastlinerow] = lastline

            
            
            rows = dataframe_to_rows(emp1, index=False, header=False)

            for r_idx, row in enumerate(rows, 14):
                for c_idx, value in enumerate(row, 2):
                    sheet1.cell(row=r_idx, column=c_idx, value=value)
                    sheet1.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                    sheet1.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    sheet1.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            sheet1['H5']=key
            sheet1['F7']=name
            sheet1['F8']=fathername

            sheet1.sheet_properties.pageSetUpPr.fitToPage = True


        emp_count = len(data_formH.index)
        emp_dic = dict()
        for i in range(0,emp_count):
            key = (data_formH).iloc[i]['Employee Code']
            emp_dic[key] = attandance_data(data_formH,i)
            sheet_key = form+'_'+str(key)
            name= data_formH[data_formH['Employee Code']==key]['Employee Name'].values[0]
            fathername= data_formH[data_formH['Employee Code']==key]["Father's Name"].values[0]
            logging.info(name)
            logging.info(fathername)
            prepare_emp_sheet(emp_dic[key],sheet_key,key,name,fathername)
            logging.info(key)
            logging.info(sheet_key)
        if form=='FORM H':
            formHfinalfile = os.path.join(filelocation,'FormH.xlsx')
        if form=='FORM F':
            formHfinalfile = os.path.join(filelocation,'FormF.xlsx')
        
        formHfile.remove(sheetformh)
        formHfile.save(filename=formHfinalfile)

    
    def create_muster():

        musterfilepath = os.path.join(karnatakafilespath,'Muster.xlsx')
        musterfile = load_workbook(filename=musterfilepath)
        logging.info('muster file has sheet: '+str(musterfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_muster = data.drop_duplicates(subset=['Employee Code']).copy()
        data_muster.fillna(value=0, inplace=True)

        data_muster['S.no'] = list(range(1,len(data_muster)+1))

        data_muster['datelast'] ='---'

        first3columns = ["S.no",'Employee Code','Employee Name']
        last2columns = ["datelast","Days Paid"]

        columnstotake =[]
        days = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31']
        for day in days:
            for col in data_muster.columns:
                if col[5:7]==day:
                    columnstotake.append(col)
        if len(columnstotake)==28:

            columnstotake.append('29')
            columnstotake.append('30')
            columnstotake.append('31')
            data_muster['29'] = ''
            data_muster['30'] = ''
            data_muster['31'] = ''
            
        elif len(columnstotake)==29:
            columnstotake.append('30')
            columnstotake.append('31')
            data_muster['30'] = ''
            data_muster['31'] = ''

        elif len(columnstotake)==30:
            columnstotake.append('31')
            data_muster['31'] = ''

        muster_columns = first3columns+columnstotake+last2columns

        muster_data = data_muster[muster_columns]

        mustersheet = musterfile['Muster']

        mustersheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for muster is ready')

        
        rows = dataframe_to_rows(muster_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 18):
            for c_idx, value in enumerate(row, 2):
                mustersheet.cell(row=r_idx, column=c_idx, value=value)
                mustersheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                mustersheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                mustersheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        logging.info('')

        contractline = mustersheet['B10'].value
        B10_data = contractline+' '+contractor_name+', '+contractor_address
        mustersheet['B10'] = B10_data

        if str(data_muster['Nature of work'].dtype)[0:3] != 'obj':
            data_muster['Nature of work'] = data_muster['Nature of work'].astype(str)

        if str(data_muster['Location'].dtype)[0:3] != 'obj':
            data_muster['Location'] = data_muster['Location'].astype(str)

        if str(data_muster['Company Name'].dtype)[0:3] != 'obj':
            data_muster['Company Name'] = data_muster['Company Name'].astype(str)

        if str(data_muster['Company Address'].dtype)[0:3] != 'obj':
            data_muster['Company Address'] = data_muster['Company Address'].astype(str)

        if str(data_muster['Unit'].dtype)[0:3] != 'obj':
            data_muster['Unit'] = data_muster['Unit'].astype(str)

        if str(data_muster['Address'].dtype)[0:3] != 'obj':
            data_muster['Address'] = data_muster['Address'].astype(str)

        locationline = mustersheet['B11'].value
        B11_data = locationline+' '+data_muster['Nature of work'][0]+', '+data_muster['Location'][0]
        mustersheet['B11'] = B11_data

        establine = mustersheet['B12'].value
        if data_muster['PE_or_contract'][0]== 'PE':
            B12_data = establine+' '+data_muster['Company Name'][0]+', '+data_muster['Company Address'][0]
        else:
            B12_data = establine+' '+data_muster['Unit'][0]+', '+data_muster['Address'][0]
        mustersheet['B12'] = B12_data

        peline = mustersheet['B13'].value
        if data_muster['PE_or_contract'][0]== 'PE':
            B13_data = peline+' '+data_muster['Company Name'][0]+', '+data_muster['Company Address'][0]
        else:
            B13_data = peline+' '+data_muster['Unit'][0]+', '+data_muster['Address'][0]
        mustersheet['B13'] = B13_data

        mustersheet['B4'] = 'Combined Muster Roll-cum-Register of Wages in lieu of '+month+' '+str(year)

        musterfinalfile = os.path.join(filelocation,'Muster.xlsx')
        musterfile.save(filename=musterfinalfile)

    def create_formXIX():

        formXIXfilepath = os.path.join(karnatakafilespath,'FormXIX.xlsx')
        formXIXfile = load_workbook(filename=formXIXfilepath)
        logging.info('Form XIX file has sheet: '+str(formXIXfile.sheetnames))
        sheetformXIX = formXIXfile['FORM XIX']

        
        logging.info('create columns which are now available')

        data_formXIX = data.drop_duplicates(subset=['Employee Code']).copy()

        data_formXIX.fillna(value=0, inplace=True)

        emp_count = len(data_formXIX.index)
        
        for i in range(0,emp_count):
            key = (data_formXIX).iloc[i]['Employee Code']
            sheet_key = 'FORM XIX_'+str(key)

            emp_data = (data_formXIX).iloc[i]

            sheet1 = formXIXfile.copy_worksheet(sheetformXIX)
            sheet1.title = sheet_key
            sheet1['D7'] = contractor_name+', '+contractor_address
            sheet1['D8'] = emp_data['Nature of work']+', '+emp_data['Location']
            if emp_data['PE_or_contract'][0]== 'PE':
                sheet1['D9'] = emp_data['Company Name']+', '+emp_data['Company Address']
                sheet1['D10'] = emp_data['Company Name']+', '+emp_data['Company Address']
            else:
                sheet1['D9'] = emp_data['Unit']+', '+emp_data['Address']
                sheet1['D10'] = emp_data['Unit']+', '+emp_data['Address']
            sheet1['D11'] = emp_data['Employee Name']
            sheet1['D12'] = emp_data['Gender']
            sheet1['D13'] = month+'-'+str(year)
            sheet1['D14'] = key
            sheet1['D15'] = emp_data['Days Paid']
            sheet1['D16'] = emp_data['Earned Basic']
            sheet1['D17'] = emp_data['DA']
            sheet1['D18'] = emp_data['HRA']
            sheet1['D19'] = emp_data['Tel and Int Reimb']
            sheet1['D20'] = emp_data['Bonus']
            sheet1['D21'] = emp_data['Fuel Reimb']
            sheet1['D22'] = emp_data['Corp Attire Reimb']
            sheet1['D23'] = emp_data['CCA']
            sheet1['D24'] = emp_data['Conveyance']+emp_data['Medical Allowance']+emp_data['Telephone Reimb']+emp_data['Other Allowance']+emp_data['Prof Dev Reimb']+emp_data['Meal Allowance']+emp_data['Special Allowance']+emp_data['Personal Allowance']+emp_data['Other Reimb']+emp_data['Arrears']+emp_data['Variable Pay']+emp_data['Other Earning']+emp_data['Leave Encashment']+emp_data['Stipend']
            sheet1['D25'] = emp_data['Total Earning']
            sheet1['D26'] = emp_data['Insurance']
            sheet1['D27'] = emp_data['PF']
            sheet1['D28'] = emp_data['ESIC']
            sheet1['D29'] = emp_data['P.Tax']
            sheet1['D30'] = emp_data['TDS']
            sheet1['D31'] = emp_data['CSR']+emp_data['VPF']+emp_data['LWF EE']+emp_data['Salary Advance']+emp_data['Loan Deduction']+emp_data['Loan Interest']+emp_data['Other Deduction']
            sheet1['D32'] = emp_data['Total Deductions']
            sheet1['D33'] = emp_data['Net Paid']

        formXIXfinalfile = os.path.join(filelocation,'FormXIX.xlsx')
        formXIXfile.remove(sheetformXIX)
        formXIXfile.save(filename=formXIXfinalfile)

    def create_ecard():

        ecardfilepath = os.path.join(karnatakafilespath,'Employment card.xlsx')
        ecardfile = load_workbook(filename=ecardfilepath)
        logging.info('Employment card file has sheet: '+str(ecardfile.sheetnames))
        sheetecard = ecardfile['Employment card']

        
        logging.info('create columns which are now available')

        data_ecard = data.drop_duplicates(subset=['Employee Code']).copy()
        data_ecard.fillna(value=0, inplace=True)

        emp_count = len(data_ecard.index)
        
        for i in range(0,emp_count):
            key = (data_ecard).iloc[i]['Employee Code']
            sheet_key = 'Employment card_'+str(key)

            emp_data = (data_ecard).iloc[i]
            emp_data.fillna(value='', inplace=True)

            sheet1 = ecardfile.copy_worksheet(sheetecard)
            sheet1.title = sheet_key
            sheet1['B4'] = contractor_name
            sheet1['B5'] = str(emp_data['Contractor_LIN'])+' / '+str(emp_data['Contractor_PAN'])
            sheet1['B6'] = emp_data['Contractor_email']
            sheet1['B7'] = emp_data['Contractor_mobile']
            sheet1['B7'].number_format= numbers.FORMAT_NUMBER
            sheet1['B8'] = emp_data['Nature of work']
            sheet1['B9'] = contractor_address
            if emp_data['PE_or_contract'][0]== 'PE':
                sheet1['B10'] = emp_data['Company Name']
            else:
                sheet1['B10'] = emp_data['Unit']
            sheet1['B11'] = str(emp_data['Unit_LIN'])+' / '+str(emp_data['Unit_PAN'])
            sheet1['B12'] = emp_data['Unit_email']
            sheet1['B13'] = emp_data['Unit_mobile']
            sheet1['B13'].number_format= numbers.FORMAT_NUMBER
            sheet1['B14'] = emp_data['Employee Name']
            sheet1['B15'] = emp_data['Aadhar Number']
            sheet1['B15'].number_format= numbers.FORMAT_NUMBER
            sheet1['B16'] = emp_data['Mobile Tel No.']
            sheet1['B16'].number_format= numbers.FORMAT_NUMBER
            sheet1['B17'] = key
            sheet1['B18'] = emp_data['Designation']
            sheet1['B19'] = emp_data['FIXED MONTHLY GROSS']
            sheet1['B20'] = emp_data['Date Joined']
            sheet1['B21'] = '-'
            

        ecardfinalfile = os.path.join(filelocation,'Employment card.xlsx')
        ecardfile.remove(sheetecard)
        ecardfile.save(filename=ecardfinalfile)
            

    create_form_A()
    return
    create_form_B()
    create_form_XXI()
    create_form_XXII()
    create_form_XXIII()
    create_form_XX()
    create_wages()
    create_form_H_F('FORM H')
    create_form_H_F('FORM F')
    create_muster()
    create_formXIX()
    create_ecard()

def Central_Process(data,filelocation,month,year):
    Centralfilespath = os.path.join(Statefolder,'Central')
    logging.info('Central files path is :'+str(Centralfilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]
    def Form_C():
        formCfilepath = os.path.join(Centralfilespath,'Form C Format of register of loan.xlsx')
        formCfile = load_workbook(filename=formCfilepath)
        logging.info('Form C file has sheet: '+str(formCfile.sheetnames))
        logging.info('create columns which are now available')

        data_formC = data.copy(deep=True)
        data_formC=data_formC.drop_duplicates(subset="Employee Name", keep="last")
        columns=['Employee Code',"Employee Name","Recovery_Type","Particulars","Date of payment","amount","whether_show_cause_issue","explaination_heard_in_presence_of",
                                    "num_installments","first_month_year","last_month_year","Date_of_complete_recovery","remarks"]
        

        Recovery_Type_columns_name=['Other Deduction','OtherDeduction1', 'OtherDeduction2',
                                                        'OtherDeduction3', 'OtherDeduction4', 'OtherDeduction5','Damage or Loss','Fine','Salary Advance']

        data_formC["Recovery_Type"]=data_formC.loc[:,Recovery_Type_columns_name].sum(axis=1)
        data_formC["amount"]=data_formC["Recovery_Type"]
        data_formC[["Particulars","whether_show_cause_issue","explaination_heard_in_presence_of",
                    "num_installments","first_month_year","last_month_year","Date_of_complete_recovery"]]="---"
        
        formC_data=data_formC[columns]
        formCsheet = formCfile['Sheet1']
        formCsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formC_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 9):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formCsheet.cell(row=r_idx, column=c_idx, value=value)
                formCsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formCsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formCsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formI):
                    formCsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formCsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formCsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formCsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formI):
                    formCsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formCsheet.row_dimensions[r_idx].height = 20
                else:
                    formCsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formCsheet.row_dimensions[r_idx].height = 20

        formCsheet['A4']=formCsheet['A4'].value+" : "+str(data_formC['UnitName'].unique()[0])
        formCfinalfile = os.path.join(filelocation,'Form C Format of register of loan.xlsx')
        formCfile.save(filename=formCfinalfile)
    
    def Form_I():
        formIfilepath = os.path.join(Centralfilespath,'Form I register of Fine.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formI = data.copy(deep=True)
        data_formI=data_formI.drop_duplicates(subset="Employee Name", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","name&date_of_offence","cause_against_fine","FIXED MONTHLY GROSS",
                                        "Date of payment_fine_released","Date of payment_fine_imposed","remarks"]

        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        data_formI[["name&date_of_offence","cause_against_fine"]]="---"
        
        data_formI['Fine']=data_formI['Fine'].astype(float)
        data_formI["Date of payment_fine_released"]=data_formI['Date of payment']
        data_formI["Date of payment_fine_imposed"]=data_formI['Date of payment']
        # data_formI.loc[data_formI['Fine']==0,["FIXED MONTHLY GROSS","Date of payment_fine_released","Date of payment_fine_imposed"]]="---"
        data_formI.loc[:,["FIXED MONTHLY GROSS","Date of payment_fine_released","Date of payment_fine_imposed"]]="---"
        
        data_formI["remarks"]=""

        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 8):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formI):
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIsheet.row_dimensions[r_idx].height = 20
                else:
                    formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIsheet.row_dimensions[r_idx].height = 20

        formIsheet['A4']=formIsheet['A4'].value+" : "+str(data_formXII['UnitName'].unique()[0])
        formIfinalfile = os.path.join(filelocation,'Form I register of Fine.xlsx')
        formIfile.save(filename=formIfinalfile)

    def Form_II_reg_damage_loss():
        formIIfilepath = os.path.join(Centralfilespath,'Form II Register of deductions for damage or loss.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy(deep=True)
        data_formII=data_formII.drop_duplicates(subset="Employee Name", keep="last")

        data_formII.fillna(value=0, inplace=True)
        #print(sorted(data_formII.columns))
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","Damage or Loss","whether_work_showed_cause",
                                        "Date of payment & amount of deduction","num_instalments","Date of payment","remarks"]
        
        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII[["Damage or Loss","whether_work_showed_cause","Date of payment & amount of deduction","num_instalments","Date of payment","remarks"]]="---"
        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formII_data, index=False, header=False)
        border_sides_thick = Side(style='thick')       
        border_sides_thin = Side(style='thin')
        
        logging.info('rows taken out from data')
        for r_idx, row in enumerate(rows, 9):
            for c_idx, value in enumerate(row, 1):
                formIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                if len(row)==c_idx and int(row[0])==len(data_formII):
                       formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                elif len(row)==c_idx:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                elif int(row[0])==len(data_formII):
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                else:
                    formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                #border_sides = Side(style='thin')
                #formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        
        formIIsheet['A5']=formIIsheet['A5'].value+str(data_formC['UnitName'].unique()[0])
        formIIsheet['A6']="PERIOD "+str(month)+" "+str(year)
        formIIfinalfile = os.path.join(filelocation,'Form II Register of deductions for damage or loss.xlsx')
        formIIfile.save(filename=formIIfinalfile)

    def Form_IV():
        formIVfilepath = os.path.join(Delhifilespath,'Form IV Overtime register.xlsx')
        formIVfile = load_workbook(filename=formIVfilepath)
        logging.info('Form IV file has sheet: '+str(formIVfile.sheetnames))
        logging.info('create columns which are now available')

        data_formIV = data.copy(deep=True)
        data_formIV=data_formIV.drop_duplicates(subset="Employee Name", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Gender","Designation_Dept","Date_overtime_worked",
                                        "Extent of over-time","Total over-time","Normal hrs ",
                                        "FIXED MONTHLY GROSS","overtime rate",
                                        "normal_earning","Overtime",'Total Earning',"date_overtime_paid"]
        
        # data_formIV['Total\r\nOT Hrs']=data_formIV[['Total\r\nOT Hrs',"Overtime",'Total Earning']].astype(float)
        # data_formIV["Total over-time"]=data_formIV['Total\r\nOT Hrs']
        # data_formIV["normal_earning"]=data_formIV['FIXED MONTHLY GROSS']-data_formIV["Overtime"]
        # data_formIV.loc[data_formIV['Total\r\nOT Hrs']==0,["Total over-time","Normal hrs ",
        #                                 "FIXED MONTHLY GROSS","overtime rate",
        #                                 "normal_earning","Overtime",'Total Earning']]="---"

        # data_formIV["date_overtime_paid"]=data_formIV['Date of payment']
        # data_formIV.loc[data_formIV["Overtime"]==0,"date_overtime_paid"]="---"
        # data_formIV.loc[data_formIV['Total\r\nOT Hrs']==0,"date_overtime_paid"]="---"
        # data_formIV["Extent of over-time"]="-----"
        # data_formIV["Date_overtime_worked"]="-----"
        
        data_formIV["Date_overtime_worked","Extent of over-time","Total over-time","Normal hrs ",
                    "FIXED MONTHLY GROSS","overtime rate","normal_earning","Overtime",'Total Earning',"date_overtime_paid"]="---"

        data_formIV['S.no'] = list(range(1,len(data_formIV)+1))
        data_formIV['Designation_Dept']=data_formIV["Designation"]+"_"+data_formIV["Department"]
        # data_formIV["Date of payment & amount of deduction"]=data_formIV['Date of payment']+"\n"+data_formIV["Total Deductions"]
        formIV_data=data_formIV[columns]
        formIVsheet = formIVfile['Sheet1']
        formIVsheet.sheet_properties.pageSetUpPr.fitToPage = True
        # for column in  range(ord('A'), ord('O') + 1):
        #     formIVsheet.unmerge_cells(chr(column)+"7:"+chr(column)+"14")

        logging.info('data for form IV is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formIV_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 10):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formIVsheet.cell(row=r_idx, column=c_idx, value=value)
                formIVsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIVsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formIV):
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formIVsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formIVsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formIV):
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formIVsheet.row_dimensions[r_idx].height = 20
                else:
                    formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formIVsheet.row_dimensions[r_idx].height = 20

        # formIVsheet['A4']=formIVsheet['A4'].value+"  "+data_formIV['Company Name'].unique()[0]+"  "+data_formIV['Company Address'].unique()[0]+"                                Month Ending: "+month+" "+str(year)
        # formIVsheet.cell(row=4, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
        formIVsheet['A4']="Month Ending: "+month+" "+str(year)
        formIVfinalfile = os.path.join(filelocation,'Form IV Overtime register.xlsx')
        formIVfile.save(filename=formIVfinalfile)

    
Stateslist = ['Karnataka','Maharashtra','Delhi','Telangana','Uttar Pradesh','Tamilnadu','Goa','Gujarat','Kerala','Madhya Pradesh','Rajasthan','Haryana',
'West Bengal','Uttarakhand']

State_Process = {'karnataka':Karnataka,'maharashtra':Maharashtra,'delhi':Delhi,'telangana':Telangana,'uttar pradesh':Uttar_Pradesh,'tamilnadu':Tamilnadu,'goa':Goa,
                'gujarat':Gujarat,'kerala':Kerala,'madhya pradesh':Madhya_Pradesh,'rajasthan':Rajasthan,'haryana':Haryana,
                'west bengal':West_Bengal,'uttarakhand':Uttarakhand}

companylist = ['SVR LTD','PRY Wine Ltd','CDE Technology Ltd']

def Type5(inputfolder,month,year):
    logging.info('type5 data process running')

def Type4(inputfolder,month,year):
    logging.info('type4 data process running')

def Type3(inputfolder,month,year):
    logging.info('type3 data process running')

def Type2(inputfolder,month,year):
    logging.info('type2 data process running')

    



def Type1(inputfolder,month,year):
    global nomatch
    nomatch=''
    logging.info('type1 data process running')

    emp_df_columns = ['Employee Code_master', 'Employee Name_master', 'Company Name','Company Address', 'Grade', 'Branch_master',
       'Department', 'Designation_master', 'Division', 'Group', 'Category', 'Unit',
       'Location Code', 'State', 'Date of Birth', 'Date Joined_master',
       'Date of Confirmation', 'Date Left', 'Title', 'Last Inc. Date',
       'Ticket Number', 'Local Address 1', 'Local Address 2',
       'Local Address 3', 'Local Address 4', 'Local City Name',
       'Local District Name', 'Local PinCode', 'Local State Name',
       'Residence Tel No.', 'Permanent Address 1', 'Permanent Address 2',
       'Permanent Address 3', 'Permanent Address 4', 'UAN Number_master',
       'Permanent Tel No.', 'Office Tel No.', 'Extension Tel No.',
       'Mobile Tel No.', "Father's Name", 'Gender', 'Age', 'Number of Months',
       'Marital Status', 'PT Number', 'PF Number (Old Version)', 'PF Number',
       'PF Number (WithComPrefix)', 'PAN Number', 'ESIC Number (Old Version)',
       'ESIC Number_master', 'ESIC Number (CompPrefix)', 'FPF Number', 'PF Flag',
       'ESIC Flag', 'PT Flag', 'Bank A/c Number_master', 'Bank Name', 'Mode',
       'Account Code_master', 'E-Mail_master', 'Remarks_master', 'PF Remarks', 'ESIC Remarks',
       'ESIC IMP Code', 'ESIC IMP Name', 'Employee Type (For PF)',
       'Freeze Account', 'Freeze Date', 'Freeze Reason', 'Type of House (In)',
       'Comp. adn.', 'Staying In (Metro Type)', 'Children (For CED)',
       'TDS Rate', 'Resignation Date', 'Reason for Leaving', 'Bank A/C No.1',
       'Bank A/C No.2', 'Bank A/C No.3', 'Alt.Email', 'Emp Status',
       'Probation Date', 'Surcharge Flag', 'Gratuity Code',
       'Resign Offer Date', 'Permanent City', 'Permanent District',
       'Permanent Pin Code', 'Permanent State', 'Spouse Name',
       'PF Joining Date', 'PRAN Number', 'Group Joining Date', 'Aadhar Number',
       'Child in Hostel (For CED)', 'Total Exp in Years', 'P', 'L',
       'Identification mark','Nationality',	'Education Level',	'Category Address',
       'Type of Employment',	'Service Book No',	'Nature of work']

    salary_df_columns = ['Sr', 'DivisionName', 'Sal Status', 'Emp Code_salary', 'Emp Name_salary', 'Designation_salary',
       'Date Joined', 'UnitName', 'Branch_salary', 'Days Paid', 'Earned Basic','DA', 'HRA',
       'Conveyance', 'Medical Allowance', 'Telephone Reimb',
       'Tel and Int Reimb', 'Bonus', 'Other Allowance', 'Fuel Reimb',
       'Prof Dev Reimb', 'Corp Attire Reimb', 'Meal Allowance',
       'Special Allowance', 'Personal Allowance','Overtime', 'CCA', 'Other Reimb',
       'Arrears', 'Other Earning', 'Variable Pay', 'Leave Encashment',
       'Stipend', 'Consultancy Fees', 'OtherAllowance1', 'OtherAllowance2', 'OtherAllowance3', 'OtherAllowance4', 'OtherAllowance5'
       'Total Earning', 'Insurance', 'CSR',
       'PF', 'ESIC','VPF', 'P.Tax', 'LWF EE', 'Salary Advance', 'Loan Deduction',
       'Loan Interest', 'Fine',	'Damage or Loss','Other Deduction', 'TDS', 'OtherDeduction1', 'OtherDeduction2', 'OtherDeduction3', 'OtherDeduction4', 'OtherDeduction5'
       'Total Deductions','Net Paid', 'BankName', 'Bank A/c Number_salary', 'Account Code_salary', 'Remarks_salary_salary',
       'PF Number (Old)', 'UAN Number_salary', 'ESIC Number_salary', 'Personal A/c Number',
       'E-Mail_salary', 'Mobile No.', 'FIXED MONTHLY GROSS', 'CHECK CTC Gross','Date of payment',	'Arrears salary', 'Cheque No - NEFT date']

    atten_df_columns = ['Emp Code', 'Employee Name', 'Branch', 'Designation', 'Sat\r\n01/02',
       'Sun\r\n02/02', 'Mon\r\n03/02', 'Tue\r\n04/02', 'Wed\r\n05/02',
       'Thu\r\n06/02', 'Fri\r\n07/02', 'Sat\r\n08/02', 'Sun\r\n09/02',
       'Mon\r\n10/02', 'Tue\r\n11/02', 'Wed\r\n12/02', 'Thu\r\n13/02',
       'Fri\r\n14/02', 'Sat\r\n15/02', 'Sun\r\n16/02', 'Mon\r\n17/02',
       'Tue\r\n18/02', 'Wed\r\n19/02', 'Thu\r\n20/02', 'Fri\r\n21/02',
       'Sat\r\n22/02', 'Sun\r\n23/02', 'Mon\r\n24/02', 'Tue\r\n25/02',
       'Wed\r\n26/02', 'Thu\r\n27/02', 'Fri\r\n28/02', 'Sat\r\n29/02',
       'Total\r\nDP', 'Total\r\nABS', 'Total\r\nLWP', 'Total\r\nCL',
       'Total\r\nSL', 'Total\r\nPL', 'Total\r\nL1', 'Total\r\nL2',
       'Total\r\nL3', 'Total\r\nL4', 'Total\r\nL5', 'Total\r\nCO-',
       'Total\r\nCO+', 'Total\r\nOL', 'Total\r\nWO', 'Total\r\nPH',
       'Total\r\nEO', 'Total\r\nWOP', 'Total\r\nPHP', 'Total\r\nOT Hrs',
       'Total\r\nLT Hrs']

    leave_df_columns = ['Emp. Code', 'Emp. Name', 'Leave Type', 'Opening', 'Monthly Increment',
       'Used', 'Closing', 'Leave Accrued', 'Encash']

    leftemp_df_columns = ['Employee Name', 'Employee Code', 'Date Joined', 'Date Left',
       'UAN Number']

    unit_df_columns = ['Unit', 'Location_code','Location', 'Address', 'Registration_no','Unit_PAN','Unit_LIN','Unit_email','Unit_mobile', 'PE_or_contract',
       'State_or_Central', 'start_time', 'end_time', 'rest_interval','Contractor_name','Contractor_Address','Contractor_PAN', 'Contractor_LIN', 'Contractor_email',	'Contractor_mobile','Normal hrs', 'overtime rate']

    logging.info('column variables set')

    

    
    file_list = os.listdir(inputfolder)
    logging.info('input folder is '+str(inputfolder))
    for f in file_list:
        if f[0:6].upper()=='MASTER':
            masterfilename = f
            logging.info('masterfilename is :'+f)
        if f[0:6].upper()=='SALARY':
            salaryfilename = f
            logging.info('salaryfilename is :'+f)
        if f[0:10].upper()=='ATTENDANCE':
            attendancefilename = f
            logging.info('attendancefilename is :'+f)
        if f[0:5].upper()=='LEAVE':
            leavefilename = f
            logging.info('leavefilename is :'+f)
        if f[0:14].upper()=='LEFT EMPLOYEES':
            leftempfilename = f
            logging.info('leftempfilename is :'+f)
        if f[0:5].upper()=='UNITS':
            unitfilename = f
            logging.info('unitfilename is :'+f)
    
    logging.info('file names set')
    
    if 'masterfilename' in locals():
        masterfile = os.path.join(inputfolder,masterfilename)
        employee_data = pd.read_excel(masterfile)
        employee_data.dropna(subset=['Employee Code','Location Code'], inplace=True)
        employee_data.dropna(how='all', inplace=True)
        employee_data.reset_index(drop=True, inplace=True)
        employee_data.rename(columns={"Employee Code": "Employee Code_master", "Employee Name": "Employee Name_master", "Designation": "Designation_master", "Branch": "Branch_master", "Date Joined": "Date Joined_master", "UAN Number": "UAN Number_master",
                           "ESIC Number": "ESIC Number_master", "Bank A/c Number": "Bank A/c Number_master", "Account Code": "Account Code_master",
                           "E-Mail": "E-Mail_master", "Remarks": "Remarks_master"}, inplace=True)
        logging.info('employee data loaded')
    else:
        employee_data = pd.DataFrame(columns = emp_df_columns)
        logging.error('employee data not available setting empty dataset')
    if 'salaryfilename' in locals():
        salaryfile = os.path.join(inputfolder,salaryfilename)
        salary_data = pd.read_excel(salaryfile)
        salary_data.dropna(subset=['Emp Code'], inplace=True)
        salary_data.dropna(how='all', inplace=True)
        salary_data.reset_index(drop=True, inplace=True)
        salary_data.rename(columns={"Emp Code": "Emp Code_salary", "Emp Name": "Emp Name_salary","DesigName": "Designation_salary", "Branch": "Branch_salary", "Date Joined": "Date Joined_salary", "UAN Number": "UAN Number_salary",
                           "ESIC Number": "ESIC Number_salary", "Bank A/c Number": "Bank A/c Number_salary", "Account Code": "Account Code_salary",
                           "E-Mail": "E-Mail_salary", "Remarks": "Remarks_salary"}, inplace=True)
        logging.info('salary data loaded')
    else:
        salary_data = pd.DataFrame(columns = salary_df_columns)
        logging.info('salary data not available setting empty dataset')
    if 'attendancefilename' in locals():
        attendancefile = os.path.join(inputfolder,attendancefilename)
        attendance_data = pd.read_excel(attendancefile)
        attendance_data.dropna(subset=['Emp Code'], inplace=True)
        attendance_data.dropna(how='all', inplace=True)
        attendance_data.reset_index(drop=True, inplace=True)
        logging.info('attendance data loaded')
    else:
        attendance_data = pd.DataFrame(columns = atten_df_columns)
        logging.info('attendance data not available setting empty dataset')
    if 'leavefilename' in locals():
        leavefile = os.path.join(inputfolder,leavefilename)
        leave_data = pd.read_excel(leavefile)
        leave_data.dropna(subset=['Emp. Code'], inplace=True)
        leave_data.dropna(how='all', inplace=True)
        leave_data.reset_index(drop=True, inplace=True)
        logging.info('leave data loaded')
    else:
        leave_data = pd.DataFrame(columns = leave_df_columns)
        logging.info('leave data not available setting empty dataset')
    if 'leftempfilename' in locals():
        leftempfile = os.path.join(inputfolder,leftempfilename)
        leftemp_data = pd.read_excel(leftempfile)
        leftemp_data.dropna(subset=['Employee Code'], inplace=True)
        leftemp_data.dropna(how='all', inplace=True)
        leftemp_data.reset_index(drop=True, inplace=True)
        leftemp_data.rename(columns={"Employee Code": "Employee Code_left"}, inplace=True)
        logging.info('left employees data loaded')
    else:
        leftemp_data = pd.DataFrame(columns = leftemp_df_columns)
        logging.info('left employees data not available setting empty dataset')
    if 'unitfilename' in locals():
        unitfile = os.path.join(inputfolder,unitfilename)
        unit_data = pd.read_excel(unitfile)
        unit_data.dropna(subset=['Location Code'], inplace=True)
        unit_data.dropna(how='all', inplace=True)
        unit_data.reset_index(drop=True, inplace=True)
        logging.info('unit data loaded')
    else:
        unit_data = pd.DataFrame(columns = unit_df_columns)
        logging.info('unit data not available setting empty dataset')

    employee_data.drop(columns='Date Left', inplace=True)

    logging.info(type(employee_data['Location Code'][0]))
    logging.info(unit_data.head())

    if str(employee_data['Location Code'].dtype)[0:3] != 'int':
        employee_data['Location Code'] = employee_data['Location Code'].astype(int)
    
    if str(employee_data['Employee Code_master'].dtype)[0:3] != 'obj':
        employee_data['Employee Code_master'] = employee_data['Employee Code_master'].astype(str)

    if str(unit_data['Location Code'].dtype)[0:3] != 'int':
        unit_data['Location Code'] = unit_data['Location Code'].astype(int)


    employee_data.drop(columns=list(employee_data.columns.intersection(salary_data.columns)), inplace=True)

    if str(salary_data['Emp Code_salary'].dtype)[0:3] != 'obj':
        salary_data['Emp Code_salary'] = salary_data['Emp Code_salary'].astype(str)

    attendance_data.drop(columns=['Employee Name', 'Branch', 'Designation'], inplace=True)

    if str(attendance_data['Emp Code'].dtype)[0:3] != 'obj':
        attendance_data['Emp Code'] = attendance_data['Emp Code'].astype(str)

    if str(leave_data['Emp. Code'].dtype)[0:3] != 'obj':
        leave_data['Emp. Code'] = leave_data['Emp. Code'].astype(str)

    leftemp_data.drop(columns=['Employee Name', 'Date Joined', 'UAN Number'],inplace=True)

    if str(leftemp_data['Employee Code_left'].dtype)[0:3] != 'obj':
        leftemp_data['Employee Code_left'] = leftemp_data['Employee Code_left'].astype(str)
    

    CDE_Data = salary_data.merge(employee_data,how='left',left_on='Emp Code_salary', right_on='Employee Code_master').merge(
        unit_data,how='left',on='Location Code').merge(
            attendance_data,how='left',left_on='Emp Code_salary', right_on='Emp Code').merge(
                leave_data, how='left', left_on='Emp Code_salary', right_on='Emp. Code').merge(
                    leftemp_data, how='left', left_on='Emp Code_salary', right_on='Employee Code_left')
    
    '''
    CDE_Data = employee_data.merge(unit_data,how='left',on='Location Code').merge(
        salary_data,how='left',left_on='Employee Code',right_on='Emp Code').merge(
            attendance_data,how='left',left_on='Employee Code', right_on='Emp Code').merge(
                leave_data, how='left', left_on='Employee Code', right_on='Emp. Code').merge(
                    leftemp_data, how='left', on='Employee Code')
    '''
    

    CDE_Data['Employee Code'] = CDE_Data['Emp Code_salary']
    CDE_Data['Employee Name'] = CDE_Data['Emp Name_salary'].combine_first(CDE_Data['Employee Name_master'])
    CDE_Data['Designation'] = CDE_Data['Designation_salary'].combine_first(CDE_Data['Designation_master'])
    CDE_Data['Branch'] = CDE_Data['Branch_salary'].combine_first(CDE_Data['Branch_master'])
    CDE_Data['Date Joined'] = CDE_Data['Date Joined_salary'].combine_first(CDE_Data['Date Joined_master'])
    CDE_Data['UAN Number'] = CDE_Data['UAN Number_salary'].combine_first(CDE_Data['UAN Number_master'])
    CDE_Data['ESIC Number'] = CDE_Data['ESIC Number_salary'].combine_first(CDE_Data['ESIC Number_master'])
    CDE_Data['Bank A/c Number'] = CDE_Data['Bank A/c Number_salary'].combine_first(CDE_Data['Bank A/c Number_master'])
    CDE_Data['Account Code'] = CDE_Data['Account Code_salary'].combine_first(CDE_Data['Account Code_master'])
    CDE_Data['E-Mail'] = CDE_Data['E-Mail_salary'].combine_first(CDE_Data['E-Mail_master'])
    CDE_Data['Remarks'] = CDE_Data['Remarks_salary'].combine_first(CDE_Data['Remarks_master'])

    
    logging.info('merged all data sets')

    logging.info(len(salary_data))
    logging.info(len(CDE_Data))



    rename_list=[]
    renamed=[]
    drop_list=[]
    for x in list(CDE_Data.columns):
        if x[-2:]=='_x':
            rename_list.append(x)
            renamed.append(x[0:-2])
        if x[-2:]=='_y':
            drop_list.append(x)
    
    rename_dict = dict(zip(rename_list,renamed))

    CDE_Data.rename(columns=rename_dict, inplace=True)

    logging.info('columns renamed correctly')

    CDE_Data.drop(columns=drop_list, inplace=True)

    logging.info('dropped duplicate columns')

    print(CDE_Data['Date of payment'].dtype)

    if str(CDE_Data['Date of payment'].dtype)[0:8] == 'datetime':
        CDE_Data['Date of payment'] = CDE_Data['Date of payment'].dt.date

    monthyear = month+' '+str(year)
    if monthyear.upper() in masterfilename.upper():
        logging.info('month year matches with data')
        #for all state employees(PE+contractor)
        statedata = CDE_Data[CDE_Data['State_or_Central']=='State'].copy()
        # statedata.State='Gujarat'
        CDE_States = list(statedata['State'].unique())
        implemented_state_list=[x.lower() for x in State_Process.keys()]
        for state in CDE_States:
            state=state.lower()
            if state not in implemented_state_list:
                logging.info('State {} not implemented in our set,that is {} hence continuing'.format(state,implemented_state_list))
                print('State {} not implemented in our set,that is {} hence continuing'.format(state,implemented_state_list))
                continue

            unit_with_location = list((statedata[statedata.State==state]['Unit']+';'+statedata[statedata.State==state]['Location']).unique())
            for UL in unit_with_location:
                inputdata = statedata[(statedata['State']==state) & (statedata['Unit']==UL.split(';')[0]) & (statedata['Location']==UL.split(';')[1])].copy()
                inputdata['Contractor_name'] = inputdata['Contractor_name'].fillna(value='')
                inputdata['Contractor_Address'] = inputdata['Contractor_Address'].fillna(value='')
                if UL.strip()[-1] == '.':
                    ULis = UL.strip()[0:-1]
                else:
                    ULis = UL.strip()
                inpath = os.path.join(inputfolder,'Registers','States',state,ULis)
                logging.info('folder for forms path is'+str(inpath))
                if os.path.exists(inpath):
                    logging.info('running state process')
                    logging.info(inputdata)
                    contractor_name= inputdata['Contractor_name'].unique()[0]
                    contractor_address= inputdata['Contractor_Address'].unique()[0]
                    State_Process[state](data=inputdata,contractor_name=contractor_name,contractor_address=contractor_address,filelocation=inpath,month=month,year=year)
                else:
                    logging.info('making directory')
                    os.makedirs(inpath)
                    logging.info('directory created')
                    logging.info(inputdata)
                    contractor_name= inputdata['Contractor_name'].unique()[0]
                    contractor_address= inputdata['Contractor_Address'].unique()[0]
                    State_Process[state](data=inputdata,contractor_name=contractor_name,contractor_address=contractor_address,filelocation=inpath,month=month,year=year)

        #for contractors form
        contractdata = CDE_Data[(CDE_Data['State_or_Central']=='State') & (CDE_Data['PE_or_contract']=='Contract')].copy()
        contractor_units = list((contractdata['Unit']+';'+contractdata['Location']).unique())
        for UL in contractor_units:
            inputdata = contractdata[(contractdata['Unit']==UL.split(';')[0]) & (contractdata['Location']==UL.split(';')[1])]
            if UL.strip()[-1] == '.':
                ULis = UL.strip()[0:-1]
            else:
                ULis = UL.strip()
            inpath = os.path.join(inputfolder,'Registers','Contractors',ULis)
            if os.path.exists(inpath):
                logging.info('running contractor process')
            else:
                logging.info('making directory')
                os.makedirs(inpath)
                logging.info('directory created')
            if not inputdata.empty:
                Contractor_Process(data=inputdata,filelocation=inpath,month=month,year=year)
            

        #for central form
        centraldata = CDE_Data[CDE_Data['State_or_Central']=='Central'].copy()
        central_units = list((centraldata['Unit']+','+centraldata['Location']).unique())
        for UL in central_units:
            inputdata = centraldata[(centraldata['Unit']==UL.split(',')[0]) & (centraldata['Location']==UL.split(',')[1])]
            inpath = os.path.join(inputfolder,'Registers','Central',UL)
            if os.path.exists(inpath):
                logging.info('running contractor process')
            else:
                logging.info('making directory')
                os.makedirs(inpath)
                logging.info('directory created')
    
    else:
        nomatch = "Date you mentioned doesn't match with Input data"
        logging.error(nomatch)

    

DataProcess = {'Type1':Type1,'Type2':Type2,'Type3':Type3,'Type4':Type4,'Type5':Type5}


def CompanyDataProcessing(companytype,inputfolder,month,year):
    inputfolder = Path(inputfolder)
    yr = int(year)
    DataProcess[companytype](inputfolder,month,yr)

#backend code ends here

Types = ['Type1','Type2','Type3','Type4','Type5']


Months = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']

Years = ['2017','2018','2019','2020']

Typeis = tk.StringVar()

companyname = tk.StringVar()

month = tk.StringVar()
year = tk.StringVar()


folderLabel = ttk.LabelFrame(master, text="Select the Company")
folderLabel.grid(column=0,row=1,padx=20,pady=20)

TypeLabel = Label(folderLabel,text="Company Type")
TypeLabel.grid(column=1,row=0,padx=20,pady=20)

TypeEntry = ttk.Combobox(folderLabel,values=Types,textvariable=Typeis)
TypeEntry.grid(column=2, row=0, padx=20,pady=20)

companynameLabel = Label(folderLabel, text="Company Name")
companynameLabel.grid(column=1, row=1, padx=20,pady=20)

comapnynameEntry = tk.Entry(folderLabel,textvariable=companyname)
comapnynameEntry.grid(column=2, row=1, padx=20,pady=20)

MonthLabel = Label(folderLabel, text="Month and Year")
MonthLabel.grid(column=1, row=3, padx=20,pady=20)

MonthEntry = ttk.Combobox(folderLabel,values=Months,textvariable=month)
MonthEntry.grid(column=2, row=3, padx=20,pady=20)

YearEntry = ttk.Combobox(folderLabel,values=Years,textvariable=year)
YearEntry.grid(column=3, row=3, padx=20,pady=20)

def disfo():
    foldername = filedialog.askdirectory()
    logging.info(foldername)
    logging.info(type(foldername))
    foldernamelabel.configure(text=foldername)


button = ttk.Button(folderLabel, text = "Select Company Folder", command=disfo)
button.grid(column=1, row=2, columnspan=2,padx=20, pady=20)

foldernamelabel = Label(folderLabel, text="")
foldernamelabel.grid(column=1, row=4, columnspan=2,padx=20,pady=20)




def generateforms(comptype,mn,yr):
    companytype=comptype.get()

    month = mn.get()
    year = yr.get()


    getfolder = foldernamelabel.cget("text")



    logging.info(type(companytype))
    logging.info(companytype)

    logging.info(type(getfolder))
    logging.info(getfolder)

    if (companytype =="" and getfolder =="" and (month =="" or year =="")):
        report.configure(text="Please select month year, company folder and company type")
    elif (companytype=="" and getfolder =="" and not(month =="" or year =="")):
        report.configure(text="Please select company folder and company type")
    elif (companytype=="" and getfolder !="" and not(month =="" or year =="")):
        report.configure(text="Please select company type")
    elif (companytype!="" and getfolder =="" and not(month =="" or year =="")):
        report.configure(text="Please select company folder")
    elif (companytype =="" and getfolder !="" and (month =="" or year=="")):
        report.configure(text="Please select month year and company type")
    elif (companytype!="" and getfolder=="" and (month =="" or year=="")):
        report.configure(text="Please select month year and company folder")
    elif (companytype!="" and getfolder!="" and (month =="" or year=="")):
        report.configure(text="Please select month year")
    else:
        logging.info(companytype, getfolder,  month,  year)
        report.configure(text="Processing")
        try:
            CompanyDataProcessing(companytype,getfolder,month,year)
        except Exception as e:
            logging.info('Failed')
            report.configure('Failed')
        else:
            if nomatch=='':
                logging.info('Completed Form Creation')
                report.configure(text='Completed Form Creation')
            else:
                logging.info(nomatch)
                report.configure(text=nomatch)
        finally:
            logging.info('done')
        
def convert_forms_to_pdf():

    getfolder = foldernamelabel.cget("text")

    if getfolder=="":
        report.configure(text="Please select company folder")
    else:
        registerfolder = os.path.join(Path(getfolder),'Registers')
        if os.path.exists(registerfolder):
            for root, dirs, files in os.walk(registerfolder):
                for fileis in files:
                    if fileis.endswith(".xlsx"):
                        try:
                            create_pdf(root,fileis)
                        except Exception as e:
                            logging.info('Failed pdf Conversion')
                            report.configure(text="Failed")
                        else:
                            logging.info('Completed pdf Conversion')
                            report.configure(text="Completed")
                        finally:
                            logging.info('done')
        else:
            report.configure(text="Registers not available")
                        



generateforms = partial(generateforms,Typeis,month,year)

button = ttk.Button(master, text = "Generate Forms", command=generateforms)
button.grid(column=1, row=1, columnspan=2,padx=20, pady=20)

Detailbox = ttk.LabelFrame(master, text="")
Detailbox.grid(column=0,row=2,padx=20,pady=20)

report = Label(Detailbox, text="                                                            ")
report.grid(column=0, row=0, padx=20,pady=20)



button2 = ttk.Button(master, text = "Convert forms to PDF", command=convert_forms_to_pdf)
button2.grid(column=0, row=3, columnspan=2,padx=20, pady=20)


mainloop()


