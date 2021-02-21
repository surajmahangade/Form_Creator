from states import logging,monthdict,Statefolder
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill, numbers

def Hyderabad(data,contractor_name,contractor_address,filelocation,month,year,report,master):
    
    logging.info("Hyderabad form creation")
    Hyderabadfilespath = os.path.join(Statefolder,'Hyderabad')
    logging.info('Hyderabad files path is :'+str(Hyderabadfilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]
    #wages reg
    # input_filelocation=filelocation.split("Registers")[0]
    # min_wages_goa=read_min_wages_file("GOA","SEMI-SKILLED",input_filelocation)
    
    def Form_X():
        formXfilepath = os.path.join(Hyderabadfilespath,'Form X Register of fine.xlsx')
        formXfile = load_workbook(filename=formXfilepath)
        logging.info('Form I file has sheet: '+str(formXfile.sheetnames))
        logging.info('create columns which are now available')

        data_formX = data.copy(deep=True)
        data_formX=data_formX.drop_duplicates(subset="Employee Code", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Act of Commission",
                                        "cause_against_fine","Total wages fine imposed",
                                        "amount of fine","Date of payment_fine_imposed","Date of payment_fine_realized","remarks"]

        data_formX['S.no'] = list(range(1,len(data_formX)+1))
        # data_formX[["name&date_of_offence","cause_against_fine","remarks"]]="---"
        
        # data_formX['Fine']=data_formX['Fine'].astype(float)
        # data_formX["Date of payment_fine_released"]=data_formX['Date of payment']
        # data_formX["Date of payment_fine_imposed"]=data_formX['Date of payment']
        # data_formX.loc[data_formX['Fine']==0,["FIXED MONTHLY GROSS","Date of payment_fine_released","Date of payment_fine_imposed","remarks"]]="---"
        data_formX[["Act of Commission","cause_against_fine","Total wages fine imposed","amount of fine",
                        "Date of payment_fine_imposed","Date of payment_fine_realized","remarks"]]=""
        formX_data=data_formX[columns]
        formXsheet = formXfile['Sheet1']
        formXsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form X is ready')

        
        rows = dataframe_to_rows(formX_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 8):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formXsheet.cell(row=r_idx, column=c_idx, value=value)
                formXsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formXsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formX):
                    formXsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formXsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formXsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formXsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formX):
                    formXsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formXsheet.row_dimensions[r_idx].height = 20
                else:
                    formXsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formXsheet.row_dimensions[r_idx].height = 20

        formXsheet['A5']=formXsheet['A5'].value+" : "+str(data_formX['UnitName'].unique()[0])
        formXfinalfile = os.path.join(filelocation,'Form X Register of fine.xlsx')
        formXfile.save(filename=formXfinalfile)

    def Form_XI():
        formXIfilepath = os.path.join(Hyderabadfilespath,'Form XI register of damage or loss.xlsx')
        formXIfile = load_workbook(filename=formXIfilepath)
        logging.info('Form XI file has sheet: '+str(formXIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formXI = data.copy(deep=True)
        data_formXI=data_formXI.drop_duplicates(subset="Employee Code", keep="last")
        columns=['S.no',"Employee Name","Father's Name","Damage_loss_cause","whether_work_showed_cause",
                                        "amount of deduction","date_deduction_imposed","num_instalments","Date on which total amount realised","remarks"]
        
        data_formXI['S.no'] = list(range(1,len(data_formXI)+1))
        # data_formXI["attendancefile"]="---"
        # data_formXI[["whether_work_showed_cause","num_instalments"]]="-----"
        # data_formXI["remarks"]=""
        # ######################################
        # data_formXI['Date of payment']=data_formXI['Date of payment'].apply(lambda x: x.strftime('%d-%m-%Y'))
        # data_formXI['Damage or Loss']=data_formXI['Damage or Loss'].astype(float)
        # data_formXI['Damage or Loss']=data_formXI['Damage or Loss'].fillna(0)
        # data_formXI["Damage_loss_with_date"]=data_formXI['Date of payment']+" & "+data_formII['Damage or Loss'].astype(str)
        # data_formXI.loc[data_formXI["Damage or Loss"]==0,"Damage_loss_with_date"]="---"

        # data_formXI["Date of payment & amount of deduction"]=data_formXI["Damage_loss_with_date"]
        # data_formXI["num_instalments"]="1"
        # data_formXI["Date on which total amount realised"]=data_formXI['Date of payment']
        # data_formXI.loc[data_formXI["Damage or Loss"]==0,["Date on which total amount realised","num_instalments"]]="---"
        ###################################
        data_formXI[["Damage_loss_cause","whether_work_showed_cause",
                        "amount of deduction","date_deduction_imposed","num_instalments",
                        "Date on which total amount realised","remarks"]]=""
        formXI_data=data_formXI[columns]
        formXIsheet = formXIfile['Sheet1']
        formXIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form XI is ready')

        
        rows = dataframe_to_rows(formXI_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 9):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formXIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formXIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formXI):
                    formXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formXIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formXIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formXI):
                    formXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formXIsheet.row_dimensions[r_idx].height = 20
                else:
                    formXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formXIsheet.row_dimensions[r_idx].height = 20

        formXIsheet['A5']=formXIsheet['A5'].value+"  :  "+str(data_formXI['UnitName'].unique()[0])
        formXIfinalfile = os.path.join(filelocation,'Form XI register of damage or loss.xlsx')
        formXIfile.save(filename=formXIfinalfile)

    def Form_XII_reg_advance():
        formXIIfilepath = os.path.join(Hyderabadfilespath,'Form XII Register of advance.xlsx')
        formXIIfile = load_workbook(filename=formXIIfilepath)
        logging.info('Form XII file has sheet: '+str(formXIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formXII = data.copy(deep=True)
        data_formXII=data_formXII.drop_duplicates(subset="Employee Code", keep="last")

        data_formXII.fillna(value=0, inplace=True)
        columns=['S.no',"Employee Name","Father's Name","Salary Advance","date_advance_given","purpose_advance",
                                        "num_installments_advance","Postponement_granted",
                                        "Date repaid","remarks"]
                                        
                                        
        data_formXII['S.no'] = list(range(1,len(data_formXII)+1))
        data_formXII["Salary Advance"]=data_formXII["Salary Advance"].astype(str)
        
        data_formXII[["date_advance_given","purpose_advance","num_installments_advance","Postponement_granted",
                                        "Date repaid","remarks"]]=""

        formXII_data=data_formXII[columns]
        formXIIsheet = formXIIfile['Sheet1']
        formXIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        
        logging.info('data for form XII is ready')

        
        rows = dataframe_to_rows(formXII_data, index=False, header=False)

        logging.info('rows taken out from data')
        border_sides_thick = Side(style='thick')       
        border_sides_thin = Side(style='thin')
        for r_idx, row in enumerate(rows, 7):
            for c_idx, value in enumerate(row, 1):
                formXIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formXIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                if len(row)==c_idx and int(row[0])==len(data_formXII):
                       formXIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                       formXIIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formXIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formXIIsheet.row_dimensions[r_idx].height = 20
                elif int(row[0])==len(data_formXII):
                    formXIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formXIIsheet.row_dimensions[r_idx].height = 20
                else:
                    formXIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formXIIsheet.row_dimensions[r_idx].height = 20
                #border_sides = Side(style='thin')
                #formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        #formIVsheet['A4']=formIVsheet['A4'].value+" : "+data_formIV['Unit'].unique()[0]
        formXIIsheet['A4']=formXIIsheet['A4'].value+str(data_formXII['UnitName'].unique()[0])
        # formXIIsheet['A7']="PERIOD "+str(month)+" "+str(year)

        formXIIfinalfile = os.path.join(filelocation,'Form XII Register of advance.xlsx')
        formXIIfile.save(filename=formXIIfinalfile)
    
    def Form_XXII_reg_employee():
        formXXIIfilepath = os.path.join(Hyderabadfilespath,'Form XXII register of employees.xlsx')
        formXXIIfile = load_workbook(filename=formXXIIfilepath)
        logging.info('Form XXII file has sheet: '+str(formXXIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formXXII = data.copy(deep=True)
        data_formXXII=data_formXXII.drop_duplicates(subset="Employee Code", keep="last")
        
        columns=['S.no','Employee Code',"Employee Name",'Gender','Date of Birth',"start_time","end_time",
                "rest_interval"]

        # data_formXXII_columns=list(data_formXXII.columns)
        # start=data_formXXII_columns.index('Emp Code')
        # end=data_formXXII_columns.index('Total\r\nDP')
        columnstotake =[]
        days = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31']
        for day in days:
            for col in data_formXXII.columns:
                if col[5:7]==day:
                    columnstotake.append(col)
        if len(columnstotake)==28:
            columnstotake.append('29')
            columnstotake.append('30')
            columnstotake.append('31')
            data_formXXII['29'] = ''
            data_formXXII['30'] = ''
            data_formXXII['31'] = ''
            
        elif len(columnstotake)==29:
            columnstotake.append('30')
            columnstotake.append('31')
            data_formXXII['30'] = ''
            data_formXXII['31'] = ''

        elif len(columnstotake)==30:
            columnstotake.append('31')
            data_formXXII['31'] = ''
        
        columns.extend(columnstotake)
        
        columns.extend(["Date","from","to","extent","remarks"])

        data_formXXII[["Date","from","to","extent","remarks"]]=""

        data_formXXII['S.no'] = list(range(1,len(data_formXXII)+1))

        formXXII_data=data_formXXII[columns]
        formXXIIsheet = formXXIIfile['Sheet1']
        formXXIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form XXII is ready')

        
        rows = dataframe_to_rows(formXXII_data, index=False, header=False)

        logging.info('rows taken out from data')
        row_num=0
        for r_idx, row in enumerate(rows, 11):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                formXXIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formXXII):
                    formXXIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formXXIIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formXXIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formXXIIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formXXII):
                    formXXIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formXXIIsheet.row_dimensions[r_idx].height = 20
                else:
                    formXXIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formXXIIsheet.row_dimensions[r_idx].height = 20

        
        #formXXIIsheet['AE4']=formXXIIsheet['AE4'].value+"   "+str(data_formP['Registration_no'].unique()[0])
        
        formXXIIsheet['A4']=formXXIIsheet['A4'].value+str(data_formXXII['UnitName'].unique()[0])
        
        # formXXIIsheet['A6']="From:  01"+"-"+str(month)+"-"+str(year)+"       "+"From:  01"+"-"+str(month)+"-"+str(year)
        formXXIIsheet['A6']=formXXIIsheet['A6'].value+str(data_formXXII["Address"].unique()[0])
        formXXIIsheet['Z4']=formXXIIsheet['Z4'].value+str(month)+" "+str(year)
        formXXIIsheet['Z6']=formXXIIsheet['Z6'].value+str(data_formXXII['Registration_no'].unique()[0])
        
        formXXIIfinalfile = os.path.join(filelocation,'Form XXII register of employees.xlsx')
        formXXIIfile.save(filename=formXXIIfinalfile)

    def Form_XXIII():
        formXXIIIfilepath = os.path.join(Hyderabadfilespath,'Form XXIII Register of wages.xlsx')
        formXXIIIfile = load_workbook(filename=formXXIIIfilepath)
        logging.info('Form XXIII file has sheet: '+str(formXXIIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formXXIII = data.copy(deep=True)
        data_formXXIII=data_formXXIII.drop_duplicates(subset="Employee Code", keep="last")
        columns=["Employee Name","Date Joined","FIXED MONTHLY GROSS",'Total Earning',"Overtime","basic_and_allo",
                'HRA','Bonus','Fuel Reimb','Prof Dev Reimb', 'Corp Attire Reimb','Telephone Reimb',
                'CCA','Other Earning','Leave Encashment',"Overtime",'Total Earning','Insurance',
                "CSR",'PF','ESIC','P.Tax',"LWF EE",'Salary Advance',"Loan Deduction","Loan Interest",
                "Fine","Damage or Loss","all_Other_deductions", 'TDS',"all_Other_deductions",'Total Deductions',
                'Net Paid','Date of payment',"sign"]
        
        data_formXXIII['DA']= data_formXXIII['DA'].replace("",0).astype(float)
        data_formXXIII['Earned Basic']=data_formXXIII['Earned Basic'].replace("",0).astype(float)
        data_formXXIII['basic_and_allo'] = data_formXXIII['Earned Basic']+ data_formXXIII['DA']
        all_Other_deductions_columns=['Other Deduction','OtherDeduction1', 'OtherDeduction2','OtherDeduction3', 
                                        'OtherDeduction4', 'OtherDeduction5']
        
        data_formXXIII[all_Other_deductions_columns]=data_formXXIII[all_Other_deductions_columns].replace("",0).astype(float)
        data_formXXIII[all_Other_deductions_columns]=data_formXXIII[all_Other_deductions_columns].fillna(0)

        data_formXXIII["all_Other_deductions"]=data_formXXIII.loc[:,all_Other_deductions_columns].sum(axis=1)

        data_formXXIII[["sign"]]=""
        

        # data_formXXIII['S.no'] = list(range(1,len(data_formXXIII)+1))

        formXXIII_data=data_formXXIII[columns]
        formXXIIIsheet = formXXIIIfile['Sheet1']
        formXXIIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form XXIII is ready')

        
        rows = dataframe_to_rows(formXXIII_data, index=False, header=False)
        rows_copy = list(dataframe_to_rows(formXXIII_data, index=False, header=False))
        logging.info('rows taken out from data')
        
        row_num=0
        for r_idx, row in enumerate(rows, 10):
            row_num+=1
            for c_idx, value in enumerate(row, 1):
                #formXXIIIsheet.cell(row=r_idx, column=c_idx).value=value
                formXXIIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                border_sides_thick = Side(style='thick')       
                border_sides_thin = Side(style='thin')
                if len(row)==c_idx and row_num==len(data_formXXIII):
                    formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thick)
                    formXXIIIsheet.row_dimensions[r_idx].height = 20
                elif len(row)==c_idx:
                    formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thick, bottom=border_sides_thin) 
                    formXXIIIsheet.row_dimensions[r_idx].height = 20
                elif row_num==len(data_formXXIII):
                    formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thick)
                    formXXIIIsheet.row_dimensions[r_idx].height = 20
                else:
                    formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides_thin, bottom=border_sides_thin)
                    formXXIIIsheet.row_dimensions[r_idx].height = 20
        

        
        # if data["PE_or_contract"].unique()[0].upper()=="PE":
        formXXIIIsheet['A5']=formXXIIIsheet['A5'].value+" "+str(data_formXXIII['UnitName'].unique()[0])
        formXXIIIsheet['A6']=formXXIIIsheet['A6'].value+" "+str(data_formXXIII['Address'].unique()[0])
        formXXIIIsheet['A7']=formXXIIIsheet['A7'].value+"   "+str(data_formXXIII['Registration_no'].unique()[0])
        formXXIIIsheet['AE5']=formXXIIIsheet['AE5'].value+" "+str(month)+" "+str(year)
        monthstart = datetime.date(year,month_num,1)
        monthend = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
            
        formXXIIIsheet['AE6']="From : "+str(monthstart)+"  To : "+str(monthend)
        

        
        # else:
        #     formXXIIIsheet['A4']=" Name of Establishment:-   "+str(data_formXXIII['UnitName'].unique()[0])+" "+str(data_formXXIII['Address'].unique()[0])
        #     formXXIIIsheet['A5']="Name of Employer and address:-   "+str(data_formXXIII['Contractor_name'].unique()[0])+","+str(data_formXXIII['Contractor_Address'].unique()[0])
        
        
        formXXIIIfinalfile = os.path.join(filelocation,'Form XXIII Register of wages.xlsx')
        formXXIIIfile.save(filename=formXXIIIfinalfile)

    def From_XXV():
        formXXVfilepath = os.path.join(Hyderabadfilespath,'Form XXV Register of leave.xlsx')
        formXXVfile = load_workbook(filename=formXXVfilepath)
        logging.info('Form XXV file has sheet: '+str(formXXVfile.sheetnames))
        
        logging.info('create columns which are now available')

        data_formXXV = data.copy(deep=True)
        leave_file_data=data_formXXV[["Employee Code","Employee Name","Leave Type","Opening","Monthly Increment","Leave Accrued","Used","Encash","Closing"]]
        
        data_formXXV=data_formXXV.drop_duplicates(subset="Employee Code", keep="last")

        data_formXXV["Employee Name & Code"]=data_formXXV["Employee Name"].astype(str)+"||"+data_formXXV["Employee Code"].astype(str)

        columns=["Employee Name & Code","Registration_no","Father's Name","Date Joined"]
        data_formXXV_columns=list(data_formXXV.columns)
        start=data_formXXV_columns.index('Arrears salary')
        end=data_formXXV_columns.index('Total\r\nDP')
        columns.extend(data_formXXV_columns[start+1:end])

        formXXV_data=data_formXXV[columns]
        formXXVsheet = formXXVfile['Sheet1']

        formXXVsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XXV is ready')

        
        #rows_copy = list(dataframe_to_rows(formXII_data, index=False, header=False))
        def cell_write(sheet,r_idx,c_idx,value):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            
        def PL_write(row_index,target,start,end,is_abs_num,name):
            cell_write(target,row_index,1,str(month)+" "+str(year))
            cell_write(target,row_index,2,"")
            cell_write(target,row_index,3,start)
            cell_write(target,row_index,4,end)
            cell_write(target,row_index,5,is_abs_num)
            
            cell_write(target,row_index,6,"")
            
            cell_write(target,row_index,7,start)
            cell_write(target,row_index,8,end)
            cell_write(target,row_index,9,is_abs_num)

            emp_details=leave_file_data.loc[leave_file_data["Employee Name"]==name,:]
            
            closing=emp_details["Closing"].loc[emp_details["Leave Type"]=="PL"]
            # if not closing.empty:
            cell_write(target,row_index,7,closing.to_string(index=False))
            # else:
            #     cell_write(target,row_index,7,emp_details["Closing"])

            cell_write(target,row_index,10,"")
            cell_write(target,row_index,11,"")
            cell_write(target,row_index,12,"")
            cell_write(target,row_index,13,"")
            cell_write(target,row_index,14,"")
            cell_write(target,row_index,15,"")
            cell_write(target,row_index,16,"")


        form_write={'PL':PL_write}
        
        def start_end_date_attendance(rows,absent_label,row_offset,initial_offset):  
           # print("infunction",row_offset)
            is_abs_num=0
            row_index=0
            added={}
            for sheet_idx, row in enumerate(rows, 10):
                row_index=0
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        name=value.split("||")[0]
                        code=value.split("||")[1]
                        if code =="nan":
                            code=name
                        try:
                            target=formXXVfile[code]
                            added[target.title]=0
                        except:
                            target = formXXVfile.copy_worksheet(formXXVsheet)
                            target.title=code
                            #initial offset
                            row_offset[code]=initial_offset
                            added[target.title]=0
                        target['A4']="Name and address of the Establishment:- "+" "+str(data_formXXV['UnitName'].unique()[0])+", "+str(data_formXXV['Address'].unique()[0])
                        
                        # if data_formXII['PE_or_contract'].unique()[0]=="Contractor":
                        #     target["A5"]="Name of Employer and address:-  "+str(data_formXII['UnitName'].unique()[0])
                        # else:
                        #     target["A5"]="Name of Employer and address:- ---"
                        target['A6']="Name of Employee : "+str(name)
                        added[target.title]=0
                    elif c_idx==2:
                        target['A5']="Registration No. :- "+str(value)
                    elif c_idx==3:
                        target['A7']="Father's Name : "+str(value)
                    elif c_idx==4:
                        target['A8']="Date of appointment:- "+str(value)

                    elif is_abs_num==0 and value==absent_label:
                        is_abs_num=1
                        start=columns[c_idx-1]
                        end=columns[c_idx-1]
                    elif value==absent_label:
                        is_abs_num+=1
                        end=columns[c_idx-1]
                    elif is_abs_num:
                        #target.cell(row=row_index+13, column=1+column_offset, value=is_abs_num)
                     #   print("write",row_index,row_offset,row_index+row_offset[target.title])
                        start_date=start.split("\n")[1].replace("/","-")+"-"+str(year)
                        end_date=end.split("\n")[1].replace("/","-")+"-"+str(year)
                        form_write[absent_label](row_index+row_offset[target.title],target,start_date,end_date,is_abs_num,name)
                        target.insert_rows(row_index+row_offset[target.title]+1)
                        is_abs_num=0
                        row_index+=1
                        added[target.title]+=1

            return added
        offset={}
        initial_offset=12
        start_end_date_attendance(dataframe_to_rows(formXXV_data, index=False, header=False),"PL",offset,initial_offset)
        formXXVfile.remove(formXXVfile["Sheet1"])
        formXXVfile.remove(formXXVfile["Sheet2"])
        formXXVfile.remove(formXXVfile["Sheet3"])
        formXXVfinalfile = os.path.join(filelocation,'Form XXV Register of leave.xlsx')
        formXXVfile.save(filename=formXXVfinalfile)
    try:
        Form_X()
        master.update()
        Form_XI()
        master.update()
        Form_XII_reg_advance()
        master.update()
        Form_XXII_reg_employee()
        master.update()
        Form_XXIII()
        master.update()
        From_XXV()
        master.update()
    except KeyError as e:
        logging.info("Key error : Check if {} column exsists".format(e))
        print("Key error {}".format(e))
        report.configure(text="Failed: Check input file format  \n column {} not found".format(e))
        master.update()
        raise KeyError
    except FileNotFoundError as e:
        logging.info("File not found : Check if {} exsists".format(e))
        report.configure(text="Failed: File {} not found".format(e))
        master.update()
        raise FileNotFoundError